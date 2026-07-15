"""Read .xlsx/.csv/.tsv files into raw cell grids, tolerating bank realities.

The reader only gets values out of the file; recovering table structure is
``sheet_analyzer``'s job. Formulas are read as their last cached value when
the workbook has one (files saved by Excel do); a formula with no cached
value comes through as ``None`` and is surfaced downstream rather than
silently computed.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from app.adapters.excel_csv.sheet_analyzer import MergedRange

SUPPORTED_SUFFIXES = (".xlsx", ".csv", ".tsv")


class WorkbookReadError(ValueError):
    pass


@dataclass(frozen=True)
class SheetGrid:
    name: str
    grid: list[list[Any]]
    merged_ranges: list[MergedRange]


def read_source(path: Path) -> list[SheetGrid]:
    if not path.exists():
        raise WorkbookReadError(f"Source file does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix in {".csv", ".tsv"}:
        return [_read_delimited(path, delimiter="\t" if suffix == ".tsv" else ",")]
    if suffix == ".xls":
        raise WorkbookReadError(
            "Legacy .xls workbooks are not supported; save the file as .xlsx and retry."
        )
    raise WorkbookReadError(
        f"Unsupported file type {suffix!r}; supported: {', '.join(SUPPORTED_SUFFIXES)}."
    )


def _read_xlsx(path: Path) -> list[SheetGrid]:
    # Imported lazily: openpyxl is only needed when an .xlsx workbook is read.
    import openpyxl  # noqa: PLC0415

    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        message = str(exc)
        if "encrypted" in message.lower() or "password" in message.lower():
            raise WorkbookReadError(
                f"Workbook {path.name} is password-protected; provide an unprotected copy."
            ) from exc
        raise WorkbookReadError(f"Cannot open workbook {path.name}: {message}") from exc

    sheets: list[SheetGrid] = []
    try:
        for worksheet in workbook.worksheets:
            grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
            merged = [
                (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
                for rng in worksheet.merged_cells.ranges
            ]
            sheets.append(SheetGrid(name=worksheet.title, grid=grid, merged_ranges=merged))
    finally:
        workbook.close()
    return sheets


def _read_delimited(path: Path, *, delimiter: str) -> SheetGrid:
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            grid: list[list[Any]] = [list(row) for row in csv.reader(handle, delimiter=delimiter)]
    except UnicodeDecodeError as exc:
        raise WorkbookReadError(f"File {path.name} is not valid UTF-8 text.") from exc
    return SheetGrid(name=path.stem, grid=grid, merged_ranges=[])
