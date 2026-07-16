"""Manual upload templates (market_data_adapter.md §8.2).

One downloadable template per scope category. Each template ships:

- a header row with the exact column names the parser expects,
- a single example row showing the format (its first cell carries the
  comment marker so the parser ignores it if the operator leaves it in),
- a legend block explaining every column with its validation rules
  (ISO 4217 currency codes, ISO 8601 dates, allowed enum values).

Human-friendliness rule: yields in the yield-curve template are entered as
PERCENT (``15.80`` meaning 15.80%); the parser converts them to the decimal
fractions the canonical model stores (``0.1580``). FX rates and macro values
are absolute levels and pass through unchanged.

Templates are generated programmatically so headers can never drift from the
parser: both sides read :data:`TEMPLATE_HEADERS`.
"""

from __future__ import annotations

import io
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

type TemplateKind = Literal["yield_curve", "fx_rates", "credit_ratings", "macro_forecasts"]

TEMPLATE_KINDS: tuple[TemplateKind, ...] = (
    "yield_curve",
    "fx_rates",
    "credit_ratings",
    "macro_forecasts",
)

# Rows whose first cell starts with this marker are ignored by the parser:
# the example row, the legend block, and any operator notes.
COMMENT_PREFIX = "#"

TEMPLATE_HEADERS: dict[TemplateKind, tuple[str, ...]] = {
    "yield_curve": ("currency", "curve_name", "as_of_date", "tenor_months", "rate_percent"),
    "fx_rates": (
        "base_currency",
        "quote_currency",
        "rate_type",
        "tenor_months",
        "rate",
        "as_of_date",
    ),
    "credit_ratings": ("issuer", "agency", "rating", "watch_status", "rating_date"),
    "macro_forecasts": ("index_code", "value", "scenario", "horizon_months", "as_of_date"),
}

# The §8.2 example row per template. Written with the first cell comment
# prefixed so an un-deleted example never becomes canonical data.
TEMPLATE_EXAMPLE_ROWS: dict[TemplateKind, tuple[Any, ...]] = {
    "yield_curve": ("GHS", "GHS_GOV_BOND", "2026-06-30", 3, 15.80),
    "fx_rates": ("USD", "GHS", "spot", None, 12.85, "2026-06-30"),
    "credit_ratings": ("GHANA_SOVEREIGN", "fitch", "B-", "stable", "2026-06-15"),
    "macro_forecasts": ("GHANA_GDP_FORECAST", 5.80, "base", 12, "2026-06-30"),
}

_COMMON_LEGEND: tuple[str, ...] = (
    "# HOW TO USE THIS TEMPLATE",
    "# Enter one record per row directly under the header row.",
    "# Rows whose first cell starts with '#' (like the example row above) are ignored on upload.",
    "# Dates are ISO 8601 (YYYY-MM-DD). Currency codes are 3-letter ISO 4217 (GHS, USD, EUR, ...).",
)

_COLUMN_LEGENDS: dict[TemplateKind, tuple[str, ...]] = {
    "yield_curve": (
        "# currency — 3-letter ISO 4217 code of the curve. Supported: GHS, USD, EUR, GBP, "
        "NGN, KES, ZAR.",
        "# curve_name — curve identifier, e.g. GHS_GOV_BOND. Rows sharing currency and "
        "curve_name form one curve.",
        "# as_of_date — business date of the observation (ISO 8601); must match the upload "
        "as-of date.",
        "# tenor_months — positive whole number of months (1, 3, 6, 12, 24, 36, 60, 84, 120).",
        "# rate_percent — annualized yield entered as a PERCENT: 15.80 means 15.80% "
        "(stored as 0.1580).",
    ),
    "fx_rates": (
        "# base_currency / quote_currency — 3-letter ISO 4217 codes. The rate is quote "
        "currency per 1 unit of base currency: USD/GHS 12.85 means 12.85 GHS per USD.",
        "# rate_type — 'spot' or 'forward'.",
        "# tenor_months — leave blank for spot; positive whole months for forward (1, 3, 6, 12).",
        "# rate — positive decimal FX level (not a percentage).",
        "# as_of_date — business date of the observation (ISO 8601); must match the upload "
        "as-of date.",
    ),
    "credit_ratings": (
        "# issuer — GHANA_SOVEREIGN or NIGERIA_SOVEREIGN.",
        "# agency — 'moodys', 'sp', or 'fitch'.",
        "# rating — the agency's rating symbol, e.g. B-, Caa1.",
        "# watch_status — optional: 'positive', 'negative', 'stable', or 'developing'; "
        "leave blank if none.",
        "# rating_date — the agency's action date (ISO 8601); may be earlier than the upload "
        "as-of date.",
    ),
    "macro_forecasts": (
        "# index_code — GHANA_GDP_FORECAST, GHANA_INFLATION_FORECAST, or GHANA_POLICY_RATE_PATH.",
        "# value — forecast value in the indicator's natural unit, e.g. 5.80 for 5.8% GDP growth.",
        "# scenario — 'base', 'adverse', or 'severely_adverse'.",
        "# horizon_months — forecast horizon in positive whole months; leave blank for a "
        "spot observation.",
        "# as_of_date — business date of the forecast (ISO 8601); must match the upload "
        "as-of date.",
    ),
}

_MIN_COLUMN_WIDTH = 16
_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def template_filename(kind: TemplateKind) -> str:
    """The download filename per §8.2, e.g. ``yield_curve_template.xlsx``."""
    return f"{kind}_template.xlsx"


def template_media_type() -> str:
    return _XLSX_MEDIA_TYPE


def build_template(kind: TemplateKind) -> bytes:
    """Build the .xlsx template for one scope category and return its bytes."""
    headers = TEMPLATE_HEADERS[kind]
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # a fresh Workbook always has an active sheet
    sheet.title = kind

    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    example = TEMPLATE_EXAMPLE_ROWS[kind]
    sheet.append([f"{COMMENT_PREFIX} {example[0]}", *example[1:]])

    sheet.append([])
    for line in (*_COMMON_LEGEND, *_COLUMN_LEGENDS[kind]):
        sheet.append([line])

    for index, header in enumerate(headers, start=1):
        column = sheet.column_dimensions[get_column_letter(index)]
        column.width = max(len(header) + 4, _MIN_COLUMN_WIDTH)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
