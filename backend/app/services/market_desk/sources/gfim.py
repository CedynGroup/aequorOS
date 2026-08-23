"""GFIM secondary-market parsers (fixtures README §4).

Daily trading report XLSX (one workbook per trading day, ~7 sheets) ->
per-security observations:

- ``GHS.GFIM.{ISIN}.YIELD``  closing yield (pct)
- ``GHS.GFIM.{ISIN}.PRICE``  end-of-day closing price (index, per 100 face)
- ``GHS.GFIM.{ISIN}.VOLUME`` face volume traded (ghs)
- ``GHS.GFIM.{ISIN}.TRADES`` number of trades (index)

Sell/buy-back trades are repos, not outright trades, and the same ISIN can
appear in both — SBB rows get a ``GHS.GFIM.SBB.{ISIN}.*`` stem so they can
never collide with (or supersede) the outright series.

Workbook quirks encoded from the captured files: sheet names carry TRAILING
SPACES (``'OLD GOG NOTES AND BONDS '``, ``'CORPORATE  '``); header cells
contain literal newlines (``'CLOSING\\n YIELD'``); the TREASURY BILLS sheet
drags leftover junk columns (a 2021 date cell, duplicate TENOR headers) so
columns are located BY HEADER NAME within the row that contains ``ISIN``;
security rows with no trades leave VOLUME/NUMBER TRADED empty while still
publishing an indicative mark (flagged ``no_trades``, never dropped).

Monthly Status Report (text layer of the ~11.5MB PDF) -> whole-month summary
series from section B: ``GHS.GFIM.MONTHLY.{VOLUME,VALUE_GHS,TRADES}``.
"""

from __future__ import annotations

import io
import re
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

from app.services.market_desk.sources.core import (
    QF_NO_TRADES,
    ObservationDraft,
    ParseContext,
    ParseResult,
)

_SHEET_SECURITY_TYPES: tuple[tuple[str, str], ...] = (
    ("NEW GOG NOTES AND BONDS", "new_gog_bond"),
    ("DDEP BONDS", "ddep_bond"),
    ("OLD GOG NOTES AND BONDS", "old_gog_bond"),
    ("CORPORATE", "corporate"),
    ("TREASURY BILLS", "bill"),
    ("SELL BUY BACK TRADES- GOG BONDS", "sell_buy_back"),
)

_ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")
_SHEET_DATE_RE = re.compile(r"Date:\s*[A-Za-z]+,\s*(\d{1,2})\s+([A-Za-z]+),?\s+(\d{4})")


def _norm_header(cell: Any) -> str:
    """Collapse the literal newlines/spaces GFIM leaves in header cells."""
    return re.sub(r"\s+", " ", str(cell)).strip().upper() if cell is not None else ""


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "N/A"}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _sheet_as_of(rows: list[tuple[Any, ...]]) -> date | None:
    for row in rows[:6]:
        for cell in row:
            if cell is None:
                continue
            match = _SHEET_DATE_RE.search(str(cell))
            if match:
                return datetime.strptime(
                    f"{match.group(1)} {match.group(2)} {match.group(3)}", "%d %B %Y"
                ).date()
    return None


def _header_columns(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """Locate the header row (the one containing ISIN) and map header names
    to column indexes — never trust column positions on these sheets."""
    for index, row in enumerate(rows[:8]):
        headers = [_norm_header(cell) for cell in row]
        if "ISIN" not in headers:
            continue
        mapping: dict[str, int] = {}
        for column, header in enumerate(headers):
            if header and header not in mapping:  # first wins over junk dupes
                mapping[header] = column
        return index, mapping
    return None


def _pick(mapping: dict[str, int], *candidates: str) -> int | None:
    for candidate in candidates:
        if candidate in mapping:
            return mapping[candidate]
    return None


def parse_gfim_daily_xlsx(raw: bytes, *, context: ParseContext) -> ParseResult:  # noqa: PLR0912, PLR0915
    """Every per-security sheet of one daily trading report workbook."""
    result = ParseResult()
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            security_type = None
            for known, type_slug in _SHEET_SECURITY_TYPES:
                if sheet_name.strip().upper() == known:
                    security_type = type_slug
                    break
            if security_type is None:
                if sheet_name.strip().upper() != "SUMMARY":
                    result.warnings.append(f"unrecognized sheet skipped: {sheet_name!r}")
                continue
            sheet = workbook[sheet_name]
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            located = _header_columns(rows)
            if located is None:
                result.warnings.append(f"sheet {sheet_name!r}: no ISIN header row found")
                continue
            header_index, columns = located
            as_of = _sheet_as_of(rows) or context.as_of_date
            isin_col = columns["ISIN"]
            yield_col = _pick(columns, "CLOSING YIELD", "YIELD")
            price_col = _pick(
                columns,
                "END OF DAY CLOSING PRICE",
                "CLOSING PRICE",
                "WEIGHTED AVERAGE CLOSING PRICE",
            )
            volume_col = _pick(columns, "VOLUME", "VOLUME TRADED")
            trades_col = _pick(columns, "NUMBER TRADED")
            tenor_col = _pick(columns, "TENOR")
            description_col = _pick(columns, "SECURITY DESCRIPTION")
            maturity_col = _pick(columns, "MATURITY DATE")

            stem = "GHS.GFIM.SBB" if security_type == "sell_buy_back" else "GHS.GFIM"
            emitted = 0
            for row in rows[header_index + 1 :]:
                if isin_col >= len(row):
                    continue
                isin_cell = str(row[isin_col]).strip() if row[isin_col] is not None else ""
                if not _ISIN_RE.match(isin_cell):
                    continue  # TOTAL rows, section banners, blank padding
                attributes: dict[str, Any] = {"security_type": security_type}
                if tenor_col is not None and row[tenor_col] is not None:
                    attributes["tenor"] = str(row[tenor_col]).strip()
                if description_col is not None and row[description_col] is not None:
                    attributes["security_description"] = str(row[description_col]).strip()
                if maturity_col is not None and row[maturity_col] is not None:
                    attributes["maturity_date"] = str(row[maturity_col])[:10]

                volume = _decimal(row[volume_col]) if volume_col is not None else None
                trades = _decimal(row[trades_col]) if trades_col is not None else None
                no_trades = not trades or trades == 0
                cells: list[tuple[str, Decimal | None, str]] = [
                    ("YIELD", _decimal(row[yield_col]) if yield_col is not None else None, "pct"),
                    (
                        "PRICE",
                        _decimal(row[price_col]) if price_col is not None else None,
                        "index",
                    ),
                    ("VOLUME", volume, "ghs"),
                    ("TRADES", trades, "index"),
                ]
                for leg, value, unit in cells:
                    if value is None:
                        continue
                    draft = ObservationDraft(
                        series_code=f"{stem}.{isin_cell}.{leg}",
                        as_of_date=as_of,
                        value=value,
                        unit=unit,
                        attributes=dict(attributes),
                    )
                    if no_trades and leg in {"YIELD", "PRICE"}:
                        draft.flag(QF_NO_TRADES)  # indicative mark, not a traded level
                    result.observations.append(draft)
                    emitted += 1
            if emitted == 0:
                result.warnings.append(f"sheet {sheet_name!r}: no security rows emitted")
    finally:
        workbook.close()
    if not result.observations:
        result.errors.append("no observations parsed from GFIM daily workbook")
    return result


# The TOC also lists this heading (with dotted leaders); the body occurrence
# is the LAST match. '-GHS' may or may not ride on the heading itself.
_STATS_HEADING_RE = re.compile(r"B\.\s+SUMMARY\s+OF\s+TRADE\s+STATISTICS(?:-?\s*GHS)?")

_MONTH_NAMES = (
    "JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|"
    "NOVEMBER|DECEMBER"
)
# A month name directly followed by a 4-digit year. ``\s*`` (not ``\s+``)
# tolerates every layout the reconstructor produces: the spaced headings
# ("SECURITIES TRADED FOR JULY 2026"), the double-spaced ones ("JAN- JULY
#  2026") and the drop-cap title banner where the glyphs collapse
# ("JUNE2026"). The report month saturates the document — TOC, section
# headings and every table repeat it — while a stray prior-year comparison
# ("... AS AT JULY 2025") shows up once, so the resolver takes the majority
# vote rather than the first hit (verified 37:1 on the real July-2026 report).
_MONTH_YEAR_RE = re.compile(rf"\b({_MONTH_NAMES})\s*(\d{{4}})\b")
# Filename fallback: the FileBird upload name reliably carries the month
# ("…/GFIM-Status-Report-July-2026.pdf"), the last-ditch anchor when a future
# layout change shatters the in-document banner.
_URL_MONTH_RE = re.compile(rf"({_MONTH_NAMES})[-_ ]+(\d{{4}})", re.IGNORECASE)
_STAT_ROWS: tuple[tuple[str, str, str], ...] = (
    ("Volume Traded", "GHS.GFIM.MONTHLY.VOLUME", "ghs"),
    ("Value Traded (GHS)", "GHS.GFIM.MONTHLY.VALUE_GHS", "ghs"),
    ("No. of Trades", "GHS.GFIM.MONTHLY.TRADES", "index"),
)


def _status_report_text(raw: bytes) -> str:
    """Decode a GFIM monthly status artifact: PDF (live captures) or plain text (fixtures)."""
    if raw[:4] == b"%PDF":
        # Deferred on purpose: PDF extraction is an optional dependency and the
        # fixture path (plain text) must import and run without it.
        from app.services.market_desk.sources.pdf_text import page_lines  # noqa: PLC0415

        pages = page_lines(raw)
        return "\n".join(line for page in pages for line in page)
    return raw.decode("utf-8", errors="replace")


def _month_year_to_date(month_token: str, year_token: str) -> date | None:
    try:
        return datetime.strptime(f"01 {month_token.title()} {year_token}", "%d %B %Y").date()
    except ValueError:
        return None


def _resolve_report_month(text: str, *, source_url: str | None = None) -> date | None:
    """The report month, robust to case / spacing / drop-cap format variance.

    Majority vote over every ``<MONTH> <YEAR>`` in the document (the report
    month dominates; stray comparison years lose), falling back to the month
    encoded in the source filename when the text carries no banner at all.
    """
    counts: Counter[tuple[str, str]] = Counter(_MONTH_YEAR_RE.findall(text.upper()))
    if counts:
        (month_token, year_token), _ = counts.most_common(1)[0]
        resolved = _month_year_to_date(month_token, year_token)
        if resolved is not None:
            return resolved
    if source_url:
        url_match = _URL_MONTH_RE.search(source_url)
        if url_match:
            return _month_year_to_date(url_match.group(1), url_match.group(2))
    return None


def parse_gfim_status_report(raw: bytes, *, context: ParseContext) -> ParseResult:
    """Section B of the monthly status report -> whole-month GHS summary
    series (as-of = first of the report month, ``period`` in attributes).

    Only the GHS trade-statistics block is extracted; the ranking tables,
    US$ statistics, and outstanding-securities sections are recorded as
    unsupported rather than half-parsed.

    Live captures are PDFs (FileBird); fixtures ship pre-extracted ``.txt``.
    Both paths are supported.
    """
    result = ParseResult()
    text = _status_report_text(raw)

    report_month = _resolve_report_month(text, source_url=context.source_url)
    if report_month is None:
        result.errors.append(
            "report month not found: no '<MONTH> <YEAR>' banner in the document "
            "and no month in the source filename"
        )
        return result

    headings = list(_STATS_HEADING_RE.finditer(text))
    if not headings:
        result.errors.append("section 'B. SUMMARY OF TRADE STATISTICS' not found")
        return result
    heading = headings[-1]  # earlier hits are the table of contents
    block = text[heading.end() : heading.end() + 4000]

    for label, series_code, unit in _STAT_ROWS:
        row_match = re.search(
            re.escape(label) + r"\s+([\d,]+(?:\.\d+)?)", block
        )
        if row_match is None:
            result.warnings.append(f"status report row not found: {label!r}")
            continue
        value = _decimal(row_match.group(1))
        if value is None:
            result.warnings.append(f"unparseable value for {label!r}: {row_match.group(1)!r}")
            continue
        result.observations.append(
            ObservationDraft(
                series_code=series_code,
                as_of_date=report_month,
                value=value,
                unit=unit,
                attributes={"period": report_month.strftime("%Y-%m"), "row": label},
            )
        )
    result.warnings.append(
        "sections not extracted (unsupported): US$ statistics, outstanding securities, "
        "primary-market activities, member rankings, yield-curve chart"
    )
    if not result.observations:
        result.errors.append("no summary rows parsed from GFIM status report")
    return result
