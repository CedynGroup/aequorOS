"""XLSX rendering of a resolved return (docs/regulatory_reporting.md §5).

One metadata sheet, one sheet per template section, one fidelity/provenance
footer sheet. Styling is deliberately regulator-neutral: bold headers on a
light-grey fill, bold totals rows, ``#,##0;(#,##0)`` number formats (thousands
separators, parenthesised negatives per research §11) — no brand colors.

Byte output is deterministic for a given package so re-exports keep a stable
checksum: workbook document properties are pinned to the package's
``generated_at`` and the container zip is re-written with fixed entry
timestamps (openpyxl stamps entries with wall-clock time).
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.regulatory_reporting.templates import (
    ColumnSpec,
    RenderedCell,
    RenderedReturn,
    RenderedRow,
)

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))
_GHS_FORMAT = "#,##0;(#,##0)"
_PCT_FORMAT = "0.00"
_NUMBER_FORMAT = "#,##0.00;(#,##0.00)"
_ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)
_MAX_SHEET_TITLE = 31
_WORKING_COPY_BANNER = "WORKING COPY — FOR INTERNAL REVIEW · not a filing artifact"


@dataclass(frozen=True)
class _SectionLocation:
    sheet_title: str
    first_data_row: int
    row_count: int
    columns: dict[str, int]


def _sheet_title(title: str, used: set[str]) -> str:
    cleaned = "".join(ch for ch in title if ch not in "[]:*?/\\")[:_MAX_SHEET_TITLE]
    candidate = cleaned
    suffix = 2
    while candidate in used:
        stem = cleaned[: _MAX_SHEET_TITLE - len(f" ({suffix})")]
        candidate = f"{stem} ({suffix})"
        suffix += 1
    used.add(candidate)
    return candidate


def _write_cell(sheet: Worksheet, row_idx: int, col_idx: int, cell: RenderedCell) -> None:
    if cell.value is None:
        return
    if cell.kind == "bool":
        target = sheet.cell(row=row_idx, column=col_idx, value="Yes" if cell.value else "No")
        target.alignment = Alignment(horizontal="center")
        return
    if cell.kind == "text":
        sheet.cell(row=row_idx, column=col_idx, value=str(cell.value))
        return
    # Decimal — openpyxl stores it as a real number, formatting handles display.
    target = sheet.cell(row=row_idx, column=col_idx, value=cell.value)
    if cell.kind == "ghs":
        target.number_format = _GHS_FORMAT
    elif cell.kind == "pct":
        target.number_format = _PCT_FORMAT
    else:
        target.number_format = _NUMBER_FORMAT


def _write_table_row(
    sheet: Worksheet,
    row_idx: int,
    rendered_row: RenderedRow,
    columns: tuple[ColumnSpec, ...],
    *,
    bold: bool = False,
) -> None:
    """One grid row on the human-facing sheet.

    A section total carries a snapshot key as its ``code``
    (``on_balance_mismatch_total_ghs``) because the ``equals_sum_of_rows``
    validation and ``_snapshot_total`` bind to it. Written into LMTD Table 2's
    "#" column, beneath BoG's own row numbers 1 to 17, it reads as an internal
    variable on a return — so it is left out here exactly as it is on the PDF,
    which this workbook is the audit twin of. The key itself is untouched in the
    snapshot and in the CSV, the machine interchange format.
    """
    for col_idx, (spec, cell) in enumerate(zip(columns, rendered_row.cells, strict=True), start=1):
        if not (rendered_row.is_total and spec.key == "code"):
            _write_cell(sheet, row_idx, col_idx, cell)
        styled = sheet.cell(row=row_idx, column=col_idx)
        if bold:
            styled.font = Font(bold=True)
            styled.fill = _TOTAL_FILL


def _autosize(sheet: Worksheet, column_count: int) -> None:
    for col_idx in range(1, column_count + 1):
        longest = 0
        for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is not None:
                longest = max(longest, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 12), 60)


def _metadata_sheet(workbook: Workbook, rendered: RenderedReturn) -> None:
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Return Metadata"
    for row_idx, (label, value) in enumerate(rendered.metadata_pairs, start=1):
        label_cell = sheet.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True)
        sheet.cell(row=row_idx, column=2, value=value)
    row_idx = len(rendered.metadata_pairs) + 2
    sheet.cell(row=row_idx, column=1, value="Attestation").font = Font(bold=True)
    for offset, line in enumerate(rendered.attestation_lines, start=1):
        sheet.cell(row=row_idx + offset, column=2, value=line)
    row_idx += len(rendered.attestation_lines) + 2
    if rendered.template.notes:
        sheet.cell(row=row_idx, column=1, value="Template notes").font = Font(bold=True)
        for offset, note in enumerate(rendered.template.notes, start=1):
            sheet.cell(row=row_idx + offset, column=2, value=note)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110


def _section_sheets(
    workbook: Workbook, rendered: RenderedReturn, used_titles: set[str]
) -> dict[str, _SectionLocation]:
    locations: dict[str, _SectionLocation] = {}
    for section in rendered.sections:
        sheet_title = _sheet_title(section.title, used_titles)
        sheet = workbook.create_sheet(sheet_title)
        title_cell = sheet.cell(row=1, column=1, value=section.title)
        title_cell.font = Font(bold=True, size=12)
        unit_cell = sheet.cell(row=2, column=1, value=rendered.template.currency_unit)
        unit_cell.font = Font(italic=True)
        fidelity_cell = sheet.cell(
            row=3,
            column=1,
            value=(
                f"Fidelity: {section.layout.fidelity} · Layout: {section.layout.layout_id} · "
                f"Source: {section.layout.source_citation}"
            ),
        )
        fidelity_cell.font = Font(italic=True, size=9)
        header_row = 5
        for col_idx, column in enumerate(section.layout.columns, start=1):
            header = sheet.cell(row=header_row, column=col_idx, value=column.header)
            header.font = Font(bold=True)
            header.fill = _HEADER_FILL
            header.border = _THIN_BORDER
        row_idx = header_row + 1
        for rendered_row in section.rows:
            _write_table_row(sheet, row_idx, rendered_row, section.layout.columns)
            row_idx += 1
        locations[section.layout.section_code] = _SectionLocation(
            sheet_title=sheet_title,
            first_data_row=header_row + 1,
            row_count=len(section.rows),
            columns={
                column.key: column_idx
                for column_idx, column in enumerate(section.layout.columns, start=1)
            },
        )
        if section.total_row is not None:
            _write_table_row(sheet, row_idx, section.total_row, section.layout.columns, bold=True)
            row_idx += 1
        if section.layout.notes:
            row_idx += 1
            for note in section.layout.notes:
                note_cell = sheet.cell(row=row_idx, column=1, value=f"Note: {note}")
                note_cell.font = Font(italic=True, size=9)
                row_idx += 1
        _autosize(sheet, len(section.layout.columns))
    return locations


def _quote_sheet_title(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def _report_cell(location: _SectionLocation, key: str, row_offset: int) -> str | None:
    column = location.columns.get(key)
    if column is None or row_offset >= location.row_count:
        return None
    sheet = _quote_sheet_title(location.sheet_title)
    return f"{sheet}!{get_column_letter(column)}{location.first_data_row + row_offset}"


def _snapshot_total(snapshot: dict[str, Any], code: str) -> Decimal | None:
    for row in snapshot.get("totals", []):
        if row.get("code") != code or row.get("value") in (None, ""):
            continue
        try:
            return Decimal(str(row["value"])) / Decimal("1000")
        except (InvalidOperation, TypeError, ValueError):
            return None
    return None


def _working_header(sheet: Worksheet, title: str) -> None:
    sheet.cell(row=1, column=1, value=_WORKING_COPY_BANNER).font = Font(bold=True, color="9C0006")
    sheet.cell(row=2, column=1, value=title).font = Font(bold=True, size=12)
    sheet.cell(
        row=3,
        column=1,
        value=(
            "Inputs are copied from the sealed report sheets. Edit only this workbook for "
            "review; it does not change the immutable package or any filing artifact."
        ),
    ).font = Font(italic=True, size=9)


def _working_lmt(sheet: Worksheet, locations: dict[str, _SectionLocation]) -> bool:
    inputs = locations.get("prudential_ratio_inputs")
    reported = locations.get("prudential_ratio_percentages")
    if inputs is None or reported is None or "value" not in inputs.columns:
        return False
    _working_header(sheet, "SDI LMTD Table 1 — Working Ratio Calculations")
    sheet.append([])
    headers = ("Input", "Sealed report value", "Calculation", "Reported ratio", "Check")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    input_names = (
        "Narrow liquid assets",
        "Broad liquid assets",
        "Volatile liabilities",
        "Total deposits",
        "Short-term liabilities",
        "Total assets",
    )
    for offset, label in enumerate(input_names):
        row = 6 + offset
        source = _report_cell(inputs, "value", offset)
        sheet.cell(row=row, column=1, value=label)
        if source is not None:
            sheet.cell(row=row, column=2, value=f"={source}").number_format = _GHS_FORMAT

    ratios = (
        ("Narrow liquid assets / volatile liabilities", 6, 8, 0),
        ("Broad liquid assets / volatile liabilities", 7, 8, 1),
        ("Narrow liquid assets / short-term liabilities", 6, 10, 2),
        ("Broad liquid assets / short-term liabilities", 7, 10, 3),
        ("Narrow liquid assets / total deposits", 6, 9, 4),
        ("Broad liquid assets / total deposits", 7, 9, 5),
        ("Narrow liquid assets / total assets", 6, 11, 6),
        ("Broad liquid assets / total assets", 7, 11, 7),
    )
    for offset, (label, numerator, denominator, reported_offset) in enumerate(ratios, start=13):
        sheet.cell(row=offset, column=1, value=label)
        calculated = sheet.cell(
            row=offset, column=3, value=f'=IFERROR(B{numerator}/B{denominator}*100,"")'
        )
        calculated.number_format = _PCT_FORMAT
        source = _report_cell(reported, "value", reported_offset)
        if source is not None:
            reported_cell = sheet.cell(row=offset, column=4, value=f"={source}")
            reported_cell.number_format = _PCT_FORMAT
            sheet.cell(
                row=offset,
                column=5,
                value=f'=IF(ABS(C{offset}-D{offset})<0.0001,"MATCH","CHECK")',
            )
    return True


def _working_large_exposures(
    sheet: Worksheet, locations: dict[str, _SectionLocation], snapshot: dict[str, Any]
) -> bool:
    nof = _snapshot_total(snapshot, "nof_ghs")
    sections = [
        locations.get(code) for code in ("template_1", "template_2", "template_3", "template_4")
    ]
    available = [location for location in sections if location is not None and location.row_count]
    if nof is None or not available:
        return False
    _working_header(sheet, "SDI Large Exposures — Working % of Net Own Funds")
    sheet.cell(row=5, column=1, value="Net Own Funds (GHS '000)").font = Font(bold=True)
    sheet.cell(row=5, column=2, value=nof).number_format = _GHS_FORMAT
    sheet.cell(row=6, column=1, value="Large-exposure threshold (10% of NOF)").font = Font(
        bold=True
    )
    sheet.cell(row=6, column=2, value="=B5*10%").number_format = _GHS_FORMAT
    row_idx = 8
    for location in available:
        if "value" not in location.columns or "pct_nof" not in location.columns:
            continue
        sheet.cell(row=row_idx, column=1, value=location.sheet_title).font = Font(bold=True)
        row_idx += 1
        for column, header in enumerate(
            ("Counterparty key", "Exposure", "Calculated % NOF", "Reported % NOF", "Check"), start=1
        ):
            cell = sheet.cell(row=row_idx, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
        row_idx += 1
        for offset in range(location.row_count):
            code = _report_cell(location, "code", offset)
            exposure = _report_cell(location, "value", offset)
            reported = _report_cell(location, "pct_nof", offset)
            if code is not None:
                sheet.cell(row=row_idx, column=1, value=f"={code}")
            if exposure is not None:
                sheet.cell(row=row_idx, column=2, value=f"={exposure}").number_format = _GHS_FORMAT
                sheet.cell(
                    row=row_idx, column=3, value=f'=IFERROR(B{row_idx}/$B$5*100,"")'
                ).number_format = _PCT_FORMAT
            if reported is not None:
                sheet.cell(row=row_idx, column=4, value=f"={reported}").number_format = _PCT_FORMAT
                sheet.cell(
                    row=row_idx,
                    column=5,
                    value=f'=IF(ABS(C{row_idx}-D{row_idx})<0.0001,"MATCH","CHECK")',
                )
            row_idx += 1
        row_idx += 1
    return True


def _working_stress(sheet: Worksheet, locations: dict[str, _SectionLocation]) -> bool:
    positions = locations.get("t1_summary_positions")
    if positions is None or not positions.row_count:
        return False
    _working_header(sheet, "SDI Annual Stress — Working CAR Calculations")
    headers = (
        "Period",
        "Net Own Funds",
        "Risk-Weighted Assets",
        "Calculated CAR %",
        "Reported CAR %",
        "Check",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for offset in range(positions.row_count):
        row_idx = 6 + offset
        period = _report_cell(positions, "description", offset)
        nof = _report_cell(positions, "value", offset)
        rwa = _report_cell(positions, "total_rwa", offset)
        reported = _report_cell(positions, "car_pct", offset)
        if period is not None:
            sheet.cell(row=row_idx, column=1, value=f"={period}")
        if nof is not None:
            sheet.cell(row=row_idx, column=2, value=f"={nof}").number_format = _NUMBER_FORMAT
        if rwa is not None:
            sheet.cell(row=row_idx, column=3, value=f"={rwa}").number_format = _NUMBER_FORMAT
            sheet.cell(
                row=row_idx, column=4, value=f'=IFERROR(B{row_idx}/C{row_idx}*100,"")'
            ).number_format = _PCT_FORMAT
        if reported is not None:
            sheet.cell(row=row_idx, column=5, value=f"={reported}").number_format = _PCT_FORMAT
            sheet.cell(
                row=row_idx,
                column=6,
                value=f'=IF(ABS(D{row_idx}-E{row_idx})<0.0001,"MATCH","CHECK")',
            )
    return True


def _working_irrbb(sheet: Worksheet, locations: dict[str, _SectionLocation]) -> bool:
    gap = locations.get("repricing_gap")
    if gap is None or not gap.row_count:
        return False
    _working_header(sheet, "SDI IRRBB — Working Repricing-Gap Calculations")
    headers = (
        "Time Bucket",
        "Rate-Sensitive Assets",
        "Rate-Sensitive Liabilities",
        "Calculated Gap",
        "Reported Gap",
        "Check",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for offset in range(gap.row_count):
        row_idx = 6 + offset
        bucket = _report_cell(gap, "code", offset)
        rsa = _report_cell(gap, "rsa_ghs", offset)
        rsl = _report_cell(gap, "rsl_ghs", offset)
        reported = _report_cell(gap, "value", offset)
        if bucket is not None:
            sheet.cell(row=row_idx, column=1, value=f"={bucket}")
        if rsa is not None:
            sheet.cell(row=row_idx, column=2, value=f"={rsa}").number_format = _GHS_FORMAT
        if rsl is not None:
            sheet.cell(row=row_idx, column=3, value=f"={rsl}").number_format = _GHS_FORMAT
        if rsa is not None and rsl is not None:
            sheet.cell(
                row=row_idx, column=4, value=f"=B{row_idx}-C{row_idx}"
            ).number_format = _GHS_FORMAT
        if reported is not None:
            sheet.cell(row=row_idx, column=5, value=f"={reported}").number_format = _GHS_FORMAT
            sheet.cell(
                row=row_idx,
                column=6,
                value=f'=IF(ABS(D{row_idx}-E{row_idx})<0.0001,"MATCH","CHECK")',
            )
    return True


def _working_calculations_sheet(
    workbook: Workbook,
    rendered: RenderedReturn,
    locations: dict[str, _SectionLocation],
    snapshot: dict[str, Any],
    used_titles: set[str],
) -> None:
    sheet = workbook.create_sheet(_sheet_title("Working Calculations", used_titles))
    code = rendered.template.return_code
    written = (
        _working_lmt(sheet, locations)
        if code == "SDI-LMT-MONTHLY"
        else _working_large_exposures(sheet, locations, snapshot)
        if code == "SDI-LE-MONTHLY"
        else _working_stress(sheet, locations)
        if code == "SDI-STRESS-ANNUAL"
        else _working_irrbb(sheet, locations)
        if code == "SDI-IRRBB-QUARTERLY"
        else False
    )
    if not written:
        workbook.remove(sheet)
        return
    for column in range(1, 7):
        sheet.column_dimensions[get_column_letter(column)].width = 28


def _provenance_sheet(workbook: Workbook, rendered: RenderedReturn, used_titles: set[str]) -> None:
    sheet = workbook.create_sheet(_sheet_title("Fidelity & Provenance", used_titles))
    row_idx = 1
    for line in rendered.provenance_lines:
        sheet.cell(row=row_idx, column=1, value=line).font = Font(bold=(row_idx == 1))
        row_idx += 1
    row_idx += 1
    headers = ("Module", "Run ID", "Input Hash", "Engine Version")
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    row_idx += 1
    for module, run_id, input_hash, engine_version in rendered.provenance_runs:
        for col_idx, value in enumerate((module, run_id, input_hash, engine_version), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
    row_idx += 1
    sheet.cell(row=row_idx, column=1, value="Per-section fidelity").font = Font(bold=True)
    row_idx += 1
    for section in rendered.sections:
        sheet.cell(row=row_idx, column=1, value=section.layout.layout_id)
        sheet.cell(row=row_idx, column=2, value=section.layout.fidelity)
        sheet.cell(row=row_idx, column=3, value=section.layout.source_citation)
        row_idx += 1
    _autosize(sheet, 4)


def _normalize_zip(data: bytes, pinned: datetime) -> bytes:
    """Re-write the xlsx container with fixed entry timestamps — and pin the
    ``dcterms:modified`` document property, which openpyxl force-stamps with
    wall-clock time at save — so identical workbook content always yields
    identical bytes."""
    stamp = pinned.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    normalized = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(data)) as source,
        zipfile.ZipFile(normalized, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + stamp + rb"\g<2>",
                    payload,
                )
            info = zipfile.ZipInfo(name, date_time=_ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            target.writestr(info, payload)
    return normalized.getvalue()


def render_xlsx(
    rendered: RenderedReturn,
    *,
    generated_at: datetime,
    working_copy: bool = False,
    snapshot: dict[str, Any] | None = None,
) -> bytes:
    workbook = Workbook()
    pinned = generated_at.replace(tzinfo=None, microsecond=0)
    workbook.properties.created = pinned
    workbook.properties.modified = pinned
    workbook.properties.creator = "AequorOS Regulatory Reporting"
    workbook.properties.lastModifiedBy = "AequorOS Regulatory Reporting"
    if working_copy:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    used_titles: set[str] = {"Return Metadata"}
    _metadata_sheet(workbook, rendered)
    locations = _section_sheets(workbook, rendered, used_titles)
    if working_copy and snapshot is not None:
        _working_calculations_sheet(workbook, rendered, locations, snapshot, used_titles)
    _provenance_sheet(workbook, rendered, used_titles)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return _normalize_zip(buffer.getvalue(), pinned)


__all__ = ["render_xlsx"]
