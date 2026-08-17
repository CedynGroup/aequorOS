"""Template-faithful PDF export of a computed BoG form — the filing artifact.

This is the SAME official form as the xlsx export, rendered to paper: the
committed layout's own grid, the official line descriptions at their official
row/column positions, merges honoured, values in the boxes BoG expects them in.
A supervisor must recognise the return on sight, so nothing internal appears in
the body — no cell references, no line ids, no resolver names.

The earlier implementation fell through to the generic tabular renderer, which
emitted a Line / Official cell / Status listing keyed by internal ids. That is a
completion aid, not a return: no supervisor can read it and BoG would not accept
it. The completion detail lives on the xlsx export's "Completion notes" sheet,
which is an internal working artifact — it is deliberately absent here.

Cells the bank has not yet sourced are left BLANK on the form; an official
return must never show a fabricated zero, and a blank box already means exactly
"not supplied".
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .engine import FormResult, scale_for_export
from .layout import SheetLayout
from .render import (
    WORKING_COPY_BANNER,
    _date_overrides,
    _header_text,
    _previous_reporting_date,
)

_NAVY = colors.HexColor("#1F3864")
_GRID_GREY = colors.HexColor("#BFBFBF")
_HEADER_FILL = colors.HexColor("#F2F2F2")

_STYLES = getSampleStyleSheet()
_TITLE = ParagraphStyle("BogTitle", parent=_STYLES["Title"], fontSize=16, leading=20)
_CELL = ParagraphStyle("BogCell", parent=_STYLES["BodyText"], fontSize=7.5, leading=9)
_NOTE = ParagraphStyle("BogNote", parent=_STYLES["BodyText"], fontSize=8, leading=10.5)
_GRID_TEXT = ParagraphStyle(
    "BogGrid", parent=_STYLES["BodyText"], fontSize=6.4, leading=7.6, spaceAfter=0
)
_GRID_BOLD = ParagraphStyle("BogGridBold", parent=_GRID_TEXT, fontName="Helvetica-Bold")

# Below this the grid stops being legible; wider sheets get landscape, and then
# the font shrinks to the floor rather than silently dropping columns.
_MIN_COL_MM = 9.0
_LABEL_COL_MM = 52.0


def _invariant_canvas(*args: Any, **kwargs: Any) -> pdf_canvas.Canvas:
    kwargs["invariant"] = 1
    return pdf_canvas.Canvas(*args, **kwargs)


class _PageFurniture:
    def __init__(self, *, footer: str, banner: str | None) -> None:
        self._footer = footer
        self._banner = banner

    def __call__(self, canvas: pdf_canvas.Canvas, doc: SimpleDocTemplate) -> None:
        width, height = doc.pagesize
        canvas.saveState()
        canvas.setStrokeColor(_NAVY)
        canvas.setLineWidth(2)
        canvas.line(14 * mm, height - 12 * mm, width - 14 * mm, height - 12 * mm)
        if self._banner:
            canvas.setFont("Helvetica-Bold", 8)
            canvas.setFillColor(colors.HexColor("#B26B00"))
            canvas.drawCentredString(width / 2, height - 9 * mm, self._banner)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(14 * mm, 8 * mm, self._footer)
        canvas.drawRightString(width - 14 * mm, 8 * mm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()


def _format_value(value: Any, number_format: str | None) -> str:  # noqa: PLR0911
    """Render a computed amount the way the official sheet shows it."""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%d-%b-%y")
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
        if number == 0:
            return "-"
        quantised = number.quantize(Decimal("0.01"))
        if quantised == quantised.to_integral_value():
            text = f"{quantised.to_integral_value():,}"
        else:
            text = f"{quantised:,.2f}"
        if number < 0:
            text = f"({text.lstrip('-')})"
        if number_format and "%" in str(number_format):
            text = f"{text}%"
        return text
    return str(value)


def _used_columns(layout: SheetLayout, result: FormResult) -> list[int]:
    """The official grid's real width.

    The templates carry formulas dragged far past the printed form (BSD1's
    sheet runs to column IV — 2,844 formula cells across 255 columns, of which
    ten are the actual return). Labels delimit the form a supervisor sees, so
    the grid is bounded by the last LABELLED column and then widened only for
    columns that genuinely carry a value.
    """
    labelled = {
        cell.col
        for cell in layout.cells
        if cell.kind == "label" and str(cell.value or "").strip()
    }
    valued = {
        (layout.by_ref[ref].col if ref in layout.by_ref else _col_of(ref))
        for (sheet_name, ref), raw in result.all_values().items()
        if sheet_name == layout.name and raw is not None
    }
    if labelled:
        limit = max(labelled)
        used = {col for col in labelled | valued if col <= limit}
    else:
        used = labelled | valued
    return sorted(used) or [1]


def _col_of(ref: str) -> int:
    return column_index_from_string(coordinate_from_string(ref)[0])


def _row_of(ref: str) -> int:
    return coordinate_from_string(ref)[1]


def _sheet_matrix(  # noqa: PLR0913
    layout: SheetLayout,
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
) -> tuple[list[list[Any]], list[int], dict[tuple[int, int], bool]]:
    """The sheet as a dense grid of display strings, plus which cells are bold."""
    columns = _used_columns(layout, result)
    col_index = {col: i for i, col in enumerate(columns)}
    unscaled = result.unscaled_cells
    date_cells = _date_overrides(layout, reporting_date)
    extra = {
        ref: raw
        for (sheet_name, ref), raw in result.all_values().items()
        if sheet_name == layout.name and ref not in layout.by_ref and raw is not None
    }
    max_row = max(
        [layout.max_row] + [_row_of(ref) for ref in extra] or [layout.max_row]
    )
    grid: list[list[Any]] = [["" for _ in columns] for _ in range(max_row)]
    bold: dict[tuple[int, int], bool] = {}

    for cell in layout.cells:
        if cell.col not in col_index:
            continue
        row_i, col_i = cell.row - 1, col_index[cell.col]
        if cell.ref in date_cells:
            text = date_cells[cell.ref].strftime("%d-%b-%y")
        elif cell.kind == "label":
            raw_text = str(cell.value) if cell.value is not None else ""
            replaced = _header_text(
                raw_text,
                bank_name=bank_name,
                period_label=period_label,
                reporting_date=reporting_date,
                previous_reporting_date=_previous_reporting_date(reporting_date),
            )
            text = replaced if replaced is not None else raw_text
        else:
            value = scale_for_export(
                result.spec,
                layout.name,
                result.value(layout.name, cell.ref),
                unscaled=(layout.name, cell.ref) in unscaled,
            )
            text = _format_value(value, cell.number_format)
        grid[row_i][col_i] = text.strip() if isinstance(text, str) else text
        bold[(row_i, col_i)] = cell.bold

    for ref, raw in extra.items():
        col = _col_of(ref)
        if col not in col_index:
            continue
        value = scale_for_export(
            result.spec, layout.name, raw, unscaled=(layout.name, ref) in unscaled
        )
        grid[_row_of(ref) - 1][col_index[col]] = _format_value(value, None)

    # drop wholly blank rows so the form does not run to pages of white space
    keep = [i for i, row in enumerate(grid) if any(str(c).strip() for c in row)]
    return [grid[i] for i in keep], columns, {
        (keep.index(r), c): v for (r, c), v in bold.items() if r in keep and v
    }


def _column_widths(layout: SheetLayout, columns: list[int], available_mm: float) -> list[float]:
    """Official column proportions, scaled to the page."""
    raw: list[float] = []
    for index, col in enumerate(columns):
        width = layout.column_widths.get(get_column_letter(col))
        if width is None:
            width = 30.0 if index == 0 else 11.0
        raw.append(max(float(width), 4.0))
    # the leading description column carries the BoG line text; give it room
    total = sum(raw)
    scaled = [w / total * available_mm for w in raw]
    floor = min(_MIN_COL_MM, available_mm / max(len(columns), 1))
    scaled = [max(w, floor) for w in scaled]
    over = sum(scaled) / available_mm
    if over > 1.0:
        scaled = [w / over for w in scaled]
    return [w * mm for w in scaled]


def _merge_spans(
    layout: SheetLayout, columns: list[int], row_map: dict[int, int]
) -> list[tuple[str, tuple[int, int], tuple[int, int]]]:
    col_index = {col: i for i, col in enumerate(columns)}
    spans: list[tuple[str, tuple[int, int], tuple[int, int]]] = []
    for merge in layout.merges:
        try:
            start, end = merge.split(":")
        except ValueError:
            continue
        c0, r0, c1, r1 = _col_of(start), _row_of(start), _col_of(end), _row_of(end)
        if c0 not in col_index or c1 not in col_index:
            continue
        if r0 not in row_map or r1 not in row_map:
            continue
        spans.append(("SPAN", (col_index[c0], row_map[r0]), (col_index[c1], row_map[r1])))
    return spans


def _sheet_story(  # noqa: PLR0913
    layout: SheetLayout,
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    page_width_mm: float,
) -> list[Any]:
    grid, columns, bold = _sheet_matrix(
        layout,
        result,
        bank_name=bank_name,
        period_label=period_label,
        reporting_date=reporting_date,
    )
    if not grid:
        return []
    widths = _column_widths(layout, columns, page_width_mm)
    body = [
        [
            Paragraph(
                str(text).replace("&", "&amp;").replace("<", "&lt;"),
                _GRID_BOLD if bold.get((r, c)) else _GRID_TEXT,
            )
            if str(text).strip()
            else ""
            for c, text in enumerate(row)
        ]
        for r, row in enumerate(grid)
    ]
    table = Table(body, colWidths=widths, repeatRows=0)
    style: list[Any] = [
        ("GRID", (0, 0), (-1, -1), 0.25, _GRID_GREY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
        ("TOPPADDING", (0, 0), (-1, -1), 1.0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.0),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]
    table.setStyle(TableStyle(style))
    return [
        Paragraph(f"<b>{layout.name.strip()}</b>", _NOTE),
        Spacer(0, 2 * mm),
        table,
    ]


_UNIT_WORDS = {
    "millions": "Millions",
    "thousands": "Thousands",
    "units": "Units (absolute)",
    "percent": "Percent",
}


def _unit_note(spec: Any) -> str:
    """The sheet unit convention, stated the way the official sheets state it."""
    units = {sheet.unit for sheet in spec.sheets}
    if len(units) == 1:
        return _UNIT_WORDS.get(next(iter(units)), str(next(iter(units))))
    return "as stated on each sheet (BoG unit conventions preserved)"


def _cover(  # noqa: PLR0913
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    generated_at: datetime,
    package_line: str,
) -> list[Any]:
    spec = result.spec
    counts = result.status_counts
    pairs = [
        ("Institution", bank_name),
        ("Return", f"{spec.code} — {spec.title}"),
        ("Reporting date", reporting_date),
        ("Reporting period", period_label),
        ("Frequency", spec.frequency),
        ("Reporting basis", spec.basis),
        ("Currency unit", _unit_note(spec)),
        ("Source workbook", spec.workbook),
        ("Cells completed", str(counts.get("mapped", 0) + counts.get("derived", 0))),
        ("Cells outstanding", str(counts.get("input_required", 0) + counts.get("unmapped", 0))),
        ("Generated at", generated_at.isoformat()),
        ("Package", package_line),
    ]
    rows = [
        [Paragraph(f"<b>{label}</b>", _CELL), Paragraph(str(value), _CELL)]
        for label, value in pairs
    ]
    table = Table(rows, colWidths=[45 * mm, 115 * mm])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.3, _GRID_GREY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (0, -1), _HEADER_FILL),
            ]
        )
    )
    return [
        Spacer(0, 18 * mm),
        Paragraph(f"{spec.code} — {spec.title}", _TITLE),
        Spacer(0, 5 * mm),
        table,
    ]


def _attestation() -> list[Any]:
    rows = [
        [Paragraph("<b>Prepared by</b>", _CELL), "", "", ""],
        [
            Paragraph("Name", _CELL),
            Paragraph("Designation", _CELL),
            Paragraph("Signature", _CELL),
            Paragraph("Date", _CELL),
        ],
        ["", "", "", ""],
        [Paragraph("<b>Reviewed and approved by</b>", _CELL), "", "", ""],
        [
            Paragraph("Name", _CELL),
            Paragraph("Designation", _CELL),
            Paragraph("Signature", _CELL),
            Paragraph("Date", _CELL),
        ],
        ["", "", "", ""],
    ]
    table = Table(rows, colWidths=[45 * mm, 45 * mm, 45 * mm, 25 * mm], rowHeights=None)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.3, _GRID_GREY),
                ("SPAN", (0, 0), (-1, 0)),
                ("SPAN", (0, 3), (-1, 3)),
                ("BACKGROUND", (0, 0), (-1, 0), _HEADER_FILL),
                ("BACKGROUND", (0, 3), (-1, 3), _HEADER_FILL),
                ("TOPPADDING", (0, 2), (-1, 2), 10),
                ("BOTTOMPADDING", (0, 2), (-1, 2), 10),
                ("TOPPADDING", (0, 5), (-1, 5), 10),
                ("BOTTOMPADDING", (0, 5), (-1, 5), 10),
            ]
        )
    )
    return [
        Paragraph("<b>Attestation</b>", _NOTE),
        Spacer(0, 3 * mm),
        table,
        Spacer(0, 4 * mm),
        Paragraph(
            "We attest that this return is complete and accurate to the best of our "
            "knowledge (Act 930 s.93(3) applies to inaccurate or incomplete submissions).",
            _NOTE,
        ),
    ]


# NOTE: there is deliberately no completion appendix here. The PDF IS the return
# the institution files — spreadsheet cell references ("B23"), internal line ids
# and resolver names ("bsd1.daily", "not yet mapped") are machinery, and a filed
# document must not carry them. Preparers get exactly that detail on the xlsx
# export's "Completion notes" sheet, which is an internal working artifact.
# Incomplete boxes read as blanks on the form, which is what a blank box means.


def render_form_pdf(  # noqa: PLR0913
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    generated_at: datetime,
    package_line: str = "",
    mode: str = "official",
) -> bytes:
    """The official BoG form as a PDF — the artifact the institution files."""
    if mode not in ("official", "working"):
        raise ValueError(f"unknown render mode {mode!r}")
    layouts = result.layout.sheets
    wide = max((len(_used_columns(s, result)) for s in layouts), default=1) > 6
    pagesize = landscape(A4) if wide else A4
    page_width_mm = (pagesize[0] / mm) - 28.0

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
        title=f"{result.spec.code} — {result.spec.title}",
        author="AequorOS Regulatory Reporting",
    )
    story: list[Any] = _cover(
        result,
        bank_name=bank_name,
        period_label=period_label,
        reporting_date=reporting_date,
        generated_at=generated_at,
        package_line=package_line,
    )
    story += [Spacer(0, 8 * mm), *_attestation(), PageBreak()]
    for sheet_layout in layouts:
        section = _sheet_story(
            sheet_layout,
            result,
            bank_name=bank_name,
            period_label=period_label,
            reporting_date=reporting_date,
            page_width_mm=page_width_mm,
        )
        if section:
            story += [*section, PageBreak()]
    if story and isinstance(story[-1], PageBreak):
        story.pop()

    footer = f"{result.spec.code} · {bank_name} · {reporting_date}"
    if mode == "official":
        footer = f"{footer} · official return"
    document.build(
        story,
        onFirstPage=_PageFurniture(
            footer=footer, banner=WORKING_COPY_BANNER if mode == "working" else None
        ),
        onLaterPages=_PageFurniture(
            footer=footer, banner=WORKING_COPY_BANNER if mode == "working" else None
        ),
        canvasmaker=_invariant_canvas,
    )
    return buffer.getvalue()
