"""Template-faithful xlsx export of a computed BoG form — two modes.

The workbook is rebuilt from the committed official layout — same sheet set (in
the official order, empty placeholder tabs included), same labels at the same
cells, same merges / column widths / bold / number formats — so a supervisor
recognises the return on sight.

``mode="official"`` (artifact kind ``xlsx``): every cell carries its evaluated
NUMBER, never a live formula, and every sheet is protected — the sealed Excel
twin of the submission PDF, immutable, hashed, maker-checked. The template's
formula text is preserved on the "Completion notes" sheet for auditability.

``mode="working"`` (artifact kind ``xlsx_working``): the ALM / Finance review
copy — leaf/input amounts as values, and every formula cell carries the
template's OWN formula (SUM, Domestic + Foreign → Total, cross-sheet annex
links) so Excel recalculates when a reviewer challenges an input. Cross-WORKBOOK
links (BSD8 → [1]BSD2) cannot resolve inside one file and are written as their
evaluated value with a note. The file is labelled WORKING COPY — FOR INTERNAL
REVIEW (workbook title, print header on every sheet, Completion notes banner)
and is never a filing artifact.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from .engine import FormResult, scale_for_export
from .layout import SheetLayout
from .spec import LineValue

# Header prompts, enumerated from EVERY official template (bog_forms discovery,
# 2026-08-15). Two traps drive the design: (1) some templates ship with a real
# bank's name / a stale date typed after the prompt (BSD1B: "Reporting
# Institution: FIRST ATLANTIC MERCHANT BANK LTD", "Reporting as at 1 Sept.
# 2003") — an export must never carry those, so everything AFTER a bank-name
# prompt is discarded; (2) look-alike COLUMN headers ("NAME OF BANK /
# INSTITUTION" in BSD2A, "FINAL MATURITY DATE", "Expiry Date", "DATE APPOINTED")
# must be left verbatim, hence the explicit exclusions.
_BANK_PROMPT = re.compile(
    r"^\s*(NAME\s+OF\s+(?:THE\s+)?(?:REPORTING\s+)?(?:BANK|INSTITUTION)|"
    r"REPORT(?:ING)?\s+INSTITUTION|BANK)\s*(:|[…\.]{2,})",
    re.IGNORECASE,
)
_PERIOD_PROMPT = re.compile(
    r"(DATA\s+AS\s+OF\s+END|\bAS\s+AT\b|\bAS\s+OF\b|FOR\s+THE\s+(?:WEEK|MONTH|QUARTER|PERIOD|YEAR)"
    r"\s+END(?:ED|ING)?|WEEK\s+ENDING|MONTH\s+ENDING|REPORT\s+FOR\s+PERIOD|REPORTING\s+DATE|"
    r"DATE\s+OF\s+REPORTING|^\s*PERIOD\s*(?::|\s+AS\s+AT)|POSITION\(S\)\s+AS\s+AT|"
    r"PLEASE\s+PUT\s+DATE\s+HERE)",
    re.IGNORECASE,
)
_PREVIOUS_WEEK_PROMPT = re.compile(r"DATE\s*\(PREVIOUS\s+WEEK\)\s*:", re.IGNORECASE)
_NOT_A_HEADER = re.compile(
    r"(MATURITY|EXPIRY|APPOINTED|GRANTED|ACQUISITION|AUDIT|NAME\s+OF\s+(?:BANK|INSTITUTION)\s*/|"
    r"NAME\s+OF\s+CUSTOMER|NET\s+WORTH|COUNTERPARTY)",
    re.IGNORECASE,
)
_TRAILING_PROMPT_TAIL = re.compile(
    r"(\s*[…\.]{2,}.*$|\s*-\s*$|\s+\d{1,2}(?:\s+\w+\.?)?(?:\s+\d{4})?\s*$)", re.UNICODE
)


def _header_text(  # noqa: PLR0911
    label: str,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    previous_reporting_date: str | None = None,
) -> str | None:
    """Fill a header prompt; return None for every other label (kept verbatim)."""
    text = label.strip()
    if _NOT_A_HEADER.search(text):
        return None
    m = _BANK_PROMPT.match(text)
    if m:
        prompt = m.group(1).strip()
        return f"{prompt}: {bank_name}"
    if _PREVIOUS_WEEK_PROMPT.search(text):
        base = text.split(":", 1)[0].strip()
        return f"{base}: {previous_reporting_date or ''}".rstrip()
    if _PERIOD_PROMPT.search(text):
        if re.match(r"^\s*PLEASE\s+PUT\s+DATE\s+HERE", text, re.IGNORECASE):
            return reporting_date
        if re.match(r"^\s*PERIOD\s*:", text, re.IGNORECASE):
            return f"PERIOD: {period_label}"
        if re.match(r"^\s*(REPORTING\s+DATE|DATE\s+OF\s+REPORTING)\s*:", text, re.IGNORECASE):
            base = text.split(":", 1)[0].strip()
            return f"{base}: {reporting_date}"
        # Sentence-style prompt ("DATA AS OF END………. 20……", "AS AT ………",
        # "Reporting as at 1 Sept. 2003", "FOR THE WEEK ENDING…"): strip the
        # dotted/stale tail and append the reporting date.
        base = _TRAILING_PROMPT_TAIL.sub("", text).rstrip(" :")
        return f"{base} {reporting_date}"
    return None


def _previous_reporting_date(reporting_date: str) -> str | None:
    """The prior weekly reporting date (BSD1 'DATE (PREVIOUS WEEK)'), when parseable."""
    try:
        from datetime import date, timedelta  # noqa: PLC0415

        return (date.fromisoformat(reporting_date) - timedelta(days=7)).isoformat()
    except ValueError:
        return None


WORKING_COPY_BANNER = "WORKING COPY — FOR INTERNAL REVIEW · not a filing artifact"


def _is_external_link(formula: str | None) -> bool:
    return bool(formula) and "[" in str(formula) and "]" in str(formula)


def official_width(layout: SheetLayout) -> int:
    """Last column of the PRINTED form.

    The official workbooks carry formulas dragged far beyond the form a
    supervisor reads — BSD1's sheet runs to column IV, 2,844 formula cells of
    which ten columns are the return. Those strays evaluate to nothing and
    would export as a wall of ``0`` across every subtotal row. Labels delimit
    the real grid, so the printed form ends at the last labelled column.
    """
    labelled = [
        cell.col
        for cell in layout.cells
        if cell.kind == "label" and str(cell.value or "").strip()
    ]
    return max(labelled) if labelled else 0


def _is_offgrid_stray(layout: SheetLayout, cell: Any, value: Any) -> bool:
    """A dragged formula outside the printed form with nothing to say."""
    limit = official_width(layout)
    if not limit or cell.col <= limit or cell.kind == "label":
        return False
    return value is None or value == 0


# Date cells in the official templates are a trap. The weekday columns of BSD1
# ('DATE (PREVIOUS WEEK)') and the 'DATE OF REPORTING' box are FORMULA cells
# carrying a date number format; the workbook they referenced is gone, so they
# evaluate to ~0 and Excel renders the 1899 epoch (observed as
# "1899-12-29 23:59:58.87"). Others are LABEL cells holding a stale literal the
# original filer typed (BSD1 B3: 2004-05-26). Neither may reach an export, so
# the renderer supplies the real dates.
_DATE_FORMAT_TOKENS = re.compile(r"(?<!\\)(yy|mmm|ddd?)", re.IGNORECASE)
_WEEKDAY_LABELS = {
    "MON": 0, "MONDAY": 0, "TUE": 1, "TUES": 1, "TUESDAY": 1,
    "WED": 2, "WEDS": 2, "WEDNESDAY": 2, "THU": 3, "THUR": 3, "THURS": 3, "THURSDAY": 3,
    "FRI": 4, "FRIDAY": 4, "SAT": 5, "SATURDAY": 5, "SUN": 6, "SUNDAY": 6,
}


def _is_date_format(number_format: str | None) -> bool:
    """True when the cell's number format renders a DATE rather than an amount."""
    if not number_format:
        return False
    stripped = re.sub(r"\[\$?-?[0-9A-Fa-f]+\]", "", str(number_format))
    return bool(_DATE_FORMAT_TOKENS.search(stripped))


def _weekday_header(layout: SheetLayout) -> tuple[int, list[tuple[int, int]]]:
    """The DAY header row and its (column, weekday) pairs in column order."""
    found: dict[int, list[tuple[int, int]]] = {}
    for cell in layout.cells:
        if cell.kind != "label" or cell.value is None:
            continue
        weekday = _WEEKDAY_LABELS.get(str(cell.value).strip().upper().rstrip("."))
        if weekday is not None:
            found.setdefault(cell.row, []).append((cell.col, weekday))
    if not found:
        return 0, []
    row = max(found, key=lambda r: len(found[r]))
    return row, sorted(found[row])


def _weekday_column_dates(layout: SheetLayout, reporting_date: str) -> dict[int, date]:
    """Map each weekday-headed COLUMN to its actual date.

    The day columns are POSITIONAL, not calendar weekdays: the week ENDS on the
    reporting date and each earlier column counts back one day. This must stay
    identical to how the line maps resolve their data — ``sources_ext/bsd1.py``
    ``DAY_COLUMNS`` reads ``wed = period_end - 0 … thu = period_end - 6`` — or
    the form would print correctly-sourced figures under the wrong dates.
    Anchoring on real weekdays instead was tried and is wrong: it shifts the
    whole grid whenever the reporting date is not the week's closing weekday.
    (When a bank files a weekly return on its proper week-ending date, the two
    coincide and the printed dates are real weekdays too.)
    """
    try:
        as_of = date.fromisoformat(reporting_date)
    except ValueError:
        return {}
    _, pairs = _weekday_header(layout)
    if not pairs:
        return {}
    span = len(pairs)
    return {
        col: as_of - timedelta(days=span - 1 - position)
        for position, (col, _) in enumerate(pairs)
    }


def _date_overrides(layout: SheetLayout, reporting_date: str) -> dict[str, date]:
    """Real dates for every date-formatted cell the template cannot supply."""
    try:
        as_of = date.fromisoformat(reporting_date)
    except ValueError:
        return {}
    by_column = _weekday_column_dates(layout, reporting_date)
    header_row, _ = _weekday_header(layout)
    out: dict[str, date] = {}
    for cell in layout.cells:
        if not _is_date_format(cell.number_format):
            continue
        if cell.kind == "label" and not _looks_like_a_date(cell.value):
            # a real text label that merely inherited a date format — leave it
            continue
        # Only the day GRID takes its column's date; the boxes above the DAY
        # header (PERIOD, DATE OF REPORTING) are the reporting date itself.
        if cell.row <= header_row:
            out[cell.ref] = as_of
            continue
        column_date = by_column.get(cell.col)
        if column_date is None:
            out[cell.ref] = as_of
            continue
        # BSD1 states the week twice — 'DATE (PREVIOUS WEEK)' above the deposit
        # block and 'DATE (CURRENT WEEK)' above liquid assets. Both rows sit in
        # the same day columns, so the row's own label decides which week it is.
        if _PREVIOUS_WEEK_PROMPT.search(layout.label_for_row(cell.row)):
            column_date = column_date - timedelta(days=7)
        out[cell.ref] = column_date
    return out


def _looks_like_a_date(value: Any) -> bool:
    """True for a stale template date literal, however the layout stored it."""
    if isinstance(value, (datetime, date)):
        return True
    text = str(value or "").strip()
    if not text:
        return False
    try:
        datetime.fromisoformat(text)
    except ValueError:
        return False
    return True


def _write_sheet(  # noqa: PLR0912, PLR0913
    ws: Any,
    layout: SheetLayout,
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    mode: str = "official",
) -> None:
    for letter, width in layout.column_widths.items():
        ws.column_dimensions[letter].width = width
    unscaled = result.unscaled_cells
    date_cells = _date_overrides(layout, reporting_date)
    for cell in layout.cells:
        col_letter, row = coordinate_from_string(cell.ref)
        target = ws[cell.ref]
        if cell.ref in date_cells:
            target.value = date_cells[cell.ref]
            target.number_format = cell.number_format or "d-mmm-yy"
            if cell.bold:
                target.font = Font(bold=True)
            if cell.align:
                target.alignment = Alignment(horizontal=cell.align)
            continue
        if cell.kind == "label":
            text = str(cell.value)
            replaced = _header_text(
                text,
                bank_name=bank_name,
                period_label=period_label,
                reporting_date=reporting_date,
                previous_reporting_date=_previous_reporting_date(reporting_date),
            )
            target.value = replaced if replaced is not None else text
        else:
            raw = result.value(layout.name, cell.ref)
            if _is_offgrid_stray(layout, cell, raw):
                continue
            if mode == "working" and cell.kind == "formula" and not _is_external_link(cell.formula):
                # the template's own formula, live — Excel recalculates
                target.value = cell.formula
            else:
                target.value = scale_for_export(
                    result.spec, layout.name, raw, unscaled=(layout.name, cell.ref) in unscaled
                )
            if cell.number_format:
                target.number_format = cell.number_format
            elif isinstance(target.value, (int, float)):
                target.number_format = "#,##0.00;(#,##0.00);-"
        if cell.bold:
            target.font = Font(bold=True)
        if cell.align:
            target.alignment = Alignment(horizontal=cell.align)
    # Blank data grids: bound cells that the official template leaves EMPTY
    # (no placeholder) are not in the layout — write their values too.
    limit = official_width(layout)
    for (sheet_name, ref), raw in result.all_values().items():
        if sheet_name != layout.name or ref in layout.by_ref or raw is None:
            continue
        # same bound as the layout pass: a dragged formula to the right of the
        # printed form contributes nothing but a column of zeros
        if limit and column_index_from_string(coordinate_from_string(ref)[0]) > limit and not raw:
            continue
        target = ws[ref]
        target.value = scale_for_export(
            result.spec, layout.name, raw, unscaled=(sheet_name, ref) in unscaled
        )
        if isinstance(target.value, (int, float)):
            target.number_format = "#,##0.00;(#,##0.00);-"
    for merge in layout.merges:
        try:
            ws.merge_cells(merge)
        except ValueError:
            continue
    if mode == "working":
        ws.oddHeader.center.text = WORKING_COPY_BANNER
        ws.oddFooter.center.text = (
            f"{result.spec.code} · {bank_name} · {reporting_date} · working copy"
        )
        ws.sheet_properties.tabColor = "FFC000"
    else:
        # sealed: lock the sheet so downloaded totals cannot be edited in place
        ws.protection.sheet = True
        ws.oddFooter.center.text = (
            f"{result.spec.code} · {bank_name} · {reporting_date} · sealed official export"
        )
    # header cells declared explicitly by the sheet spec win over regex fills
    sheet_spec = result.spec.sheet(layout.name)
    if sheet_spec is not None:
        h = sheet_spec.header
        if h.bank_name:
            ws[h.bank_name].value = bank_name
        if h.period:
            ws[h.period].value = period_label
        if h.reporting_date:
            ws[h.reporting_date].value = reporting_date
        for ref, text in h.extras.items():
            ws[ref].value = text


def _completion_notes(
    wb: Workbook, result: FormResult, generated_at: datetime, *, mode: str = "official"
) -> None:
    ws = wb.create_sheet("Completion notes")
    ws.append([f"{result.spec.code} — completion notes (AequorOS)"])
    ws["A1"].font = Font(bold=True)
    if mode == "working":
        ws.append([WORKING_COPY_BANNER])
        ws["A2"].font = Font(bold=True, color="C00000")
        ws.append(
            [
                f"Generated {generated_at.isoformat()} · live template formulas (values for "
                "inputs; cross-workbook links written as evaluated values) — for ALM/Finance "
                "review; the sealed official export and the submission PDF are the governed "
                "artifacts."
            ]
        )
    else:
        ws.append(
            [f"Generated {generated_at.isoformat()} · values-only sealed export (sheets protected)"]
        )
    ws.append([f"Basis: {result.spec.basis} · Workbook: {result.spec.workbook}"])
    counts = result.status_counts
    ws.append(
        [
            f"Lines mapped: {counts['mapped']} · input required: {counts['input_required']} · "
            f"unmapped cells: {counts['unmapped']} · "
            f"template formulas evaluated: {counts['derived']}"
        ]
    )
    for note in result.spec.scope_notes:
        ws.append([note])
    ws.append([])
    ws.append(["Sheet", "Cell", "Line", "Label", "Status", "Source", "Notes"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for lv in result.lines:
        if lv.status != "mapped":
            ws.append([lv.sheet, lv.cell, lv.code, lv.label, lv.status, lv.source or "", lv.notes])
    for sheet, ref, label in result.unmapped_cells:
        ws.append(
            [sheet, ref, "", label, "unmapped", "", "official input cell with no line map yet"]
        )
    if result.missing_dependencies:
        ws.append([])
        ws.append([f"Missing dependency forms: {', '.join(result.missing_dependencies)}"])
    if result.errors:
        ws.append([])
        ws.append(["Errors"])
        for err in result.errors:
            ws.append([err])
    ws.append([])
    ws.append(["Template formulas (evaluated; retained for audit)"])
    ws.append(["Sheet", "Cell", "Formula"])
    for sheet in result.layout.sheets:
        for cell in sheet.formula_cells:
            ws.append([sheet.name, cell.ref, cell.formula])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["G"].width = 50


def render_form_xlsx(  # noqa: PLR0913
    result: FormResult,
    *,
    bank_name: str,
    period_label: str,
    reporting_date: str,
    generated_at: datetime,
    include_completion_notes: bool = True,
    mode: str = "official",
) -> bytes:
    """``mode`` = ``"official"`` (sealed values, protected) or ``"working"`` (live formulas)."""
    if mode not in ("official", "working"):
        msg = f"unknown export mode {mode!r} (official|working)"
        raise ValueError(msg)
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    for layout in result.layout.sheets:
        ws = wb.create_sheet(layout.name[:31])
        _write_sheet(
            ws,
            layout,
            result,
            bank_name=bank_name,
            period_label=period_label,
            reporting_date=reporting_date,
            mode=mode,
        )
    if include_completion_notes:
        _completion_notes(wb, result, generated_at, mode=mode)
    wb.properties.title = (
        f"{result.spec.code} — {WORKING_COPY_BANNER}"
        if mode == "working"
        else f"{result.spec.code} — official sealed export"
    )
    wb.properties.created = generated_at
    wb.properties.modified = generated_at
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def lines_by_sheet(result: FormResult) -> dict[str, list[LineValue]]:
    grouped: dict[str, list[LineValue]] = {}
    for lv in result.lines:
        grouped.setdefault(lv.sheet, []).append(lv)
    return grouped
