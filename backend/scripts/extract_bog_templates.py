"""Extract the official BoG return templates into committed layout JSON.

    uv run python scripts/extract_bog_templates.py \
        --source ../docs/reporting \
        --out app/services/regulatory_reporting/bog_forms/layouts

For every ``FORM BSD*.xls`` under ``--source`` this converts the legacy workbook to
xlsx with LibreOffice (headless; formulas preserved), then records EVERY sheet and
EVERY non-empty cell — labels, numeric input cells, and formula cells with their
formula text — plus merged ranges, column widths, and bold/number-format hints.
The JSON is the single layout the template-faithful exporter reproduces, so a
regenerated file must be byte-identical unless the official template changed.

LibreOffice is needed only to (re)generate the JSON; the runtime exporter reads
the committed JSON with openpyxl alone.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

FORM_FROM_FILE = re.compile(r"FORM\s+(BSD\s?[0-9]+[AB]?)", re.IGNORECASE)


def form_code(filename: str) -> str:
    match = FORM_FROM_FILE.search(filename)
    if not match:
        msg = f"cannot derive form code from {filename!r}"
        raise ValueError(msg)
    return match.group(1).replace(" ", "").upper()


def convert_to_xlsx(source: Path, workdir: Path) -> Path:
    soffice = shutil.which("soffice") or "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    subprocess.run(  # noqa: S603 - fixed binary, controlled args
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(workdir), str(source)],
        check=True,
        capture_output=True,
    )
    out = workdir / (source.stem + ".xlsx")
    if not out.exists():
        msg = f"LibreOffice produced no xlsx for {source.name}"
        raise RuntimeError(msg)
    return out


def _cell_kind(value: Any) -> str:
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    if isinstance(value, bool):
        return "label"
    if isinstance(value, (int, float)):
        return "input"
    return "label"


def extract_workbook(xlsx: Path, original_name: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        cells: list[dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None or (isinstance(value, str) and not value.strip()):
                    continue
                kind = _cell_kind(value)
                entry: dict[str, Any] = {"ref": cell.coordinate, "kind": kind}
                if kind == "formula":
                    entry["formula"] = value
                elif kind == "input":
                    entry["value"] = value
                else:
                    entry["value"] = str(value)
                if cell.font is not None and cell.font.bold:
                    entry["bold"] = True
                if cell.number_format and cell.number_format != "General":
                    entry["number_format"] = cell.number_format
                if cell.alignment is not None and cell.alignment.horizontal:
                    entry["align"] = cell.alignment.horizontal
                cells.append(entry)
        widths = {
            letter: round(dim.width, 2) for letter, dim in ws.column_dimensions.items() if dim.width
        }
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "column_widths": widths,
                "merges": [str(m) for m in ws.merged_cells.ranges],
                "cells": cells,
            }
        )
    return {
        "form": form_code(original_name),
        "workbook": original_name,
        "sheets": sheets,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args(argv)
    args.out.mkdir(parents=True, exist_ok=True)

    files = sorted(
        [p for p in args.source.iterdir() if p.suffix.lower() == ".xls"], key=lambda p: p.name
    )
    if not files:
        print(f"no .xls templates under {args.source}", file=sys.stderr)
        return 1
    with tempfile.TemporaryDirectory() as tmp:
        workdir = Path(tmp)
        for source in files:
            xlsx = convert_to_xlsx(source, workdir)
            layout = extract_workbook(xlsx, source.name)
            target = args.out / f"{layout['form']}.json"
            target.write_text(json.dumps(layout, indent=1, ensure_ascii=False) + "\n")
            n_cells = sum(len(s["cells"]) for s in layout["sheets"])
            print(
                f"{source.name:36} -> {target.name:12} "
                f"sheets={len(layout['sheets']):2} cells={n_cells}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
