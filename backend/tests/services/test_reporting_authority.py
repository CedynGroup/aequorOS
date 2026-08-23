"""Reporting authority, provenance and eligibility (forensic audit §8, §10 item 3, ARCH-8).

The audit's reporting findings, made executable:

1. ``bog_form`` returns returned ``source_runs=[]`` with nothing saying why. A
   package now carries an explicit authority record; the template-authoritative
   case says so in words and backs it with the official template digest, the
   line-map digest and the formula-evaluator version.
2. BSD2A computed a ratio "on render" with no input binding and no formula cell
   (CF-3). Its authority is now declared: the official BSD2A grid carries no
   formula cell at all, so the Guide's own paragraph is the authority and the
   package says that rather than implying BoG's template produced it.
3. Return eligibility was answered in two places on different criteria. There
   is now one authority, and both the calendar and the package-mint site
   consume it.

None of this changes a financial figure. The BoG workbook formulas remain
authoritative and untouched, and the BSD5A CAR inequality
(``bog_forms/test_bsd5.py``) stays pinned — this suite deliberately asserts no
equality between two methodologies of the same metric.
"""

from __future__ import annotations

import io
from dataclasses import replace
from datetime import UTC, date, datetime
from typing import Any, cast

import openpyxl
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from loguru import logger as loguru_logger
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.core.observability import Condition
from app.domain.authority.registry import REGISTRY as authority_registry
from app.domain.authority.registry import Regime
from app.models import (
    Bank,
    BankFinancialFact,
    BankReportingPeriod,
    RegulatoryPackage,
    RegulatoryRun,
)
from app.schemas.regulatory_liquidity import RegulatoryRunCreate
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.services import regulatory_liquidity
from app.services.regulatory_reporting import calendar, eligibility, generation
from app.services.regulatory_reporting.bog_forms.catalog import all_form_specs, form_spec
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import bsd2a as bsd2a_linemap
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from app.services.regulatory_reporting.provenance import (
    ENGINE_BACKED_RESOLVERS,
    FORMULA_EVALUATOR_VERSION,
    RESOLVER_AUTHORITY,
    ReportAuthority,
    declared_methodology_notes,
    mapping_digest,
    template_digest,
)
from app.services.regulatory_reporting.registry import REGISTRY, ReturnDefinition
from tests.api.helpers import ORG_1, USER_1, headers
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

from app.db.session import get_sessionmaker  # isort: skip

_AS_OF = date(2026, 6, 30)
REPORTING_DATE = date(2026, 3, 31)
MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
_FIXED_TS = datetime(2026, 8, 15, tzinfo=UTC)


def _sample_bank(db: Session) -> Bank:
    bank = db.get(Bank, SAMPLE_BANK_ID)
    assert bank is not None
    return bank


def _seed_with_baseline_run(db: Session) -> None:
    materialize_canonical_test_book(db)
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    run = regulatory_liquidity.create_liquidity_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=period_id, scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded"


def _generate_service(db: Session, return_code: str) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=return_code, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _numeric_cells(payload: bytes) -> dict[str, float]:
    """Every numeric cell of a rendered workbook, keyed sheet!ref."""
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    values: dict[str, float] = {}
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    values[f"{name}!{cell.coordinate}"] = float(cell.value)
    return values


# ---------------------------------------------------------------------------
# helpers (mirroring test_bog_forms_framework's package pipeline)
# ---------------------------------------------------------------------------


def _materialize() -> None:
    session = get_sessionmaker()()
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()


def _latest_period_end(db_client: TestClient) -> str:
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    return periods[0]["period_end"]


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:400]}"
    return response.json()


def _detail(db_client: TestClient, package_id: str) -> dict:
    return db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package_id}", headers=headers()
    ).json()


def _make_bank(db: Session, *, institution_type: str) -> Bank:
    bank = Bank(
        organization_id=ORG_1,
        name=f"{institution_type} eligibility tenant",
        short_name="Elig",
        currency="GHS",
        jurisdiction_code="GH",
        license_type="x",
        institution_type=institution_type,
    )
    db.add(bank)
    db.flush()
    return bank


# ---------------------------------------------------------------------------
# 1. every figure states its authority — and the classification cannot lapse
# ---------------------------------------------------------------------------


def test_every_bog_resolver_declares_a_reporting_authority() -> None:
    """No resolver may feed a filed cell without saying who owns the value.

    This is the gate that keeps the audit's finding closed: adding a resolver
    without classifying it fails here rather than quietly shipping a figure with
    no stated authority.
    """
    from app.services.regulatory_reporting.bog_forms import sources_ext  # noqa: F401, PLC0415
    from app.services.regulatory_reporting.bog_forms.sources import (  # noqa: PLC0415
        registered_resolvers,
    )

    registered = set(registered_resolvers())
    unclassified = sorted(registered - set(RESOLVER_AUTHORITY))
    assert not unclassified, (
        "these bog_forms resolvers declare no reporting authority: "
        f"{unclassified}. Add them to provenance.RESOLVER_AUTHORITY."
    )
    stale = sorted(set(RESOLVER_AUTHORITY) - registered)
    assert not stale, f"provenance.RESOLVER_AUTHORITY names unregistered resolvers: {stale}"


def test_every_line_map_binding_resolves_to_a_declared_authority() -> None:
    """Sweep every declared line of every registered form, not just the registry."""
    for spec in all_form_specs():
        for sheet in spec.sheets:
            for line in sheet.lines:
                if line.source is None:
                    continue
                assert line.source in RESOLVER_AUTHORITY, (
                    f"{spec.code}/{sheet.name} line {line.code} binds resolver "
                    f"{line.source!r}, which declares no reporting authority"
                )


def test_engine_backed_resolvers_are_the_declared_form_engine_overlap() -> None:
    """The audit's "which BoG cells overlap an engine?" answered as data.

    These are exactly the bindings where a template cell and a calculation
    engine both produce a figure, so they are the population an equivalence
    proof has to cover. Pinned so the set cannot grow unnoticed.
    """
    assert {
        "run.metric",
        "bsd5.run_line",
        "bsd5.avg_gross_income",
        "bsd13.nop",
        "bsd1b.nop",
    } == ENGINE_BACKED_RESOLVERS
    for name in ENGINE_BACKED_RESOLVERS:
        assert RESOLVER_AUTHORITY[name] is ReportAuthority.ENGINE_RUN


# ---------------------------------------------------------------------------
# 2. BSD2A — audit CF-3
# ---------------------------------------------------------------------------


def test_bsd2a_official_grid_carries_no_formula_cell_to_use() -> None:
    """The premise of BSD2A's disposition, verified against the official layout.

    BSD2A's committed layout is a blank grid: BoG published no formula for the
    "%age of exposure to net worth" column, so there is nothing to evaluate and
    nothing may be invented. That is *why* the ratio is computed by the platform
    under the Guide's own words rather than by the workbook evaluator.
    """
    layout = load_layout("BSD2A")
    formula_cells = [(s.name, c.ref) for s in layout.sheets for c in s.formula_cells]
    assert formula_cells == [], (
        "BSD2A now has template formula cells; if BoG published them, the ratio must be "
        f"evaluated from the template rather than computed: {formula_cells[:5]}"
    )


def test_bsd2a_ratio_binding_is_guide_authoritative() -> None:
    """The ratio's authority is the Guide paragraph, declared, not implied."""
    assert RESOLVER_AUTHORITY["bsd2a.form_cells_ratio_pct"] is ReportAuthority.GUIDE_INSTRUCTION
    spec = form_spec("BSD2A")
    ratio_lines = [
        line
        for sheet in spec.sheets
        for line in sheet.lines
        if line.source == "bsd2a.form_cells_ratio_pct"
    ]
    assert ratio_lines, "BSD2A no longer binds the percentage-of-net-worth column"
    for line in ratio_lines:
        # The citation travels with the binding, so the authority is auditable
        # from the line map alone.
        assert "5(vii)" in line.notes, line.notes
        # A percentage is already in the sheet's unit and must not be scaled.
        assert line.unscaled is True


def test_bsd2a_sum_binding_is_only_a_selection_of_computed_cells() -> None:
    """``form_cells_sum`` aggregates BSD2 cells; it applies no new BoG rule."""
    assert RESOLVER_AUTHORITY["bsd2a.form_cells_sum"] is ReportAuthority.TEMPLATE_INPUT_MAPPING
    # Every category row it serves names BSD2 foreign-column cells, i.e. figures
    # BSD2's own template already produced.
    assert bsd2a_linemap.BSD2_FOREIGN_CELLS
    for refs in bsd2a_linemap.BSD2_FOREIGN_CELLS.values():
        assert all(ref.startswith("C") for ref in refs), refs


# ---------------------------------------------------------------------------
# 3. the package record — template-authoritative made explicit
# ---------------------------------------------------------------------------


def test_bog_form_package_declares_template_authority_not_an_empty_lineage(
    db_client: TestClient,
) -> None:
    _materialize()
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, "BSD2", reporting_date)
    detail = _detail(db_client, package["id"])

    # source_runs stays empty — that is the truth, not a hole.
    assert detail["source_runs"] == []

    prov = detail["snapshot"]["provenance"]
    assert prov["authority"] == ReportAuthority.TEMPLATE_FORMULA.value
    assert "template-authoritative" in prov["source_runs_rationale"]
    assert "no calculation run" in prov["source_runs_rationale"]

    # ... and the versioned identity of everything between data and filed cell.
    assert prov["template_hash"] == template_digest("BSD2")
    assert prov["template_hash"].startswith("sha256:")
    assert prov["mapping_version"] == mapping_digest("BSD2")
    assert prov["mapping_version"].startswith("sha256:")
    assert prov["formula_evaluator_version"] == FORMULA_EVALUATOR_VERSION
    assert prov["official_workbook"] == form_spec("BSD2").workbook
    assert prov["effective_date"] == reporting_date
    assert prov["regulator"] == "BOG"
    assert prov["policy_resolution"]["jurisdiction_code"] == "GH"
    assert prov["policy_resolution"]["base_currency"] == "GHS"

    # The per-field tally: BoG's own formulas own the derived cells.
    counts = prov["authority_counts"]
    layout = load_layout("BSD2")
    assert counts[ReportAuthority.TEMPLATE_FORMULA.value] == sum(
        len(s.formula_cells) for s in layout.sheets
    )
    assert counts[ReportAuthority.TEMPLATE_INPUT_MAPPING.value] > 0


def test_every_bog_form_row_and_section_states_who_owns_it(db_client: TestClient) -> None:
    _materialize()
    reporting_date = _latest_period_end(db_client)
    detail = _detail(db_client, _generate(db_client, "BSD2A", reporting_date)["id"])
    valid = {a.value for a in ReportAuthority}
    seen: set[str] = set()
    for section in detail["snapshot"]["sections"]:
        assert section["authority"] in valid, section["code"]
        for row in section["rows"]:
            assert row["authority"] in valid, row["code"]
            seen.add(row["authority"])
    # BSD2A is the form that proves the point: one sheet, three authorities.
    assert ReportAuthority.GUIDE_INSTRUCTION.value in seen
    assert ReportAuthority.TEMPLATE_INPUT_MAPPING.value in seen
    assert ReportAuthority.NOT_YET_SOURCED.value in seen


def test_engine_backed_package_records_the_full_run_provenance(db_session: Session) -> None:
    """A run-backed return records everything WS-A's CalculationProvenance knows.

    Crucially the DB column keeps its historical four-key shape — adopting the
    formal provenance interface moves no stored lineage — while the rich record
    (parameter digest, schema versions, scenario, actor, computed_at) rides in
    the snapshot.
    """
    _seed_with_baseline_run(db_session)
    package = _generate_service(db_session, "LCR-NSFR")

    assert package.source_runs, "a run-backed return must carry its lineage"
    for entry in package.source_runs:
        assert set(entry) == {"module", "run_id", "input_hash", "engine_version"}

    prov = package.snapshot["provenance"]
    assert prov["authority"] == ReportAuthority.ENGINE_RUN.value
    assert prov["source_runs_rationale"] is None
    assert prov["source_runs"]
    for entry in prov["source_runs"]:
        assert entry["parameter_digest"]
        assert entry["input_schema_version"]
        assert entry["output_schema_version"]
        assert entry["computed_at"]
        assert entry["provenance_complete"] is True
        assert entry["filable"] is True
    assert prov["fact_generation"]["input_hashes"]
    assert prov["parameter_version"] != "not_applicable:no_governed_parameters"
    assert prov["calculation_version"]
    assert prov["effective_date"] == REPORTING_DATE.isoformat()
    # Audit CF-1: the package states WHICH lcr_pct it is reporting — and the
    # id it states RESOLVES. Asserting the string against itself (what this
    # line did until 2026-08-22, audit D-8) cannot tell a real methodology
    # from a typo, which is how ``basel_bog_bsd3`` survived.
    declared = {n["metric_id"]: n for n in prov["declared_methodologies"]}
    lcr_note = declared["lcr_pct"]
    assert lcr_note["registry_status"] == "registered", lcr_note
    authority = authority_registry.get(
        "lcr_pct", regime=Regime.CRD_BASEL, methodology_id=lcr_note["methodology_id"]
    )
    assert authority.is_primary is True
    assert "LCR-NSFR" in authority.reporting_mappings
    # Every section inherits the package authority, so no field lacks one.
    assert all(
        section["authority"] == ReportAuthority.ENGINE_RUN.value
        for section in package.snapshot["sections"]
    )


def test_the_source_run_entry_shape_is_byte_identical_to_the_legacy_writer(
    db_session: Session,
) -> None:
    """Adopting WS-A's primitive must move no stored lineage.

    ``generation._source_run_entry`` now delegates to
    ``CalculationProvenance.source_run_entry()``. If that ever diverged from the
    historical dict, every package's ``snapshot_sha256`` and ``content_digest``
    would shift silently.
    """
    _seed_with_baseline_run(db_session)
    run = db_session.scalars(
        select(RegulatoryRun).where(
            RegulatoryRun.organization_id == DEMO_ORG_ID,
            RegulatoryRun.bank_id == SAMPLE_BANK_ID,
            RegulatoryRun.module == "liquidity",
            RegulatoryRun.status == "succeeded",
        )
    ).first()
    assert run is not None
    assert generation.source_run_entry(run) == {
        "module": run.module,
        "run_id": str(run.id),
        "input_hash": run.input_hash,
        "engine_version": run.engine_version,
    }


# ---------------------------------------------------------------------------
# 3b. package immutability — the export must render the snapshot, never recompute
# ---------------------------------------------------------------------------


def test_bog_form_export_renders_the_sealed_snapshot_and_never_recomputes(
    db_session: Session,
) -> None:
    """The audit's package-immutability chain, proved by removing the inputs.

    Generate a BoG form, then delete every canonical fact the form was resolved
    from, then export. If the exporter recomputed, the workbook would come back
    blank. It comes back with the approved figures, because
    ``FormResult.from_snapshot`` reads the sealed cells — and the snapshot seal
    is unchanged by the export.
    """
    materialize_canonical_test_book(db_session)
    package = _generate_service(db_session, "BSD2")
    seal = package.snapshot_sha256
    assert seal == generation.snapshot_content_hash(package.snapshot)
    before = render_bog_form_xlsx("BSD2", package.snapshot, _sample_bank(db_session), _FIXED_TS)
    cells_before = _numeric_cells(before)
    assert cells_before, "the fixture book must produce at least one figure"

    deleted = (
        db_session.query(BankFinancialFact)
        .filter(
            BankFinancialFact.organization_id == DEMO_ORG_ID,
            BankFinancialFact.bank_id == SAMPLE_BANK_ID,
        )
        .delete(synchronize_session=False)
    )
    assert deleted, "the fixture book must have had facts to remove"
    db_session.flush()

    after = render_bog_form_xlsx("BSD2", package.snapshot, _sample_bank(db_session), _FIXED_TS)
    assert _numeric_cells(after) == cells_before
    assert package.snapshot_sha256 == seal
    assert generation.snapshot_content_hash(package.snapshot) == seal


def test_every_declared_methodology_resolves_in_the_authority_registry() -> None:
    """The completeness gate the CF-1 declarations never had (audit D-8).

    A ``declared_methodologies`` entry is only worth the record it writes if the
    id names a methodology the authority registry actually knows. When it does
    not, ``declared_methodology_notes`` degrades silently to
    ``registry_status: not_registered`` — no regime, no engine, and, decisively,
    NO divergence block — and the filed package discloses nothing while every
    test stays green. That is exactly what happened: LCR-NSFR declared
    ``basel_bog_bsd3``, an id registered for no metric, and the two tests
    covering it compared the literal to the same literal.

    This walks the whole reporting registry, so a THIRD return cannot repeat it.
    """
    unresolved: list[str] = []
    checked = 0
    for definition in REGISTRY.values():
        for note in declared_methodology_notes(definition.declared_methodologies):
            checked += 1
            if note["registry_status"] != "registered":
                unresolved.append(
                    f"{definition.code}: {note['metric_id']} -> "
                    f"{note['methodology_id']} ({note['registry_status']}; "
                    f"registered alternates: {note['alternate_methodologies']})"
                )
    assert checked, "no return declares a methodology — the gate would be vacuous"
    assert not unresolved, (
        "every declared (metric_id, methodology_id) must name a methodology in "
        "app.domain.authority.registry, or the declaration discloses nothing: "
        + "; ".join(unresolved)
    )


def test_the_two_lcr_methodologies_are_declared_and_never_equated() -> None:
    """Audit CF-1 recorded as data, with no equality assertion anywhere.

    BOTH methodologies cap inflows — LCR-NSFR once across the whole book at the
    governed ``lcr_inflow_cap_pct`` threshold, LMT Table 11 separately per
    currency at a hard-coded 75%. Both are correct under their own authority,
    and neither is "the uncapped one". This test pins
    that they are declared as DIFFERENT methodologies — it deliberately does not
    compare the two numbers, and no test should.

    Rewritten 2026-08-22 (audit D-8). It previously asserted each declared id
    against a hard-coded copy of itself, which proved nothing and let
    ``basel_bog_bsd3`` — registered nowhere — pass as the flagship LCR
    methodology. Both sides now RESOLVE against the authority registry, and the
    divergence between them is read from the registry rather than assumed.
    """
    lcr_nsfr = dict(REGISTRY["LCR-NSFR"].declared_methodologies)
    lmt = dict(REGISTRY["LMT"].declared_methodologies)
    assert lcr_nsfr["lcr_pct"] != lmt["lcr_pct"]

    uncapped = authority_registry.get(
        "lcr_pct", regime=Regime.CRD_BASEL, methodology_id=lcr_nsfr["lcr_pct"]
    )
    capped = authority_registry.get("lcr_pct", regime=Regime.LMTD, methodology_id=lmt["lcr_pct"])
    assert uncapped.is_primary is True
    assert capped.is_primary is False
    # The registry — not this test — states that the two differ and why.
    assert capped.divergence is not None
    assert capped.divergence.versus_methodology_id == uncapped.methodology_id


def test_bsd5a_declares_the_form_ratio_as_its_own_methodology() -> None:
    """BoG's E70 = E25/E69 is a declared alternate, not a broken engine tie-out.

    Its inequality with the capital engine's ``car_pct`` is pinned by
    ``tests/services/bog_forms/test_bsd5.py`` ("by construction, not by
    accident"). Declaring the methodology here is what makes that inequality a
    documented divergence rather than an unexplained disagreement.

    Rewritten 2026-08-22 (audit D-8) for the same reason as the LCR pair: the
    declared id is now resolved against the authority registry instead of being
    compared to a copy of itself.
    """
    declared = dict(REGISTRY["BSD5A"].declared_methodologies)
    assert set(declared) == {"car_pct"}
    authority = authority_registry.get(
        "car_pct", regime=Regime.CRD_BASEL, methodology_id=declared["car_pct"]
    )
    assert authority.is_primary is False
    assert authority.divergence is not None
    assert any(cell.startswith("BSD5A!") for cell in authority.reporting_mappings)


# ---------------------------------------------------------------------------
# 4. the unit contract (P0-24 backend half — consumed by WS-D-ui)
# ---------------------------------------------------------------------------


def test_generic_sections_and_totals_carry_a_unit(db_session: Session) -> None:
    """Every section and every total carries ``unit``, in the GENERIC families.

    P0-24: the UI's promise that units are shown per section held only for BSD
    forms, whose sheets each declare an official unit. The generic builders
    emitted no unit key at all, so headline ratios lost their ``%`` on the way
    into the snapshot — and the prior-period comparative section, which already
    read ``total["unit"]``, always found nothing. LCR-NSFR is the case that
    matters: a snapshot mixing cedi amounts with a percentage ratio.
    """
    _seed_with_baseline_run(db_session)
    snapshot = _generate_service(db_session, "LCR-NSFR").snapshot
    valid_kinds = {"currency", "percent", "count", "text", "ratio", "years", "mixed", ""}

    for section in snapshot["sections"]:
        assert "unit" in section, section["code"]
        assert section["unit_kind"] in valid_kinds, section["code"]
        if section["total"] is not None:
            assert "unit" in section["total"], section["code"]
            assert "unit_kind" in section["total"], section["code"]

    totals = {total["code"]: total for total in snapshot["totals"]}
    assert totals, "LCR-NSFR must publish headline totals"
    for total in totals.values():
        assert "unit" in total, total["code"]
        assert "unit_kind" in total, total["code"]
    # The specific regression: the headline ratio keeps its percentage identity.
    assert totals["lcr_pct"]["unit"] == "pct"
    assert totals["lcr_pct"]["unit_kind"] == "percent"
    assert totals["hqla_total_ghs"]["unit"] == "ghs"
    assert totals["hqla_total_ghs"]["unit_kind"] == "currency"

    # A cross-footed section total now carries its unit too, which is what the
    # comparative section reads.
    hqla = next(s for s in snapshot["sections"] if s["code"] == "hqla")
    assert hqla["total"]["equals_sum_of_rows"] is True
    assert hqla["total"]["unit_kind"] in valid_kinds


def test_bog_form_sections_keep_the_official_unit_convention(db_client: TestClient) -> None:
    """The BSD half must not regress: the Guide's own unit survives verbatim."""
    _materialize()
    reporting_date = _latest_period_end(db_client)
    detail = _detail(db_client, _generate(db_client, "BSD2", reporting_date)["id"])
    for section in detail["snapshot"]["sections"]:
        assert section["unit"], section["code"]
        assert section["unit_kind"], section["code"]


def test_snapshot_builders_emit_the_unit_and_authority_keys() -> None:
    """The contract itself, at the builder — independent of any one return."""
    rows = [
        generation.snapshot_row("a", "A", "1", unit="ghs"),
        generation.snapshot_row("b", "B", "2", unit="ghs"),
    ]
    section = generation.snapshot_section("s", "S", rows)
    assert section["unit"] == "ghs"  # inferred unanimously from the rows
    assert section["authority"] is None  # stamped at package level

    mixed = generation.snapshot_section(
        "m",
        "M",
        [
            generation.snapshot_row("a", "A", "1", unit="ghs"),
            generation.snapshot_row("r", "R", "2", unit="pct"),
        ],
    )
    assert mixed["unit"] == "mixed"

    assert generation.snapshot_section("e", "E", [])["unit"] == ""
    assert generation.snapshot_section("x", "X", rows, unit="pct")["unit"] == "pct"

    total = generation.snapshot_total("t", "T", "13.5", unit="pct")
    assert total["unit"] == "pct"
    assert total["unit_kind"] == "percent"
    assert generation.snapshot_total("u", "U", "1")["unit"] == ""
    assert generation.snapshot_total("u", "U", "1")["unit_kind"] == ""


def test_the_two_unit_vocabularies_normalise_onto_one_kind() -> None:
    """NEW-8: one key a renderer can switch on, without losing official scale.

    The generic families say WHAT is measured; a BoG sheet says at WHAT SCALE
    the official form reports. Both survive in ``unit``; ``unit_kind`` is the
    normalisation. An unrecognised unit must resolve to unknown, never to
    currency — guessing money for an unlabelled figure is how a percentage ends
    up with a currency symbol on a filed return.
    """
    assert generation.unit_kind("ghs") == "currency"
    assert generation.unit_kind("millions") == "currency"
    assert generation.unit_kind("thousands") == "currency"
    assert generation.unit_kind("units") == "currency"
    assert generation.unit_kind("pct") == "percent"
    assert generation.unit_kind("percent") == "percent"
    assert generation.unit_kind("count") == "count"
    assert generation.unit_kind("text") == "text"
    assert generation.unit_kind("mixed") == "mixed"
    assert generation.unit_kind("") == ""
    assert generation.unit_kind(None) == ""
    assert generation.unit_kind("furlongs") == ""


def test_appendix_ii_does_not_default_its_unit_to_cedis() -> None:
    """A currency literal must never be defaulted onto a filed artifact."""
    from pathlib import Path  # noqa: PLC0415

    source = (
        Path(__file__).resolve().parents[2]
        / "app"
        / "services"
        / "regulatory_reporting"
        / "generation.py"
    )
    text = source.read_text()
    assert 'appendix.get("unit", "GHS' not in text


# ---------------------------------------------------------------------------
# 5. return eligibility — ONE authority (audit ARCH-8)
# ---------------------------------------------------------------------------


def test_calendar_and_generation_read_the_same_eligibility_authority(
    db_session: Session,
) -> None:
    """The point of ARCH-8: one decision function, two consumers."""
    bank = _make_bank(db_session, institution_type="universal_bank")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    resolved = eligibility.resolve_eligibility(db_session, ctx, bank, as_of=_AS_OF)
    assert resolved.institution_class == "bank"
    assert resolved.jurisdiction_code == "GH"

    eligible_codes = {d.code for d in resolved.eligible_definitions()}
    calendar_codes = {
        item.return_code
        for item in calendar.list_obligations(db_session, ctx, bank.id, as_of=_AS_OF).obligations
    }
    # The calendar is a subset: event-driven packs are eligible but mint no
    # periodic obligation. Nothing may appear on the calendar that the
    # eligibility authority did not admit.
    assert calendar_codes <= eligible_codes
    assert calendar_codes


def test_an_ineligible_return_cannot_be_generated(db_session: Session) -> None:
    """Structurally impossible, not merely hidden: the mint site gates on it."""
    sdi = _make_bank(db_session, institution_type="savings_and_loans")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    bank_only = next(d for d in REGISTRY.values() if "sdi" not in d.institution_classes)

    resolved = eligibility.resolve_eligibility(db_session, ctx, sdi, as_of=_AS_OF)
    decision = resolved.decide(bank_only, reporting_date=_AS_OF)
    assert decision.eligible is False
    assert any(c.code == "institution_class" and not c.satisfied for c in decision.criteria)

    with pytest.raises(HTTPException) as exc:
        generation.generate_package(
            db_session,
            ctx,
            sdi.id,
            RegulatoryPackageCreate(return_code=bank_only.code, reporting_date=_AS_OF),
        )
    assert exc.value.status_code == 403
    # Starlette types ``HTTPException.detail`` as ``str | None``; this refusal
    # deliberately carries the STRUCTURED decision, which is the property under
    # test. Cast (the convention in test_report_comparison.py) rather than assert
    # a shape the reader would then have to trust twice.
    detail = cast("dict[str, Any]", exc.value.detail)
    assert detail["error_code"] == "return_not_eligible"
    # Every failed dimension is named, not a single opaque refusal.
    assert detail["decision"]["criteria"]
    assert "class" in detail["message"].lower()


def test_a_refused_package_is_reported_not_silent(db_session: Session) -> None:
    """A refusal writes no audit event and no row — the log line is the only record.

    Package generation used to fail with a bare ``HTTPException``: no log, no
    audit event, no row (``CONDITION_SOURCES[PACKAGE_FAILED]`` read
    ``NOT EMITTED``). A bank blocked on the last day of a filing window left
    nothing an operator could find. The mint site now emits
    ``reporting.package_failed`` carrying the refusal's OWN error code, so the
    reason is queryable rather than reconstructed from a 4xx in an access log.
    """
    sdi = _make_bank(db_session, institution_type="savings_and_loans")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    bank_only = next(d for d in REGISTRY.values() if "sdi" not in d.institution_classes)

    emitted: list[dict[str, Any]] = []

    def _capture(message: Any) -> None:
        extra = dict(message.record["extra"])
        if extra.get("condition") == Condition.PACKAGE_FAILED.value:
            emitted.append(extra)

    sink_id = loguru_logger.add(_capture, level="DEBUG")
    try:
        with pytest.raises(HTTPException):
            generation.generate_package(
                db_session,
                ctx,
                sdi.id,
                RegulatoryPackageCreate(return_code=bank_only.code, reporting_date=_AS_OF),
            )
    finally:
        loguru_logger.remove(sink_id)

    assert len(emitted) == 1, "a refused package must report exactly once"
    extra = emitted[0]
    assert extra["reason"] == "return_not_eligible"
    assert extra["status_code"] == 403
    assert extra["return_code"] == bank_only.code
    assert extra["bank_id"] == sdi.id
    assert extra["organization_id"] == ORG_1
    # No credential-shaped field, and nothing unbounded.
    assert all(isinstance(value, (str, int)) for value in extra.values())


def test_an_eligible_sdi_return_is_not_silently_excluded(db_session: Session) -> None:
    """The other half of ARCH-8: the filter must not swallow a real SDI return.

    The registry contains two public-directive SDI packets. Both must reach the
    calendar and neither may leak to a universal-bank tenant.
    """
    sdi = _make_bank(db_session, institution_type="savings_and_loans")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)

    resolved = eligibility.resolve_eligibility(db_session, ctx, sdi, as_of=_AS_OF)
    assert resolved.coverage_note() is None
    assert {definition.code for definition in resolved.eligible_definitions()} == {
        "SDI-LMT-MONTHLY",
        "SDI-LE-MONTHLY",
        "SDI-STRESS-ANNUAL",
        "SDI-IRRBB-QUARTERLY",
    }

    obligations = calendar.list_obligations(db_session, ctx, sdi.id, as_of=_AS_OF)
    assert {obligation.return_code for obligation in obligations.obligations} == {
        "SDI-LMT-MONTHLY",
        "SDI-LE-MONTHLY",
        "SDI-STRESS-ANNUAL",
        "SDI-IRRBB-QUARTERLY",
    }
    assert obligations.coverage_note is None

    bank = _make_bank(db_session, institution_type="universal_bank")
    bank_view = eligibility.resolve_eligibility(db_session, ctx, bank, as_of=_AS_OF)
    assert not {
        "SDI-LMT-MONTHLY",
        "SDI-LE-MONTHLY",
        "SDI-STRESS-ANNUAL",
        "SDI-IRRBB-QUARTERLY",
    } & {
        definition.code for definition in bank_view.eligible_definitions()
    }
    with pytest.raises(HTTPException) as exc:
        generation.generate_package(
            db_session,
            ctx,
            bank.id,
            RegulatoryPackageCreate(return_code="SDI-LMT-MONTHLY", reporting_date=_AS_OF),
        )
    assert exc.value.status_code == 403


def test_a_return_from_another_jurisdiction_is_refused(db_session: Session) -> None:
    """The jurisdiction dimension is real, not decorative."""
    bank = _make_bank(db_session, institution_type="universal_bank")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    resolved = eligibility.resolve_eligibility(db_session, ctx, bank, as_of=_AS_OF)
    foreign: ReturnDefinition = replace(
        REGISTRY["LMT"], code="TEST-NG-RETURN", jurisdictions=("NG",)
    )
    decision = resolved.decide(foreign, reporting_date=_AS_OF)
    assert decision.eligible is False
    assert any(c.code == "jurisdiction" and not c.satisfied for c in decision.criteria)


def test_a_not_yet_effective_return_is_refused(db_session: Session) -> None:
    """Effective dating is evaluated where the registry establishes a date."""
    bank = _make_bank(db_session, institution_type="universal_bank")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    resolved = eligibility.resolve_eligibility(db_session, ctx, bank, as_of=_AS_OF)
    future = replace(REGISTRY["LMT"], code="TEST-FUTURE-RETURN", effective_from=date(2027, 1, 1))
    decision = resolved.decide(future, reporting_date=_AS_OF)
    assert decision.eligible is False
    assert any(c.code == "effective_date" and not c.satisfied for c in decision.criteria)
    # ... and an unestablished effective date says so rather than passing mute.
    live = resolved.decide(REGISTRY["LMT"], reporting_date=_AS_OF)
    effective = next(c for c in live.criteria if c.code == "effective_date")
    assert effective.satisfied is True
    assert eligibility.NOT_ESTABLISHED in effective.detail


def test_cadence_is_advisory_and_never_refuses_generation(db_session: Session) -> None:
    """A return generated off its cadence anchor is flagged, not refused.

    Cadence establishes WHEN an obligation arises, not WHETHER the institution
    is subject to the return, and banks legitimately dry-run off-anchor. Making
    it blocking would be a new restriction dressed as a correctness fix.
    """
    bank = _make_bank(db_session, institution_type="universal_bank")
    ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
    resolved = eligibility.resolve_eligibility(db_session, ctx, bank, as_of=_AS_OF)
    mid_month = date(2026, 6, 15)
    decision = resolved.decide(REGISTRY["LMT"], reporting_date=mid_month)
    frequency = next(c for c in decision.criteria if c.code == "frequency")
    assert frequency.satisfied is False
    assert decision.eligible is True
    assert decision.advisories


def test_every_registered_return_declares_its_eligibility_dimensions() -> None:
    """No dimension may be inherited silently from a dataclass default.

    The audit found every entry sitting on the default institution class. The
    BSD pack now declares its class and jurisdiction explicitly, and the SDI
    coverage number is asserted so that the day an SDI return is registered, the
    honest-deferral note stops being emitted.
    """
    for definition in REGISTRY.values():
        assert definition.institution_classes, definition.code
        assert definition.jurisdictions == ("GH",), definition.code
        assert definition.regulator == "BOG", definition.code
    coverage = eligibility.registry_class_coverage()
    assert coverage["bank"] == len(REGISTRY) - 4
    assert coverage["sdi"] == 4
