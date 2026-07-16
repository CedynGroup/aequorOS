"""Export engine tests: real packages via the backbone, all three kinds,
storage via the in-memory client (same monkeypatch seam as the market-data
adapter tests)."""

from __future__ import annotations

import io
import zipfile
from datetime import date
from decimal import Decimal
from uuid import UUID

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.models import (
    Bank,
    BankReportingPeriod,
    RegulatoryPackage,
    RegulatoryPackageArtifact,
)
from app.schemas.regulatory_liquidity import RegulatoryRunCreate
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.services import regulatory_liquidity
from app.services.regulatory_reporting import generation
from app.services.regulatory_reporting.exports import export_package
from app.services.regulatory_reporting.registry import REGISTRY
from app.services.regulatory_reporting.templates import (
    CURRENCY_UNIT_DIVISOR,
    CURRENCY_UNIT_NOTE,
    get_template,
)
from app.services.sample_bank_seed import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    seed_sample_bank,
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
REPORTING_DATE = date(2026, 3, 31)


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _seed_with_baseline_run(db: Session) -> None:
    seed_sample_bank(db)
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    run = regulatory_liquidity.create_liquidity_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=period_id, scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded"


def _generate(db: Session, return_code: str = "BSD3") -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=return_code, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _artifact_count(db: Session, package_id: UUID) -> int:
    return (
        db.scalar(
            select(func.count())
            .select_from(RegulatoryPackageArtifact)
            .where(RegulatoryPackageArtifact.package_id == package_id)
        )
        or 0
    )


def _read_output(db: Session, storage: InMemoryStorageClient, object_path: str) -> bytes:
    slug = db.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug, "export_package assigns the bank storage slug on first use"
    for obj in storage.list(slug, "outputs"):
        if obj.location.object_path == object_path:
            _, stream = storage.read(obj.location)
            return stream.read()
    raise AssertionError(f"No stored object at {object_path}")


def test_export_all_kinds_upserts_artifacts_with_stable_checksums(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)

    artifacts = {
        kind: export_package(db_session, MAKER, package, kind) for kind in ("xlsx", "csv", "pdf")
    }
    db_session.commit()

    assert _artifact_count(db_session, package.id) == 3
    prefix = f"bog_returns/{REPORTING_DATE.isoformat()}/{package.id}/BSD3"
    assert artifacts["xlsx"].object_path == f"{prefix}.xlsx"
    # BSD3 is multi-section, so the csv-kind artifact is a zip container.
    assert artifacts["csv"].object_path == f"{prefix}.zip"
    assert artifacts["pdf"].object_path == f"{prefix}.pdf"
    for kind, artifact in artifacts.items():
        assert artifact.kind == kind
        assert artifact.size_bytes > 0
        assert len(artifact.checksum_sha256) == 64
        stored = _read_output(db_session, storage, artifact.object_path)
        assert len(stored) == artifact.size_bytes

    first_checksums = {kind: artifact.checksum_sha256 for kind, artifact in artifacts.items()}
    first_ids = {kind: artifact.id for kind, artifact in artifacts.items()}

    # Re-export replaces in place: same rows, same checksums, no duplicates.
    for kind in ("xlsx", "csv", "pdf"):
        again = export_package(db_session, MAKER, package, kind)
        assert again.id == first_ids[kind]
        assert again.checksum_sha256 == first_checksums[kind]
    db_session.commit()
    assert _artifact_count(db_session, package.id) == 3


def test_xlsx_round_trip_metadata_headers_and_totals(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "xlsx")
    payload = _read_output(db_session, storage, artifact.object_path)
    workbook = load_workbook(io.BytesIO(payload))

    metadata = workbook["Return Metadata"]
    pairs = {
        row[0].value: row[1].value
        for row in metadata.iter_rows(min_col=1, max_col=2)
        if row[0].value is not None
    }
    assert pairs["Institution"] == "Sample Bank Ltd"
    assert pairs["Reporting date"] == REPORTING_DATE.isoformat()
    assert pairs["Currency unit"] == CURRENCY_UNIT_NOTE
    assert pairs["Template fidelity"] == "PARTIAL"
    assert pairs["Package version"] == str(package.version)
    assert "solo" in str(pairs["Reporting basis"]).lower()

    template = get_template("bog-bsd3-liquidity-v1")
    assert template is not None
    hqla_layout = template.sections[0]
    assert hqla_layout.section_code == "hqla"
    sheet = workbook["Stock of HQLA"]
    headers = [sheet.cell(row=5, column=idx).value for idx in range(1, 4)]
    assert headers == [column.header for column in hqla_layout.columns]

    snapshot_sections = {s["code"]: s for s in package.snapshot["sections"]}
    hqla = snapshot_sections["hqla"]
    expected_total = Decimal(hqla["total"]["value"]) / CURRENCY_UNIT_DIVISOR
    total_row_idx = 5 + len(hqla["rows"]) + 1
    total_cell = sheet.cell(row=total_row_idx, column=3)
    assert Decimal(str(total_cell.value)) == expected_total
    assert total_cell.font.bold
    assert total_cell.number_format == "#,##0;(#,##0)"

    # Provenance footer sheet carries the source runs and input hashes.
    provenance = workbook["Fidelity & Provenance"]
    text = "\n".join(
        str(cell.value) for row in provenance.iter_rows() for cell in row if cell.value
    )
    assert f"package {package.id}" in text
    assert package.source_runs[0]["input_hash"] in text
    assert package.source_runs[0]["engine_version"] in text


def test_xlsx_export_bytes_are_deterministic(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    first = export_package(db_session, MAKER, package, "xlsx").checksum_sha256
    second = export_package(db_session, MAKER, package, "xlsx").checksum_sha256
    assert first == second


def test_pdf_export_is_valid_and_nonempty(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "pdf")
    payload = _read_output(db_session, storage, artifact.object_path)
    assert payload.startswith(b"%PDF")
    assert len(payload) == artifact.size_bytes
    assert artifact.size_bytes > 1000


def test_csv_zip_parses_with_metadata_sections_and_provenance(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "csv")
    payload = _read_output(db_session, storage, artifact.object_path)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        names = archive.namelist()
        assert "00_metadata.csv" in names
        assert "01_hqla.csv" in names
        assert "99_provenance.csv" in names
        # One numbered file per template section (BSD3 has 7) + metadata + provenance.
        assert len(names) == 9

        metadata_rows = archive.read("00_metadata.csv").decode("utf-8").splitlines()
        assert any("Sample Bank Ltd" in row for row in metadata_rows)
        assert any(CURRENCY_UNIT_NOTE in row for row in metadata_rows)

        hqla_rows = archive.read("01_hqla.csv").decode("utf-8").splitlines()
        assert hqla_rows[0] == "#section,hqla"
        header_idx = next(i for i, row in enumerate(hqla_rows) if row.startswith("Row,"))
        assert hqla_rows[header_idx] == "Row,Item,Amount (GHS '000)"
        snapshot_sections = {s["code"]: s for s in package.snapshot["sections"]}
        expected_total = (
            Decimal(snapshot_sections["hqla"]["total"]["value"]) / CURRENCY_UNIT_DIVISOR
        )
        total_cells = hqla_rows[-1].split(",")
        assert Decimal(total_cells[-1]) == expected_total

        provenance_rows = archive.read("99_provenance.csv").decode("utf-8").splitlines()
        assert any(package.source_runs[0]["input_hash"] in row for row in provenance_rows)


def test_missing_snapshot_section_raises_409(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _ = storage
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    snapshot = dict(package.snapshot)
    snapshot["sections"] = [
        section for section in snapshot["sections"] if section["code"] != "hqla"
    ]
    package.snapshot = snapshot
    db_session.commit()

    with pytest.raises(HTTPException) as exc_info:
        export_package(db_session, MAKER, package, "xlsx")
    assert exc_info.value.status_code == 409
    assert "snapshot_section_missing" in str(exc_info.value.detail)
    assert "hqla" in str(exc_info.value.detail)

    package.snapshot = {}
    db_session.commit()
    with pytest.raises(HTTPException) as empty_info:
        export_package(db_session, MAKER, package, "pdf")
    assert empty_info.value.status_code == 409
    assert "snapshot_empty" in str(empty_info.value.detail)


def test_lmt_return_reuses_liquidity_generator_and_exports_subset(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session, return_code="LMT")
    assert package.snapshot["template_id"] == "bog-lmt-liquidity-v1"
    assert package.snapshot["fidelity"] == "PARTIAL"

    artifact = export_package(db_session, MAKER, package, "xlsx")
    payload = _read_output(db_session, storage, artifact.object_path)
    workbook = load_workbook(io.BytesIO(payload))
    # Metadata + the four LCR-subset sections + provenance; the NSFR sheets
    # and the unfillable LMTD Tables 1-10 are honestly absent (TODO(RR-6)).
    assert workbook.sheetnames == [
        "Return Metadata",
        "Stock of HQLA",
        "Cash Outflows (30 days)",
        "Cash Inflows (30 days)",
        "Liquidity Coverage Ratio Summar",
        "Fidelity & Provenance",
    ]
    assert artifact.object_path.endswith("/LMT.xlsx")


def test_every_registry_entry_has_a_template_with_matching_sections() -> None:
    generator_sections = {
        "liquidity": {
            "hqla",
            "outflows",
            "inflows",
            "lcr_summary",
            "nsfr_asf",
            "nsfr_rsf",
            "nsfr_summary",
        },
        "capital": {
            "cet1",
            "at1",
            "tier2",
            "credit_rwa",
            "market_rwa",
            "operational_rwa",
            "capital_ratios",
        },
        "irrbb": {"repricing_gap", "eve_scenarios", "earnings_at_risk", "summary"},
        "fx": {
            "currency_positions",
            "standalone_var",
            "hedges",
            "scenario_nop",
            "nop_summary",
        },
        "icaap_stress": {"forecast_summary", "forecast_path", "stress_summary"},
    }
    for definition in REGISTRY.values():
        template = get_template(definition.template_id)
        assert template is not None, definition.template_id
        assert template.return_code == definition.code
        assert template.fidelity in ("CONFIRMED", "PARTIAL", "REPRESENTATIVE")
        assert template.source_citation
        section_codes = {layout.section_code for layout in template.sections}
        # Every layout must map to a section the generator actually emits —
        # templates never reference (or fabricate) data that does not exist.
        assert section_codes <= generator_sections[definition.generator]
        for layout in template.sections:
            assert layout.fidelity in ("CONFIRMED", "PARTIAL", "REPRESENTATIVE")
            assert layout.source_citation
            if layout.fidelity == "REPRESENTATIVE":
                assert layout.layout_id.endswith("_representative")
