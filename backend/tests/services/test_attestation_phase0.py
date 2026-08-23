"""Attestation Phase 0 prerequisites (docs/attestation_esignature.md §1.6, §3.1).

Five gaps, all exercised against real packages through the real generator and
the real exporter (in-memory storage seam, same as
test_regulatory_reporting_exports.py):

* **G13** — every package carries a ``content_digest`` that is stable across
  regenerations of identical figures, unlike ``snapshot_sha256``.
* **G16** — the corporate (LRT) packs, which bind no engine run, carry a
  ``register_state_digest`` that moves when reportable register data moves and
  stands still when only bookkeeping changes.
* **G3** — a drifted snapshot is refused at export instead of rendering.
* **G2** — every export appends an immutable artifact-version row pinning the
  object-store version, and a signed version blocks re-export.
* **G9 / G14** — a consolidated package renders as consolidated in all three
  formats, and a granted solo resubmission does not bump the consolidated
  ORASS revision.
"""

from __future__ import annotations

import base64
import io
import re
import zipfile
import zlib
from copy import deepcopy
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from unittest.mock import patch
from uuid import UUID, uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.models import (
    AttestationSignature,
    Bank,
    BankLicense,
    BankNameHistory,
    BankProduct,
    BankReportingPeriod,
    InstitutionProfile,
    Outlet,
    RegulatoryArtifactVersion,
    RegulatoryPackage,
    RegulatoryResubmissionRequest,
    RelatedParty,
    RelatedPartyRole,
    Shareholding,
)
from app.schemas.institution_profile import OutletCreate, OutletUpdate
from app.schemas.regulatory_liquidity import RegulatoryRunCreate
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.services import institution_profile, regulatory_liquidity
from app.services.attestation import digests, register_state
from app.services.regulatory_reporting import exports as reporting_exports
from app.services.regulatory_reporting import generation, workflow
from app.services.regulatory_reporting.exports import export_package
from app.services.regulatory_reporting.registry import REGISTRY
from app.services.regulatory_reporting.templates import CONSOLIDATED_BASIS
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)
from tests.services.test_lrt_packs import (
    _seed_full_register,  # pyright: ignore[reportPrivateUsage]
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
REPORTING_DATE = date(2026, 3, 31)


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _seed_with_baseline_run(db: Session) -> None:
    materialize_canonical_test_book(db)
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    run = regulatory_liquidity.create_liquidity_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=period_id, scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded"


def _rerun_liquidity(db: Session) -> None:
    """Re-execute the baseline liquidity engine over the ALREADY-seeded book."""
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    run = regulatory_liquidity.create_liquidity_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=period_id, scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded"


def _generate(
    db: Session, return_code: str = "LCR-NSFR", *, basis: str = "solo"
) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(
            return_code=return_code,
            reporting_date=REPORTING_DATE,
            basis=basis,  # pyright: ignore[reportArgumentType]
        ),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _artifact_versions(db: Session, package_id: UUID) -> list[RegulatoryArtifactVersion]:
    return list(
        db.scalars(
            select(RegulatoryArtifactVersion)
            .where(RegulatoryArtifactVersion.package_id == package_id)
            .order_by(RegulatoryArtifactVersion.created_at, RegulatoryArtifactVersion.id)
        )
    )


def _bank_slug(db: Session) -> str:
    slug = db.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    return slug


def _read_output(db: Session, storage: InMemoryStorageClient, object_path: str) -> bytes:
    for obj in storage.list(_bank_slug(db), "outputs"):
        if obj.location.object_path == object_path:
            _, stream = storage.read(obj.location)
            return stream.read()
    raise AssertionError(f"No stored object at {object_path}")


def _pdf_text(payload: bytes) -> bytes:
    """Concatenate a reportlab PDF's page streams as decoded text.

    reportlab writes page content through ``ASCII85Decode`` + ``FlateDecode``,
    so both filters have to be undone before the drawn strings are readable.
    """
    text = bytearray()
    for body in re.findall(rb"stream\r?\n(.*?)endstream", payload, re.DOTALL):
        stream = bytes(body).strip(b"\r\n")
        try:
            text.extend(zlib.decompress(base64.a85decode(stream, adobe=True)))
        except (ValueError, zlib.error):
            text.extend(stream)
    return bytes(text)


# ---------------------------------------------------------------------------
# G13 — content_digest on every return
# ---------------------------------------------------------------------------


def test_content_digest_is_sealed_and_survives_regeneration(db_session: Session) -> None:
    _seed_with_baseline_run(db_session)
    first = _generate(db_session)
    assert first.content_digest == digests.content_digest(first.snapshot)
    # Engine-backed returns bind runs, so they carry no register-state digest.
    assert first.source_runs != []
    assert first.register_state_digest is None

    second = _generate(db_session)
    assert second.version == first.version + 1
    # Identical figures, a later generated_at: the version seal moves, the
    # content fingerprint does not. That difference is the whole point of G13.
    assert second.snapshot_sha256 != first.snapshot_sha256
    assert second.content_digest == first.content_digest


def test_content_digest_survives_a_source_run_rerun(db_session: Session) -> None:
    """The OTHER volatility axis: re-executing the engines over unchanged data.

    Audit 2026-08-22 D-16. ``_stamp_provenance`` writes ``snapshot["provenance"]``
    before the package is sealed, and every ``source_runs`` entry carries the
    executing run's ``run_id``, ``computed_at`` and ``actor_id``. A rerun over an
    unchanged book mints a new run row, so those three move while the
    value-based ``input_hash`` and every reported figure stand still. Before the
    fix ``strip_volatile_fields`` reached only ``snapshot["metadata"]``, so the
    content fingerprint moved with the execution — destroying the one property
    it exists for, on the value signers sign.
    """
    _seed_with_baseline_run(db_session)
    first = _generate(db_session)
    first_run_ids = [entry["run_id"] for entry in first.snapshot["provenance"]["source_runs"]]
    first_hashes = [entry["input_hash"] for entry in first.snapshot["provenance"]["source_runs"]]
    first_digest = first.content_digest
    assert first_run_ids, "LCR-NSFR binds engine runs — that is the axis under test"

    # A genuinely new execution of the same engine over the same canonical book:
    # the run row is re-minted, the book is not re-seeded.
    _rerun_liquidity(db_session)
    second = _generate(db_session)
    second_prov = second.snapshot["provenance"]["source_runs"]

    assert [entry["run_id"] for entry in second_prov] != first_run_ids
    assert [entry["input_hash"] for entry in second_prov] == first_hashes
    assert second.snapshot_sha256 != first.snapshot_sha256  # the version seal moves
    assert second.content_digest == first_digest  # the content fingerprint does not


def test_content_digest_excludes_run_identity_and_nothing_else(db_session: Session) -> None:
    """Both directions of the exclusion, on a REAL package snapshot.

    The stored snapshot keeps the full provenance block — run identity, timing
    and actor are the evidence a supervisor needs. Only the digest INPUT drops
    them, and it drops nothing else: ``input_hash`` is value-based and is
    therefore content, so tampering with it must still move the digest.
    """
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    snapshot = deepcopy(package.snapshot)
    baseline = digests.content_digest(snapshot)
    assert baseline == package.content_digest

    entries = snapshot["provenance"]["source_runs"]
    assert entries
    for entry in entries:
        # The stored evidence is present — it is excluded, not deleted.
        assert {"run_id", "computed_at", "actor_id"} <= set(entry)
        entry["run_id"] = str(uuid4())
        entry["computed_at"] = "2099-01-01T00:00:00+00:00"
        entry["actor_id"] = str(uuid4())
    assert digests.content_digest(snapshot) == baseline

    entries[0]["input_hash"] = "0" * 64
    assert digests.content_digest(snapshot) != baseline


def test_content_digest_moves_when_a_reported_figure_moves(db_session: Session) -> None:
    """The exclusion did not blunt the fingerprint: figures still move it."""
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    baseline = digests.content_digest(package.snapshot)

    moved = deepcopy(package.snapshot)
    row = next(
        row
        for section in moved["sections"]
        for row in section.get("rows", [])
        if _is_numeric(row.get("value"))
    )
    row["value"] = str(Decimal(str(row["value"])) + Decimal("1"))
    assert digests.content_digest(moved) != baseline


def _is_numeric(value: object) -> bool:
    if value is None:
        return False
    try:
        Decimal(str(value))
    except InvalidOperation:
        return False
    return True


def test_every_registered_return_family_seals_a_content_digest(db_session: Session) -> None:
    _seed_full_register(db_session)
    period_id = db_session.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    run = regulatory_liquidity.create_liquidity_run(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=period_id, scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded"

    # One engine-backed return plus all five master-data packs: the two binding
    # classes of §1.7 that Phase 0 has to seal.
    for return_code in (
        "LCR-NSFR",
        "LRT-PROFILE",
        "LRT-OUTLET",
        "LRT-PARTY",
        "LRT-CAPITAL",
        "LRT-PRODUCT",
    ):
        package = _generate(db_session, return_code)
        assert package.content_digest == digests.content_digest(package.snapshot), return_code
        if package.source_runs:
            assert package.register_state_digest is None, return_code
        else:
            assert package.register_state_digest is not None, return_code


def test_all_registered_returns_route_through_the_single_sealing_site() -> None:
    """Structural guard for G13: there is one place a package is minted.

    ``generate_package`` is the only ``RegulatoryPackage(...)`` construction in
    ``app/``, and it seals ``content_digest`` unconditionally — so proving every
    registered return dispatches through it proves all of them are sealed,
    without seeding an engine run per return. Count pinned so a new return
    family must consciously join this guard: 39 = the 15 pre-BSD entries (the
    original thirteen incl. the recoded CAR-RWA / LCR-NSFR, STRESS-PACK (item 6),
    LAS-QUARTERLY (item 14, still template-gated); BSD-MONTHLY retired) + the
    23 official Bank of Ghana BSD forms (family "bsd", generator "bog_form",
    2026-08-15 — every one dispatches through generate_package and is sealed) +
    ICAAP-STRESS-APPENDIX2 (family "icaap_stress", generator "icaap_stress",
    2026-08-20 — the annual enterprise-wide stress return in the Appendix II
    Tables 1–6 formats; docs/stress.md §3.6, also routed through generate_package).
    """
    generators = generation._GENERATORS  # pyright: ignore[reportPrivateUsage]
    # 39 bank returns + the 4 SDI reports added 2026-08-22 (family "sdi":
    # SDI-LMT-MONTHLY, SDI-IRRBB-QUARTERLY, SDI-LE-MONTHLY, SDI-STRESS-ANNUAL).
    assert len(REGISTRY) == 43  # noqa: PLR2004
    assert sum(1 for d in REGISTRY.values() if d.family == "bsd") == 23  # noqa: PLR2004
    assert sum(1 for d in REGISTRY.values() if d.family == "sdi") == 4  # noqa: PLR2004
    for definition in REGISTRY.values():
        assert definition.generator in generators, definition.code


# ---------------------------------------------------------------------------
# G16 — register_state_digest for the master-data packs
# ---------------------------------------------------------------------------


def test_register_state_rows_carry_business_fields_only(db_session: Session) -> None:
    _seed_full_register(db_session)
    rows = register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID)
    tables = {row["table"] for row in rows}
    assert tables == {
        "institution_profiles",
        "related_parties",
        "related_party_roles",
        "shareholdings",
        "outlets",
        "bank_products",
        "bank_licenses",
        "bank_name_history",
    }
    # The declared table list must match what the projection actually emits.
    assert set(register_state.REGISTER_TABLES) == tables
    for row in rows:
        assert set(row) == {"table", "id", "updated_at", "values"}
        values: dict[str, Any] = row["values"]
        # Bookkeeping must not ride in ``values``: the digest has to move only
        # when reportable data moves.
        assert not {"id", "organization_id", "bank_id", "created_at", "updated_at"} & set(values)
    # Every projected value is JSON-native — canonical_json has no default=.
    assert digests.register_state_digest(rows)


def test_projection_covers_every_business_column_of_every_register_table(
    db_session: Session,
) -> None:
    """A column added to the register but not to the projection would be an
    invisible change: reportable data could move without moving the digest.
    This compares the projection against the mapped columns so a future schema
    addition fails here rather than silently weakening a signature."""
    _seed_full_register(db_session)
    projected: dict[str, set[str]] = {}
    for row in register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID):
        projected.setdefault(str(row["table"]), set()).update(row["values"])

    bookkeeping = {"id", "organization_id", "bank_id", "created_at", "updated_at"}
    for model in (
        InstitutionProfile,
        RelatedParty,
        RelatedPartyRole,
        Shareholding,
        Outlet,
        BankProduct,
        BankLicense,
        BankNameHistory,
    ):
        table = model.__tablename__
        business = {column.key for column in model.__table__.columns} - bookkeeping
        assert projected[table] == business, table


def test_register_state_digest_tracks_reportable_register_changes(db_session: Session) -> None:
    _seed_full_register(db_session)
    pack = _generate(db_session, "LRT-OUTLET")
    sealed = pack.register_state_digest
    assert sealed is not None
    assert sealed == digests.register_state_digest(
        register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID)
    )

    outlet_id = db_session.scalar(
        select(Outlet.id).where(
            Outlet.organization_id == DEMO_ORG_ID,
            Outlet.bank_id == SAMPLE_BANK_ID,
            Outlet.outlet_number == "BR-014",
        )
    )
    assert outlet_id is not None
    institution_profile.update_outlet(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        outlet_id,
        OutletUpdate(
            reason="Correct the branch name after signage change",
            outlet_type="branch",
            name="Kumasi Adum",
            outlet_number="BR-014",
            address={"city": "Kumasi", "street": "Adum High Street"},
            opened_on=date(2026, 2, 1),
        ),
    )
    # A mutated register is detectable against the sealed pack — which is the
    # provenance Class B otherwise lacks entirely.
    assert (
        digests.register_state_digest(
            register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID)
        )
        != sealed
    )


def test_register_state_digest_ignores_row_order_and_covers_new_rows(
    db_session: Session,
) -> None:
    _seed_full_register(db_session)
    rows = register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID)
    assert digests.register_state_digest(rows) == digests.register_state_digest(
        list(reversed(rows))
    )

    before = digests.register_state_digest(rows)
    institution_profile.create_outlet(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        OutletCreate(
            reason="Open the Takoradi branch",
            outlet_type="branch",
            name="Takoradi Market Circle Branch",
            outlet_number="BR-021",
            address={"city": "Takoradi"},
            opened_on=date(2026, 3, 2),
        ),
    )
    after = digests.register_state_digest(
        register_state.register_state_rows(db_session, MAKER, SAMPLE_BANK_ID)
    )
    assert after != before


def test_register_state_rows_are_tenant_scoped(db_session: Session) -> None:
    _seed_full_register(db_session)
    other_tenant = TenantContext(organization_id="OR-NOSUCH1", actor_user_id=DEMO_USER_ID)
    assert register_state.register_state_rows(db_session, other_tenant, SAMPLE_BANK_ID) == []


def test_master_data_pack_binds_as_master_data_class(db_session: Session) -> None:
    _seed_full_register(db_session)
    pack = _generate(db_session, "LRT-PROFILE")
    assert pack.content_digest is not None
    assert pack.register_state_digest is not None
    # The certification digest a signer will cover: master_data binding is only
    # constructible because the register-state digest now exists (G16).
    digest = digests.certification_digest(
        organization_id=pack.organization_id,
        bank_id=pack.bank_id,
        package_id=str(pack.id),
        package_version=pack.version,
        return_code=pack.return_code,
        reporting_date=pack.reporting_date.isoformat(),
        basis=pack.basis,
        content_digest_value=pack.content_digest,
        binding_class="master_data",
        register_state_digest_value=pack.register_state_digest,
    )
    assert len(digest) == 64
    # The same package cannot bind as engine_run — there is no run to bind to.
    with pytest.raises(ValueError, match="requires at least one source run"):
        digests.certification_digest(
            organization_id=pack.organization_id,
            bank_id=pack.bank_id,
            package_id=str(pack.id),
            package_version=pack.version,
            return_code=pack.return_code,
            reporting_date=pack.reporting_date.isoformat(),
            basis=pack.basis,
            content_digest_value=pack.content_digest,
            binding_class="engine_run",
        )


# ---------------------------------------------------------------------------
# G3 — the snapshot seal is verified at export
# ---------------------------------------------------------------------------


def test_export_refuses_a_snapshot_that_drifted_from_its_seal(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _ = storage
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    drifted = dict(package.snapshot)
    sections = [dict(section) for section in drifted["sections"]]
    hqla = next(section for section in sections if section["code"] == "hqla")
    hqla["rows"] = [{**row, "value": "1"} for row in hqla["rows"]]
    drifted["sections"] = sections
    package.snapshot = drifted
    db_session.commit()

    for kind in ("xlsx", "csv", "pdf"):
        with pytest.raises(HTTPException) as exc_info:
            export_package(db_session, MAKER, package, kind)  # pyright: ignore[reportArgumentType]
        assert exc_info.value.status_code == 409
        assert "snapshot_integrity_failed" in str(exc_info.value.detail)
    assert _artifact_versions(db_session, package.id) == []


def test_export_warns_but_proceeds_when_no_seal_was_ever_recorded(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    package.snapshot_sha256 = None
    db_session.commit()

    with patch.object(
        reporting_exports.logger,
        "warning",
        wraps=reporting_exports.logger.warning,
    ) as warning:
        artifact = export_package(db_session, MAKER, package, "xlsx")
    db_session.commit()
    assert artifact.size_bytes > 0
    assert _read_output(db_session, storage, artifact.object_path)
    warning.assert_called_once()
    assert "snapshot_sha256" in warning.call_args.args[0]


# ---------------------------------------------------------------------------
# G2 — append-only artifact versions
# ---------------------------------------------------------------------------


def test_every_export_appends_an_artifact_version_with_storage_version(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "xlsx")
    db_session.commit()

    versions = _artifact_versions(db_session, package.id)
    assert len(versions) == 1
    version = versions[0]
    assert version.kind == "xlsx"
    assert version.object_path == artifact.object_path
    assert version.checksum_sha256 == artifact.checksum_sha256
    assert version.size_bytes == artifact.size_bytes
    assert version.created_by == DEMO_USER_ID
        # The backend reports an object-store version id, so "the artifact as
    # filed" is resolvable from the database rather than only from the bucket.
    assert version.storage_version_id
    stored = next(
        obj
        for obj in storage.list(_bank_slug(db_session), "outputs")
        if obj.location.object_path == artifact.object_path
    )
    assert stored.version_id == version.storage_version_id

    # Re-export appends rather than overwriting: the upserted artifact row keeps
    # its identity while the version history grows.
    again = export_package(db_session, MAKER, package, "xlsx")
    db_session.commit()
    assert again.id == artifact.id
    assert len(_artifact_versions(db_session, package.id)) == 2


def test_a_signature_pinning_a_version_blocks_re_export(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    """Signedness is DERIVED from a signature referencing the version row.

    There is deliberately no mutable ``signed`` flag: the table is append-only,
    so such a flag could never be written on Postgres and would be a lie. The
    signature reference is the authoritative statement that these exact bytes
    were signed.
    """
    _ = storage
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    export_package(db_session, MAKER, package, "pdf")
    db_session.commit()
    version = _artifact_versions(db_session, package.id)[0]
    
    db_session.add(_signature_over(package, version))
    db_session.commit()

    with pytest.raises(HTTPException) as exc_info:
        export_package(db_session, MAKER, package, "pdf")
    assert exc_info.value.status_code == 409
    assert "artifact_version_signed" in str(exc_info.value.detail)


def _signature_over(
    package: RegulatoryPackage, version: RegulatoryArtifactVersion
) -> AttestationSignature:
    """A minimal signature row pinning one artifact version."""
    assert package.content_digest is not None
    return AttestationSignature(
        organization_id=package.organization_id,
        bank_id=package.bank_id,
        package_id=package.id,
        package_version=package.version,
        signing_role="approver",
        signer_id="SGN-TESTTESTTESTTE",
        signer_user_id=DEMO_USER_ID,
        binding_class="engine_run",
        certification_digest="a" * 64,
        content_digest=package.content_digest,
        statement="Test attestation statement.",
        attestation_payload={"schema": "aequoros-signature-v1"},
        payload_digest="b" * 64,
        signature_method="pades_b_lta",
        signature_value=b"test-signature",
        certificate_pem="-----BEGIN CERTIFICATE-----test-----END CERTIFICATE-----",
        certificate_sha256="c" * 64,
        declared_at=datetime.now(UTC),
        artifact_version_id=version.id,
        prev_hash="0" * 64,
        entry_hash="d" * 64,
    )


# ---------------------------------------------------------------------------
# G9 — the rendered basis is the package's actual basis
# ---------------------------------------------------------------------------


def test_consolidated_package_renders_as_consolidated_in_every_format(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session, basis="consolidated")
    assert package.snapshot["institution"]["basis"] == "consolidated"

    xlsx = export_package(db_session, MAKER, package, "xlsx")
    csv_artifact = export_package(db_session, MAKER, package, "csv")
    pdf = export_package(db_session, MAKER, package, "pdf")
    db_session.commit()

    workbook = load_workbook(io.BytesIO(_read_output(db_session, storage, xlsx.object_path)))
    pairs = {
        row[0].value: row[1].value
        for row in workbook["Return Metadata"].iter_rows(min_col=1, max_col=2)
        if row[0].value is not None
    }
    assert pairs["Reporting basis"] == CONSOLIDATED_BASIS
    assert "solo" not in str(pairs["Reporting basis"]).lower()

    with zipfile.ZipFile(
        io.BytesIO(_read_output(db_session, storage, csv_artifact.object_path))
    ) as archive:
        metadata_csv = archive.read("00_metadata.csv").decode("utf-8")
    assert CONSOLIDATED_BASIS in metadata_csv

    pdf_bytes = _read_output(db_session, storage, pdf.object_path)
    assert pdf_bytes.startswith(b"%PDF")
    # The cover page carries the basis; reportlab compresses the page stream,
    # so the assertion is on the rendered text extracted from the artifact.
    assert b"Consolidated" in _pdf_text(pdf_bytes)


def test_solo_package_keeps_the_act_930_solo_wording(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_with_baseline_run(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "xlsx")
    db_session.commit()
    workbook = load_workbook(io.BytesIO(_read_output(db_session, storage, artifact.object_path)))
    pairs = {
        row[0].value: row[1].value
        for row in workbook["Return Metadata"].iter_rows(min_col=1, max_col=2)
        if row[0].value is not None
    }
    assert str(pairs["Reporting basis"]).startswith("Solo")
    assert "Act 930 s.91(1)" in str(pairs["Reporting basis"])


# ---------------------------------------------------------------------------
# G14 — the revision chain is per basis
# ---------------------------------------------------------------------------


def _grant_resubmission(db: Session, package: RegulatoryPackage) -> None:
    """A granted, unconsumed resubmission request on ``package``.

    Written directly rather than through ``workflow.request_resubmission`` so
    the revision arithmetic is tested without dragging in the whole
    approve → submit → acknowledge path.
    """
    db.add(
        RegulatoryResubmissionRequest(
            organization_id=package.organization_id,
            package_id=package.id,
            reason="The solo return used a stale HQLA extract.",
            status="granted",
            requested_by=DEMO_USER_ID,
            decided_at=datetime.now(UTC),
        )
    )
    db.commit()


def test_granted_solo_resubmission_does_not_bump_the_consolidated_revision(
    db_session: Session,
) -> None:
    _seed_with_baseline_run(db_session)
    solo = _generate(db_session)
    consolidated = _generate(db_session, basis="consolidated")

    revision = workflow._submission_revision  # pyright: ignore[reportPrivateUsage]
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, solo) == "1.0"
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, consolidated) == "1.0"

    _grant_resubmission(db_session, solo)
    # The grant belongs to the solo chain only — before the G14 fix it bumped
    # the consolidated return's ORASS revision as well.
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, solo) == "1.1"
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, consolidated) == "1.0"

    # A later solo version inherits its own chain's revision, not a fresh one.
    solo_v2 = _generate(db_session)
    assert solo_v2.version == 2
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, solo_v2) == "1.1"
    assert revision(db_session, MAKER, SAMPLE_BANK_ID, consolidated) == "1.0"
