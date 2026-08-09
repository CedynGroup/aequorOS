"""W6 return tests (docs/submission_pipeline_plan.md §W6.2–4).

Large Exposures (LE-MONTHLY): canonical counterparty exposures crossing the
10% NOF line, connected-counterparty grouping, exempt classification, the
top-100 cap, %-of-NOF math against the capital run's Tier 1, validate +
export round-trip, and the 409 paths (no baseline capital run / no canonical
positions).

LMT monitoring tools: the contractual maturity-mismatch ladder, top-10
depositor funding concentration and available unencumbered assets appear as
additional sections when canonical data exists.

IRRBB BoG ±450 bp parameterization: the param rows are seeded effective-dated
but the engine computes only the Basel set, so the ±450 return rows appear
ONLY when run metrics actually carry them — asserted both ways.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any
from uuid import UUID

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.db.base import utc_now
from app.models import (
    Bank,
    BankReportingPeriod,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    CanonicalProduct,
    IngestionBatch,
    LineageRecord,
    ParamLiquidityHaircut,
    ParamLiquidityThreshold,
    ParamStressShock,
    RegulatoryPackage,
    RegulatoryRun,
    RelatedParty,
)
from app.schemas.regulatory_irr import IrrScenarioBatchCreate
from app.schemas.regulatory_liquidity import RegulatoryRunCreate
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.services import regulatory_capital, regulatory_irr, regulatory_liquidity
from app.services.regulatory_reporting import calendar, generation, validation
from app.services.regulatory_reporting.exports import export_package
from app.services.regulatory_reporting.registry import REGISTRY
from app.services.sample_bank_seed import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    seed_sample_bank,
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
REPORTING_DATE = date(2026, 3, 31)
PCT = Decimal("0.0001")


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _period_id(db: Session) -> UUID:
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    return period_id


def _run_capital_baseline(db: Session) -> None:
    run = regulatory_capital.create_capital_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="capital", reporting_period_id=_period_id(db), scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded", run
    db.expire_all()


def _run_liquidity_baseline(db: Session) -> None:
    run = regulatory_liquidity.create_liquidity_run(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryRunCreate(
            module="liquidity", reporting_period_id=_period_id(db), scenario_code="baseline"
        ),
    )
    assert run.status == "succeeded", run
    db.expire_all()


def _tier1(db: Session) -> Decimal:
    preview = regulatory_capital.get_bsd2_preview(db, MAKER, SAMPLE_BANK_ID, _period_id(db))
    return Decimal(str(preview.tier1_total.value))


class _CanonicalSeeder:
    """Minimal canonical scaffold (batch + lineage + row builders) at as-of."""

    def __init__(self, db: Session, as_of: date = REPORTING_DATE) -> None:
        self.db = db
        batch = IngestionBatch(
            organization_id=DEMO_ORG_ID,
            bank_id=SAMPLE_BANK_ID,
            source_system="EXCEL_CSV",
            adapter_version="1.0",
            extraction_mode="full",
            status="accepted",
            as_of_date=as_of,
        )
        db.add(batch)
        db.flush()
        lineage = LineageRecord(
            organization_id=DEMO_ORG_ID,
            ingestion_batch_id=batch.id,
            operation_type="ADAPTER_TRANSLATE",
            operation_ref="w6-fixture",
            input_lineage_ids=[],
        )
        db.add(lineage)
        db.flush()
        self.common: dict[str, Any] = {
            "organization_id": DEMO_ORG_ID,
            "bank_id": SAMPLE_BANK_ID,
            "as_of_date": as_of,
            "source_system": "EXCEL_CSV",
            "ingestion_batch_id": batch.id,
            "lineage_id": lineage.id,
            "validation_status": "accepted",
        }

    def counterparty(  # noqa: PLR0913 - keyword-only fixture builder
        self,
        ref: str,
        name: str,
        counterparty_type: str,
        *,
        group_reference: str | None = None,
        tin: str | None = None,
        rating: str | None = None,
        resident: bool | None = None,
    ) -> CanonicalCounterparty:
        row = CanonicalCounterparty(
            **self.common,
            source_reference=ref,
            name=name,
            counterparty_type=counterparty_type,
            group_reference=group_reference,
            rating=rating,
            resident=resident,
            external_identifiers={"tin": tin} if tin else {},
        )
        self.db.add(row)
        self.db.flush()
        return row

    def product(self, code: str, regulatory_category: str | None) -> CanonicalProduct:
        row = CanonicalProduct(
            **self.common,
            source_reference=f"PRODUCT/{code}",
            product_code=code,
            name=code,
            regulatory_category=regulatory_category,
        )
        self.db.add(row)
        self.db.flush()
        return row

    def position(  # noqa: PLR0913 - keyword-only fixture builder
        self,
        ref: str,
        position_type: str,
        balance_ghs: Decimal | str,
        *,
        counterparty: CanonicalCounterparty | None = None,
        product: CanonicalProduct | None = None,
        maturity: date | None = None,
        ifrs9_stage: int | None = None,
        currency: str = "GHS",
        interest_rate: Decimal | None = None,
        deposit_account_type: str | None = None,
        encumbered: bool | None = None,
        pledged_as_collateral: bool | None = None,
        operational_purpose: bool | None = None,
        redeemable_within_two_days: bool | None = None,
        extra_attributes: dict[str, Any] | None = None,
    ) -> None:
        position = CanonicalPosition(
            **self.common,
            source_reference=ref,
            position_type=position_type,
            currency=currency,
        )
        self.db.add(position)
        self.db.flush()
        attributes: dict[str, Any] = {"balance_ghs": str(balance_ghs)}
        if extra_attributes:
            attributes.update(extra_attributes)
        self.db.add(
            CanonicalPositionSnapshot(
                **self.common,
                source_reference=ref,
                position_id=position.id,
                counterparty_id=counterparty.id if counterparty is not None else None,
                product_id=product.id if product is not None else None,
                balance=Decimal(str(balance_ghs)),
                interest_rate=interest_rate,
                contractual_maturity=maturity,
                ifrs9_stage=ifrs9_stage,
                deposit_account_type=deposit_account_type,
                encumbered=encumbered,
                pledged_as_collateral=pledged_as_collateral,
                operational_purpose=operational_purpose,
                redeemable_within_two_days=redeemable_within_two_days,
                attributes=attributes,
            )
        )
        self.db.flush()


def _generate(db: Session, return_code: str) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=return_code, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _sections(package: RegulatoryPackage) -> dict[str, dict[str, Any]]:
    return {section["code"]: section for section in package.snapshot["sections"]}


def _pct_of(value: Decimal, nof: Decimal) -> Decimal:
    return (value / nof * Decimal("100")).quantize(PCT)


def _seed_le_book(db: Session, tier1: Decimal) -> dict[str, Decimal]:
    """Canonical exposures sized relative to Tier 1 (= the NOF proxy)."""
    seeder = _CanonicalSeeder(db)
    sovereign_product = seeder.product("SEC.GOG.5Y", "SOVEREIGN_LOCAL_CCY")

    big_corp = seeder.counterparty("CP/BIG", "Akwaaba Industries Ltd", "CORPORATE", tin="C-001")
    alpha = seeder.counterparty(
        "CP/ALPHA", "Volta Alpha Ltd", "CORPORATE", group_reference="VOLTA-GROUP", tin="C-002"
    )
    beta = seeder.counterparty(
        "CP/BETA", "Volta Beta Ltd", "CORPORATE", group_reference="VOLTA-GROUP", tin="C-003"
    )
    small = seeder.counterparty("CP/SMALL", "Adum Traders", "SME", tin="C-004")
    gov = seeder.counterparty("CP/GOG", "Government of Ghana", "SOVEREIGN")

    drawn_big = (tier1 * Decimal("0.10")).quantize(Decimal("0.01"))
    notional_big = (tier1 * Decimal("0.04")).quantize(Decimal("0.01"))
    undrawn_big = notional_big * Decimal("0.5")
    member_each = (tier1 * Decimal("0.06")).quantize(Decimal("0.01"))
    small_amount = (tier1 * Decimal("0.04")).quantize(Decimal("0.01"))
    gov_amount = (tier1 * Decimal("0.15")).quantize(Decimal("0.01"))
    issuer_amount = (tier1 * Decimal("0.11")).quantize(Decimal("0.01"))

    seeder.position(
        "LOAN/BIG",
        "LOAN",
        drawn_big,
        counterparty=big_corp,
        maturity=date(2028, 3, 31),
        ifrs9_stage=2,
        extra_attributes={
            "notional_ghs": str(notional_big),
            "credit_conversion_factor": "0.5",
            "ecl_provision_ghs": "250000",
        },
    )
    seeder.position(
        "LOAN/ALPHA",
        "LOAN",
        member_each,
        counterparty=alpha,
        ifrs9_stage=1,
        maturity=date(2027, 3, 31),
    )
    seeder.position(
        "LOAN/BETA",
        "LOAN",
        member_each,
        counterparty=beta,
        ifrs9_stage=1,
        maturity=date(2027, 9, 30),
    )
    seeder.position("LOAN/SMALL", "LOAN", small_amount, counterparty=small, ifrs9_stage=1)
    seeder.position(
        "SEC/GOV", "SECURITY_HOLDING", gov_amount, counterparty=gov, product=sovereign_product
    )
    # Counterparty-less sovereign security: the issuer attribute is the identity.
    seeder.position(
        "SEC/ISSUER",
        "SECURITY_HOLDING",
        issuer_amount,
        product=sovereign_product,
        extra_attributes={"issuer": "Bank of Ghana"},
    )
    return {
        "big_total": drawn_big + undrawn_big,
        "drawn_big": drawn_big,
        "undrawn_big": undrawn_big,
        "group_total": member_each * 2,
        "member_each": member_each,
        "small": small_amount,
        "gov": gov_amount,
        "issuer": issuer_amount,
    }


# ---------------------------------------------------------------------------
# LE-MONTHLY — registry, derivation, 409s, export round-trip
# ---------------------------------------------------------------------------


def test_le_registry_entry_is_periodic_monthly_confirmed() -> None:
    definition = REGISTRY["LE-MONTHLY"]
    assert definition.family == "large_exposures"
    assert definition.frequency == "monthly"
    assert definition.event_driven is False
    assert definition.fidelity == "CONFIRMED"
    assert definition.default_channel == "orass_sandbox"
    assert definition.generator == "large_exposures"
    # Day 9 is the observed BoG monthly convention (day-count unstated in the
    # directive draft Part VI) — overridable at onboarding.
    assert definition.deadline_rule(REPORTING_DATE) == date(2026, 4, 9)
    assert "Part VI" in definition.directive_citation


def test_le_calendar_expands_monthly_obligation(db_session: Session) -> None:
    seed_sample_bank(db_session)
    obligations = calendar.list_obligations(
        db_session, MAKER, SAMPLE_BANK_ID, 1, as_of=date(2026, 4, 5)
    ).obligations
    le_items = [item for item in obligations if item.return_code == "LE-MONTHLY"]
    assert le_items, "the calendar must expand LE-MONTHLY like any periodic return"
    march = [item for item in le_items if item.reporting_date == REPORTING_DATE]
    assert march and march[0].due_date == date(2026, 4, 9)
    assert march[0].return_family == "large_exposures"


def test_le_generation_derives_templates_from_canonical_exposures(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_capital_baseline(db_session)
    tier1 = _tier1(db_session)
    amounts = _seed_le_book(db_session, tier1)

    package = _generate(db_session, "LE-MONTHLY")
    assert package.return_family == "large_exposures"
    assert package.snapshot["template_id"] == "bog-le-monthly-v1"
    assert package.snapshot["fidelity"] == "CONFIRMED"
    assert [entry["module"] for entry in package.source_runs] == ["capital"]

    sections = _sections(package)
    assert set(sections) == {
        "template_1",
        "template_1a",
        "template_2",
        "template_3",
        "template_4",
    }

    # Template 1: only the NON-exempt exposures ≥10% NOF, largest first.
    t1_rows = sections["template_1"]["rows"]
    assert [row["description"] for row in t1_rows] == [
        "Akwaaba Industries Ltd",
        "VOLTA-GROUP",
    ]
    big = t1_rows[0]
    assert Decimal(big["value"]) == amounts["big_total"]
    assert Decimal(big["drawn_ghs"]) == amounts["drawn_big"]
    assert Decimal(big["undrawn_ccf_ghs"]) == amounts["undrawn_big"]
    assert Decimal(big["pct_nof"]) == _pct_of(amounts["big_total"], tier1)
    assert big["connection"] == "single"
    assert big["tin"] == "C-001"
    assert big["ifrs9_stage"] == "2"
    group = t1_rows[1]
    assert group["connection"] == "group"
    assert Decimal(group["value"]) == amounts["group_total"]
    assert Decimal(group["pct_nof"]) == _pct_of(amounts["group_total"], tier1)
    total = sections["template_1"]["total"]
    assert total["equals_sum_of_rows"] is True
    assert Decimal(total["value"]) == amounts["big_total"] + amounts["group_total"]

    # Template 1a: membership rows for the group reported in Template 1.
    t1a_rows = sections["template_1a"]["rows"]
    assert {row["description"] for row in t1a_rows} == {"Volta Alpha Ltd", "Volta Beta Ltd"}
    for row in t1a_rows:
        assert row["group_reference"] == "VOLTA-GROUP"
        assert Decimal(row["value"]) == amounts["member_each"]
        assert "group reference" in row["basis_of_connection"]

    # Template 2: every entity (exempt included), largest first, within the cap.
    t2_names = [row["description"] for row in sections["template_2"]["rows"]]
    assert t2_names == [
        "Government of Ghana",
        "Akwaaba Industries Ltd",
        "VOLTA-GROUP",
        "Bank of Ghana",
        "Adum Traders",
    ]
    t2_by_name = {row["description"]: row for row in sections["template_2"]["rows"]}
    assert Decimal(t2_by_name["Akwaaba Industries Ltd"]["provisions_ghs"]) == Decimal("250000")

    # Template 3: exempt (sovereign counterparty + sovereign-category issuer).
    t3_rows = sections["template_3"]["rows"]
    assert {row["description"] for row in t3_rows} == {"Government of Ghana", "Bank of Ghana"}
    assert all("canonical category" in row["exempt_basis"] for row in t3_rows)

    # Template 4 is empty by construction: no CRM data, pre-CRM == post-CRM.
    assert sections["template_4"]["rows"] == []
    assert Decimal(sections["template_4"]["total"]["value"]) == Decimal("0")

    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert Decimal(totals["tier1_ghs"]) == tier1
    assert Decimal(totals["nof_ghs"]) == tier1  # documented Tier-1 proxy
    assert totals["large_exposure_count"] == "2"
    assert Decimal(totals["largest_exposure_pct_nof"]) == _pct_of(amounts["gov"], tier1)

    metadata = package.snapshot["metadata"]
    assert "Tier 1" in metadata["nof_basis"]
    assert metadata["large_exposure_threshold_pct_nof"] == "10"


def test_le_top_100_cap_truncates_with_info_finding(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_capital_baseline(db_session)
    seeder = _CanonicalSeeder(db_session)
    for index in range(103):
        counterparty = seeder.counterparty(
            f"CP/N{index:03d}", f"Counterparty {index:03d}", "CORPORATE"
        )
        seeder.position(
            f"LOAN/N{index:03d}",
            "LOAN",
            Decimal(1_000_000 + index),
            counterparty=counterparty,
        )

    package = _generate(db_session, "LE-MONTHLY")
    sections = _sections(package)
    assert len(sections["template_2"]["rows"]) == 100
    # Largest first: the highest-balance counterparty leads the table.
    assert sections["template_2"]["rows"][0]["description"] == "Counterparty 102"
    findings = package.snapshot["metadata"]["generation_findings"]
    assert any(item["rule"] == "le.top100_truncated" for item in findings)

    # The truncation note rides the validation report as an INFO finding.
    read = validation.validate_package(db_session, MAKER, SAMPLE_BANK_ID, package.id)
    assert read.status == "validated"
    assert read.validation_report is not None
    truncation = [
        finding
        for finding in read.validation_report.findings
        if finding.rule == "le.top100_truncated"
    ]
    assert truncation and truncation[0].severity == "INFO"


def test_le_validates_and_exports_round_trip(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    seed_sample_bank(db_session)
    _run_capital_baseline(db_session)
    _seed_le_book(db_session, _tier1(db_session))
    package = _generate(db_session, "LE-MONTHLY")

    read = validation.validate_package(db_session, MAKER, SAMPLE_BANK_ID, package.id)
    assert read.status == "validated", read.validation_report
    assert read.validation_report is not None and read.validation_report.passed is True

    xlsx = export_package(db_session, MAKER, package, "xlsx")
    pdf = export_package(db_session, MAKER, package, "pdf")
    db_session.commit()
    assert xlsx.size_bytes > 0 and pdf.size_bytes > 0
    slug = db_session.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    payloads = {}
    for obj in storage.list(slug, "outputs"):
        _, stream = storage.read(obj.location)
        payloads[obj.location.object_path] = stream.read()
    assert payloads[pdf.object_path].startswith(b"%PDF")
    workbook = load_workbook(io.BytesIO(payloads[xlsx.object_path]))
    assert workbook.sheetnames == [
        "Return Metadata",
        "LE Template 1 Large Exposures",
        "LE Template 1a Connected",
        "LE Template 2 Top 100",
        "LE Template 3 Exempted",
        "LE Template 4 Other Pre-CRM",
        "Fidelity & Provenance",
    ]


def test_le_409_without_baseline_capital_run(db_session: Session) -> None:
    seed_sample_bank(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LE-MONTHLY")
    assert exc_info.value.status_code == 409
    assert "no_baseline_run" in str(exc_info.value.detail)


def test_le_409_without_canonical_positions(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_capital_baseline(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LE-MONTHLY")
    assert exc_info.value.status_code == 409
    assert "no_canonical_positions" in str(exc_info.value.detail)


# ---------------------------------------------------------------------------
# LMT — canonical monitoring-tool sections
# ---------------------------------------------------------------------------


def _seed_lmt_book(db: Session) -> None:
    """A book exercising every Table 2 mechanism (as-of 2026-03-31).

    Advances 15 days out (row 2, 15d-1mth), a security in 1-2 yrs (row 3), a
    derivative asset and a derivative liability (sign split rows 3/8), a
    volatile CALL deposit (row 7, non-contractual), a stable FIXED deposit
    45 days out (row 6, 1-2 mths), an unclassified deposit (stable by
    complement, non-contractual), an undrawn commitment (row 15 default),
    an LC classified via obs_category (row 16) and a bare guarantee
    (row 17 default, balance fallback, non-contractual).
    """
    seeder = _CanonicalSeeder(db)
    depositor_a = seeder.counterparty("CP/DEP-A", "Kanda Pensions Trust", "NBFI")
    depositor_b = seeder.counterparty("CP/DEP-B", "Osu Manufacturing Ltd", "CORPORATE")
    seeder.position("LOAN/L1", "LOAN", Decimal("8000000"), maturity=date(2026, 4, 15))
    seeder.position("SEC/S1", "SECURITY_HOLDING", Decimal("5000000"), maturity=date(2027, 6, 30))
    seeder.position(
        "SWAP/D1", "DERIVATIVE", Decimal("-400000"), maturity=date(2026, 9, 30)
    )  # liability MTM → row 8, 6-9 mths
    seeder.position(
        "SWAP/D2", "DERIVATIVE", Decimal("250000"), maturity=date(2026, 6, 30)
    )  # asset MTM → row 3, 2-3 mths
    seeder.position(
        "DEP/A",
        "DEPOSIT",
        Decimal("6000000"),
        counterparty=depositor_a,
        deposit_account_type="CALL",  # volatile (row 7)
    )
    seeder.position(
        "DEP/B",
        "DEPOSIT",
        Decimal("3000000"),
        counterparty=depositor_b,
        maturity=date(2026, 5, 15),
        deposit_account_type="FIXED",  # stable (row 6), 1-2 mths
    )
    seeder.position("DEP/ANON", "DEPOSIT", Decimal("1000000"))  # unclassified → stable
    seeder.position(
        "COMMIT/C1",
        "COMMITMENT_UNDRAWN",
        Decimal("0"),
        maturity=date(2026, 12, 31),
        extra_attributes={"notional_ghs": "2000000"},  # default → row 15, 9m-1yr
    )
    seeder.position(
        "LCG/G1",
        "LC_GUARANTEE",
        Decimal("0"),
        maturity=date(2026, 4, 5),
        extra_attributes={"notional_ghs": "900000", "obs_category": "letter_of_credit"},
    )  # row 16, 2-7 days
    seeder.position("LCG/G2", "LC_GUARANTEE", Decimal("600000"))  # default → row 17, NC


def test_lmt_carries_ladder_concentration_and_unencumbered_sections(  # noqa: PLR0915 - one linear pass over the printed grid
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    _seed_lmt_book(db_session)

    package = _generate(db_session, "LMT")
    sections = _sections(package)
    assert {"maturity_ladder", "funding_concentration", "unencumbered_assets"} <= set(sections)

    ladder = {row["code"]: row for row in sections["maturity_ladder"]["rows"]}
    assert set(ladder) == {str(n) for n in range(1, 18)}  # the printed rows 1-17

    # Category placement across the published buckets.
    assert Decimal(ladder["2"]["d15_1m_ghs"]) == Decimal("8000000")  # loan, 15 days out
    assert Decimal(ladder["3"]["y1_2_ghs"]) == Decimal("5000000")  # 2027 security
    assert Decimal(ladder["3"]["m2_3_ghs"]) == Decimal("250000")  # derivative asset MTM
    assert Decimal(ladder["8"]["m6_9_ghs"]) == Decimal("400000")  # derivative liability, abs
    assert Decimal(ladder["6"]["m1_2_ghs"]) == Decimal("3000000")  # stable FIXED, 45 days
    # Undated positions report non-contractual, never Next Day.
    assert Decimal(ladder["6"]["non_contractual_ghs"]) == Decimal("1000000")  # unclassified
    assert Decimal(ladder["7"]["non_contractual_ghs"]) == Decimal("6000000")  # CALL deposit
    assert Decimal(ladder["7"]["next_day_ghs"]) == Decimal("0")

    # Derived rows: 1 = 2+3+4, 5 = 6+7+8+9, 10 = 1-5 (Total column = `value`).
    assert Decimal(ladder["1"]["value"]) == Decimal("13250000")
    assert Decimal(ladder["5"]["value"]) == Decimal("10400000")
    assert Decimal(ladder["10"]["value"]) == Decimal("2850000")

    # Cumulative row 11 runs across dated buckets only: by 9m-1yr it has seen
    # +8.0M -3.0M +0.25M -0.4M; the non-contractual cell stays blank as printed.
    assert Decimal(ladder["11"]["m9_12_ghs"]) == Decimal("4850000")
    assert Decimal(ladder["11"]["value"]) == Decimal("9850000")  # through >5 yrs
    assert "non_contractual_ghs" not in ladder["11"]

    # Off-balance block: 15 (commitment default), 16 (obs_category LC),
    # 17 (guarantee default, balance fallback), 14 = 15+16+17, 12 = 13+14.
    assert Decimal(ladder["15"]["m9_12_ghs"]) == Decimal("2000000")
    assert Decimal(ladder["16"]["d2_7_ghs"]) == Decimal("900000")
    assert Decimal(ladder["17"]["non_contractual_ghs"]) == Decimal("600000")
    assert Decimal(ladder["13"]["value"]) == Decimal("0")
    assert Decimal(ladder["14"]["value"]) == Decimal("3500000")
    assert Decimal(ladder["12"]["value"]) == Decimal("3500000")

    total = sections["maturity_ladder"]["total"]
    assert total["equals_sum_of_rows"] is False  # a grid, not a summable column
    assert Decimal(total["value"]) == Decimal("2850000")  # on-balance mismatch

    concentration = {row["code"]: row for row in sections["funding_concentration"]["rows"]}
    # Both depositors exceed 1% of the 13.25M asset base — the directive's
    # population rule, not a Top-N cut.
    assert concentration["1"]["description"] == "Kanda Pensions Trust"
    assert Decimal(concentration["1"]["value"]) == Decimal("6000000")
    assert Decimal(concentration["1"]["pct_total_liabilities"]) == Decimal("57.6923")
    assert concentration["1"]["related_party"] == "No"
    assert concentration["2"]["description"] == "Osu Manufacturing Ltd"
    assert Decimal(concentration["top20_total"]["value"]) == Decimal("9000000")
    assert Decimal(concentration["top20_pct_of_deposits"]["value"]) == Decimal("90")
    assert {"maturity_of_exposures", "deposit_funding_concentration"} <= set(sections)

    # Table 9: the unencumbered security itemizes in section A (non-sovereign,
    # secondary-market marketable) with no calibrated haircut yet.
    unencumbered = {row["code"]: row for row in sections["unencumbered_assets"]["rows"]}
    assert Decimal(unencumbered["A1"]["value"]) == Decimal("5000000")
    assert unencumbered["A1"]["haircut_source"] == "unset"
    assert Decimal(unencumbered["A1"]["monetized_value_ghs"]) == Decimal("5000000")
    assert Decimal(unencumbered["A_total"]["value"]) == Decimal("5000000")
    assert Decimal(unencumbered["C_GHS"]["value"]) == Decimal("5000000")

    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert Decimal(totals["on_balance_mismatch_total_ghs"]) == Decimal("2850000")
    assert Decimal(totals["top20_deposits_pct"]) == Decimal("90")

    # The unattributed deposit is counted in the total but flagged as unrankable.
    findings = package.snapshot["metadata"]["generation_findings"]
    assert any(item["rule"] == "lmt.unattributed_funding" for item in findings)

    read = validation.validate_package(db_session, MAKER, SAMPLE_BANK_ID, package.id)
    assert read.status == "validated", read.validation_report

    artifact = export_package(db_session, MAKER, package, "xlsx")
    db_session.commit()
    slug = db_session.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    payload = b""
    for obj in storage.list(slug, "outputs"):
        if obj.location.object_path == artifact.object_path:
            _, stream = storage.read(obj.location)
            payload = stream.read()
    workbook = load_workbook(io.BytesIO(payload))
    assert "Contractual Maturity Mismatch" in workbook.sheetnames
    # xlsx sheet titles truncate at 31 characters.
    assert any(name.startswith("Funding from Significant") for name in workbook.sheetnames)
    assert "Available Unencumbered Assets" in workbook.sheetnames


def _seed_table1_book(db: Session) -> None:
    """A book exercising every LMTD ¶5 Narrow/Broad classification leg.

    Narrow = 2.0 vault + 3.0 BoG + 1.0 operational correspondent + 4.0 AAA
    non-resident placement + 2.5 domestic-bank claim + 5.0 sovereign bill
    ≤1yr + 0.8 two-day redeemable = 18.3M. Broad adds the 6.0M GoG bond >1yr
    and equities capped at other-broad/9 (3.0M holding → 2.7M) = 27.0M.
    The encumbered 1.0M bill counts in neither. Volatile 10.0M; total
    deposits 27.0M; short-term 26.5M (by-nature 18.0 + 4.0 short fixed +
    3.0 interbank + 1.5 contingent); total assets 28.3M.
    """
    seeder = _CanonicalSeeder(db)
    bog = seeder.counterparty("CP/BOG", "Central Bank", "CENTRAL_BANK")
    corr_de = seeder.counterparty(
        "CP/CORR-DE", "Frankfurt Corr Bank", "BANK_OECD", resident=False
    )
    aaa_bank = seeder.counterparty(
        "CP/AAA", "Zurich Prime Bank", "BANK_OECD", rating="AAA", resident=False
    )
    local_bank = seeder.counterparty("CP/GH-BANK", "Volta Bank", "BANK_NON_OECD", resident=True)
    tbill = seeder.product("SEC.GOG.TBILL", "SOVEREIGN_GOG_TBILL_0RW")
    bond = seeder.product("SEC.GOG.BOND", "SOVEREIGN_GOG_BOND_0RW")
    equity = seeder.product("SEC.GSE.EQ", "EQUITY_GSE_LISTED")

    seeder.position("CASH/VAULT", "CASH", Decimal("2000000"))  # leg (a)
    seeder.position("CASH/BOG", "CASH", Decimal("3000000"), counterparty=bog)  # leg (d)
    seeder.position(
        "CASH/CORR", "CASH", Decimal("1000000"), counterparty=corr_de, operational_purpose=True
    )  # leg (b)
    seeder.position("IBP/AAA", "INTERBANK_PLACEMENT", Decimal("4000000"), counterparty=aaa_bank)
    seeder.position("IBP/GH", "INTERBANK_PLACEMENT", Decimal("2500000"), counterparty=local_bank)
    seeder.position(
        "SEC/TBILL", "SECURITY_HOLDING", Decimal("5000000"), product=tbill,
        maturity=date(2026, 9, 30),
    )  # leg (e)
    seeder.position(
        "SEC/TBILL-ENC", "SECURITY_HOLDING", Decimal("1000000"), product=tbill,
        maturity=date(2026, 9, 30), encumbered=True,
    )  # excluded: encumbered
    seeder.position(
        "SEC/MDB", "SECURITY_HOLDING", Decimal("800000"),
        maturity=date(2026, 12, 31), redeemable_within_two_days=True,
    )  # leg (f)
    seeder.position(
        "SEC/GOGBOND", "SECURITY_HOLDING", Decimal("6000000"), product=bond,
        maturity=date(2030, 6, 30),
    )  # broad only
    seeder.position("SEC/EQ", "SECURITY_HOLDING", Decimal("3000000"), product=equity)
    seeder.position(
        "DEP/CUR", "DEPOSIT", Decimal("10000000"), deposit_account_type="CURRENT"
    )  # volatile + by-nature short-term
    seeder.position(
        "DEP/SAV", "DEPOSIT", Decimal("8000000"), deposit_account_type="SAVINGS"
    )  # stable, by-nature short-term
    seeder.position(
        "DEP/FIX-LONG", "DEPOSIT", Decimal("5000000"), deposit_account_type="FIXED",
        maturity=date(2028, 3, 31),
    )  # NOT short-term
    seeder.position(
        "DEP/FIX-SHORT", "DEPOSIT", Decimal("4000000"), deposit_account_type="FIXED",
        maturity=date(2026, 6, 30),
    )  # short-term by maturity
    seeder.position(
        "IBB/1", "INTERBANK_BORROWING", Decimal("3000000"), maturity=date(2026, 4, 30)
    )
    seeder.position(
        "LCG/ST", "LC_GUARANTEE", Decimal("0"), maturity=date(2026, 8, 31),
        extra_attributes={"notional_ghs": "1500000"},
    )  # contingent ≤ 1 yr


def test_lmt_table1_prudential_ratios(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    _seed_table1_book(db_session)

    package = _generate(db_session, "LMT")
    sections = _sections(package)
    assert {"prudential_ratio_inputs", "prudential_ratio_percentages"} <= set(sections)

    inputs = {row["code"]: row for row in sections["prudential_ratio_inputs"]["rows"]}
    assert Decimal(inputs["narrow"]["value"]) == Decimal("18300000")
    assert Decimal(inputs["broad"]["value"]) == Decimal("27000000")  # equity capped at 2.7M
    assert Decimal(inputs["volatile"]["value"]) == Decimal("10000000")
    assert Decimal(inputs["total_deposits"]["value"]) == Decimal("27000000")
    assert Decimal(inputs["short_term"]["value"]) == Decimal("26500000")
    assert Decimal(inputs["total_assets"]["value"]) == Decimal("28300000")
    # No prior canonical book: the Previous Month column is absent, not zero.
    assert "previous_month_ghs" not in inputs["narrow"]

    ratios = {row["code"]: row for row in sections["prudential_ratio_percentages"]["rows"]}
    assert len(ratios) == 8
    assert Decimal(ratios["narrow_to_volatile"]["value"]) == Decimal("183")
    assert Decimal(ratios["broad_to_volatile"]["value"]) == Decimal("270")
    assert Decimal(ratios["broad_to_total_deposits"]["value"]) == Decimal("100")
    assert all(row["status"] == "ok" for row in ratios.values())
    assert all(row["threshold_source"] == "regulatory_default" for row in ratios.values())
    assert Decimal(ratios["narrow_to_volatile"]["threshold_min_pct"]) == Decimal("80")

    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert Decimal(totals["narrow_liquid_assets_ghs"]) == Decimal("18300000")
    assert Decimal(totals["broad_liquid_assets_ghs"]) == Decimal("27000000")
    assert totals["prudential_ratio_breaches"] == "0"


def test_lmt_table1_uses_board_register_floor(db_session: Session) -> None:
    """A Board-adopted floor above the observed ratio flips it to a breach."""
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    _seed_table1_book(db_session)
    db_session.add(
        ParamLiquidityThreshold(
            organization_id=DEMO_ORG_ID,
            jurisdiction_code="GH",
            institution_class="bank",
            threshold_code="narrow_to_volatile",
            threshold_pct=Decimal("200"),
            effective_from=date(2026, 1, 1),
            approved_by="Board minute 2026-03",
            approval_timestamp=utc_now(),
        )
    )
    db_session.flush()

    package = _generate(db_session, "LMT")
    sections = _sections(package)
    ratios = {row["code"]: row for row in sections["prudential_ratio_percentages"]["rows"]}
    row = ratios["narrow_to_volatile"]
    assert row["threshold_source"] == "board_register"
    assert Decimal(row["threshold_min_pct"]) == Decimal("200")
    assert row["status"] == "below_minimum"  # 183% < the Board's 200% floor

    totals = {item["code"]: item["value"] for item in package.snapshot["totals"]}
    assert totals["prudential_ratio_breaches"] == "1"
    findings = package.snapshot["metadata"]["generation_findings"]
    assert any(
        item["rule"] == "lmt.prudential_ratio_below_minimum" for item in findings
    ), findings


def _seed_currency_book(db: Session) -> None:
    """Four currencies, all significant (>= 5% of the 16.0M liability base).

    Liabilities: 10.0M GHS CALL (volatile demand), 4.0M-equivalent USD fixed
    maturing in 20 days, 1.1M GBP CURRENT, 0.9M EUR borrowing beyond 30 days.
    Assets: 8.0M GHS HQLA (6.0M sovereign bill + 2.0M vault cash), a 3.0M USD
    loan maturing in 15 days (the only 30-day inflow), 0.7M EUR security.
    All amounts are ingested cedi equivalents (balance_ghs).
    """
    seeder = _CanonicalSeeder(db)
    tbill = seeder.product("SEC.GOG.TBILL.C", "SOVEREIGN_GOG_TBILL_0RW")
    seeder.position("CCY/DEP-GHS", "DEPOSIT", Decimal("10000000"), deposit_account_type="CALL")
    seeder.position(
        "CCY/DEP-USD", "DEPOSIT", Decimal("4000000"), currency="USD",
        deposit_account_type="FIXED", maturity=date(2026, 4, 20),
        extra_attributes={"balance_ghs": "4000000"},
    )
    seeder.position(
        "CCY/DEP-GBP", "DEPOSIT", Decimal("1100000"), currency="GBP",
        deposit_account_type="CURRENT",
        extra_attributes={"balance_ghs": "1100000"},
    )
    seeder.position(
        "CCY/IBB-EUR", "INTERBANK_BORROWING", Decimal("900000"), currency="EUR",
        maturity=date(2026, 10, 31),
        extra_attributes={"balance_ghs": "900000"},
    )
    seeder.position(
        "CCY/TBILL", "SECURITY_HOLDING", Decimal("6000000"), product=tbill,
        maturity=date(2026, 9, 30),
    )
    seeder.position("CCY/VAULT", "CASH", Decimal("2000000"))
    seeder.position(
        "CCY/LOAN-USD", "LOAN", Decimal("3000000"), currency="USD",
        maturity=date(2026, 4, 15),
        extra_attributes={"balance_ghs": "3000000"},
    )
    seeder.position(
        "CCY/SEC-EUR", "SECURITY_HOLDING", Decimal("700000"), currency="EUR",
        maturity=date(2027, 6, 30),
        extra_attributes={"balance_ghs": "700000"},
    )


def test_lmt_significant_currency_tables(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    _seed_currency_book(db_session)

    package = _generate(db_session, "LMT")
    sections = _sections(package)
    assert {"assets_liabilities_by_currency", "lcr_by_currency"} <= set(sections)

    # Table 6 — rows ordered by liability base, all four >= 5% of 16.0M.
    table6 = sections["assets_liabilities_by_currency"]["rows"]
    assert [row["code"] for row in table6] == ["GHS", "USD", "GBP", "EUR"]
    by_ccy = {row["code"]: row for row in table6}
    assert Decimal(by_ccy["GHS"]["assets_ghs"]) == Decimal("8000000")
    assert Decimal(by_ccy["GHS"]["liabilities_ghs"]) == Decimal("10000000")
    assert Decimal(by_ccy["GHS"]["value"]) == Decimal("-2000000")  # mismatch (2-3)
    assert Decimal(by_ccy["GHS"]["mismatch_pct_total_liabilities"]) == Decimal("-12.5")
    assert Decimal(by_ccy["USD"]["value"]) == Decimal("-1000000")
    assert Decimal(by_ccy["EUR"]["value"]) == Decimal("-200000")
    total = sections["assets_liabilities_by_currency"]["total"]
    assert Decimal(total["value"]) == Decimal("-4300000")

    # Table 11 — fixed printed columns, per-currency LCR components.
    table11 = {row["code"]: row for row in sections["lcr_by_currency"]["rows"]}
    assert Decimal(table11["level_1"]["cedi_ghs"]) == Decimal("8000000")
    assert Decimal(table11["level_1"]["usd_ghs"]) == Decimal("0")
    assert Decimal(table11["level_2a"]["cedi_ghs"]) == Decimal("0")  # no L2 taxonomy
    assert Decimal(table11["total_cash_outflow"]["cedi_ghs"]) == Decimal("10000000")
    assert Decimal(table11["total_cash_outflow"]["usd_ghs"]) == Decimal("4000000")
    assert Decimal(table11["total_cash_outflow"]["pound_ghs"]) == Decimal("1100000")
    assert Decimal(table11["total_cash_outflow"]["euro_ghs"]) == Decimal("0")
    # USD inflow: 3.0M contractual, capped at 75% of the 4.0M outflow.
    assert Decimal(table11["total_cash_inflow"]["usd_ghs"]) == Decimal("3000000")
    assert Decimal(table11["net_cash_outflow"]["usd_ghs"]) == Decimal("1000000")
    # Net is (1)-(2): the printed (2-1) label is a recorded deviation.
    assert Decimal(table11["net_cash_outflow"]["cedi_ghs"]) == Decimal("10000000")
    assert Decimal(table11["lcr_pct"]["cedi_ghs"]) == Decimal("80")
    # Aggregate column: 8.0M L1 over 12.1M net outflow.
    assert Decimal(table11["lcr_pct"]["value"]) == Decimal("66.1157")

    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert totals["significant_currencies"] == "4"


def test_lmt_funding_concentration_netting_related_and_tables_7_8(
    db_session: Session,
) -> None:
    """Tables 5/7/8 mechanics: para-23 netting, related-party flags, blocks.

    Book: assets = one 50.0M loan (1% threshold = 0.5M). Depositors — a
    related pensions group (8.0M call, of which 2.0M pledged to secure a
    facility), a resident bank (5.0M fixed at 60 days), a government agency
    (4.0M fixed at 200 days), and an anonymous 1.0M. A 3.0M negotiable-paper
    note at 9 months fills Table 8's instrument rows.
    """
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    db_session.add(
        RelatedParty(
            organization_id=DEMO_ORG_ID,
            bank_id=SAMPLE_BANK_ID,
            party_type="legal_entity",
            full_name="Akosombo Pensions Ltd",
            status="active",
        )
    )
    db_session.flush()

    seeder = _CanonicalSeeder(db_session)
    related_cp = seeder.counterparty("CP/REL", "Akosombo Pensions Ltd", "CORPORATE")
    bank_cp = seeder.counterparty("CP/BANK", "Volta Interbank Ltd", "BANK_NON_OECD")
    gov_cp = seeder.counterparty("CP/GOV", "Cocoa Board Agency", "GOVERNMENT_ENTITY")
    seeder.position("LN/BIG", "LOAN", Decimal("50000000"), maturity=date(2029, 3, 31))
    seeder.position(
        "DEP/REL-A", "DEPOSIT", Decimal("6000000"), counterparty=related_cp,
        deposit_account_type="CALL",
    )
    seeder.position(
        "DEP/REL-B", "DEPOSIT", Decimal("2000000"), counterparty=related_cp,
        deposit_account_type="CALL", pledged_as_collateral=True,
    )
    seeder.position(
        "DEP/BANK", "DEPOSIT", Decimal("5000000"), counterparty=bank_cp,
        deposit_account_type="FIXED", maturity=date(2026, 5, 30),  # 60 days
    )
    seeder.position(
        "DEP/GOV", "DEPOSIT", Decimal("4000000"), counterparty=gov_cp,
        deposit_account_type="FIXED", maturity=date(2026, 10, 17),  # 200 days
    )
    seeder.position("DEP/ANON", "DEPOSIT", Decimal("1000000"))
    seeder.position(
        "NP/1", "OTHER_LIABILITY", Decimal("3000000"), maturity=date(2026, 12, 31),
        extra_attributes={"funding_instrument": "negotiable_paper"},
    )

    package = _generate(db_session, "LMT")
    sections = _sections(package)

    # Table 5: all three attributed depositors exceed 1% of 50.0M assets.
    concentration = {row["code"]: row for row in sections["funding_concentration"]["rows"]}
    assert concentration["1"]["description"] == "Akosombo Pensions Ltd"
    assert concentration["1"]["related_party"] == "Yes"  # register name match
    assert concentration["2"]["related_party"] == "No"
    # Netting (para 23): attributed top-20 deposits are 17.0M gross; the
    # 2.0M pledged leaves BOTH sides — (17-2) / (18-2) = 93.75%.
    assert Decimal(concentration["top20_total"]["value"]) == Decimal("17000000")
    assert Decimal(concentration["top20_pct_of_deposits"]["value"]) == Decimal("93.75")

    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert Decimal(totals["top20_deposits_pct"]) == Decimal("93.75")

    # Table 7: block A places the call deposits in <1 month (on demand), the
    # 60-day fixed in 1-3 months and the 200-day fixed in 6-12 months.
    table7 = {row["code"]: row for row in sections["maturity_of_exposures"]["rows"]}
    assert Decimal(table7["A"]["m_lt1_ghs"]) == Decimal("8000000")
    assert Decimal(table7["A"]["m1_3_ghs"]) == Decimal("5000000")
    assert Decimal(table7["A"]["m6_12_ghs"]) == Decimal("4000000")
    assert Decimal(table7["A"]["value"]) == Decimal("17000000")
    assert Decimal(table7["B_total"]["value"]) == Decimal("17000000")

    # Table 8: associates = the related group's gross deposits; FI and
    # government blocks pick their counterparty types; negotiable paper
    # lands in 6-12 months with the <=12m of-which row filled.
    table8 = {row["code"]: row for row in sections["deposit_funding_concentration"]["rows"]}
    assert Decimal(table8["associates"]["value"]) == Decimal("8000000")
    assert Decimal(table8["associates"]["next_day_ghs"]) == Decimal("8000000")  # both CALL
    assert Decimal(table8["top20_financial"]["value"]) == Decimal("5000000")
    assert Decimal(table8["top20_government"]["value"]) == Decimal("4000000")
    assert Decimal(table8["negotiable_paper"]["m6_12_ghs"]) == Decimal("3000000")
    assert Decimal(table8["negotiable_paper_lte_12m"]["value"]) == Decimal("3000000")
    assert Decimal(table8["negotiable_paper_gt_5y"]["value"]) == Decimal("0")


def test_lmt_collateral_and_no_maturity_tables(db_session: Session) -> None:
    """Tables 3, 4, 9 (with a calibrated haircut) and 10.

    Book: a sovereign bond (BoG-eligible, section 9B, haircut 10% from the
    schedule), an equity holding (section 9A, uncalibrated), an undated
    equity position for Table 3, and a loan carrying received customer
    collateral (re-hypothecable government paper, partly re-pledged, partly
    unavailable) plus own debt securities for Table 10.
    """
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    db_session.add(
        ParamLiquidityHaircut(
            organization_id=DEMO_ORG_ID,
            jurisdiction_code="GH",
            asset_class="SOVEREIGN",
            haircut_pct=Decimal("10"),
            effective_from=date(2026, 1, 1),
            approved_by="ALCO review 2026-01",
            approval_timestamp=utc_now(),
        )
    )
    db_session.flush()

    seeder = _CanonicalSeeder(db_session)
    bond = seeder.product("SEC.GOG.BOND.T9", "SOVEREIGN_GOG_BOND_0RW")
    equity = seeder.product("SEC.GSE.EQ.T9", "EQUITY_GSE_LISTED")
    seeder.position(
        "T9/BOND", "SECURITY_HOLDING", Decimal("4000000"), product=bond,
        maturity=date(2029, 6, 30),
    )
    seeder.position(
        "T9/EQ", "SECURITY_HOLDING", Decimal("1000000"), product=equity,
    )  # undated equity → Table 3 row too
    seeder.position(
        "T4/LOAN", "LOAN", Decimal("9000000"), maturity=date(2027, 3, 31),
        extra_attributes={
            "collateral_instrument": "GoG bonds received",
            "collateral_asset_class": "debt_government",
            "collateral_received_ghs": "6000000",
            "collateral_rehypothecable": "yes",
            "collateral_rehypothecated_ghs": "2500000",
            "collateral_unavailable_ghs": "500000",
            "collateral_bog_eligible": "yes",
        },
    )
    seeder.position(
        "T10/OWN", "OTHER_LIABILITY", Decimal("0"), maturity=date(2028, 1, 1),
        extra_attributes={
            "own_debt_available_ghs": "1200000",
            "own_debt_unavailable_ghs": "300000",
        },
    )

    package = _generate(db_session, "LMT")
    sections = _sections(package)
    assert {
        "items_no_contractual_maturity",
        "collateral_rehypothecation",
        "unencumbered_assets",
        "collateral_received",
    } <= set(sections)

    # Table 3: only the undated equity (deposits excluded by rule).
    table3 = sections["items_no_contractual_maturity"]["rows"]
    assert [(row["description"], row["value"]) for row in table3] == [
        ("SEC.GSE.EQ.T9", "1000000")
    ]

    # Table 4: A=6.0M, B=2.5M, C=3.5M.
    table4 = sections["collateral_rehypothecation"]["rows"]
    (entry,) = table4
    assert entry["description"] == "GoG bonds received"
    assert Decimal(entry["total_amounts_ghs"]) == Decimal("6000000")
    assert Decimal(entry["hypothecated_ghs"]) == Decimal("2500000")
    assert Decimal(entry["value"]) == Decimal("3500000")
    total4 = sections["collateral_rehypothecation"]["total"]
    assert Decimal(total4["value"]) == Decimal("3500000")

    # Table 9: bond → section B with the calibrated 10% haircut; equity → A.
    table9 = {row["code"]: row for row in sections["unencumbered_assets"]["rows"]}
    assert Decimal(table9["B1"]["value"]) == Decimal("4000000")
    assert table9["B1"]["haircut_source"] == "schedule"
    assert Decimal(table9["B1"]["haircut_pct"]) == Decimal("10")
    assert Decimal(table9["B1"]["monetized_value_ghs"]) == Decimal("3600000")
    assert Decimal(table9["A1"]["value"]) == Decimal("1000000")
    assert table9["A1"]["haircut_source"] == "unset"
    # Section C aggregates A∪B by significant currency with monetized values.
    assert Decimal(table9["C_GHS"]["value"]) == Decimal("5000000")
    assert Decimal(table9["C_GHS"]["monetized_value_ghs"]) == Decimal("4600000")

    # Table 10: government-debt class row + own debt + grand total.
    table10 = {row["code"]: row for row in sections["collateral_received"]["rows"]}
    gov = table10["debt_government"]
    assert Decimal(gov["value"]) == Decimal("6000000")
    assert Decimal(gov["bog_eligible_ghs"]) == Decimal("6000000")
    assert Decimal(gov["group_issued_ghs"]) == Decimal("0")
    assert Decimal(gov["unavailable_ghs"]) == Decimal("500000")
    own = table10["own_debt"]
    assert Decimal(own["value"]) == Decimal("1200000")
    assert Decimal(own["unavailable_ghs"]) == Decimal("300000")
    grand = table10["grand_total"]
    assert Decimal(grand["value"]) == Decimal("7200000")
    assert Decimal(grand["unavailable_ghs"]) == Decimal("800000")


def test_lmt_without_canonical_positions_omits_position_tools(db_session: Session) -> None:
    seed_sample_bank(db_session)
    _run_liquidity_baseline(db_session)
    package = _generate(db_session, "LMT")
    sections = _sections(package)
    # No canonical positions: the ladder and depositor ranking are honestly
    # absent; the HQLA-fact tool still renders from the seeded facts.
    assert "maturity_ladder" not in sections
    assert "funding_concentration" not in sections
    # Table 9 is canonical-based as of P2-6: absent without positions.
    assert "unencumbered_assets" not in sections
    assert {"hqla", "outflows", "inflows", "lcr_summary"} <= set(sections)


# ---------------------------------------------------------------------------
# IRRBB — BoG ±450 bp parameterization and conditional return rows
# ---------------------------------------------------------------------------


def _irr_baseline_run(db: Session) -> RegulatoryRun:
    run = db.scalar(
        select(RegulatoryRun)
        .where(
            RegulatoryRun.organization_id == DEMO_ORG_ID,
            RegulatoryRun.bank_id == SAMPLE_BANK_ID,
            RegulatoryRun.module == "irr",
            RegulatoryRun.scenario_code == "baseline",
            RegulatoryRun.status == "succeeded",
        )
        .order_by(RegulatoryRun.created_at.desc())
        .limit(1)
    )
    assert run is not None
    return run


def test_irrbb_bog_450_params_flow_into_engine_outputs(db_session: Session) -> None:
    seed_sample_bank(db_session)
    shocks = db_session.scalars(
        select(ParamStressShock).where(
            ParamStressShock.module == "irr",
            ParamStressShock.scenario_code.in_(("parallel_up_450", "parallel_down_450")),
        )
    ).all()
    assert {(row.scenario_code, row.shock_key) for row in shocks} == {
        ("parallel_up_450", "parallel_bp"),
        ("parallel_down_450", "parallel_bp"),
    }
    assert {row.shock_value for row in shocks} == {Decimal("450"), Decimal("-450")}

    # GAP-5 closed the former engine gap: when the active parameter set
    # carries the BoG GHS ±450 bp shocks, the engine computes them as
    # INFORMATIONAL scenarios — present in the outputs, excluded from the
    # Basel outlier test (worst scenario/breach stay Basel-only).
    regulatory_irr.run_all_irr_scenarios(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        IrrScenarioBatchCreate(reporting_period_id=_period_id(db_session)),
    )
    metrics = _irr_baseline_run(db_session).metrics
    scenario_codes = {entry["scenario_code"] for entry in metrics["eve_by_scenario"]}
    assert {"parallel_up_450", "parallel_down_450"} <= scenario_codes
    assert "eve_up_450_ghs" in metrics and "ear_up_450_ghs" in metrics
    # The supervisory outlier verdict is unchanged by the add-ons.
    assert metrics["worst_scenario"] not in ("parallel_up_450", "parallel_down_450")
    for entry in metrics["eve_by_scenario"]:
        if entry["scenario_code"] in ("parallel_up_450", "parallel_down_450"):
            assert entry["breach"] is False


def test_irrbb_450_rows_render_only_when_metrics_carry_them(db_session: Session) -> None:
    seed_sample_bank(db_session)
    regulatory_irr.run_all_irr_scenarios(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        IrrScenarioBatchCreate(reporting_period_id=_period_id(db_session)),
    )

    # Simulate an uncalibrated bank: the engine now computes ±450 whenever the
    # param set carries the shocks (GAP-5), so strip the keys to model a param
    # set without the BoG calibration — the rows must be honestly absent.
    run = _irr_baseline_run(db_session)
    run.metrics = {
        key: value
        for key, value in run.metrics.items()
        if key
        not in ("eve_up_450_ghs", "eve_down_450_ghs", "ear_up_450_ghs", "ear_down_450_ghs")
    }
    db_session.commit()
    package = _generate(db_session, "IRRBB-PILOT")
    sections = _sections(package)
    eve_codes = {row["code"] for row in sections["eve_scenarios"]["rows"]}
    ear_codes = {row["code"] for row in sections["earnings_at_risk"]["rows"]}
    assert not {"eve_up_450_ghs", "eve_down_450_ghs"} & eve_codes
    assert not {"ear_up_450_ghs", "ear_down_450_ghs"} & ear_codes
    assert package.snapshot["metadata"]["bog_ghs_450_rows_present"] is False

    # Once a run's metrics actually carry the BoG ±450 outputs, the rows render.
    run = _irr_baseline_run(db_session)
    run.metrics = {
        **run.metrics,
        "eve_up_450_ghs": "-25000000",
        "eve_down_450_ghs": "26000000",
        "ear_up_450_ghs": "-16000000",
        "ear_down_450_ghs": "16000000",
    }
    db_session.commit()

    regenerated = _generate(db_session, "IRRBB-PILOT")
    sections = _sections(regenerated)
    eve_rows = {row["code"]: row for row in sections["eve_scenarios"]["rows"]}
    ear_rows = {row["code"]: row for row in sections["earnings_at_risk"]["rows"]}
    assert eve_rows["eve_up_450_ghs"]["value"] == "-25000000"
    assert eve_rows["eve_down_450_ghs"]["value"] == "26000000"
    assert ear_rows["ear_up_450_ghs"]["value"] == "-16000000"
    assert ear_rows["ear_down_450_ghs"]["value"] == "16000000"
    assert "BoG GHS calibration" in eve_rows["eve_up_450_ghs"]["description"]
    assert regenerated.snapshot["metadata"]["bog_ghs_450_rows_present"] is True
