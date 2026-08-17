"""LRT corporate return-pack tests (plan W5).

Register fixtures go through the W4 institution-profile service (mirroring
tests/api/test_institution_profile.py data); per pack: generate → snapshot
section/row assertions → validation passes → real xlsx/pdf export (same
in-memory storage seam as test_regulatory_reporting_exports.py) →
event-driven codes stay out of the calendar → 409s for missing master data.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.models import Bank, RegulatoryPackage
from app.schemas.institution_profile import (
    BankLicenseCreate,
    BankNameHistoryCreate,
    BankProductCreate,
    InstitutionProfilePut,
    OutletCreate,
    RelatedPartyCreate,
    RelatedPartyRoleInput,
    ShareholdingCreate,
)
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.services import institution_profile
from app.services.regulatory_reporting import calendar, generation, validation
from app.services.regulatory_reporting import packages as packages_service
from app.services.regulatory_reporting.exports import export_package
from app.services.regulatory_reporting.registry import REGISTRY
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
REPORTING_DATE = date(2026, 3, 31)
CORPORATE_CODES = {"LRT-PROFILE", "LRT-OUTLET", "LRT-PARTY", "LRT-CAPITAL", "LRT-PRODUCT"}

PROFILE_PAYLOAD = InstitutionProfilePut(
    reason="Initial corporate profile capture",
    institution_type="Universal Bank",
    legal_entity_structure="Public Limited Company",
    authorisation_date=date(2005, 8, 15),
    approved_capital=Decimal("400000000.00"),
    incorporation_date=date(2004, 11, 2),
    tin="C0001234567",
    registration_number="CS-2004-11-0042",
    orass_institution_code="GH-UB-0042",
    traded_on_exchange=True,
    exchange_name="Ghana Stock Exchange",
    isin="GH0000000042",
    ownership_local_pct=Decimal("60.00"),
    ownership_foreign_pct=Decimal("40.00"),
    parent_country_code="GH",
)


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _seed_profile(db: Session) -> None:
    institution_profile.upsert_profile(db, MAKER, SAMPLE_BANK_ID, PROFILE_PAYLOAD)
    institution_profile.create_license(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        BankLicenseCreate(
            reason="Record universal banking license",
            license_name="Universal Banking License",
            license_class="Class 1",
            issued_on=date(2005, 8, 15),
        ),
    )
    institution_profile.create_name_history_entry(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        BankNameHistoryCreate(
            reason="Record pre-merger name",
            previous_name="Sample Savings & Loans Ltd",
            changed_on=date(2010, 4, 1),
            change_reason="Upgrade to universal banking license",
        ),
    )


def _seed_parties(db: Session) -> dict[str, Any]:
    """Mirrors tests/api/test_institution_profile.py register data."""
    holdco = institution_profile.create_related_party(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RelatedPartyCreate(
            reason="Register corporate shareholder",
            party_type="legal_entity",
            full_name="Golden Coast Holdings Ltd",
            roles=[RelatedPartyRoleInput(role="shareholder")],
        ),
    )
    owner = institution_profile.create_related_party(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RelatedPartyCreate(
            reason="Register ultimate beneficial owner",
            party_type="individual",
            full_name="Kwame Owusu",
            roles=[RelatedPartyRoleInput(role="ultimate_beneficial_owner")],
        ),
    )
    institution_profile.add_shareholding(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        holdco.id,
        ShareholdingCreate(
            reason="Record ordinary shareholding with UBO",
            share_type="Ordinary",
            share_subtype="Class A",
            shareholder_rights="voting",
            number_of_shares=Decimal("1250000.00"),
            pct_shareholding=Decimal("12.5000"),
            ubo_party_id=owner.id,
        ),
    )
    director = institution_profile.create_related_party(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RelatedPartyCreate(
            reason="Appoint non-executive director",
            party_type="individual",
            full_name="Ama Mensah",
            contact={"email": "ama.mensah@example.test"},
            roles=[
                RelatedPartyRoleInput(
                    role="director",
                    appointed_on=date(2025, 6, 1),
                    term_of_appointment="3 years renewable once",
                    sitting_allowance=Decimal("1500.00"),
                    travel_allowance=Decimal("500.00"),
                    annual_fees=Decimal("24000.00"),
                )
            ],
        ),
    )
    auditor = institution_profile.create_related_party(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RelatedPartyCreate(
            reason="Register external auditor",
            party_type="legal_entity",
            full_name="Assurance Partners Chartered Accountants",
            regulated_elsewhere=True,
            regulated_jurisdiction="Ghana (ICAG)",
            roles=[
                RelatedPartyRoleInput(role="external_auditor", icag_registration="ICAG/F/2026/042")
            ],
        ),
    )
    ceo = institution_profile.create_related_party(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RelatedPartyCreate(
            reason="Register chief executive officer",
            party_type="individual",
            full_name="Efua Boateng",
            roles=[
                RelatedPartyRoleInput(
                    role="chief_executive_officer", appointed_on=date(2024, 1, 15)
                )
            ],
        ),
    )
    return {"holdco": holdco, "owner": owner, "director": director, "auditor": auditor, "ceo": ceo}


def _seed_outlets(db: Session) -> None:
    institution_profile.create_outlet(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        OutletCreate(
            reason="Open the Kumasi branch",
            outlet_type="branch",
            name="Kumasi Adum Branch",
            outlet_number="BR-014",
            address={"city": "Kumasi", "street": "Adum High Street"},
            opened_on=date(2026, 2, 1),
        ),
    )
    institution_profile.create_outlet(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        OutletCreate(
            reason="Record closed agency",
            outlet_type="agency",
            name="Tema Harbour Agency",
            outlet_number="AG-003",
            address={"city": "Tema"},
            status="closed",
            opened_on=date(2020, 5, 1),
            closed_on=date(2026, 1, 31),
            relocated_from="Tema Fishing Harbour",
        ),
    )


def _seed_products(db: Session) -> None:
    institution_profile.create_product(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        BankProductCreate(
            reason="Propose SME overdraft product",
            name="SME Flex Overdraft",
            product_type="credit",
        ),
    )
    institution_profile.create_product(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        BankProductCreate(
            reason="Record approved savings product",
            name="Akiba Savings",
            product_type="deposit",
            status="approved",
            approval_reference="BOG/PRD/2026/017",
        ),
    )


def _seed_full_register(db: Session) -> dict[str, Any]:
    materialize_canonical_test_book(db)
    _seed_profile(db)
    parties = _seed_parties(db)
    _seed_outlets(db)
    _seed_products(db)
    return parties


def _generate(db: Session, return_code: str) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=return_code, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _sections(package: RegulatoryPackage) -> dict[str, dict[str, Any]]:
    return {section["code"]: section for section in package.snapshot["sections"]}


def _validate_passes(db: Session, package: RegulatoryPackage) -> None:
    read = validation.validate_package(db, MAKER, SAMPLE_BANK_ID, package.id)
    assert read.status == "validated", read.validation_report
    assert read.validation_report is not None
    assert read.validation_report.passed is True


def _export_kinds(db: Session, storage: InMemoryStorageClient, package: RegulatoryPackage) -> bytes:
    """Export xlsx + pdf through the real exporter; returns the xlsx bytes."""
    xlsx = export_package(db, MAKER, package, "xlsx")
    pdf = export_package(db, MAKER, package, "pdf")
    db.commit()
    assert xlsx.size_bytes > 0
    assert pdf.size_bytes > 0
    slug = db.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    payloads: dict[str, bytes] = {}
    for obj in storage.list(slug, "outputs"):
        _, stream = storage.read(obj.location)
        payloads[obj.location.object_path] = stream.read()
    assert payloads[pdf.object_path].startswith(b"%PDF")
    return payloads[xlsx.object_path]


def test_registry_carries_event_driven_corporate_packs() -> None:
    for code in CORPORATE_CODES:
        definition = REGISTRY[code]
        assert definition.family == "corporate"
        assert definition.regulator == "BOG"
        assert definition.default_channel == "orass_sandbox"
        assert definition.event_driven is True
        assert definition.fidelity == "CONFIRMED"
        assert "draft" in definition.directive_citation
    # Periodic returns stay non-event-driven.
    assert REGISTRY["LCR-NSFR"].event_driven is False


def test_lrt_profile_pack_generates_validates_and_exports(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_full_register(db_session)
    package = _generate(db_session, "LRT-PROFILE")

    assert package.return_family == "corporate"
    assert package.snapshot["template_id"] == "bog-lrt-profile-v1"
    assert package.snapshot["fidelity"] == "CONFIRMED"
    assert package.source_runs == []  # master-data pack, not run-derived
    assert package.snapshot["institution"]["orass_institution_code"] == "GH-UB-0042"

    sections = _sections(package)
    assert set(sections) == {
        "general_details",
        "business_activities",
        "exchange_membership",
        "ownership",
        "licenses",
        "name_history",
    }
    general = {row["code"]: row["value"] for row in sections["general_details"]["rows"]}
    assert general["institution_type"] == "Universal Bank"
    assert general["orass_institution_code"] == "GH-UB-0042"
    assert general["approved_capital"] == "400000000.00"
    assert general["authorisation_date"] == "2005-08-15"
    assert general["incorporation_date"] == "2004-11-02"
    ownership = {row["code"]: row["value"] for row in sections["ownership"]["rows"]}
    assert ownership["ownership_local_pct"] == "60.00"
    assert ownership["parent_country_code"] == "GH"
    exchange = {row["code"]: row["value"] for row in sections["exchange_membership"]["rows"]}
    assert exchange["traded_on_exchange"] == "Yes"
    assert [row["description"] for row in sections["licenses"]["rows"]] == [
        "Universal Banking License"
    ]
    assert [row["description"] for row in sections["name_history"]["rows"]] == [
        "Sample Savings & Loans Ltd"
    ]
    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert totals["approved_capital_ghs"] == "400000000.00"
    assert totals["active_license_count"] == "1"

    _validate_passes(db_session, package)
    xlsx_payload = _export_kinds(db_session, storage, package)
    workbook = load_workbook(io.BytesIO(xlsx_payload))
    assert "General Details" in workbook.sheetnames
    assert "Name History" in workbook.sheetnames


def test_lrt_profile_409_without_profile(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LRT-PROFILE")
    assert exc_info.value.status_code == 409
    assert "no_institution_profile" in str(exc_info.value.detail)


def test_lrt_outlet_pack_lists_open_closed_and_checklist(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_full_register(db_session)
    package = _generate(db_session, "LRT-OUTLET")

    sections = _sections(package)
    assert set(sections) == {"outlets_open", "outlets_closed", "required_documents"}
    open_rows = sections["outlets_open"]["rows"]
    assert [row["description"] for row in open_rows] == ["Kumasi Adum Branch"]
    assert open_rows[0]["code"] == "BR-014"
    assert open_rows[0]["opened_on"] == "2026-02-01"
    assert open_rows[0]["address"] == "Kumasi, Adum High Street"
    closed_rows = sections["outlets_closed"]["rows"]
    assert [row["description"] for row in closed_rows] == ["Tema Harbour Agency"]
    assert closed_rows[0]["closed_on"] == "2026-01-31"
    assert closed_rows[0]["relocated_from"] == "Tema Fishing Harbour"
    assert [row["code"] for row in sections["required_documents"]["rows"]] == [
        "RES",
        "RD",
        "SOP",
    ]
    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert totals["outlets_open_count"] == "1"
    assert totals["outlets_closed_count"] == "1"

    _validate_passes(db_session, package)
    xlsx_payload = _export_kinds(db_session, storage, package)
    workbook = load_workbook(io.BytesIO(xlsx_payload))
    assert "Opening of Outlets (OO)" in workbook.sheetnames


def test_lrt_party_pack_resolves_ubo_and_grades_due_diligence(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    parties = _seed_full_register(db_session)
    package = _generate(db_session, "LRT-PARTY")

    sections = _sections(package)
    # One section per category PRESENT (no outsourced/liquidators seeded).
    assert set(sections) == {
        "shareholders",
        "directors",
        "key_management",
        "auditors",
        "ubos",
        "due_diligence_checklist",
    }

    shareholder_rows = sections["shareholders"]["rows"]
    assert len(shareholder_rows) == 1
    row = shareholder_rows[0]
    assert row["description"] == "Golden Coast Holdings Ltd"
    assert row["value"] == "12.5000"
    assert row["share_type"] == "Ordinary"
    assert row["rights"] == "voting"
    assert row["ubo_name"] == "Kwame Owusu"  # UBO party id resolved to a name

    director_rows = sections["directors"]["rows"]
    assert director_rows[0]["description"] == "Ama Mensah"
    assert director_rows[0]["value"] == "director"
    assert director_rows[0]["sitting_allowance"] == "1500.00"
    assert director_rows[0]["annual_fees"] == "24000.00"

    assert sections["key_management"]["rows"][0]["description"] == "Efua Boateng"
    auditor_rows = sections["auditors"]["rows"]
    assert auditor_rows[0]["value"] == "ICAG/F/2026/042"
    assert auditor_rows[0]["regulated_jurisdiction"] == "Ghana (ICAG)"
    assert sections["ubos"]["rows"][0]["description"] == "Kwame Owusu"

    dd_rows = sections["due_diligence_checklist"]["rows"]
    by_kind: dict[str, list[str]] = {"CDD": [], "EDD": [], "PNF": []}
    for item in dd_rows:
        by_kind[item["code"].split("-")[0]].append(item["description"])
    # CDD for every party; EDD + PNF for individuals ONLY.
    assert len(by_kind["CDD"]) == 5
    individuals = {
        parties["owner"].full_name,
        parties["director"].full_name,
        parties["ceo"].full_name,
    }
    for kind in ("EDD", "PNF"):
        assert len(by_kind[kind]) == 3
        for name in individuals:
            assert any(name in description for description in by_kind[kind])
    legal_entities = (parties["holdco"].full_name, parties["auditor"].full_name)
    for kind in ("EDD", "PNF"):
        for name in legal_entities:
            assert not any(name in description for description in by_kind[kind])

    _validate_passes(db_session, package)
    _export_kinds(db_session, storage, package)


def test_lrt_party_409_without_parties(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    _seed_profile(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LRT-PARTY")
    assert exc_info.value.status_code == 409
    assert "no_related_parties" in str(exc_info.value.detail)


def test_lrt_capital_pack_cross_foots_pct_total(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_full_register(db_session)
    package = _generate(db_session, "LRT-CAPITAL")

    sections = _sections(package)
    assert set(sections) == {"shareholder_register", "capital_summary", "capital_checklist"}
    register = sections["shareholder_register"]
    assert [row["value"] for row in register["rows"]] == ["12.5000"]
    assert register["rows"][0]["number_of_shares"] == "1250000.00"
    assert register["rows"][0]["ubo_name"] == "Kwame Owusu"
    assert register["total"]["code"] == "pct_total"
    assert register["total"]["equals_sum_of_rows"] is True
    assert Decimal(register["total"]["value"]) == Decimal("12.5000")

    summary = {row["code"]: row for row in sections["capital_summary"]["rows"]}
    assert summary["approved_capital_ghs"]["value"] == "400000000.00"
    assert summary["approved_capital_ghs"]["unit"] == "ghs"
    assert summary["total_pct_shareholding"]["unit"] == "pct"
    assert [row["code"] for row in sections["capital_checklist"]["rows"]] == [
        "URP",
        "USI",
        "RES",
        "SOP",
    ]

    # The declared pct_total is the ONLY cross-foot in the corporate packs;
    # validation must pass it (and everything else) cleanly.
    _validate_passes(db_session, package)
    _export_kinds(db_session, storage, package)


def test_lrt_capital_409_without_profile(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LRT-CAPITAL")
    assert exc_info.value.status_code == 409
    assert "no_institution_profile" in str(exc_info.value.detail)


def test_lrt_product_pack_orders_by_status_and_exports(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_full_register(db_session)
    package = _generate(db_session, "LRT-PRODUCT")

    sections = _sections(package)
    assert set(sections) == {"products", "product_checklist"}
    rows = sections["products"]["rows"]
    # Sorted by (status, name): approved before proposed.
    assert [(row["description"], row["value"]) for row in rows] == [
        ("Akiba Savings", "approved"),
        ("SME Flex Overdraft", "proposed"),
    ]
    assert rows[0]["approval_reference"] == "BOG/PRD/2026/017"
    assert [row["code"] for row in sections["product_checklist"]["rows"]] == [
        "DN",
        "MOU",
        "RES",
        "SOP",
    ]
    totals = {row["code"]: row["value"] for row in package.snapshot["totals"]}
    assert totals["product_count"] == "2"
    assert totals["approved_product_count"] == "1"

    _validate_passes(db_session, package)
    _export_kinds(db_session, storage, package)


def test_lrt_product_409_without_products(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    _seed_profile(db_session)
    with pytest.raises(HTTPException) as exc_info:
        _generate(db_session, "LRT-PRODUCT")
    assert exc_info.value.status_code == 409
    assert "no_products" in str(exc_info.value.detail)


def test_event_driven_returns_absent_from_calendar(db_session: Session) -> None:
    _seed_full_register(db_session)
    # Even with a generated corporate package on the books, the calendar
    # carries only periodic obligations.
    _generate(db_session, "LRT-PROFILE")
    obligations = calendar.list_obligations(
        db_session, MAKER, SAMPLE_BANK_ID, 12, as_of=date(2026, 4, 5)
    ).obligations
    assert obligations
    codes = {item.return_code for item in obligations}
    assert codes.isdisjoint(CORPORATE_CODES)
    assert "corporate" not in {item.return_family for item in obligations}
    # Periodic returns are unaffected by the skip.
    assert {"LCR-NSFR", "CAR-RWA", "IRRBB-PILOT", "FX-NOP", "ICAAP-STRESS", "LMT"} <= codes


def test_corporate_package_appears_in_package_list(db_session: Session) -> None:
    _seed_full_register(db_session)
    package = _generate(db_session, "LRT-OUTLET")

    listed = packages_service.list_packages(
        db_session, MAKER, SAMPLE_BANK_ID, return_family="corporate"
    )
    assert [item.id for item in listed.packages] == [package.id]
    assert listed.packages[0].return_code == "LRT-OUTLET"

    # read_package tolerates the empty source_runs of a master-data pack.
    read = packages_service.get_package(db_session, MAKER, SAMPLE_BANK_ID, package.id)
    assert read.source_runs == []
