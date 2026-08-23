"""Out-of-band ML-ETL dedup jobs: the handler backfills a deferred batch's
etl_report with linkage/anomaly metadata, is idempotent, is enqueued only when
inline dedup was skipped, and never touches canonical rows.

Tests call the handler directly (never the poll loop), mirroring the other job
suites. Ingestion runs against the in-memory storage client and the handler is
pointed at the same instance so the persisted raw artifact is re-extractable.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.domain.ingestion.contracts import EntityMapping, MappingConfig
from app.models import (
    AuditEvent,
    Bank,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    IngestionBatch,
    Job,
    LineageRecord,
)
from app.schemas.ingestion import IngestionBatchCreate, MappingConfigCreate
from app.services import etl_dedup_jobs, ingestion
from tests.api.helpers import ORG_1, USER_1
from tests.storage.inmemory import InMemoryStorageClient

AS_OF = date(2026, 6, 30)

# A counterparty pair whose shared national id + near-identical name yields one
# CROSS_SOURCE linkage, plus a duplicated loan row the fingerprint detector flags
# as an anomaly — so the dedup pass has real linkage AND anomaly output.
MAPPING = MappingConfig(
    field_mappings={
        "counterparty": EntityMapping(
            source_table="Customers",
            fields={
                "source_reference": "CustomerId",
                "name": "CustomerName",
                "counterparty_type": "Segment",
                "country_code": "Country",
            },
        ),
        "position": EntityMapping(
            source_table="Loans",
            fields={
                "source_reference": "AccountRef",
                "position_type": "Type",
                "currency": "Ccy",
                "balance": "Outstanding",
                "counterparty_reference": "Customer",
                "contractual_maturity": "Maturity",
            },
        ),
    },
    enum_mappings={"counterparty_type": {"RETAIL": "RETAIL_INDIVIDUAL", "CORP": "CORPORATE"}},
)


def _ctx() -> TenantContext:
    return TenantContext(organization_id=ORG_1, actor_user_id=USER_1)


def _bank(db_session: Session) -> Bank:
    bank = Bank(
        organization_id=ORG_1,
        name="ETL Dedup Bank",
        short_name="etl-dedup",
        currency="GHS",
        jurisdiction_code="GH",
        license_type="universal",
        institution_type="universal_bank",
    )
    db_session.add(bank)
    db_session.flush()
    return bank


def _mapping_id(db_session: Session, bank: Bank) -> str:
    created = ingestion.create_mapping_config(
        db_session,
        _ctx(),
        bank.id,
        MappingConfigCreate(
            source_system="EXCEL_CSV",
            name="Dedup test mapping",
            config=MAPPING,
            activate=True,
            reason="etl dedup job tests",
        ),
    )
    return str(created.id)


def _workbook(path: Path) -> Path:
    workbook = Workbook()
    customers = workbook.active
    assert customers is not None
    customers.title = "Customers"
    customers.append(["CustomerId", "CustomerName", "Segment", "Country", "NationalId"])
    customers.append(["C-001", "ACME TRADING LTD", "CORP", "GH", "GHA-000111"])
    customers.append(["C-002", "Acme Trading Limited", "CORP", "GH", "GHA-000111"])
    customers.append(["C-003", "Kwame Mensah", "RETAIL", "GH", "GHA-999"])
    loans = workbook.create_sheet("Loans")
    loans.append(["AccountRef", "Type", "Ccy", "Outstanding", "Customer", "Maturity"])
    loans.append(["LN-0001", "LOAN", "GHS", 1000, "C-001", date(2031, 3, 15)])
    loans.append(["LN-0001", "LOAN", "GHS", 1000, "C-001", date(2031, 3, 15)])
    workbook.save(path)
    return path


def _ingest(
    db_session: Session, storage: InMemoryStorageClient, bank: Bank, location: Path
) -> IngestionBatch:
    result = ingestion.start_ingestion(
        db_session,
        _ctx(),
        bank.id,
        IngestionBatchCreate(
            source_system="EXCEL_CSV",
            as_of_date=AS_OF,
            location=str(location),
            mapping_config_id=None,
            reason="dedup job test ingestion",
        ),
        storage,
    )
    batch = db_session.get(IngestionBatch, result.batch.id)
    assert batch is not None
    return batch


def _dedup_job(db_session: Session, batch: IngestionBatch) -> Job:
    jobs = db_session.scalars(
        select(Job).where(
            Job.organization_id == ORG_1,
            Job.job_type == etl_dedup_jobs.ETL_DEDUP,
        )
    ).all()
    matches = [job for job in jobs if job.payload.get("batch_id") == str(batch.id)]
    assert matches, "expected an etl_dedup job to be enqueued"
    return matches[0]


def _force_defer(monkeypatch: pytest.MonkeyPatch, storage: InMemoryStorageClient) -> None:
    """Skip inline dedup for any batch, and point the handler at ``storage``."""
    monkeypatch.setattr(ingestion, "_ETL_INLINE_DEDUP_MAX_RECORDS", 0)
    monkeypatch.setattr(etl_dedup_jobs, "get_storage_client", lambda: storage)


def _dedup_lineage_nodes(db_session: Session, batch: IngestionBatch) -> list[LineageRecord]:
    """The out-of-band ML_ETL_DEDUP nodes (deferred_pass) this batch carries."""
    nodes = db_session.scalars(
        select(LineageRecord).where(
            LineageRecord.organization_id == ORG_1,
            LineageRecord.ingestion_batch_id == batch.id,
            LineageRecord.operation_type == "ML_ETL_DEDUP",
        )
    ).all()
    return [node for node in nodes if node.details.get("deferred_pass") is True]


def test_deferred_batch_enqueues_job_and_marks_report(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)

    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))

    assert batch.etl_report is not None
    assert batch.etl_report["dedup_status"] == "deferred"
    # No dedup/anomaly pass ran inline — the linkage keys are zeroed placeholders.
    assert batch.etl_report["linkage_count"] == 0
    assert batch.etl_report["sample_linkages"] == []
    # A job is queued to backfill it.
    job = _dedup_job(db_session, batch)
    assert job.payload["batch_id"] == str(batch.id)


def test_run_etl_dedup_backfills_report_lineage_and_audit(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))
    assert batch.etl_report is not None
    preprocess_keys = {
        key: batch.etl_report[key]
        for key in ("record_count", "operation_count", "sanctioned_count", "flagged_count")
    }
    job = _dedup_job(db_session, batch)

    etl_dedup_jobs.run_etl_dedup(db_session, job)

    db_session.refresh(batch)
    report = batch.etl_report
    assert report is not None
    # Dedup + anomaly metadata is now present.
    assert report["dedup_status"] == "completed"
    assert report["linkage_count"] == 1
    assert report["linkages_by_match_type"]["CROSS_SOURCE"] == 1
    assert report["anomaly_count"] == 2
    assert "sample_anomalies" in report
    # The inline preprocessing summary was merged into, not clobbered.
    for key, value in preprocess_keys.items():
        assert report[key] == value

    # An ML_ETL_DEDUP lineage node with the real counts was appended.
    nodes = _dedup_lineage_nodes(db_session, batch)
    assert len(nodes) == 1
    assert nodes[0].details["linkages"] == 1
    assert nodes[0].details["anomalies"] == 2
    assert nodes[0].details["content_hash_match"] is True

    # A dedup-completed audit event carries the merged report.
    event = db_session.scalar(
        select(AuditEvent).where(
            AuditEvent.organization_id == ORG_1,
            AuditEvent.event_type == "ml_etl.dedup_completed",
            AuditEvent.entity_id == str(batch.id),
        )
    )
    assert event is not None
    assert event.details["linkage_count"] == 1

    # The job records the real counts in its progress.
    assert job.progress["status"] == "completed"
    assert job.progress["linkage_count"] == 1
    assert job.progress["anomaly_count"] == 2


def test_run_etl_dedup_is_idempotent(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))
    job = _dedup_job(db_session, batch)

    etl_dedup_jobs.run_etl_dedup(db_session, job)
    etl_dedup_jobs.run_etl_dedup(db_session, job)  # re-run must not double-count or error

    db_session.refresh(batch)
    assert batch.etl_report is not None
    assert batch.etl_report["linkage_count"] == 1  # not doubled
    assert job.progress["status"] == "already_completed"
    # Exactly one out-of-band lineage node and one audit event, not two.
    assert len(_dedup_lineage_nodes(db_session, batch)) == 1
    events = db_session.scalars(
        select(AuditEvent).where(
            AuditEvent.organization_id == ORG_1,
            AuditEvent.event_type == "ml_etl.dedup_completed",
            AuditEvent.entity_id == str(batch.id),
        )
    ).all()
    assert len(events) == 1


def test_inline_dedup_does_not_enqueue_a_job(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # Default threshold: the 5-record batch is deduped inline, nothing deferred.
    storage = InMemoryStorageClient()
    monkeypatch.setattr(etl_dedup_jobs, "get_storage_client", lambda: storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)

    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))

    assert batch.etl_report is not None
    assert batch.etl_report["dedup_status"] == "completed"
    assert batch.etl_report["linkage_count"] == 1  # inline pass produced it
    jobs = db_session.scalars(
        select(Job).where(Job.organization_id == ORG_1, Job.job_type == etl_dedup_jobs.ETL_DEDUP)
    ).all()
    assert jobs == []


def test_run_etl_dedup_never_mutates_canonical_rows(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))

    def _canonical_fingerprint() -> tuple[set[tuple[str, str]], set[tuple[str, str]]]:
        counterparties = {
            (str(cp.id), cp.name)
            for cp in db_session.scalars(
                select(CanonicalCounterparty).where(
                    CanonicalCounterparty.organization_id == ORG_1,
                    CanonicalCounterparty.bank_id == bank.id,
                    CanonicalCounterparty.superseded_by.is_(None),
                )
            )
        }
        snapshots = {
            (str(s.id), str(s.balance))
            for s in db_session.scalars(
                select(CanonicalPositionSnapshot).where(
                    CanonicalPositionSnapshot.organization_id == ORG_1,
                    CanonicalPositionSnapshot.bank_id == bank.id,
                    CanonicalPositionSnapshot.superseded_by.is_(None),
                )
            )
        }
        return counterparties, snapshots

    before = _canonical_fingerprint()
    assert before[0]  # the batch really did land canonical rows

    etl_dedup_jobs.run_etl_dedup(db_session, _dedup_job(db_session, batch))

    assert _canonical_fingerprint() == before


# ---------------------------------------------------------------------------
# Reliability: a failed pass must be visible, bounded, and retryable
# ---------------------------------------------------------------------------
#
# Every case below reproduces something observed on the primary database, where
# etl_dedup stood at 4 succeeded / 5 failed — and since inline dedup effectively
# never runs at core-banking scale (107,704 deposit rows against a 5,000-record
# inline bound), this deferred job is the ONLY dedup path a real bank has.


def _boom(*_args: object, **_kwargs: object) -> None:
    msg = "server closed the connection unexpectedly"
    raise RuntimeError(msg)


def test_failed_dedup_marks_the_batch_instead_of_leaving_it_pending(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """A raising handler used to leave ``dedup_status="deferred"`` for ever.

    Four batches on the primary read as "deferred" weeks after their jobs had
    exhausted every attempt — indistinguishable from "queued, will run shortly".
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))
    job = _dedup_job(db_session, batch)
    monkeypatch.setattr(etl_dedup_jobs, "run_etl", _boom)

    with pytest.raises(RuntimeError):
        etl_dedup_jobs.run_etl_dedup(db_session, job)

    db_session.refresh(batch)
    assert batch.etl_report is not None
    assert batch.etl_report["dedup_status"] == etl_dedup_jobs.DEDUP_STATUS_FAILED
    assert "server closed the connection unexpectedly" in batch.etl_report["dedup_error"]
    assert batch.etl_report["dedup_failed_at"]
    event = db_session.scalar(
        select(AuditEvent).where(
            AuditEvent.organization_id == ORG_1,
            AuditEvent.event_type == etl_dedup_jobs.DEDUP_FAILED_EVENT,
            AuditEvent.entity_id == str(batch.id),
        )
    )
    assert event is not None


def test_a_failed_batch_can_still_complete_on_a_later_run(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """The marker records what happened; it never bars the batch from completing.

    One of the five primary failures was a stored mapping config that a newer
    writer had produced and the running worker's schema could not validate. The
    code healed, but the job sat at attempts 3/3 with nothing to re-drive it.
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))
    job = _dedup_job(db_session, batch)
    monkeypatch.setattr(etl_dedup_jobs, "run_etl", _boom)
    with pytest.raises(RuntimeError):
        etl_dedup_jobs.run_etl_dedup(db_session, job)
    monkeypatch.undo()
    monkeypatch.setattr(etl_dedup_jobs, "get_storage_client", lambda: storage)

    etl_dedup_jobs.run_etl_dedup(db_session, job)

    db_session.refresh(batch)
    assert batch.etl_report is not None
    assert batch.etl_report["dedup_status"] == etl_dedup_jobs.DEDUP_STATUS_COMPLETED
    assert batch.etl_report["linkage_count"] == 1


def test_over_budget_counterparty_matching_is_reported_not_silently_dropped(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """A skipped pass reporting zero must never read as a pass that found zero.

    Three primary failures were "worker presumed dead" reclaims: the pairwise,
    model-scored counterparty pass runs for hours on a core-banking customer file
    (measured ~370us per candidate pair), far past the 900s stale-job window, and
    the one job that eventually succeeded took 2h02m — by which time the reaper
    had requeued it twice and the same handler was running concurrently.
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    monkeypatch.setattr(etl_dedup_jobs, "_DEFERRED_COUNTERPARTY_MAX_RECORDS", 0)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _workbook(tmp_path / "book.xlsx"))

    etl_dedup_jobs.run_etl_dedup(db_session, _dedup_job(db_session, batch))

    db_session.refresh(batch)
    report = batch.etl_report
    assert report is not None
    assert report["dedup_status"] == etl_dedup_jobs.DEDUP_STATUS_COMPLETED
    assert report["counterparty_matching"] == "skipped_over_budget"
    assert report["counterparty_match_budget"] == 0
    # The counterparty linkage the full pass finds is absent, and the report says
    # why rather than presenting the absence as a finding of none.
    assert report["linkages_by_match_type"]["CROSS_SOURCE"] == 0
    # The linear position/anomaly passes still ran.
    assert report["anomaly_count"] == 2


# ---------------------------------------------------------------------------
# Cross-source position matching (row level; detection only)
# ---------------------------------------------------------------------------


def _clean_workbook(path: Path) -> Path:
    """The same book WITHOUT the deliberate duplicate row.

    ``_workbook`` repeats ``LN-0001`` so the fingerprint detector has an anomaly
    to find — but a duplicated row fails validation, and a row in ``error`` is
    excluded from the population every calculation (and this pass) reads. The
    cross-source tests need a book that actually lands in the current generation.
    """
    workbook = Workbook()
    customers = workbook.active
    assert customers is not None
    customers.title = "Customers"
    customers.append(["CustomerId", "CustomerName", "Segment", "Country", "NationalId"])
    customers.append(["C-001", "ACME TRADING LTD", "CORP", "GH", "GHA-000111"])
    customers.append(["C-002", "Acme Trading Limited", "CORP", "GH", "GHA-000111"])
    loans = workbook.create_sheet("Loans")
    loans.append(["AccountRef", "Type", "Ccy", "Outstanding", "Customer", "Maturity"])
    loans.append(["LN-0001", "LOAN", "GHS", 1000, "C-001", date(2031, 3, 15)])
    workbook.save(path)
    return path


def _seed_second_source_book(db_session: Session, bank: Bank, *, reference: str) -> str:
    """A SECOND source system carrying the same arrangement id as the ingested book.

    This is the shape BK-0PMD7Z5M holds on the primary: DB_DIRECT and EXCEL_CSV
    share the SAME ``source_reference`` for 150,314 positions, and supersession is
    scoped per source system, so both live books survive in full and the exposure
    is carried twice.
    """
    batch = IngestionBatch(
        organization_id=ORG_1,
        bank_id=bank.id,
        source_system="API_PUSH",
        adapter_version="1.0",
        extraction_mode="full",
        status="accepted",
        as_of_date=AS_OF,
    )
    db_session.add(batch)
    db_session.flush()
    lineage = LineageRecord(
        organization_id=ORG_1,
        ingestion_batch_id=batch.id,
        operation_type="ADAPTER_TRANSLATE",
        operation_ref="second-source-fixture",
        input_lineage_ids=[],
    )
    db_session.add(lineage)
    db_session.flush()
    common = {
        "organization_id": ORG_1,
        "bank_id": bank.id,
        "as_of_date": AS_OF,
        "source_system": "API_PUSH",
        "ingestion_batch_id": batch.id,
        "lineage_id": lineage.id,
        "validation_status": "accepted",
        "source_reference": reference,
    }
    position = CanonicalPosition(**common, position_type="LOAN", currency="GHS")
    db_session.add(position)
    db_session.flush()
    db_session.add(
        CanonicalPositionSnapshot(
            **common,
            position_id=position.id,
            balance=Decimal("1000"),
            attributes={"balance_ghs": "1000"},
        )
    )
    db_session.flush()
    return str(position.id)


def test_single_source_book_produces_no_cross_source_finding(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """A matcher that fires on a healthy book is worse than no matcher."""
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _clean_workbook(tmp_path / "book.xlsx"))

    etl_dedup_jobs.run_etl_dedup(db_session, _dedup_job(db_session, batch))

    db_session.refresh(batch)
    assert batch.etl_report is not None
    cross = batch.etl_report["cross_source"]
    assert cross["contested_position_types"] == []
    assert cross["linkage_count"] == 0
    assert "outcome" not in cross


def test_two_systems_holding_one_position_are_matched_row_by_row(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """The gap ``run_etl`` cannot close: the duplicate arrives in another batch.

    ``run_etl`` is pure and sees one extraction, and ``PositionDeduplicator``
    groups on ``source_reference`` WITHIN that extraction. The duplicate book is
    a different batch entirely — a month apart on the primary — so only a pass
    with database access can see it.
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _clean_workbook(tmp_path / "book.xlsx"))
    twin_id = _seed_second_source_book(db_session, bank, reference="LN-0001")

    etl_dedup_jobs.run_etl_dedup(db_session, _dedup_job(db_session, batch))

    db_session.refresh(batch)
    assert batch.etl_report is not None
    cross = batch.etl_report["cross_source"]
    assert cross["contested_position_types"] == ["LOAN"]
    assert cross["linkage_count"] == 1
    link = cross["sample_linkages"][0]
    assert link["match_type"] == "CROSS_SOURCE"
    assert twin_id in link["linked_source_ids"]
    assert len(link["linked_source_ids"]) == 2
    # DETECTION ONLY: the evidence names no authoritative system.
    assert link["system_of_record"] == "undetermined"
    assert link["signals"]["system_of_record_determined"] == 0.0
    # The advisory data-quality outcome, in the shared fail-closed vocabulary.
    outcome = cross["outcome"]
    assert outcome["state"] == "data_quality_block"
    assert outcome["metric_id"] == etl_dedup_jobs.CROSS_SOURCE_METRIC_ID
    assert outcome["advisory"] is True
    assert outcome["blocks_filing"] is False
    # Operator-facing prose, never enum values.
    assert "API_PUSH" not in outcome["reason"]
    assert "loans" in outcome["reason"]
    # And it is in the lineage node the pass appends.
    node = _dedup_lineage_nodes(db_session, batch)[0]
    assert node.details["cross_source_linkages"] == 1
    assert node.details["cross_source_contested_types"] == ["LOAN"]


def test_cross_source_pass_never_mutates_or_retires_a_position(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """No source row may be destroyed or rewritten: both books survive in full.

    Auto-resolution here would DELETE A REAL BOOK — a bank legitimately splits
    its book across systems, and during a core-banking migration both copies are
    live. Choosing a winner belongs to the system-of-record register; performing
    the withdrawal is a separate, separately approved act.
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    batch = _ingest(db_session, storage, bank, _clean_workbook(tmp_path / "book.xlsx"))
    _seed_second_source_book(db_session, bank, reference="LN-0001")

    def _live_positions() -> set[tuple[str, str, str]]:
        return {
            (p.source_system, p.source_reference, str(p.superseded_by))
            for p in db_session.scalars(
                select(CanonicalPosition).where(
                    CanonicalPosition.organization_id == ORG_1,
                    CanonicalPosition.bank_id == bank.id,
                )
            )
        }

    before = _live_positions()
    assert len({system for system, _, _ in before}) == 2

    etl_dedup_jobs.run_etl_dedup(db_session, _dedup_job(db_session, batch))

    assert _live_positions() == before
    withdrawn = db_session.scalars(
        select(CanonicalPositionSnapshot).where(
            CanonicalPositionSnapshot.organization_id == ORG_1,
            CanonicalPositionSnapshot.bank_id == bank.id,
            CanonicalPositionSnapshot.withdrawn_at.is_not(None),
        )
    ).all()
    assert withdrawn == []


def test_assess_cross_source_positions_hands_the_full_evidence_set_over(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """The resolution layer's entry point: every linkage, not a sample.

    A withdrawal needs the complete set of position ids that are provably held
    twice — five samples in a report is evidence of a problem, not a basis for
    removing rows. Read-only and side-effect free, like
    ``fact_derivation.diagnose_source_overlap``, so it can be called at decision
    time rather than read back from a stored verdict that may have gone stale.
    """
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    _ingest(db_session, storage, bank, _clean_workbook(tmp_path / "book.xlsx"))
    twin_id = _seed_second_source_book(db_session, bank, reference="LN-0001")
    ctx = _ctx()

    assessment = etl_dedup_jobs.assess_cross_source_positions(db_session, ctx, bank.id, AS_OF)

    assert assessment.contested_position_types == ("LOAN",)
    assert len(assessment.result.linkages) == 1
    assert twin_id in assessment.result.matched_row_ids
    # The book-level diagnosis it is bounded by is carried through unmodified, so
    # the two layers can never disagree about what is contested.
    assert assessment.overlap.overlapping is True
    assert assessment.detail is not None
    assert assessment.detail.blocks_filing is False
    assert assessment.report()["linkage_count"] == 1


def test_assess_cross_source_positions_is_silent_on_a_single_source_bank(
    db_session: Session, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    storage = InMemoryStorageClient()
    _force_defer(monkeypatch, storage)
    bank = _bank(db_session)
    _mapping_id(db_session, bank)
    _ingest(db_session, storage, bank, _clean_workbook(tmp_path / "book.xlsx"))

    assessment = etl_dedup_jobs.assess_cross_source_positions(db_session, _ctx(), bank.id, AS_OF)

    assert assessment.contested_position_types == ()
    assert assessment.result.linkages == ()
    assert assessment.detail is None
    assert "outcome" not in assessment.report()
