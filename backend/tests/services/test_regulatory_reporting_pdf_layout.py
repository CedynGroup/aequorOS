"""The filed PDF's grid layout: no figure may wrap, no word may split, and no
snapshot key may print on the page.

Why this file exists (founder review, 2026-08-23)
-------------------------------------------------
``_section_table`` used to hand reportlab no ``colWidths``, so every column was
sized by its widest content. On LMTD Table 2 — 17 columns of GHS figures across
A4 landscape — one cell decided the whole layout: the section total carries the
snapshot key ``on_balance_mismatch_total_ghs`` in the "#" column, roughly
fifteen times the width of BoG's own row numbers 1 to 17 beneath it. That column
took the space, the thirteen maturity buckets were starved, and the filed pack
came out with headers reading "15 da ys to 1 mt h" and figures broken as
"601,209." / "47".

A number split across two lines on a regulatory return is a misread waiting to
happen, and it is the kind of defect that is obvious in a rendered PDF and
invisible in a unit test of the renderer's inputs. So these tests read the
actual bytes back with pdfplumber and assert on what is on the page.
"""

from __future__ import annotations

import io
import re
from dataclasses import replace
from datetime import UTC, datetime
from decimal import Decimal

import pdfplumber
import pytest
from openpyxl import load_workbook

from app.services.regulatory_reporting.exports.csv import render_csv
from app.services.regulatory_reporting.exports.pdf import (
    _LANDSCAPE,
    _MARGINS,
    render_pdf,
)
from app.services.regulatory_reporting.exports.xlsx import render_xlsx
from app.services.regulatory_reporting.le_generation import _TABLE2_ROW_LABELS
from app.services.regulatory_reporting.templates import (
    TEMPLATES,
    RenderedCell,
    RenderedReturn,
    RenderedRow,
    RenderedSection,
    SectionLayout,
)

#: The total row's machine key. It stays in the snapshot and the CSV — the
#: ``equals_sum_of_rows`` validation and ``_snapshot_total`` bind to it — and
#: must never reach the printed page.
TOTAL_CODE = "on_balance_mismatch_total_ghs"

#: A whole money cell, as ``format_cell`` writes it.
WHOLE_FIGURE = re.compile(r"^\(?-?[\d,]*\d\.\d{2}\)?$")
#: The head of a figure cut across lines ("601,209." or "1,234,").
CUT_HEAD = re.compile(r"^\(?-?[\d,]*\d[,.]$")
#: Its continuation on the next line ("47", ".47").
CUT_TAIL = re.compile(r"^\d{1,3}\)?$|^\.\d{2}\)?$")


def _table2_layout() -> SectionLayout:
    template = TEMPLATES["bog-sdi-lmt-monthly-v1"]
    return next(s for s in template.sections if s.section_code == "maturity_ladder")


def _row(layout: SectionLayout, code: str, label: str, *, is_total: bool = False) -> RenderedRow:
    """One Table 2 row, with figures wide enough to be the binding constraint.

    Eight significant digits is a realistic GHS '000 amount for a bank of any
    size and the widest a cell has to hold; sizing the grid for it is the whole
    point.
    """
    cells: list[RenderedCell] = []
    for spec in layout.columns:
        if spec.key == "code":
            cells.append(RenderedCell(kind="text", value=code))
        elif spec.key == "description":
            cells.append(RenderedCell(kind="text", value=label))
        elif is_total and spec.key != "value":
            cells.append(RenderedCell(kind=spec.kind, value=None))
        else:
            cells.append(RenderedCell(kind="ghs", value=Decimal("98765432.10")))
    return RenderedRow(cells=tuple(cells), is_total=is_total)


@pytest.fixture
def table2() -> RenderedReturn:
    """LMTD Table 2 with the identifiers the generator really emits: BoG's
    printed numbers on the body rows, a snapshot key on the total."""
    layout = _table2_layout()
    section = RenderedSection(
        layout=layout,
        title=layout.sheet_title,
        rows=tuple(_row(layout, number, label) for number, label in _TABLE2_ROW_LABELS),
        total_row=_row(
            layout,
            TOTAL_CODE,
            "On-balance sheet contractual mismatch — total (item 1 less item 5)",
            is_total=True,
        ),
    )
    return RenderedReturn(
        template=TEMPLATES["bog-sdi-lmt-monthly-v1"],
        metadata_pairs=(("Return", "SDI-LMT-MONTHLY"), ("Institution", "Test S&L")),
        sections=(section,),
        provenance_runs=(("liquidity", "run-1", "ab" * 32, "v1"),),
        provenance_lines=("Generated from the live snapshot.",),
    )


def _grid_words(payload: bytes) -> list[list[dict[str, float | str]]]:
    """Words on each landscape page — the pages that carry grids."""
    with pdfplumber.open(io.BytesIO(payload)) as pdf:
        return [page.extract_words() for page in pdf.pages if page.width > page.height]


def _source_vocabulary(rendered: RenderedReturn) -> set[str]:
    """Every word the return hands the renderer, lowercased.

    Headings, notes, citations, metadata, attestation lines, column headers and
    cell values — anything that can legitimately appear on the page. A word the
    page shows that is not in here is a piece of a longer one.
    """
    text: list[str] = [
        # Chrome the renderer adds itself, not carried on the return: the page
        # footer, the section sub-heading, and the fidelity grade.
        "Page Source Layout Fidelity Note",
        rendered.template.currency_unit,
        rendered.template.fidelity,
        rendered.template.title,
        rendered.template.return_code,
        rendered.template.source_citation,
        *rendered.template.notes,
        *rendered.attestation_lines,
        *rendered.provenance_lines,
        *(part for pair in rendered.metadata_pairs for part in pair),
    ]
    for section in rendered.sections:
        text.extend(
            [
                section.title,
                section.layout.sheet_title,
                section.layout.source_citation,
                section.layout.fidelity,
            ]
        )
        text.extend(section.layout.notes)
        text.extend(column.header for column in section.layout.columns)
        for row in [*section.rows, *([section.total_row] if section.total_row else [])]:
            text.extend(str(cell.value) for cell in row.cells if cell.kind == "text")
    return {word.lower() for line in text for word in re.findall(r"[A-Za-z]+", line)}


def _split_figures(words: list[dict[str, float | str]]) -> list[str]:
    """Figures broken across two lines.

    Prose cites years and paragraph numbers ("Feb 2026,", "para 17."), which
    look exactly like a cut figure on their own. A real wrap leaves a PAIR — the
    head with its tail directly beneath, inside the same column — so only pairs
    count.
    """
    found = []
    for head in words:
        if not CUT_HEAD.match(str(head["text"])):
            continue
        for tail in words:
            if tail is head or not CUT_TAIL.match(str(tail["text"])):
                continue
            below = 0 < float(tail["top"]) - float(head["top"]) < 14
            overlaps = float(tail["x0"]) < float(head["x1"]) and float(head["x0"]) < float(
                tail["x1"]
            )
            if below and overlaps:
                found.append(f"{head['text']}|{tail['text']}")
                break
    return found


class TestTable2Layout:
    def test_no_figure_wraps_across_lines(self, table2: RenderedReturn) -> None:
        payload = render_pdf(table2, sandbox_watermark=False)
        splits = [split for words in _grid_words(payload) for split in _split_figures(words)]
        assert not splits, (
            f"{len(splits)} figures broke across two lines on the filed page, e.g. "
            f"{splits[:5]}. A column is narrower than the amounts it carries — check "
            "_column_bounds and _fit_font_size in exports/pdf.py."
        )

    def test_every_figure_renders_whole(self, table2: RenderedReturn) -> None:
        """17 body rows plus the total, across the 15 numeric columns.

        Counting them guards the inverse of the test above: a layout that
        dropped or overprinted cells would have no split pairs either.
        """
        payload = render_pdf(table2, sandbox_watermark=False)
        whole = sum(
            1
            for words in _grid_words(payload)
            for word in words
            if WHOLE_FIGURE.match(str(word["text"]))
        )
        numeric = sum(1 for spec in _table2_layout().columns if spec.kind == "ghs")
        assert whole == len(_TABLE2_ROW_LABELS) * numeric + 1

    def test_no_word_splits_inside_itself(self, table2: RenderedReturn) -> None:
        """Every column holds its longest word.

        Below that width reportlab stops wrapping at spaces and breaks inside
        the word — "15 da ys to 1 mt h" in a header, "provided to o /
        ff-balance-sh / eet" in a label — which reads as a broken render on a
        document an institution signs.

        The check is vocabulary-based: every word the page prints must be a word
        the return was given. A fragment never is.
        """
        payload = render_pdf(table2, sandbox_watermark=False)
        source = _source_vocabulary(table2)
        printed = {str(word["text"]) for words in _grid_words(payload) for word in words}
        fragments = sorted(
            word
            for word in printed
            if len(word) > 1 and word.isalpha() and word.lower() not in source
        )
        assert not fragments, (
            f"words split mid-token on the page: {fragments}. A column is "
            "narrower than the longest word it carries — check _column_bounds "
            "and _FIT_SLACK in exports/pdf.py."
        )

    def test_the_grid_stays_inside_the_page_frame(self, table2: RenderedReturn) -> None:
        """Nothing is printed outside the margins.

        This is the failure mode reportlab's own column sizing produces when the
        content is wider than the paper: it does not compress the table, it
        renders it past the right edge, and the last buckets are simply not on
        the page. Every figure then reads as "whole" to a text extractor while
        being invisible to the regulator.
        """
        payload = render_pdf(table2, sandbox_watermark=False)
        left = _MARGINS["leftMargin"]
        right = _LANDSCAPE[0] - _MARGINS["rightMargin"]
        overflow = [
            (str(word["text"]), round(float(word["x1"])))
            for words in _grid_words(payload)
            for word in words
            # A half-point of tolerance for the glyph-metric rounding reportlab
            # applies when it places the run.
            if float(word["x1"]) > right + 0.5 or float(word["x0"]) < left - 0.5
        ]
        assert not overflow, (
            f"{len(overflow)} words printed outside the {left:.0f}-{right:.0f}pt "
            f"frame, e.g. {overflow[:5]}. The table is wider than the page."
        )


class TestTotalRowIdentifier:
    def test_snapshot_key_is_not_printed(self, table2: RenderedReturn) -> None:
        """The total's ``code`` is a machine key, not a BoG row number.

        Printed in the "#" column beneath rows 1 to 17 it reads as an internal
        variable leaking onto a return — and it is what forced that column wide
        enough to starve the maturity buckets.
        """
        payload = render_pdf(table2, sandbox_watermark=False)
        with pdfplumber.open(io.BytesIO(payload)) as pdf:
            text = "".join((page.extract_text() or "") for page in pdf.pages)
        assert TOTAL_CODE not in text.replace(" ", "").replace("\n", "")

    def test_bog_row_numbers_are_printed(self, table2: RenderedReturn) -> None:
        """Suppressing the key must not suppress BoG's own row numbering."""
        payload = render_pdf(table2, sandbox_watermark=False)
        printed = {str(word["text"]) for words in _grid_words(payload) for word in words}
        missing = [number for number, _ in _TABLE2_ROW_LABELS if number not in printed]
        assert not missing, f"BoG row numbers absent from the '#' column: {missing}"

    def test_no_snake_case_identifier_reaches_the_page(self) -> None:
        """No word on the grid carries an underscore.

        A blunter statement of the rule above, and the one that survives the key
        being WRAPPED: once the "#" column is narrow the key splits across
        lines, so searching the extracted text for it whole finds nothing even
        though its pieces are printed. Searching for the underscore finds them.

        The layout's own notes cite ingestion fields (``deposit_account_type``),
        so they are dropped here. The sheet sub-heading prints
        ``Layout: <layout_id>`` on purpose — that is audit provenance, tying the
        printed sheet to the structure that produced it — so it is the one
        identifier allowed through.
        """
        layout = replace(_table2_layout(), notes=())
        section = RenderedSection(
            layout=layout,
            title=layout.sheet_title,
            rows=(_row(layout, "1", "Contractual maturity of assets (items 2 to 4)"),),
            total_row=_row(layout, TOTAL_CODE, "Total", is_total=True),
        )
        rendered = RenderedReturn(
            template=TEMPLATES["bog-sdi-lmt-monthly-v1"],
            metadata_pairs=(("Return", "SDI-LMT-MONTHLY"),),
            sections=(section,),
            provenance_runs=(),
            # render_pdf reads provenance_lines[0] for the page footer, so it
            # cannot be empty — every real caller supplies one.
            provenance_lines=("Generated from the live snapshot.",),
        )
        payload = render_pdf(rendered, sandbox_watermark=False)
        printed = [
            str(word["text"])
            for words in _grid_words(payload)
            for word in words
            if "_" in str(word["text"]) and str(word["text"]) != layout.layout_id
        ]
        assert not printed, f"snake_case identifiers printed on the filed page: {printed}"

    def test_the_xlsx_audit_twin_agrees_with_the_page(self, table2: RenderedReturn) -> None:
        """The workbook is the PDF's audit twin, so the two must show the same
        thing in the "#" column. A cell that differs between a filed pack and
        its twin is exactly what an examiner reconciling them would flag."""
        payload = render_xlsx(table2, generated_at=datetime(2026, 6, 30, tzinfo=UTC))
        workbook = load_workbook(io.BytesIO(payload))
        printed = {
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert TOTAL_CODE not in printed

    def test_machine_artifacts_keep_the_key(self, table2: RenderedReturn) -> None:
        """The CSV is the machine interchange format — consumers key on the
        total's code there — so the suppression belongs to the two human-facing
        artifacts alone."""
        payload, extension = render_csv(table2)
        assert extension == "csv"
        assert TOTAL_CODE in payload.decode("utf-8")
