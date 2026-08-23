"""ICAAP stress-test submission — BoG Appendix II Tables 1–6 (docs/stress.md §3.8).

Phase 5 end-to-end over the deterministic seeded book: a Board-ATTESTED
enterprise-stress run is re-tabulated into the exact Appendix II table structure,
generated on the real package lifecycle (maker-checker, immutability, content
digest, default signing policy) and exported to pdf/xlsx. Also proves the
governance gate: the return REFUSES without a Board-attested enterprise-stress
run, and the with/without-management-actions blocks appear only when the run
modelled an approved management-actions plan (¶67(f)).
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any
from uuid import UUID

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.models import Bank, BankReportingPeriod, RegulatoryPackage, User
from app.schemas.enterprise_stress import EnterpriseStressRunCreate
from app.schemas.enterprise_stress_signoff import (
    StressSignoffAttestation,
    StressSignoffCreate,
    StressSignoffTransition,
)
from app.schemas.management_actions import (
    ManagementActionPlanApproval,
    ManagementActionPlanCreate,
    ManagementActionPlanTransition,
)
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.schemas.stress import MacroScenarioApproval, MacroScenarioCreate, MacroScenarioTransition
from app.services import (
    enterprise_stress,
    enterprise_stress_signoff,
    macro_scenarios,
    management_action_plans,
)
from app.services.attestation import workflow as attestation
from app.services.regulatory_reporting import generation
from app.services.regulatory_reporting.exports import export_package
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
CHECKER_ID = UUID("bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb")
CHECKER = TenantContext(
    organization_id=DEMO_ORG_ID, actor_user_id=CHECKER_ID, roles=("approver",)
)
REPORTING_DATE = date(2026, 3, 31)
RETURN_CODE = "ICAAP-STRESS-APPENDIX2"
SDI_RETURN_CODE = "SDI-STRESS-ANNUAL"


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _period_id(db: Session) -> UUID:
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    return period_id


def _seed_checker(db: Session) -> None:
    if db.get(User, CHECKER_ID) is not None:
        return
    db.add(
        User(
            id=CHECKER_ID,
            organization_id=DEMO_ORG_ID,
            email="stress-board@aequoros.example",
            display_name="Board Member",
            role="approver",
        )
    )
    db.commit()


def _severe_paths() -> list[dict[str, str | int]]:
    levels = {
        "gdp_growth": ("0.05", "0.00"),
        "interest_rate": ("0.20", "0.25"),
        "inflation": ("0.15", "0.21"),
        "unemployment": ("0.06", "0.09"),
        "fx_usd_ghs": ("12.5", "15.0"),
        "gse_index": ("5000", "3500"),
        "gog_yield": ("0.22", "0.26"),
    }
    paths: list[dict[str, str | int]] = []
    for variable, (base, stress) in levels.items():
        for year in (1, 2, 3):
            paths.append(
                {
                    "variable": variable,
                    "year_index": year,
                    "base_value": base,
                    "stress_value": stress,
                }
            )
    return paths


def _approved_scenario(db: Session, code: str = "adverse_2027") -> UUID:
    created = macro_scenarios.create_scenario(
        db,
        MAKER,
        MacroScenarioCreate.model_validate(
            {
                "code": code,
                "name": "2027 severe downturn",
                "scenario_type": "adverse",
                "severity": "severe",
                "horizon_years": 3,
                "narrative": "GDP contraction, cedi depreciation, rate spike.",
                "source": "BoG MPC + internal desk",
                "paths": _severe_paths(),
                "reason": "Author the annual adverse scenario.",
            }
        ),
    )
    macro_scenarios.submit_scenario(
        db, MAKER, created.id, MacroScenarioTransition(reason="Ready for approval.")
    )
    macro_scenarios.approve_scenario(
        db, CHECKER, created.id, MacroScenarioApproval(reason="Reviewed and approved.")
    )
    return created.id


def _approved_plan(db: Session, code: str = "recovery_2027") -> UUID:
    created = management_action_plans.create_plan(
        db,
        MAKER,
        ManagementActionPlanCreate.model_validate(
            {
                "code": code,
                "name": "Capital-restoration plan",
                "bank_id": None,
                "actions": [
                    {
                        "action_id": "dividend_suspension",
                        "kind": "revise_dividend",
                        "label": "Suspend dividends",
                        "trigger_kind": "always",
                        "effective_year": 1,
                        "dividend_reduction_pct": "100",
                    },
                    {
                        "action_id": "capital_raise",
                        "kind": "raise_capital",
                        "label": "Equity issuance",
                        "trigger_kind": "always",
                        "effective_year": 1,
                        "sizing": "fill_residual",
                        "capital_raise_ghs": "1",
                        "counts_as_paid_up": True,
                    },
                ],
                "reason": "Author the recovery plan.",
            }
        ),
    )
    management_action_plans.submit_plan(
        db, MAKER, created.id, ManagementActionPlanTransition(reason="submit")
    )
    management_action_plans.approve_plan(
        db, CHECKER, created.id, ManagementActionPlanApproval(reason="approve")
    )
    return created.id


def _run_enterprise_stress(
    db: Session, scenario_id: UUID, plan_id: UUID | None = None
) -> UUID:
    read = enterprise_stress.run_enterprise_stress_test(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        EnterpriseStressRunCreate(
            scenario_id=scenario_id,
            reporting_period_id=_period_id(db),
            management_action_plan_id=plan_id,
            reason="Annual ICAAP stress test.",
        ),
    )
    return read.run_id


def _attested_signoff(db: Session, run_id: UUID) -> UUID:
    signoff = enterprise_stress_signoff.create_signoff(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        StressSignoffCreate(
            run_id=run_id,
            scenario_narrative="Enterprise-wide adverse scenario covering credit, "
            "liquidity, market and IRRBB across the banking book.",
            assumptions_rationale="Documented linear elasticities; expert-judgement "
            "overlays challenged by the CRO.",
            methodology_summary="Bottom-up credit migration + coherent macro fan-out.",
            reason="Prepare the stress-run sign-off.",
        ),
    )
    enterprise_stress_signoff.submit_signoff(
        db, MAKER, SAMPLE_BANK_ID, signoff.id, StressSignoffTransition(reason="submit")
    )
    enterprise_stress_signoff.attest_signoff(
        db,
        CHECKER,
        SAMPLE_BANK_ID,
        signoff.id,
        StressSignoffAttestation(
            credibility_rationale="The Board reviewed and challenged the framework and "
            "results; the assumptions and severity are credible.",
            board_challenge="Challenged the FX severity; retained as plausible.",
            reason="Board attestation.",
        ),
    )
    return signoff.id


def _generate(db: Session) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=RETURN_CODE, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _generate_sdi(db: Session) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code=SDI_RETURN_CODE, reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _section(snapshot: dict[str, Any], code: str) -> dict[str, Any] | None:
    return next((s for s in snapshot["sections"] if s["code"] == code), None)


def _prepare(db: Session, *, plan: bool = False) -> RegulatoryPackage:
    materialize_canonical_test_book(db)
    _seed_checker(db)
    scenario_id = _approved_scenario(db)
    plan_id = _approved_plan(db) if plan else None
    run_id = _run_enterprise_stress(db, scenario_id, plan_id)
    _attested_signoff(db, run_id)
    return _generate(db)


def test_generates_appendix_ii_tables_from_attested_run(db_session: Session) -> None:
    package = _prepare(db_session)
    snapshot = package.snapshot
    assert snapshot["template_id"] == "bog-icaap-stress-appendix2-v1"
    assert package.return_family == "icaap_stress"
    assert package.status == "generated"
    assert package.content_digest

    # Table 1 summary: current + 3 base + 3 stress capital positions.
    positions = _section(snapshot, "t1_summary_positions")
    assert positions is not None
    labels = [row["code"] for row in positions["rows"]]
    assert labels == [
        "current",
        "base_y1",
        "base_y2",
        "base_y3",
        "stress_y1",
        "stress_y2",
        "stress_y3",
    ]

    # Table 1 vulnerability granularity: loss by CRD exposure class per year (¶67(g)).
    impact = _section(snapshot, "t1_impact_of_adverse")
    assert impact is not None and impact["rows"]
    assert {row["year"] for row in impact["rows"]} == {"1", "2", "3"}

    # All six Appendix II tables + governance section are present.
    for code in (
        "t1_capital_required",
        "t2_capital_projection",
        "t3_profit_and_loss",
        "t4_financial_position",
        "t5_rwa",
        "t6_risk_drivers",
        "governance",
    ):
        section = _section(snapshot, code)
        assert section is not None and section["rows"], code

    # Table 6 carries the 7 macro drivers × 3 years.
    assert len(_section(snapshot, "t6_risk_drivers")["rows"]) == 21  # type: ignore[index]

    # The directive invariant: stressed Total Pillar-1 RWA (Table 5) equals
    # Table 1's stressed RWA — carried verbatim from the enterprise-stress run.
    t5_by_label = {
        row["code"]: row["value"] for row in _section(snapshot, "t5_rwa")["rows"]  # type: ignore[index]
    }
    for row in positions["rows"]:
        if row["code"].startswith("stress_y"):
            assert t5_by_label[row["code"]] == row["total_rwa"], row["code"]

    # Provenance: sourced from exactly the one enterprise-stress run.
    assert [entry["module"] for entry in package.source_runs] == ["enterprise_stress"]

    # Without a management-actions plan the with/without blocks are omitted.
    assert _section(snapshot, "t1_management_actions") is None
    assert _section(snapshot, "t1_post_capitalisation") is None
    assert snapshot["metadata"]["with_management_actions"] is False

    # Governance provenance is carried on the snapshot (¶20, ¶67(b)(c)).
    governance = snapshot["metadata"]["governance"]
    assert governance["status"] == "attested"
    assert governance["scenario_narrative"]
    assert governance["credibility_rationale"]
    assert governance["attested_by"]


def test_sdi_stress_return_uses_attested_evidence_without_basel_table2(
    db_session: Session,
) -> None:
    materialize_canonical_test_book(db_session)
    bank = db_session.get(Bank, SAMPLE_BANK_ID)
    assert bank is not None
    bank.institution_type = "savings_and_loans"
    db_session.flush()
    _seed_checker(db_session)
    scenario_id = _approved_scenario(db_session, code="sdi_adverse_2027")
    run_id = _run_enterprise_stress(db_session, scenario_id)
    _attested_signoff(db_session, run_id)

    package = _generate_sdi(db_session)

    assert package.return_family == "sdi"
    assert package.return_code == SDI_RETURN_CODE
    assert _section(package.snapshot, "t2_capital_projection") is None
    assert _section(package.snapshot, "t1_summary_positions") is not None
    assert _section(package.snapshot, "t5_rwa") is not None
    assert package.snapshot["metadata"]["basel_table2_included"] is False
    assert package.snapshot["metadata"]["report_scope"].endswith("Table 2 excluded.")
    assert [entry["module"] for entry in package.source_runs] == ["enterprise_stress"]


def test_default_signing_policy_applies_to_the_stress_return(db_session: Session) -> None:
    package = _prepare(db_session)
    # The return inherits the platform default signing policy: a signature is
    # required before it can be filed (docs/attestation_esignature.md).
    policy = attestation.package_policy(db_session, MAKER, package)
    assert policy.require_signature is True
    assert policy.require_signed_pdf is True


def test_with_and_without_management_actions_blocks(db_session: Session) -> None:
    package = _prepare(db_session, plan=True)
    snapshot = package.snapshot
    assert snapshot["metadata"]["with_management_actions"] is True
    # The with-actions blocks (¶67(f)) now render alongside the pre-action
    # (Post-Adverse) positions.
    actions = _section(snapshot, "t1_management_actions")
    post_cap = _section(snapshot, "t1_post_capitalisation")
    assert actions is not None and actions["rows"]
    assert post_cap is not None and post_cap["rows"]
    governance = snapshot["metadata"]["governance"]
    assert governance["with_actions_stays_above_all_minima"] is not None


def test_refuses_without_an_attested_stress_run(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    _seed_checker(db_session)
    scenario_id = _approved_scenario(db_session)
    run_id = _run_enterprise_stress(db_session, scenario_id)

    # No sign-off yet → refuse.
    with pytest.raises(HTTPException) as no_signoff:
        _generate(db_session)
    assert no_signoff.value.status_code == 409
    assert no_signoff.value.detail["error_code"] == "no_attested_stress_run"  # type: ignore[index]

    # A prepared-but-not-attested sign-off is still not enough.
    signoff = enterprise_stress_signoff.create_signoff(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        StressSignoffCreate(
            run_id=run_id,
            scenario_narrative="n",
            assumptions_rationale="r",
            reason="prepare",
        ),
    )
    enterprise_stress_signoff.submit_signoff(
        db_session, MAKER, SAMPLE_BANK_ID, signoff.id, StressSignoffTransition(reason="submit")
    )
    with pytest.raises(HTTPException) as pending:
        _generate(db_session)
    assert pending.value.detail["error_code"] == "no_attested_stress_run"  # type: ignore[index]


def test_exports_to_xlsx_and_pdf(db_session: Session, storage: InMemoryStorageClient) -> None:
    package = _prepare(db_session)

    xlsx = export_package(db_session, MAKER, package, "xlsx")
    assert xlsx.size_bytes > 0 and xlsx.checksum_sha256
    pdf = export_package(db_session, MAKER, package, "pdf")
    assert pdf.size_bytes > 0 and pdf.checksum_sha256

    slug = db_session.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    payload = None
    for obj in storage.list(slug, "outputs"):
        if obj.location.object_path == xlsx.object_path:
            _, stream = storage.read(obj.location)
            payload = stream.read()
    assert payload is not None
    workbook = load_workbook(io.BytesIO(payload))
    # One sheet per Appendix II table plus metadata / provenance.
    assert "Return Metadata" in workbook.sheetnames
    assert any("Table 1" in name for name in workbook.sheetnames)
    assert any("Table 6" in name for name in workbook.sheetnames)


def test_maker_cannot_attest_own_signoff(db_session: Session) -> None:
    materialize_canonical_test_book(db_session)
    _seed_checker(db_session)
    scenario_id = _approved_scenario(db_session)
    run_id = _run_enterprise_stress(db_session, scenario_id)
    signoff = enterprise_stress_signoff.create_signoff(
        db_session,
        MAKER,
        SAMPLE_BANK_ID,
        StressSignoffCreate(
            run_id=run_id,
            scenario_narrative="n",
            assumptions_rationale="r",
            reason="prepare",
        ),
    )
    enterprise_stress_signoff.submit_signoff(
        db_session, MAKER, SAMPLE_BANK_ID, signoff.id, StressSignoffTransition(reason="submit")
    )
    # Maker ≠ checker (¶16, ¶20): the preparer/submitter cannot attest.
    with pytest.raises(HTTPException) as maker_exc:
        enterprise_stress_signoff.attest_signoff(
            db_session,
            MAKER,
            SAMPLE_BANK_ID,
            signoff.id,
            StressSignoffAttestation(credibility_rationale="c", reason="self-attest"),
        )
    assert maker_exc.value.detail["error_code"] == "maker_is_checker"  # type: ignore[index]

    # A second run also cannot mint a duplicate sign-off.
    with pytest.raises(HTTPException) as dup:
        enterprise_stress_signoff.create_signoff(
            db_session,
            MAKER,
            SAMPLE_BANK_ID,
            StressSignoffCreate(
                run_id=run_id,
                scenario_narrative="n2",
                assumptions_rationale="r2",
                reason="dup",
            ),
        )
    assert dup.value.detail["error_code"] == "signoff_exists"  # type: ignore[index]
