"""BoG official returns — framework + structural gate over EVERY form.

This is the executable form of docs/bog_returns/00_full_return_registry.md:

1. every template under docs/reporting has a committed layout, is in the
   catalogue and the registry, with the Guide's frequency/limit/basis;
2. 100% of the templates' formula cells evaluate with the built-in evaluator
   (BoG's roll-ups are reproduced, never re-implemented);
3. every form generates through the REAL package pipeline (immutable
   snapshot + content hash → maker-checker → export) on the hermetic book;
4. the xlsx export mirrors the official workbook: same sheet set in the same
   order, header cells carry bank name + period, every official label sits at
   its official cell, input/formula cells hold numbers or blanks (values-only,
   sealed), and a "Completion notes" sheet lists every input_required /
   unmapped line — the structure is never dropped.
"""

from __future__ import annotations

import inspect
import io
import re
from datetime import UTC, date, datetime
from uuid import UUID

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.models import Bank, RegulatoryPackage
from app.models.regulatory_reporting import ARTIFACT_KINDS
from app.services.regulatory_reporting import workflow as reporting_workflow
from app.services.regulatory_reporting.bog_forms.catalog import (
    all_form_codes,
    all_form_specs,
    form_spec,
)
from app.services.regulatory_reporting.bog_forms.formulas import (
    WorkbookEvaluator,
    evaluate,
)
from app.services.regulatory_reporting.bog_forms.layout import (
    available_layouts,
    load_layout,
)
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.registry_entries import (
    is_bog_official_template,
    template_id_for,
)
from app.services.regulatory_reporting.bog_forms.render import (
    _date_overrides,
    _header_text,
    official_width,
)
from app.services.regulatory_reporting.exports import render_bog_form_pdf, render_bog_form_xlsx
from app.services.regulatory_reporting.registry import REGISTRY
from app.services.regulatory_reporting.templates import TEMPLATES
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)
from tests.storage.inmemory import InMemoryStorageClient

# The Guide's List of Prudential Returns (frequency, time limit) — the registry
# must state exactly this for every form.
GUIDE_LIST: dict[str, tuple[str, int]] = {
    "BSD1": ("weekly", 9),
    "BSD1A": ("weekly", 9),
    "BSD1B": ("weekly", 9),
    "BSD2": ("monthly", 14),
    "BSD2A": ("monthly", 14),
    "BSD3A": ("monthly", 14),
    "BSD3B": ("monthly", 14),
    "BSD4": ("monthly", 14),
    "BSD5A": ("monthly", 14),
    "BSD5B": ("quarterly", 14),
    "BSD6": ("monthly", 14),
    "BSD7A": ("monthly", 14),
    "BSD7B": ("quarterly", 14),
    "BSD8": ("monthly", 14),
    "BSD9": ("quarterly", 14),
    "BSD10": ("semiannual", 14),
    "BSD11": ("semiannual", 14),
    "BSD13": ("monthly", 14),
    "BSD14": ("weekly", 9),
    "BSD15A": ("weekly", 9),
    "BSD15B": ("weekly", 9),
    "BSD16": ("monthly", 14),
    "BSD17": ("monthly", 14),
}
CONSOLIDATED = {"BSD3B", "BSD5B", "BSD7B", "BSD9"}


# ---------------------------------------------------------------------------
# 1. inventory: every template ↔ layout ↔ catalogue ↔ registry
# ---------------------------------------------------------------------------


def test_every_official_template_has_a_layout_catalogue_and_registry_entry() -> None:
    layouts = set(available_layouts())
    catalogue = set(all_form_codes())
    registry = {code for code, d in REGISTRY.items() if d.family == "bsd"}
    assert layouts == catalogue == registry == set(GUIDE_LIST), (
        f"layouts={sorted(layouts)}\ncatalogue={sorted(catalogue)}\nregistry={sorted(registry)}"
    )


def test_registry_states_the_guides_frequency_limit_and_basis() -> None:
    for code, (frequency, limit) in GUIDE_LIST.items():
        definition = REGISTRY[code]
        spec = form_spec(code)
        assert definition.frequency == frequency, code
        assert spec.time_limit_days == limit, code
        due = definition.deadline_rule(date(2026, 6, 30))
        assert (due - date(2026, 6, 30)).days == limit, code
        assert spec.basis == ("consolidated" if code in CONSOLIDATED else "solo"), code
        assert definition.generator == "bog_form"
        assert definition.template_id == template_id_for(code)
        assert is_bog_official_template(definition.template_id)
        assert definition.template_id in TEMPLATES


def test_bsd1_covers_ghana_branches_only_per_guide() -> None:
    assert form_spec("BSD1").ghana_branches_only is True
    assert all(not s.ghana_branches_only for s in all_form_specs() if s.code != "BSD1")


def test_layouts_keep_every_official_sheet_including_empty_placeholders() -> None:
    expected_sheet_counts = {
        "BSD1": 4, "BSD1A": 1, "BSD1B": 3, "BSD2": 22, "BSD2A": 1, "BSD3A": 3, "BSD3B": 3,
        "BSD4": 3, "BSD5A": 3, "BSD5B": 2, "BSD6": 2, "BSD7A": 1, "BSD7B": 1, "BSD8": 2,
        "BSD9": 3, "BSD10": 1, "BSD11": 8, "BSD13": 4, "BSD14": 1, "BSD15A": 2, "BSD15B": 1,
        "BSD16": 3, "BSD17": 2,
    }  # fmt: skip
    for code, count in expected_sheet_counts.items():
        assert len(load_layout(code).sheets) == count, code


def test_legacy_miscoded_returns_are_recoded_and_placeholder_retired() -> None:
    assert "CAR-RWA" in REGISTRY and REGISTRY["CAR-RWA"].generator == "capital"
    assert "LCR-NSFR" in REGISTRY and REGISTRY["LCR-NSFR"].generator == "liquidity"
    assert "BSD-MONTHLY" not in REGISTRY
    assert REGISTRY["BSD2"].title.startswith("BSD2 — Statement of Assets")
    assert REGISTRY["BSD5A"].title.startswith("BSD5A — Capital Adequacy")


# ---------------------------------------------------------------------------
# 2. the templates' own formulas evaluate — 100%, every form
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("code", sorted(GUIDE_LIST))
def test_every_template_formula_evaluates(code: str) -> None:
    layout = load_layout(code)
    formulas = {(s.name, c.ref): c.formula or "" for s in layout.sheets for c in s.formula_cells}
    inputs = {(s.name, c.ref): c.value for s in layout.sheets for c in s.input_cells}
    evaluator = WorkbookEvaluator(formulas, inputs, external=lambda _i, _s, _r: 0.0)
    for key in formulas:
        evaluator.value(*key)  # raises UnsupportedFormulaError on any gap


def test_formula_evaluator_semantics() -> None:
    assert evaluate("=SUM(B7:B11)", lambda cr: 2.0) == 10.0
    assert evaluate("=B6+C6", lambda cr: {"B6": 1.5, "C6": 2.5}[cr.ref]) == 4.0
    assert evaluate("=100*6%", lambda cr: 0) == 6.0
    assert evaluate("=D30/D73%", lambda cr: {"D30": 50.0, "D73": 200.0}[cr.ref]) == 25.0
    assert evaluate("=", lambda cr: 0) is None
    assert evaluate('=IF(B10<0,"( S )",IF(B10=0,"-","( L )"))', lambda cr: -3.0) == "( S )"
    assert evaluate("=[1]BSD2!D38+[1]BSD2!D39", lambda cr: 3.0 if cr.external == 1 else 0) == 6.0
    assert evaluate("=SUM('BSD2-Annex 1'!B3:B5)+A1", lambda cr: 2.0) == 8.0
    assert evaluate("=A1/A2", lambda cr: {"A1": 1.0, "A2": 0.0}[cr.ref]) == 0.0  # Excel #DIV/0 → 0


# ---------------------------------------------------------------------------
# 3. line maps only bind INPUT cells (a formula cell is BoG's arithmetic)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("code", sorted(GUIDE_LIST))
def test_line_maps_bind_only_official_input_cells(code: str) -> None:
    layout = load_layout(code)
    for sheet_name, lines in line_maps_for(code).items():
        sheet = layout.sheet(sheet_name)
        for line in lines:
            for ref in line.cells.values():
                cell = sheet.by_ref.get(ref)
                if cell is None:
                    # a BLANK data cell of the official grid (no placeholder in the
                    # template) — allowed, but it must lie inside the sheet's grid
                    col, row = coordinate_from_string(ref)
                    assert (
                        row <= sheet.max_row and column_index_from_string(col) <= sheet.max_col
                    ), f"{code}/{sheet_name}!{ref} lies outside the official sheet ({line.code})"
                    continue
                assert cell.kind == "input", f"{code}/{sheet_name}!{ref} is a {cell.kind} cell"


def test_bsd2_line_map_binds_every_leaf_row_of_the_spine() -> None:
    layout = load_layout("BSD2").sheet("BSD2")
    leaf_rows = {c.row for c in layout.input_cells if c.ref.startswith("B")}
    bound = {int(line.cells["domestic"][1:]) for line in line_maps_for("BSD2")["BSD2"]}
    assert bound == leaf_rows
    assert len(bound) == 205  # noqa: PLR2004 — the official spine has 205 leaf rows


# ---------------------------------------------------------------------------
# 4. end-to-end: generate + export EVERY form through the real package pipeline
# ---------------------------------------------------------------------------


def _materialize(db_client: TestClient) -> str:
    _ = db_client
    session = get_sessionmaker()()
    try:
        summary = materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    return summary.bank_id


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    return response.json()


def _latest_period_end(db_client: TestClient) -> str:
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    return periods[0]["period_end"]


@pytest.mark.parametrize("code", sorted(GUIDE_LIST))
def test_form_generates_and_exports_template_faithful_xlsx(  # noqa: PLR0915
    db_client: TestClient, code: str
) -> None:
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, code, reporting_date)
    assert package["return_code"] == code
    assert package["status"] == "generated"

    # snapshot carries the official structure + per-line status
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    snapshot = detail["snapshot"]
    layout = load_layout(code)
    assert snapshot["metadata"]["official_sheets"] == list(layout.sheet_names)
    payload = snapshot["bog_form"]
    assert payload["code"] == code
    counts = payload["status_counts"]
    captured = {(s.name, c.ref) for s in layout.sheets for c in s.input_cells}
    declared = {
        (sheet.name, ref)
        for sheet in form_spec(code).sheets
        for line in sheet.lines
        for ref in line.cells.values()
    }
    # Every official input cell is accounted for: a DECLARED cell is mapped or
    # input_required (declared cells may include blank-grid data cells the
    # template left empty — grid_lines); every CAPTURED input nobody declared is
    # unmapped. Nothing is dropped, nothing is double counted.
    assert counts["mapped"] + counts["input_required"] == len(declared), counts
    assert counts["unmapped"] == len(captured - declared), counts
    assert counts["derived"] == sum(len(s.formula_cells) for s in layout.sheets)
    assert not payload["errors"], payload["errors"]

    # export → the OFFICIAL workbook, values only (the exact renderer the
    # package export path calls, fed with the immutable snapshot)
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload_bytes = render_bog_form_xlsx(
            code, snapshot, bank, datetime(2026, 8, 15, tzinfo=UTC)
        )
    finally:
        session.close()
    wb = openpyxl.load_workbook(io.BytesIO(payload_bytes), data_only=False)
    official_names = [name[:31] for name in layout.sheet_names]
    assert wb.sheetnames[: len(official_names)] == official_names
    assert wb.sheetnames[-1] == "Completion notes"
    for sheet in layout.sheets:
        ws = wb[sheet.name[:31]]
        date_cells = _date_overrides(sheet, snapshot["reporting_date"])
        for cell in sheet.cells:
            value = ws[cell.ref].value
            if cell.ref in date_cells:
                # A date box the template cannot fill: either a formula whose
                # source workbook is gone (renders as the 1899 epoch) or a
                # literal the original filer typed (BSD1 B3: 2004-05-26).
                # The export must carry the REAL date, never either of those.
                assert isinstance(value, (datetime, date)), (
                    f"{code}/{sheet.name}!{cell.ref}: {value!r} is not a date"
                )
                actual = value.date() if isinstance(value, datetime) else value
                assert actual == date_cells[cell.ref], (
                    f"{code}/{sheet.name}!{cell.ref}: {actual} vs {date_cells[cell.ref]}"
                )
                assert actual.year >= 2020, f"stale template date at {cell.ref}: {actual}"
                continue
            if cell.kind == "label":
                text = str(cell.value)
                filled = _header_text(
                    text,
                    bank_name="Sample Bank Ltd",
                    period_label=snapshot["reporting_period"]["label"],
                    reporting_date=snapshot["reporting_date"],
                    previous_reporting_date="x",
                )
                if filled is None:
                    # not a header prompt → the official label is verbatim
                    assert str(value) == text, (
                        f"{code}/{sheet.name}!{cell.ref}: {value!r} vs {text!r}"
                    )
                else:
                    # a header prompt → carries the bank name / reporting date,
                    # never a stale value shipped in the template (BSD1B trap)
                    assert "FIRST ATLANTIC" not in str(value)
                    assert "2003" not in str(value)
                    assert (
                        ("Sample Bank Ltd" in str(value))
                        or (snapshot["reporting_date"] in str(value))
                        or (snapshot["reporting_period"]["label"] in str(value))
                        or "PREVIOUS WEEK" in str(value)
                    ), f"{code}/{sheet.name}!{cell.ref}: {value!r}"
            else:
                assert not (isinstance(value, str) and value.startswith("=")), (
                    f"{code}/{sheet.name}!{cell.ref} exported a live formula"
                )
        assert ws.merged_cells.ranges is not None
    # bank name + period reach the header
    first = wb[layout.sheets[0].name[:31]]
    header_blob = " ".join(
        str(c.value) for row in first.iter_rows(max_row=8) for c in row if c.value
    )
    assert "Sample Bank" in header_blob or SAMPLE_BANK_ID in header_blob or "Bank" in header_blob


def test_bsd2_spine_totals_are_bogs_own_arithmetic(db_client: TestClient) -> None:
    """The evaluated BSD2 TOTAL column equals Domestic + Foreign for every row and
    the Summary sheet re-derives from BSD2 — BoG's formulas, applied to our inputs."""
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, "BSD2", reporting_date)
    snapshot = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()["snapshot"]
    cells = snapshot["bog_form"]["cells"]["BSD2"]
    layout = load_layout("BSD2").sheet("BSD2")
    checked = 0
    for cell in layout.formula_cells:
        if (
            cell.formula
            and cell.formula.startswith("=B")
            and "+C" in cell.formula
            and cell.ref.startswith("D")
        ):
            row = cell.ref[1:]
            d = cells.get(cell.ref) or 0.0
            b = cells.get(f"B{row}") or 0.0
            c = cells.get(f"C{row}") or 0.0
            assert abs(float(d) - (float(b) + float(c))) < 1e-6, cell.ref
            checked += 1
    assert checked > 150  # noqa: PLR2004 — the spine's total column
    # governance: a normal immutable package row (content digest, version 1,
    # status "generated" awaiting maker-checker) — same lifecycle as every return
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        row = session.scalar(
            select(RegulatoryPackage).where(RegulatoryPackage.id == UUID(package["id"]))
        )
    finally:
        session.close()
    assert row is not None
    assert row.content_digest
    assert row.version == 1
    assert row.status == "generated"


# ---------------------------------------------------------------------------
# 5. dual Excel modes: sealed OFFICIAL (values, protected) vs WORKING (formulas)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("code", ["BSD2", "BSD6", "BSD8", "BSD5A"])
def test_working_copy_keeps_the_templates_own_formulas_live(
    db_client: TestClient, code: str
) -> None:
    """Mode B: every in-workbook formula cell carries the template's formula text
    (SUM, Domestic+Foreign, cross-sheet annex links); inputs are values; the
    file is unmistakably a working copy. Mode A: no live formulas, protected."""
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, code, reporting_date)
    snapshot = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()["snapshot"]
    layout = load_layout(code)
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        stamp = datetime(2026, 8, 16, tzinfo=UTC)
        official = openpyxl.load_workbook(
            io.BytesIO(render_bog_form_xlsx(code, snapshot, bank, stamp)), data_only=False
        )
        working = openpyxl.load_workbook(
            io.BytesIO(render_bog_form_xlsx(code, snapshot, bank, stamp, mode="working")),
            data_only=False,
        )
    finally:
        session.close()

    live = external = 0
    for sheet in layout.sheets:
        ws_off = official[sheet.name[:31]]
        ws_wrk = working[sheet.name[:31]]
        assert ws_off.protection.sheet is True, f"{code}/{sheet.name}: official sheet not protected"
        assert ws_wrk.protection.sheet is not True
        for cell in sheet.formula_cells:
            off_value = ws_off[cell.ref].value
            wrk_value = ws_wrk[cell.ref].value
            assert not (isinstance(off_value, str) and off_value.startswith("=")), (
                f"{code}/{sheet.name}!{cell.ref}: official export carries a live formula"
            )
            if "[" in (cell.formula or ""):
                # cross-WORKBOOK link: cannot resolve inside one file → evaluated value
                assert not (isinstance(wrk_value, str) and wrk_value.startswith("="))
                external += 1
            else:
                assert wrk_value == cell.formula, (
                    f"{code}/{sheet.name}!{cell.ref}: working copy lost the template formula"
                )
                live += 1
        for cell in sheet.input_cells:
            v = ws_wrk[cell.ref].value
            assert not (isinstance(v, str) and v.startswith("=")), (
                f"{code}/{sheet.name}!{cell.ref}: input cell became a formula"
            )
    assert live > 0
    assert working.properties.title is not None and "WORKING COPY" in working.properties.title
    assert "WORKING COPY" in str(working["Completion notes"]["A2"].value)
    if code == "BSD8":
        assert external >= 1  # the [1]BSD2 link


def test_working_copy_is_a_distinct_artifact_kind_and_never_filed(
    db_client: TestClient, monkeypatch: pytest.MonkeyPatch, storage_engine: InMemoryStorageClient
) -> None:
    """The three artifacts of one sealed run through the REAL export endpoint:
    pdf (submission), xlsx / xlsx_official (sealed twin), xlsx_working (ALM copy)
    — and the working copy is excluded from what a submission would file."""
    assert "xlsx_working" in ARTIFACT_KINDS
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, "BSD2", reporting_date)
    # route the export's object writes to the in-memory storage (never a real bucket)
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: storage_engine
    )
    base = f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}"
    kinds_seen: dict[str, str] = {}
    for kind in ("pdf", "xlsx_official", "xlsx_working"):
        response = db_client.post(f"{base}/export", headers=headers(), params={"kind": kind})
        assert response.status_code == 201, f"{kind}: {response.status_code} {response.text[:200]}"
        artifact = response.json()
        kinds_seen[kind] = artifact["kind"]
    # xlsx_official is stored under the historical sealed kind "xlsx"
    assert kinds_seen == {"pdf": "pdf", "xlsx_official": "xlsx", "xlsx_working": "xlsx_working"}
    artifacts = {
        a["kind"]: a
        for a in db_client.get(f"{base}/artifacts", headers=headers()).json()["artifacts"]
    }
    assert set(artifacts) == {"pdf", "xlsx", "xlsx_working"}
    assert artifacts["xlsx_working"]["object_path"].endswith("BSD2.working.xlsx")
    # what a filing would carry: the working copy is filtered out of every filing
    source = inspect.getsource(reporting_workflow)
    assert 'artifact.kind != "xlsx_working"' in source


def _pdf_text(payload: bytes) -> str:
    from pdfminer.high_level import extract_text  # noqa: PLC0415

    return extract_text(io.BytesIO(payload))


@pytest.mark.parametrize("code", sorted(GUIDE_LIST))
def test_pdf_export_is_the_official_form_not_a_cell_listing(
    db_client: TestClient, code: str
) -> None:
    """The filing PDF must be BoG's own form.

    It regressed once to the GENERIC tabular renderer, which emitted a
    Line / Official cell / Status / Source listing keyed by internal ids
    (``BSD1.R10.thu``, ``input_required``). That is a completion aid, not a
    return — a supervisor cannot read it and BoG would not accept it.
    """
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, code, reporting_date)
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    snapshot = detail["snapshot"]

    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_pdf(code, snapshot, bank, datetime(2026, 8, 15, tzinfo=UTC))
    finally:
        session.close()

    assert payload.startswith(b"%PDF"), code
    text = " ".join(_pdf_text(payload).split())

    # the form's own line descriptions are on the page
    layout = load_layout(code)
    labels = [
        str(cell.value).strip()
        for sheet in layout.sheets
        for cell in sheet.cells
        if cell.kind == "label" and len(str(cell.value or "").strip()) > 12
    ]
    if labels:
        hits = sum(1 for label in labels if " ".join(label.split())[:40] in text)
        assert hits >= max(1, len(labels) // 4), (
            f"{code}: only {hits}/{len(labels)} official labels reached the PDF"
        )

    # Internal machinery never reaches a filing artifact. This is the whole
    # point: the PDF is what the institution sends BoG, so resolver names,
    # per-line status tokens and spreadsheet cell references are all out.
    for leak in ("input_required", "unmapped", "status_counts", "not yet mapped", "bsd1.daily"):
        assert leak not in text, f"{code}: internal token {leak!r} leaked into the filing PDF"
    # no bare spreadsheet cell references (the old appendix listed B23, C23, …)
    cell_refs = re.findall(r"\b[A-H]\d{1,3}\b", text)
    assert len(cell_refs) < 5, f"{code}: spreadsheet cell references in the PDF: {cell_refs[:10]}"


def test_printed_day_dates_match_the_dates_the_line_maps_resolve() -> None:
    """The date printed above a day column must be the date its data came from.

    The renderer and ``sources_ext/bsd1.py`` derive the week independently, so
    they can drift apart silently and publish correctly-sourced figures under
    the wrong dates — a worse defect than a blank cell, because it looks right.
    """
    from datetime import timedelta  # noqa: PLC0415

    from app.services.regulatory_reporting.bog_forms.sources_ext.bsd1 import (  # noqa: PLC0415
        DAY_COLUMNS,
        PREVIOUS_WEEK_SHIFT,
    )

    period_end = date(2026, 6, 30)
    sheet = load_layout("BSD1").sheets[0]
    printed = _date_overrides(sheet, period_end.isoformat())
    columns = dict(zip("BCDEFGH", ["thu", "fri", "sat", "sun", "mon", "tue", "wed"], strict=True))
    for letter, key in columns.items():
        assert printed[f"{letter}28"] == period_end - timedelta(days=DAY_COLUMNS[key]), key
        assert printed[f"{letter}7"] == period_end - timedelta(
            days=DAY_COLUMNS[key] + PREVIOUS_WEEK_SHIFT
        ), key


@pytest.mark.parametrize("code", sorted(GUIDE_LIST))
def test_xlsx_writes_nothing_beyond_the_printed_form(db_client: TestClient, code: str) -> None:
    """No wall of zeros to the right of the return.

    The official workbooks drag formulas far past the printed grid (BSD1 runs
    to column IV). Exporting each one as its evaluated ``0`` filled every
    subtotal row with zeros from column K onward — visible nonsense on an
    artifact that is meant to mirror the form.
    """
    _materialize(db_client)
    reporting_date = _latest_period_end(db_client)
    package = _generate(db_client, code, reporting_date)
    snapshot = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()["snapshot"]
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 15, tzinfo=UTC))
    finally:
        session.close()
    wb = openpyxl.load_workbook(io.BytesIO(payload))
    for sheet in load_layout(code).sheets:
        limit = official_width(sheet)
        if not limit:
            continue
        ws = wb[sheet.name[:31]]
        stray = [
            (c.coordinate, c.value)
            for row in ws.iter_rows(min_col=limit + 1)
            for c in row
            if c.value is not None
        ]
        assert not stray, f"{code}/{sheet.name}: wrote past column {limit}: {stray[:5]}"
