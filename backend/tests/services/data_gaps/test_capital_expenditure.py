"""``capital_expenditure`` — the fixed-asset / capex register → BSD10 + BSD2 item 12.

1. the Sample Bank register (two half-years × seven asset classes) validates
   row by row and its own arithmetic holds (NBV = cost − accumulated
   depreciation; the six-month forecast = 0–3 + 3–6 months; the roll-forward
   opening + additions − disposals − depreciation = closing NBV);
2. the line maps bind EVERY BSD10 cell and BSD2 rows 115–121 / 123 to
   ``refs.sum`` over the register (column-major on BSD10 because each column is
   an asset class; Domestic/Foreign on BSD2 by the row's booking currency);
3. one half-year pushed through the REAL API (three-call push flow, the
   generic ``scripts/ingest_push.py`` reader, ``as_of_date`` = period end) lands
   under the kind with batch lineage; BSD10 generated through
   ``POST /regulatory-packages`` carries the register's figures per class and
   BoG's own template arithmetic runs over them (``H = SUM(C:G)``,
   ``row 18 = row 16 + row 17``); BSD2 item 12 fills at COST by class, WIP and
   accumulated depreciation, so BoG's sub-total − depreciation equals the
   register's NBV (+ WIP) — never a re-implemented line; the earlier
   half-year pushed with its own as_of does not disturb the later period;
4. without a register the cells stay blank (input_required), never zero;
5. the schema rejects a malformed row honestly (bad class, non-numeric,
   NBV that does not reconcile to cost − depreciation).
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.domain.ingestion.reference_schemas import schema_for
from app.domain.ingestion.reference_schemas.capital_expenditure import (
    ASSET_CLASSES,
    BSD2_ROW_CLASSES,
    BSD10_COLUMN_CLASSES,
    SCHEMA,
    validate_capex_row,
)
from app.models import Bank, CanonicalReferenceRow
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.linemaps.bsd10 import COLUMNS, ROW_FIELDS
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from scripts.ingest_push import read_rows
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
SAMPLE_DIR = Path(__file__).resolve().parents[3] / "onboarding" / "sample_bank"
CSV = SAMPLE_DIR / "capital_expenditure.csv"
KIND = "capital_expenditure"
PERIODS = ("2025-12-31", "2026-06-30")
CLASSES = 7
M = 1_000_000
CENT = 0.5  # cedis: the register is stated to the cedi


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _prepare(db_client: TestClient) -> list[str]:
    """Materialise the hermetic book; return the period ends, latest first."""
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return sorted((p["period_end"] for p in periods), reverse=True)


def _period_rows(period_end: str, *, source: str = PERIODS[-1]) -> list[dict[str, Any]]:
    """The Sample Bank rows of one half-year, re-dated to ``period_end`` (the
    hermetic book's periods do not coincide with the Sample Bank's)."""
    rows = read_rows(CSV, entity=False)
    return [{**r, "period_end": period_end} for r in rows if r["period_end"] == source]


def _push_reference(
    db_client: TestClient, rows: list[dict[str, Any]], *, as_of: str, key: str
) -> dict[str, Any]:
    opened = db_client.post(
        f"{BASE}/push-batches",
        headers=headers(),
        json={"as_of_date": as_of, "idempotency_key": key, "reason": f"Sample Bank {KIND}"},
    )
    assert opened.status_code == 201, opened.text
    push_id = opened.json()["push_batch_id"]
    staged = db_client.post(
        f"{BASE}/push-batches/{push_id}/records",
        headers=headers(),
        json={"reference": {KIND: rows}},
    )
    assert staged.status_code == 200, staged.text
    committed = db_client.post(f"{BASE}/push-batches/{push_id}/commit", headers=headers())
    assert committed.status_code == 201, committed.text
    return committed.json()


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def _landed() -> list[CanonicalReferenceRow]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        return list(
            session.scalars(
                select(CanonicalReferenceRow)
                .where(
                    CanonicalReferenceRow.bank_id == SAMPLE_BANK_ID,
                    CanonicalReferenceRow.dataset_kind == KIND,
                )
                .order_by(CanonicalReferenceRow.as_of_date, CanonicalReferenceRow.row_index)
            )
        )
    finally:
        session.close()


def _statuses(snapshot: dict[str, Any], sheet: str) -> dict[str, str]:
    return {
        row["cell"]: row["status"]
        for section in snapshot["sections"]
        if section["title"] == sheet
        for row in section["rows"]
    }


def _sum(rows: list[dict[str, Any]], field: str, classes: tuple[str, ...] | None = None) -> float:
    return sum(
        float(r.get(field) or 0) for r in rows if classes is None or r["asset_class"] in classes
    )


# ---------------------------------------------------------------------------
# 1. the Sample Bank register
# ---------------------------------------------------------------------------


def test_sample_bank_register_is_two_half_years_of_seven_classes() -> None:
    rows = read_rows(CSV, entity=False)
    assert sorted({r["period_end"] for r in rows}) == list(PERIODS)
    for period in PERIODS:
        subset = [r for r in rows if r["period_end"] == period]
        assert len(subset) == CLASSES
        assert {r["asset_class"] for r in subset} == set(ASSET_CLASSES)
        for row in subset:
            assert validate_capex_row(row) == [], row
            f = {k: float(v) for k, v in row.items() if k.endswith("_ghs")}
            # the sub-ledger roll-forward: opening + additions − disposals − depreciation = closing
            rolled = (
                f["opening_nbv_ghs"]
                + f["additions_purchased_ghs"]
                + f["additions_finance_lease_ghs"]
                + f["additions_hire_purchase_ghs"]
                - f["disposals_nbv_ghs"]
                - f["depreciation_ghs"]
            )
            assert abs(rolled - f["closing_nbv_ghs"]) < CENT, row["asset_class"]
        # the per-period push files carry exactly that period
        period_file = SAMPLE_DIR / f"capital_expenditure_{period}.csv"
        assert {r["period_end"] for r in read_rows(period_file, entity=False)} == {period}
    # H1 2026 opens where H2 2025 closed, class by class
    h2 = {r["asset_class"]: r for r in rows if r["period_end"] == PERIODS[0]}
    h1 = {r["asset_class"]: r for r in rows if r["period_end"] == PERIODS[1]}
    for cls in ASSET_CLASSES:
        assert float(h1[cls]["opening_nbv_ghs"]) == float(h2[cls]["closing_nbv_ghs"]), cls
        # WIP balance rolls: closing = opening + the half-year's WIP expenditure (no transfers)
        assert float(h1[cls]["wip_closing_ghs"]) == pytest.approx(
            float(h2[cls]["wip_closing_ghs"]) + float(h1[cls]["capital_wip_ghs"])
        ), cls
    template = (SAMPLE_DIR / "capital_expenditure_template.csv").read_text().strip().split(",")
    assert set(template) == set(SCHEMA.columns)


# ---------------------------------------------------------------------------
# 2. the line maps
# ---------------------------------------------------------------------------


def test_bsd10_binds_every_cell_column_by_column_to_the_register() -> None:
    lines = line_maps_for("BSD10")["BSD10"]
    by_cell = {ref: line for line in lines for ref in line.cells.values()}
    assert len(by_cell) == 50  # noqa: PLR2004 — 10 leaf rows × 5 asset-class columns
    for column, letter in COLUMNS.items():
        for row, field in ROW_FIELDS.items():
            line = by_cell[f"{letter}{row}"]
            assert line.source == "refs.sum"
            assert line.params == {
                "kind": KIND,
                "value_field": field,
                "filters": {"asset_class": list(BSD10_COLUMN_CLASSES[column])},
            }
            assert not line.unscaled
    # the union of column classes is the register's classes minus the BSD2-only one
    assert {c for cs in BSD10_COLUMN_CLASSES.values() for c in cs} == set(ASSET_CLASSES) - {
        "other_property_legal_rights"
    }


def test_bsd2_item_12_binds_cost_wip_and_depreciation_not_nbv() -> None:
    by_row = {int(line.cells["domestic"][1:]): line for line in line_maps_for("BSD2")["BSD2"]}
    for row, classes in BSD2_ROW_CLASSES.items():
        line = by_row[row]
        assert line.source == "refs.sum"
        assert line.params == {
            "kind": KIND,
            "value_field": "closing_cost_ghs",
            "currency_field": "currency",
            "filters": {"asset_class": list(classes)},
        }
        assert line.cells == {"domestic": f"B{row}", "foreign": f"C{row}"}
    assert by_row[121].params == {
        "kind": KIND,
        "value_field": "wip_closing_ghs",
        "currency_field": "currency",
    }
    assert by_row[123].params == {
        "kind": KIND,
        "value_field": "accumulated_depreciation_ghs",
        "currency_field": "currency",
    }
    assert {c for cs in BSD2_ROW_CLASSES.values() for c in cs} == set(ASSET_CLASSES)


# ---------------------------------------------------------------------------
# 3. push through the real API → BSD10 + BSD2 fill, BoG's formulas run over them
# ---------------------------------------------------------------------------


def test_pushed_half_year_fills_bsd10_and_bsd2_item_12(  # noqa: PLR0915 — one linear proof over two forms
    db_client: TestClient,
) -> None:
    latest, previous = _prepare(db_client)[:2]
    rows = _period_rows(latest)
    result = _push_reference(db_client, rows, as_of=latest, key=f"capex-{latest}")
    batch = result["batch"]
    assert batch["status"] in ("accepted", "accepted_with_warnings"), batch["validation_report"]
    assert batch["validation_report"]["summary"]["reference_rows"] == {KIND: CLASSES}
    landed = _landed()
    assert len(landed) == CLASSES
    assert {str(r.ingestion_batch_id) for r in landed} == {batch["id"]}
    assert all(r.lineage_id is not None for r in landed)
    assert landed[0].payload["asset_class"] == "land_buildings"
    assert landed[0].as_of_date == date.fromisoformat(latest)

    # ---- BSD10 ---------------------------------------------------------------
    snapshot = _generate(db_client, "BSD10", latest)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    assert payload["status_counts"]["mapped"] == 50  # noqa: PLR2004
    assert payload["status_counts"]["input_required"] == 0
    cells = payload["cells"]["BSD10"]
    for column, letter in COLUMNS.items():
        classes = BSD10_COLUMN_CLASSES[column]
        for row, field in ROW_FIELDS.items():
            assert cells[f"{letter}{row}"] == pytest.approx(_sum(rows, field, classes)), (
                column,
                field,
            )
    # BoG's own arithmetic: H = SUM(C:G) per row; row 18 = row 16 + row 17
    for row in ROW_FIELDS:
        assert cells[f"H{row}"] == pytest.approx(sum(cells[f"{c}{row}"] for c in "CDEFG"))
        assert cells[f"H{row}"] == pytest.approx(
            _sum(rows, ROW_FIELDS[row])
            - _sum(rows, ROW_FIELDS[row], ("other_property_legal_rights",))
        )
    for col in "CDEFGH":
        assert cells[f"{col}18"] == pytest.approx(cells[f"{col}16"] + cells[f"{col}17"])
    # the Guide's own consistency between G (row 13) and the 0–3 / 3–6 split holds in the file
    assert cells["H13"] == pytest.approx(cells["H18"])
    # concrete figures from the file: land & buildings purchased = land + staff premises
    assert cells["C7"] == pytest.approx(1_900_000 + 0)
    assert cells["E8"] == pytest.approx(1_100_000)  # computers on finance lease
    assert cells["G9"] == pytest.approx(600_000)  # motor vehicles on hire-purchase
    assert cells["H14"] == pytest.approx(60_000 + 50_000 + 20_000 + 450_000)  # disposal proceeds
    ws = _export("BSD10", snapshot)["BSD10"]
    assert ws["C7"].value == pytest.approx(1.9)  # ¢'Million
    assert ws["H7"].value == pytest.approx(cells["H7"] / M)
    assert not any(
        isinstance(c.value, str) and c.value.startswith("=") for row in ws.iter_rows() for c in row
    )

    # ---- BSD2 item 12 --------------------------------------------------------
    bsd2 = _generate(db_client, "BSD2", latest)
    assert not bsd2["bog_form"]["errors"], bsd2["bog_form"]["errors"]
    spine = bsd2["bog_form"]["cells"]["BSD2"]
    statuses = _statuses(bsd2, "BSD2")
    for row, classes in BSD2_ROW_CLASSES.items():
        assert statuses[f"B{row}"] == "mapped" and statuses[f"C{row}"] == "mapped"
        assert spine[f"B{row}"] == pytest.approx(_sum(rows, "closing_cost_ghs", classes)), row
        assert spine[f"C{row}"] == 0  # every register row is booked in the base currency
        assert spine[f"D{row}"] == pytest.approx(spine[f"B{row}"])  # =B+C, BoG's formula
    assert spine["B121"] == pytest.approx(_sum(rows, "wip_closing_ghs"))
    assert spine["B123"] == pytest.approx(_sum(rows, "accumulated_depreciation_ghs"))
    # BoG's sub-total and item 12: SUM(115:121) − 123 == the register's NBV + WIP balance
    total_cost = _sum(rows, "closing_cost_ghs") + _sum(rows, "wip_closing_ghs")
    assert spine["B122"] == pytest.approx(total_cost)
    assert spine["D122"] == pytest.approx(total_cost)
    assert spine["D114"] == pytest.approx(
        _sum(rows, "closing_nbv_ghs") + _sum(rows, "wip_closing_ghs")
    )
    assert spine["D114"] == pytest.approx(spine["D122"] - spine["D123"])
    assert spine["B118"] == pytest.approx(20_900_000 + 6_100_000)  # furniture + other office

    # ---- BSD6A row 35 (7. PPE net) aggregates the same BSD2 leaves: Total == FROM BSD2,
    # the depreciation leaf entering with its sign flipped; bands are the bank's -----
    bsd6 = _generate(db_client, "BSD6", latest)["bog_form"]
    assert not bsd6["errors"], bsd6["errors"]
    six_a = bsd6["cells"]["BSD6A"]
    assert six_a["B35"] == pytest.approx(spine["B114"])  # FROM BSD2
    assert six_a["C35"] == pytest.approx(spine["B114"])  # Total = cost + WIP − depreciation
    assert all(six_a[f"{col}35"] is None for col in "DEFGHIJK")  # no Guide band: bank allocates

    # ---- an earlier half-year pushed later does not disturb the later period ----
    earlier = _period_rows(previous, source=PERIODS[0])
    _push_reference(db_client, earlier, as_of=previous, key=f"capex-{previous}")
    again = _generate(db_client, "BSD10", latest)["bog_form"]["cells"]["BSD10"]
    assert again["C7"] == pytest.approx(1_900_000)  # latest period still reads its half-year
    before = _generate(db_client, "BSD10", previous)["bog_form"]["cells"]["BSD10"]
    assert before["C7"] == pytest.approx(1_400_000)  # earlier period reads its own


def test_forms_without_a_register_stay_input_required(db_client: TestClient) -> None:
    latest = _prepare(db_client)[0]
    payload = _generate(db_client, "BSD10", latest)["bog_form"]
    assert payload["status_counts"] == {
        "mapped": 0,
        "input_required": 50,
        "unmapped": 0,
        "derived": 16,
    }
    cells = payload["cells"]["BSD10"]
    assert cells["C7"] is None and cells["H7"] == 0  # blank input, BoG's SUM over blanks
    bsd2 = _generate(db_client, "BSD2", latest)
    statuses = _statuses(bsd2, "BSD2")
    assert all(statuses[f"B{row}"] == "input_required" for row in (*range(115, 122), 123))
    assert bsd2["bog_form"]["cells"]["BSD2"]["B115"] is None


# ---------------------------------------------------------------------------
# 4. malformed rows are reported honestly
# ---------------------------------------------------------------------------


def test_schema_reports_malformed_rows() -> None:
    schema = schema_for(KIND)
    assert schema is not None and schema is SCHEMA
    good = next(r for r in read_rows(CSV, entity=False) if r["asset_class"] == "computers")
    assert validate_capex_row(good) == []
    problems = validate_capex_row(
        {**good, "asset_class": "servers", "additions_purchased_ghs": "lots"}
    )
    assert any(p.startswith("field 'asset_class' must be one of") for p in problems)
    assert "field 'additions_purchased_ghs' must be numeric (got 'lots')" in problems
    assert schema.validate_row({k: v for k, v in good.items() if k != "closing_cost_ghs"}) == [
        "missing required field 'closing_cost_ghs'"
    ]
    # NBV that does not reconcile to cost − accumulated depreciation is refused
    problems = validate_capex_row({**good, "closing_nbv_ghs": "1"})
    assert any(p.startswith("closing_nbv_ghs must equal") for p in problems)
    # a six-month forecast that is not the sum of its split is refused
    problems = validate_capex_row({**good, "forecast_next_6m_ghs": "999"})
    assert any(p.startswith("forecast_next_6m_ghs must equal") for p in problems)
