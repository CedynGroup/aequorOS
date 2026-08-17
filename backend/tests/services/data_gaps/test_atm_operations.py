"""``atm_operations`` — monthly ATM / card operations per terminal → BSD16.

1. the Sample Bank month file pushed through the REAL API (three-call push
   flow, the generic ``scripts/ingest_push.py`` reader, ``as_of_date`` = the
   reporting month-end) lands under the kind with batch lineage;
2. BSD16 generated through ``POST /regulatory-packages`` for that period lists
   the terminals in file order — station text, cards issued (unscaled count),
   minimum / maximum withdrawal (cedis, ¢'Million on export) — and BoG's own
   template arithmetic runs over them: ``F = D + E`` per row and
   ``F57 = SUM(F7:F56)``; the TOTAL row D57 / E57 are the register's column
   totals; rows past the estate stay blank; a second month pushed later does
   not disturb the earlier period (one month per push, latest as_of ≤ period
   end);
3. the register schema rejects a malformed row honestly (missing station,
   non-numeric cards count).
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.domain.ingestion.reference_schemas import schema_for
from app.domain.ingestion.reference_schemas.atm_operations import SCHEMA
from app.models import Bank, CanonicalReferenceRow
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from scripts.ingest_push import read_rows
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
SAMPLE_DIR = Path(__file__).resolve().parents[3] / "onboarding" / "sample_bank"
CSV = SAMPLE_DIR / "atm_operations.csv"
ATM = "MONTHLY ATM OPERATIONS"
TERMINALS = 44
STATION_ROWS = 50
M = 1_000_000


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _prepare(db_client: TestClient) -> list[str]:
    """Materialise the hermetic book; return the period ends, latest first."""
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return sorted((p["period_end"] for p in periods), reverse=True)


def _month_rows(month_end: str) -> list[dict[str, Any]]:
    """The Sample Bank rows for one reporting month, re-dated to ``month_end``.

    The hermetic book's periods (2025-04 … 2026-03) only partly overlap the
    Sample Bank's twelve months (2025-07 … 2026-06); when the period has no file
    of its own the latest month's terminals are re-dated — the register is
    monthly, the terminals are the same estate.
    """
    rows = read_rows(CSV, entity=False)
    months = sorted({r["month"] for r in rows})
    source = month_end if month_end in months else months[-1]
    return [{**r, "month": month_end} for r in rows if r["month"] == source]


def _push_reference(
    db_client: TestClient, kind: str, rows: list[dict[str, Any]], *, as_of: str, key: str
) -> dict[str, Any]:
    opened = db_client.post(
        f"{BASE}/push-batches",
        headers=headers(),
        json={"as_of_date": as_of, "idempotency_key": key, "reason": f"Sample Bank {kind}"},
    )
    assert opened.status_code == 201, opened.text
    push_id = opened.json()["push_batch_id"]
    staged = db_client.post(
        f"{BASE}/push-batches/{push_id}/records",
        headers=headers(),
        json={"reference": {kind: rows}},
    )
    assert staged.status_code == 200, staged.text
    committed = db_client.post(f"{BASE}/push-batches/{push_id}/commit", headers=headers())
    assert committed.status_code == 201, committed.text
    return committed.json()


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def _landed(kind: str) -> list[CanonicalReferenceRow]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        return list(
            session.scalars(
                select(CanonicalReferenceRow)
                .where(
                    CanonicalReferenceRow.bank_id == SAMPLE_BANK_ID,
                    CanonicalReferenceRow.dataset_kind == kind,
                )
                .order_by(CanonicalReferenceRow.as_of_date, CanonicalReferenceRow.row_index)
            )
        )
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 1. the Sample Bank register and the line map
# ---------------------------------------------------------------------------


def test_sample_bank_register_is_twelve_months_of_the_same_estate() -> None:
    rows = read_rows(CSV, entity=False)
    months = sorted({r["month"] for r in rows})
    assert len(months) == 12  # noqa: PLR2004
    assert months[0] == "2025-07-31" and months[-1] == "2026-06-30"
    for month in months:
        subset = [r for r in rows if r["month"] == month]
        assert len(subset) == TERMINALS
        assert len({r["atm_id"] for r in subset}) == TERMINALS
        for row in subset:
            assert SCHEMA.validate_row(row) == [], row
            assert float(row["min_withdrawal_ghs"]) <= float(row["max_withdrawal_ghs"])
    # the per-month push files carry exactly that month
    for month in months:
        month_file = SAMPLE_DIR / f"atm_operations_{month[:7]}.csv"
        assert {r["month"] for r in read_rows(month_file, entity=False)} == {month}


def test_line_map_binds_every_station_cell_to_the_register() -> None:
    lines = line_maps_for("BSD16")[ATM]
    by_cell = {ref: line for line in lines for ref in line.cells.values()}
    for row in range(7, 57):
        index = row - 7
        assert by_cell[f"A{row}"].source == "constant"
        for column, field_name, numeric in (
            ("B", "station", False),
            ("C", "cards_issued", True),
            ("D", "min_withdrawal_ghs", True),
            ("E", "max_withdrawal_ghs", True),
        ):
            line = by_cell[f"{column}{row}"]
            assert line.source == "refs.field"
            assert line.params == {
                "kind": "atm_operations",
                "index": index,
                "field": field_name,
                "numeric": numeric,
            }
        assert by_cell[f"C{row}"].unscaled and not by_cell[f"D{row}"].unscaled
    assert by_cell["D57"].source == "refs.sum"
    assert by_cell["D57"].params == {"kind": "atm_operations", "value_field": "min_withdrawal_ghs"}
    assert by_cell["E57"].params == {"kind": "atm_operations", "value_field": "max_withdrawal_ghs"}
    assert len(by_cell) == 252  # noqa: PLR2004 — 50 × 5 + 2, nothing dropped


# ---------------------------------------------------------------------------
# 2. push through the real API → BSD16 fills, BoG's formulas run over it
# ---------------------------------------------------------------------------


def test_pushed_month_fills_bsd16_and_bogs_totals_run_over_it(  # noqa: PLR0915 — one journey
    db_client: TestClient,
) -> None:
    latest, previous = _prepare(db_client)[:2]
    rows = _month_rows(latest)
    result = _push_reference(db_client, "atm_operations", rows, as_of=latest, key=f"atm-{latest}")
    batch = result["batch"]
    assert batch["status"] in ("accepted", "accepted_with_warnings"), batch["validation_report"]
    assert batch["validation_report"]["summary"]["reference_rows"] == {"atm_operations": TERMINALS}
    landed = _landed("atm_operations")
    assert len(landed) == TERMINALS
    assert {str(r.ingestion_batch_id) for r in landed} == {batch["id"]}
    assert all(r.lineage_id is not None for r in landed)
    assert landed[0].payload["atm_id"] == "ATM-001"
    assert landed[0].as_of_date == date.fromisoformat(latest)

    snapshot = _generate(db_client, "BSD16", latest)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    counts = payload["status_counts"]
    blank_rows = STATION_ROWS - TERMINALS
    assert counts["mapped"] == STATION_ROWS + TERMINALS * 4 + 2
    assert counts["input_required"] == blank_rows * 4
    cells = payload["cells"][ATM]
    sum_min = sum_max = 0.0
    for index, row in enumerate(rows):
        r = 7 + index
        assert cells[f"A{r}"] == index + 1
        assert cells[f"B{r}"] == row["station"]
        assert cells[f"C{r}"] == int(row["cards_issued"])
        assert cells[f"D{r}"] == float(row["min_withdrawal_ghs"])
        assert cells[f"E{r}"] == float(row["max_withdrawal_ghs"])
        assert cells[f"F{r}"] == pytest.approx(cells[f"D{r}"] + cells[f"E{r}"])  # =D+E
        sum_min += float(row["min_withdrawal_ghs"])
        sum_max += float(row["max_withdrawal_ghs"])
    for r in range(7 + TERMINALS, 57):  # past the estate: blank, ordinal kept, F = 0
        assert cells[f"A{r}"] == r - 6
        assert cells[f"B{r}"] is None and cells[f"C{r}"] is None
        assert cells[f"D{r}"] is None and cells[f"E{r}"] is None
        assert cells[f"F{r}"] == 0
    assert cells["D57"] == pytest.approx(sum_min)
    assert cells["E57"] == pytest.approx(sum_max)
    assert cells["F57"] == pytest.approx(sum_min + sum_max)  # =SUM(F7:F56), BoG's formula
    assert cells["F57"] == pytest.approx(cells["D57"] + cells["E57"])
    assert cells["B7"] == "Head Office — Independence Avenue"

    # export: station text, unscaled count, ¢'Million amounts, no live formula
    ws = _export("BSD16", snapshot)[ATM]
    assert ws["B7"].value == rows[0]["station"]
    assert ws["C7"].value == int(rows[0]["cards_issued"])
    assert ws["D7"].value == pytest.approx(float(rows[0]["min_withdrawal_ghs"]) / M)
    assert ws["F57"].value == pytest.approx((sum_min + sum_max) / M)
    assert ws[f"B{7 + TERMINALS}"].value is None
    assert not any(
        isinstance(c.value, str) and c.value.startswith("=") for row in ws.iter_rows() for c in row
    )
    residual = {
        row["cell"]
        for section in snapshot["sections"]
        for row in section["rows"]
        if row["status"] == "input_required"
    }
    assert residual == {f"{c}{r}" for c in "BCDE" for r in range(7 + TERMINALS, 57)}
    notes = {
        row["notes"]
        for section in snapshot["sections"]
        for row in section["rows"]
        if row["status"] == "input_required"
    }
    assert all("fewer terminals" in note for note in notes)

    # a later month pushed with its own as_of does not disturb the earlier period
    earlier_rows = [{**r, "cards_issued": str(int(r["cards_issued"]) + 1000)} for r in rows]
    earlier_rows = [{**r, "month": previous} for r in earlier_rows]
    _push_reference(
        db_client, "atm_operations", earlier_rows, as_of=previous, key=f"atm-{previous}"
    )
    again = _generate(db_client, "BSD16", latest)["bog_form"]["cells"][ATM]
    assert again["C7"] == int(rows[0]["cards_issued"])  # latest period still reads its month
    before = _generate(db_client, "BSD16", previous)["bog_form"]["cells"][ATM]
    assert before["C7"] == int(rows[0]["cards_issued"]) + 1000  # earlier period reads its own


def test_form_without_a_register_stays_input_required(db_client: TestClient) -> None:
    latest = _prepare(db_client)[0]
    payload = _generate(db_client, "BSD16", latest)["bog_form"]
    assert payload["status_counts"]["mapped"] == STATION_ROWS  # the ordinals only
    assert payload["status_counts"]["input_required"] == STATION_ROWS * 4 + 2
    cells = payload["cells"][ATM]
    assert cells["F57"] == 0 and cells["D57"] is None


# ---------------------------------------------------------------------------
# 3. malformed rows are reported honestly
# ---------------------------------------------------------------------------


def test_schema_reports_malformed_rows() -> None:
    schema = schema_for("atm_operations")
    assert schema is not None and schema is SCHEMA
    good = {
        "month": "2026-06-30",
        "atm_id": "ATM-001",
        "station": "Head Office",
        "cards_issued": "120",
        "min_withdrawal_ghs": "10",
        "max_withdrawal_ghs": "2000",
    }
    assert schema.validate_row(good) == []
    problems = schema.validate_row({**good, "station": "", "cards_issued": "many"})
    assert "missing required field 'station'" in problems
    assert "field 'cards_issued' must be numeric (got 'many')" in problems
    assert schema.validate_row({k: v for k, v in good.items() if k != "month"}) == [
        "missing required field 'month'"
    ]
