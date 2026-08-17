"""``subsidiaries`` — the subsidiary register + book → BSD9 / BSD5B (and the
documented basis of BSD3B / BSD7B).

1. the Sample Bank register (three subsidiaries at one reporting date)
   validates row by row and its minority-interest workings are the
   non-controlling share of each fully consolidated subsidiary's equity;
2. the line maps: BSD9 "10. Minority interests" and every Annexure cell, BSD5B
   rows 10 / 26 read the register (``refs.sum`` / ``refs.field``); BSD7B and
   BSD3B deliberately do NOT (no P&L item can hold a subsidiary's net profit;
   rosters are a position book) and their notes say so;
3. the register pushed through the REAL API (three-call push flow, the generic
   ``scripts/ingest_push.py`` reader, ``as_of_date`` = reporting date) lands
   under the kind with batch lineage; BSD9 generated through
   ``POST /regulatory-packages`` carries minority interests (Domestic — every
   subsidiary is base-currency functional) and the Annexure's inter-company
   receivables / payables ranked by amount with date, the bank's type text and
   the subsidiary name; BoG's ``D = B + C`` runs over row 30; the balance-sheet
   lines stay the parent's BSD2 figures; BSD5B rows 10 / 26 carry the same
   minority figures; slots past the register stay blank;
4. without a register the cells stay blank (input_required), never zero;
5. the schema rejects a malformed row honestly (bad enum, ownership outside
   0–100, a fully consolidated minority-owned subsidiary without its
   minority-interest figure).
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.domain.ingestion.reference_schemas import schema_for
from app.domain.ingestion.reference_schemas.subsidiaries import (
    SCHEMA,
    validate_subsidiary_row,
)
from app.models import Bank, CanonicalReferenceRow
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.linemaps.bsd9 import (
    ANNEX_COLUMNS,
    ANNEX_PAYABLE_ROWS,
    ANNEX_RECEIVABLE_ROWS,
)
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from scripts.ingest_push import read_rows
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
SAMPLE_DIR = Path(__file__).resolve().parents[3] / "onboarding" / "sample_bank"
CSV = SAMPLE_DIR / "subsidiaries.csv"
KIND = "subsidiaries"
GROUP = "CAR FORMAT-GROUP"
SUBSIDIARIES = 3
M = 1_000_000


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _prepare(db_client: TestClient) -> list[str]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return sorted((p["period_end"] for p in periods), reverse=True)


def _register_rows(reporting_date: str) -> list[dict[str, Any]]:
    """The Sample Bank register re-dated to ``reporting_date`` (the hermetic
    book's periods do not coincide with the Sample Bank's)."""
    return [{**r, "reporting_date": reporting_date} for r in read_rows(CSV, entity=False)]


def _push_reference(
    db_client: TestClient, rows: list[dict[str, Any]], *, as_of: str, key: str
) -> dict[str, Any]:
    opened = db_client.post(
        f"{BASE}/push-batches",
        headers=headers(),
        json={"as_of_date": as_of, "idempotency_key": key, "reason": f"Sample Bank {KIND}"},
    )
    assert opened.status_code == 201, opened.text
    push_id = opened.json()["push_batch_id"]
    staged = db_client.post(
        f"{BASE}/push-batches/{push_id}/records",
        headers=headers(),
        json={"reference": {KIND: rows}},
    )
    assert staged.status_code == 200, staged.text
    committed = db_client.post(f"{BASE}/push-batches/{push_id}/commit", headers=headers())
    assert committed.status_code == 201, committed.text
    return committed.json()


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def _landed() -> list[CanonicalReferenceRow]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        return list(
            session.scalars(
                select(CanonicalReferenceRow)
                .where(
                    CanonicalReferenceRow.bank_id == SAMPLE_BANK_ID,
                    CanonicalReferenceRow.dataset_kind == KIND,
                )
                .order_by(CanonicalReferenceRow.as_of_date, CanonicalReferenceRow.row_index)
            )
        )
    finally:
        session.close()


def _statuses(snapshot: dict[str, Any], sheet: str) -> dict[str, str]:
    return {
        row["cell"]: row["status"]
        for section in snapshot["sections"]
        if section["title"] == sheet
        for row in section["rows"]
    }


# ---------------------------------------------------------------------------
# 1. the Sample Bank register
# ---------------------------------------------------------------------------


def test_sample_bank_register_is_three_consolidated_subsidiaries() -> None:
    rows = read_rows(CSV, entity=False)
    assert len(rows) == SUBSIDIARIES
    assert {r["reporting_date"] for r in rows} == {"2026-06-30"}
    assert len({r["subsidiary_id"] for r in rows}) == SUBSIDIARIES
    for row in rows:
        assert validate_subsidiary_row(row) == [], row
        assert row["consolidation_method"] == "full"
        equity = float(row["total_assets_ghs"]) - float(row["total_liabilities_ghs"])
        assert float(row["equity_ghs"]) == pytest.approx(equity)
        # the group's minority-interest working IS the non-controlling share of equity
        share = (100 - float(row["ownership_pct"])) / 100
        assert float(row["minority_interest_ghs"]) == pytest.approx(share * equity, abs=1)
        assert row["intercompany_receivable_type"] and row["intercompany_payable_type"]
    template = (SAMPLE_DIR / "subsidiaries_template.csv").read_text().strip().split(",")
    assert set(template) == set(SCHEMA.columns)


# ---------------------------------------------------------------------------
# 2. the line maps
# ---------------------------------------------------------------------------


def test_bsd9_and_bsd5b_read_the_register_bsd7b_and_bsd3b_deliberately_do_not() -> None:
    bsd9 = line_maps_for("BSD9")
    minority = next(line for line in bsd9["BSD9"] if line.cells.get("domestic") == "B30")
    assert minority.source == "refs.sum"
    assert minority.params == {
        "kind": KIND,
        "value_field": "minority_interest_ghs",
        "filters": {"consolidation_method": "full"},
        "currency_field": "functional_currency",
    }
    assert minority.cells == {"domestic": "B30", "foreign": "C30"}
    # every other BSD9 line is still the parent's BSD2 figure
    assert all(line.source == "bsd9.bsd2_lines" for line in bsd9["BSD9"] if line.code != "BSD9.R30")
    annex = bsd9["Annexure"]
    by_cell = {ref: line for line in annex for ref in line.cells.values()}
    assert set(by_cell) == {f"{c}{r}" for r in range(11, 29) for c in "ABCD"}
    for block, rows, prefix in (
        ("receivable", ANNEX_RECEIVABLE_ROWS, "intercompany_receivable"),
        ("payable", ANNEX_PAYABLE_ROWS, "intercompany_payable"),
    ):
        for row in rows:
            rank = row - rows[0]
            fields = {
                "date": "reporting_date",
                "type": f"{prefix}_type",
                "subsidiary": "name",
                "amount": f"{prefix}_ghs",
            }
            for column, letter in ANNEX_COLUMNS.items():
                line = by_cell[f"{letter}{row}"]
                assert line.source == "refs.field", (block, row, column)
                assert line.params == {
                    "kind": KIND,
                    "order_by": f"{prefix}_ghs",
                    "desc": True,
                    "index": rank,
                    "field": fields[column],
                    "numeric": column == "amount",
                }
                assert line.code == f"BSD9.ANNEX.{block}.{column}.R{row}"
    # BSD5B: the two consolidation-only rows
    group = {ref: line for line in line_maps_for("BSD5B")[GROUP] for ref in line.cells.values()}
    for ref, field in (
        ("D10", "minority_interest_ghs"),
        ("D26", "minority_interest_tier2_pref_ghs"),
    ):
        assert group[ref].source == "refs.sum"
        assert group[ref].params == {
            "kind": KIND,
            "value_field": field,
            "filters": {"consolidation_method": "full"},
        }
    assert group["D8"].source == "form.cell"  # the solo-linked lines are untouched
    # BSD7B: no cell reads the register (documented decision), the notes say why
    bsd7b = line_maps_for("BSD7B")["BSD7B"]
    assert not any(line.params.get("kind") == KIND for line in bsd7b)
    assert all(
        "subsidiaries register" in line.notes for line in bsd7b if line.code.startswith("BSD7B.R")
    )
    # BSD3B: every roster / count cell stays input_required, naming the register and the
    # subsidiary position book it is not (20 depositors × 5 + count + 10 × 5 + 50 × 5 cells)
    roster_notes = [
        line
        for lines in line_maps_for("BSD3B").values()
        for line in lines
        if line.source is None and "subsidiaries register" in line.notes
    ]
    assert sum(len(line.cells) for line in roster_notes) == 20 * 5 + 1 + 10 * 5 + 50 * 5
    assert all("subsidiary position book" in line.notes for line in roster_notes)
    assert not any(
        line.params.get("kind") == KIND
        for lines in line_maps_for("BSD3B").values()
        for line in lines
    )


# ---------------------------------------------------------------------------
# 3. push through the real API → BSD9 + BSD5B fill
# ---------------------------------------------------------------------------


def test_pushed_register_fills_bsd9_minority_and_annexure_and_bsd5b(  # noqa: PLR0915 — one linear proof
    db_client: TestClient,
) -> None:
    latest = _prepare(db_client)[0]
    rows = _register_rows(latest)
    result = _push_reference(db_client, rows, as_of=latest, key=f"subs-{latest}")
    batch = result["batch"]
    assert batch["status"] in ("accepted", "accepted_with_warnings"), batch["validation_report"]
    assert batch["validation_report"]["summary"]["reference_rows"] == {KIND: SUBSIDIARIES}
    landed = _landed()
    assert len(landed) == SUBSIDIARIES
    assert {str(r.ingestion_batch_id) for r in landed} == {batch["id"]}
    assert all(r.lineage_id is not None for r in landed)
    assert landed[0].payload["subsidiary_id"] == "SUB-001"
    assert landed[0].as_of_date == date.fromisoformat(latest)

    minority_total = sum(float(r["minority_interest_ghs"]) for r in rows)
    assert minority_total == pytest.approx(2_760_000 + 6_625_000)

    # ---- BSD9 ---------------------------------------------------------------
    snapshot = _generate(db_client, "BSD9", latest)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["missing_dependencies"] == []
    cells = payload["cells"]["BSD9"]
    statuses = _statuses(snapshot, "BSD9")
    assert statuses["B30"] == "mapped" and statuses["C30"] == "mapped"
    assert cells["B30"] == pytest.approx(minority_total)  # all base-currency subsidiaries
    assert cells["C30"] == 0
    assert cells["D30"] == pytest.approx(cells["B30"] + cells["C30"])  # =B+C, BoG's formula
    assert statuses["B23"] == "mapped"  # the parent's BSD2 lines are unchanged
    annex = payload["cells"]["Annexure"]
    annex_status = _statuses(snapshot, "Annexure")
    receivable = sorted(rows, key=lambda r: -float(r["intercompany_receivable_ghs"]))
    payable = sorted(rows, key=lambda r: -float(r["intercompany_payable_ghs"]))
    assert [r["subsidiary_id"] for r in receivable] == ["SUB-003", "SUB-001", "SUB-002"]
    assert [r["subsidiary_id"] for r in payable] == ["SUB-001", "SUB-003", "SUB-002"]
    for block_rows, ranked, prefix in (
        (ANNEX_RECEIVABLE_ROWS, receivable, "intercompany_receivable"),
        (ANNEX_PAYABLE_ROWS, payable, "intercompany_payable"),
    ):
        for index, row in enumerate(ranked):
            r = block_rows[0] + index
            assert annex[f"A{r}"] == latest
            assert annex[f"B{r}"] == row[f"{prefix}_type"]
            assert annex[f"C{r}"] == row["name"]
            assert annex[f"D{r}"] == pytest.approx(float(row[f"{prefix}_ghs"]))
            assert all(annex_status[f"{c}{r}"] == "mapped" for c in "ABCD")
        for r in block_rows[len(ranked) :]:  # slots past the register stay blank
            assert all(annex[f"{c}{r}"] is None for c in "ABCD")
            assert all(annex_status[f"{c}{r}"] == "input_required" for c in "ABCD")
    assert annex["C11"] == "Sample Leasing & Finance Ltd" and annex["D11"] == 22_000_000
    assert annex["C20"] == "Sample Investments Ltd" and annex["D20"] == 14_200_000
    assert sum(1 for s in annex_status.values() if s == "mapped") == SUBSIDIARIES * 2 * 4
    assert sum(1 for s in annex_status.values() if s == "input_required") == (
        (9 - SUBSIDIARIES) * 2 * 4
    )
    assert payload["status_counts"]["unmapped"] == 0
    # export: text lands verbatim, amounts in ¢'Million, no live formula
    wb = _export("BSD9", snapshot)
    ws = wb["Annexure"]
    assert ws["A11"].value == latest
    assert ws["B11"].value == receivable[0]["intercompany_receivable_type"]
    assert ws["C11"].value == "Sample Leasing & Finance Ltd"
    assert ws["D11"].value == pytest.approx(22.0)
    assert ws["A14"].value is None and ws["D28"].value is None
    assert wb["BSD9"]["B30"].value == pytest.approx(minority_total / M)
    assert not any(
        isinstance(c.value, str) and c.value.startswith("=") for row in ws.iter_rows() for c in row
    )

    # ---- BSD5B ---------------------------------------------------------------
    group = _generate(db_client, "BSD5B", latest)
    assert not group["bog_form"]["errors"], group["bog_form"]["errors"]
    d = group["bog_form"]["cells"][GROUP]
    group_status = _statuses(group, GROUP)
    assert group_status["D10"] == "mapped" and group_status["D26"] == "mapped"
    assert d["D10"] == pytest.approx(minority_total)
    assert d["D26"] == 0  # no subsidiary carries Tier 2 preferred shares held by minorities


def test_forms_without_a_register_stay_input_required(db_client: TestClient) -> None:
    latest = _prepare(db_client)[0]
    snapshot = _generate(db_client, "BSD9", latest)
    statuses = _statuses(snapshot, "BSD9")
    assert statuses["B30"] == "input_required" and statuses["C30"] == "input_required"
    assert snapshot["bog_form"]["cells"]["BSD9"]["B30"] is None
    annex_status = _statuses(snapshot, "Annexure")
    assert set(annex_status.values()) == {"input_required"}
    assert len(annex_status) == 72  # noqa: PLR2004 — 18 rows × 4 columns, nothing dropped
    group = _generate(db_client, "BSD5B", latest)
    group_status = _statuses(group, GROUP)
    assert group_status["D10"] == "input_required" and group_status["D26"] == "input_required"


# ---------------------------------------------------------------------------
# 4. malformed rows are reported honestly
# ---------------------------------------------------------------------------


def test_schema_reports_malformed_rows() -> None:
    schema = schema_for(KIND)
    assert schema is not None and schema is SCHEMA
    good = next(r for r in read_rows(CSV, entity=False) if r["subsidiary_id"] == "SUB-002")
    assert validate_subsidiary_row(good) == []
    problems = validate_subsidiary_row(
        {
            **good,
            "entity_type": "fintech",
            "consolidation_method": "proportional",
            "equity_ghs": "n/a",
        }
    )
    assert any(p.startswith("field 'entity_type' must be one of") for p in problems)
    assert any(p.startswith("field 'consolidation_method' must be one of") for p in problems)
    assert "field 'equity_ghs' must be numeric (got 'n/a')" in problems
    assert schema.validate_row({k: v for k, v in good.items() if k != "name"}) == [
        "missing required field 'name'"
    ]
    assert validate_subsidiary_row({**good, "ownership_pct": "120"}) == [
        "ownership_pct must be between 0 and 100 (got 120.0)"
    ]
    # a fully consolidated, 60 %-owned subsidiary must state the group's minority working
    assert validate_subsidiary_row({**good, "minority_interest_ghs": ""}) == [
        "a fully consolidated subsidiary owned below 100% must state minority_interest_ghs"
    ]
    assert validate_subsidiary_row({**good, "control_via_board": "yes"}) == [
        "field 'control_via_board' must be one of ['true', 'false'] (got 'yes')"
    ]
