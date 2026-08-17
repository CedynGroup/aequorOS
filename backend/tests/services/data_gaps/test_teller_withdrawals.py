"""``teller_withdrawals`` — over-the-counter cash withdrawals → BSD1A.

1. the Sample Bank week file (one row per teller withdrawal) pushed through the
   REAL API (three-call push flow, the generic ``scripts/ingest_push.py``
   reader, ``as_of_date`` = the week's reporting date) lands under the kind
   with batch lineage;
2. BSD1A generated through ``POST /regulatory-packages`` for that period ranks
   the week's customer accounts by weekly cedi total (largest first), prints
   CUSTOMER / BRANCH / TYPE OF A/C, places each weekday's total under the
   template's day column (¢'Million on export), and BoG's own arithmetic runs
   over it (``J = SUM(E:I)`` per row, ``J31`` grand total); rows dated outside
   the reporting week are ignored; a bank with fewer than twenty ranked
   accounts leaves the tail rows blank;
3. the dataset schema rejects a malformed row honestly (missing branch,
   non-numeric amount, unknown account type).
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.domain.ingestion.reference_schemas import schema_for
from app.domain.ingestion.reference_schemas.teller_withdrawals import SCHEMA
from app.models import Bank, CanonicalReferenceRow
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.sources_ext.bsd1a import DAY_COLUMNS
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from scripts.ingest_push import read_rows
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
SAMPLE_DIR = Path(__file__).resolve().parents[3] / "onboarding" / "sample_bank"
CSV = SAMPLE_DIR / "teller_withdrawals.csv"
SHEET = "20 LARGEST WITHDRAWALS"
KIND = "teller_withdrawals"
SAMPLE_WEEK_END = date(2026, 6, 30)
RANKED_ROWS = range(11, 31)
COLUMN_LETTERS = {"thu": "E", "fri": "F", "mon": "G", "tue": "H", "wed": "I"}
M = 1_000_000


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _prepare(db_client: TestClient) -> str:
    """Materialise the hermetic book; return the latest period end."""
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return max(p["period_end"] for p in periods)


def _week_rows(week_end: str) -> list[dict[str, Any]]:
    """The Sample Bank week re-dated so it ends on ``week_end`` (the hermetic
    book's periods do not include 2026-06-30)."""
    shift = date.fromisoformat(week_end) - SAMPLE_WEEK_END
    rows = read_rows(CSV, entity=False)
    return [
        {**r, "txn_date": (date.fromisoformat(r["txn_date"]) + shift).isoformat()} for r in rows
    ]


def _push_reference(
    db_client: TestClient, rows: list[dict[str, Any]], *, as_of: str, key: str
) -> dict[str, Any]:
    opened = db_client.post(
        f"{BASE}/push-batches",
        headers=headers(),
        json={"as_of_date": as_of, "idempotency_key": key, "reason": f"Sample Bank {KIND}"},
    )
    assert opened.status_code == 201, opened.text
    push_id = opened.json()["push_batch_id"]
    staged = db_client.post(
        f"{BASE}/push-batches/{push_id}/records",
        headers=headers(),
        json={"reference": {KIND: rows}},
    )
    assert staged.status_code == 200, staged.text
    committed = db_client.post(f"{BASE}/push-batches/{push_id}/commit", headers=headers())
    assert committed.status_code == 201, committed.text
    return committed.json()


def _generate(db_client: TestClient, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": "BSD1A", "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _export(snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx("BSD1A", snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def _landed() -> list[CanonicalReferenceRow]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        return list(
            session.scalars(
                select(CanonicalReferenceRow)
                .where(
                    CanonicalReferenceRow.bank_id == SAMPLE_BANK_ID,
                    CanonicalReferenceRow.dataset_kind == KIND,
                )
                .order_by(CanonicalReferenceRow.as_of_date, CanonicalReferenceRow.row_index)
            )
        )
    finally:
        session.close()


def _expected_ranking(rows: list[dict[str, Any]], week_end: date) -> list[dict[str, Any]]:
    """The template's rule, computed independently: accounts (customer ×
    branch × account type) of the seven days ending on ``week_end``, weekday
    rows only, ranked by weekly cedi total (largest first)."""
    start = week_end - timedelta(days=6)
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        day = date.fromisoformat(row["txn_date"])
        if day < start or day > week_end or day.weekday() > 4:  # noqa: PLR2004 — Mon..Fri
            continue
        key = (row["customer_reference"], row["branch"], row["account_type"])
        group = groups.setdefault(
            key,
            {
                "customer": row["customer_name"],
                "branch": row["branch"],
                "account_type": row["account_type"],
                "days": {},
                "total": Decimal(0),
            },
        )
        amount = Decimal(row["amount_ghs"])
        group["days"][day.weekday()] = group["days"].get(day.weekday(), Decimal(0)) + amount
        group["total"] += amount
    return sorted(
        groups.values(), key=lambda g: (-g["total"], g["customer"], g["branch"], g["account_type"])
    )


# ---------------------------------------------------------------------------
# 1. the Sample Bank week and the line map
# ---------------------------------------------------------------------------


def test_sample_bank_week_is_one_reporting_week_of_valid_rows() -> None:
    rows = read_rows(CSV, entity=False)
    assert len(rows) >= 300  # noqa: PLR2004 — a realistic week across the network
    days = sorted({r["txn_date"] for r in rows})
    assert days[0] >= "2026-06-24" and days[-1] == "2026-06-30"
    assert all(date.fromisoformat(d).weekday() < 5 for d in days)  # noqa: PLR2004 — Mon..Fri
    for row in rows:
        assert SCHEMA.validate_row(row) == [], row
        assert row["channel"] == "teller"
        if row["currency"] == "GHS":
            assert Decimal(row["amount"]) == Decimal(row["amount_ghs"])
        else:
            assert Decimal(row["amount_ghs"]) > Decimal(row["amount"])
    assert len({r["txn_reference"] for r in rows}) == len(rows)
    ranked = _expected_ranking(rows, SAMPLE_WEEK_END)
    assert len(ranked) > 20  # noqa: PLR2004 — more accounts than the twenty rows
    assert ranked[0]["total"] >= Decimal(3_000_000)  # the top withdrawers are clearly large
    assert ranked[19]["total"] >= Decimal(500_000)
    assert any(len(g["days"]) > 1 for g in ranked[:20])  # a repeat withdrawer in the top 20
    template = read_rows(SAMPLE_DIR / "teller_withdrawals_template.csv", entity=False)
    assert template == []
    header = (SAMPLE_DIR / "teller_withdrawals_template.csv").read_text().strip().split(",")
    assert set(SCHEMA.columns) <= set(header)


def test_line_map_binds_every_ranked_cell_to_the_dataset() -> None:
    lines = line_maps_for("BSD1A")[SHEET]
    by_cell = {ref: line for line in lines for ref in line.cells.values()}
    for row in RANKED_ROWS:
        assert by_cell[f"A{row}"].source == "constant"
        for column in "BCDEFGHI":
            line = by_cell[f"{column}{row}"]
            assert line.source == "bsd1a.rank"
            assert line.params == {"rank": row - 10}
            assert not line.unscaled  # cedis → ¢'Million on export (text passes through)
            assert "teller_withdrawals" in line.notes
        assert by_cell[f"E{row}"] is by_cell[f"B{row}"]  # one line per ranked row
        assert set(by_cell[f"B{row}"].cells) == {"customer", "branch", "account_type", *DAY_COLUMNS}
    assert len(by_cell) == 180  # noqa: PLR2004 — 20 serials + 20 × 8 data cells, nothing dropped
    assert "J11" not in by_cell and "J31" not in by_cell  # noqa: PT018 — template formulas


# ---------------------------------------------------------------------------
# 2. push through the real API → BSD1A ranks the week, BoG's formulas run over it
# ---------------------------------------------------------------------------


def test_pushed_week_fills_bsd1a_ranked_by_weekly_total(  # noqa: PLR0915 — one journey
    db_client: TestClient,
) -> None:
    week_end = _prepare(db_client)
    rows = _week_rows(week_end)
    result = _push_reference(db_client, rows, as_of=week_end, key=f"teller-{week_end}")
    batch = result["batch"]
    assert batch["status"] in ("accepted", "accepted_with_warnings"), batch["validation_report"]
    assert batch["validation_report"]["summary"]["reference_rows"] == {KIND: len(rows)}
    landed = _landed()
    assert len(landed) == len(rows)
    assert {str(r.ingestion_batch_id) for r in landed} == {batch["id"]}
    assert all(r.lineage_id is not None for r in landed)
    assert landed[0].as_of_date == date.fromisoformat(week_end)
    assert landed[0].payload["channel"] == "teller"

    snapshot = _generate(db_client, week_end)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    counts = payload["status_counts"]
    assert counts["mapped"] == 20 + 20 * 8  # serials + every ranked data cell
    assert counts["input_required"] == 0
    cells = payload["cells"][SHEET]

    expected = _expected_ranking(rows, date.fromisoformat(week_end))
    grand_total = Decimal(0)
    for index, row in enumerate(RANKED_ROWS):
        account = expected[index]
        assert cells[f"A{row}"] == index + 1
        assert cells[f"B{row}"] == account["customer"]
        assert cells[f"C{row}"] == account["branch"]
        assert cells[f"D{row}"] == account["account_type"].title()
        row_total = Decimal(0)
        for column, weekday in DAY_COLUMNS.items():
            amount = account["days"].get(weekday, Decimal(0))
            assert cells[f"{COLUMN_LETTERS[column]}{row}"] == pytest.approx(float(amount)), (
                f"{column}{row}"
            )
            row_total += amount
        assert row_total == account["total"]
        assert cells[f"J{row}"] == pytest.approx(float(row_total))  # =SUM(E:I), BoG's own
        grand_total += row_total
        if index > 0:  # ranked, largest first
            assert expected[index - 1]["total"] >= account["total"]
    assert cells["J31"] == pytest.approx(float(grand_total))  # =SUM(J11:J30)
    # a repeat withdrawer shows one row with several days filled (customer × day matrix)
    multi = next(i for i, g in enumerate(expected[:20]) if len(g["days"]) > 1)
    filled = [c for c in "EFGHI" if cells[f"{c}{11 + multi}"] not in (None, 0)]
    assert len(filled) == len(expected[multi]["days"])

    # export: text as-is, ¢'Million amounts, official numbering, no live formula
    ws = _export(snapshot)[SHEET]
    assert ws["A11"].value == 1
    assert ws["B11"].value == expected[0]["customer"]
    top_day = max(expected[0]["days"], key=lambda d: expected[0]["days"][d])
    top_column = next(c for c, wd in DAY_COLUMNS.items() if wd == top_day)
    assert ws[f"{COLUMN_LETTERS[top_column]}11"].value == pytest.approx(
        float(expected[0]["days"][top_day]) / M
    )
    assert ws["J31"].value == pytest.approx(float(grand_total) / M)
    assert not any(
        isinstance(c.value, str) and c.value.startswith("=") for row in ws.iter_rows() for c in row
    )


def test_rows_outside_the_reporting_week_are_ignored_and_short_weeks_stay_blank(
    db_client: TestClient,
) -> None:
    week_end = _prepare(db_client)
    end = date.fromisoformat(week_end)
    in_week = [
        {
            "txn_date": (end - timedelta(days=1)).isoformat(),
            "branch": "Osu Oxford Street",
            "customer_reference": "CIF-1",
            "customer_name": "Ama Mensah",
            "account_type": "savings",
            "currency": "GHS",
            "amount": "250000",
            "amount_ghs": "250000",
            "channel": "teller",
        },
        {
            "txn_date": end.isoformat(),
            "branch": "Osu Oxford Street",
            "customer_reference": "CIF-1",
            "customer_name": "Ama Mensah",
            "account_type": "savings",
            "currency": "GHS",
            "amount": "50000",
            "amount_ghs": "50000",
            "channel": "teller",
        },
        {
            "txn_date": end.isoformat(),
            "branch": "Kumasi Adum",
            "customer_reference": "CIF-2",
            "customer_name": "Kofi Boateng",
            "account_type": "current",
            "currency": "USD",
            "amount": "10000",
            "amount_ghs": "129400.04",
            "channel": "teller",
        },
    ]
    stale = {  # last week's largest withdrawal — outside the reporting week, ignored
        **in_week[0],
        "txn_date": (end - timedelta(days=9)).isoformat(),
        "customer_reference": "CIF-9",
        "customer_name": "Stale Ltd",
        "amount": "9000000",
        "amount_ghs": "9000000",
    }
    _push_reference(db_client, [stale, *in_week], as_of=week_end, key=f"teller-short-{week_end}")
    snapshot = _generate(db_client, week_end)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    cells = payload["cells"][SHEET]
    assert cells["B11"] == "Ama Mensah" and cells["C11"] == "Osu Oxford Street"
    assert cells["D11"] == "Savings"
    assert cells["J11"] == pytest.approx(300000.0)  # both days of the same account, one row
    assert cells["B12"] == "Kofi Boateng"
    assert cells["J12"] == pytest.approx(129400.04)  # the bank's cedi equivalent, not a rate
    assert "Stale Ltd" not in {cells[f"B{r}"] for r in RANKED_ROWS}
    # rows 13–30: fewer than twenty ranked accounts → blank, said so in the notes
    for row in range(13, 31):
        assert cells[f"A{row}"] == row - 10
        assert all(cells.get(f"{c}{row}") is None for c in "BCDEFGHI")
        assert cells[f"J{row}"] == 0
    assert payload["status_counts"]["mapped"] == 20 + 2 * 8
    assert payload["status_counts"]["input_required"] == 18 * 8
    notes = {
        row["notes"]
        for section in snapshot["sections"]
        for row in section["rows"]
        if row["status"] == "input_required"
    }
    assert notes and all("fewer than" in note for note in notes)


def test_form_without_the_dataset_stays_input_required(db_client: TestClient) -> None:
    week_end = _prepare(db_client)
    payload = _generate(db_client, week_end)["bog_form"]
    assert payload["status_counts"]["mapped"] == 20  # the serials only
    assert payload["status_counts"]["input_required"] == 160  # noqa: PLR2004 — 20 rows × 8
    cells = payload["cells"][SHEET]
    assert cells["J31"] == 0 and cells["B11"] is None


# ---------------------------------------------------------------------------
# 3. malformed rows are reported honestly
# ---------------------------------------------------------------------------


def test_schema_reports_malformed_rows() -> None:
    schema = schema_for(KIND)
    assert schema is not None and schema is SCHEMA
    good = {
        "txn_date": "2026-06-30",
        "branch": "Osu Oxford Street",
        "customer_reference": "CIF-200001",
        "customer_name": "Ama Mensah",
        "account_type": "savings",
        "currency": "GHS",
        "amount": "250000",
        "amount_ghs": "250000",
        "channel": "teller",
    }
    assert schema.validate_row(good) == []
    problems = schema.validate_row(
        {**good, "branch": "", "amount_ghs": "two-fifty", "account_type": "cheque"}
    )
    assert "missing required field 'branch'" in problems
    assert "field 'amount_ghs' must be numeric (got 'two-fifty')" in problems
    assert any(p.startswith("field 'account_type' must be one of") for p in problems)
    assert schema.validate_row({**good, "channel": "atm"}) == [
        "field 'channel' must be one of ['teller'] (got 'atm')"
    ]
    assert schema.validate_row({k: v for k, v in good.items() if k != "txn_date"}) == [
        "missing required field 'txn_date'"
    ]
