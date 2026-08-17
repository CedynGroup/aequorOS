"""``tariff_schedule`` — the charges register keyed by the official BSD15 rows.

Closes BSD15A (Domestic charges of banks · Range of pdts i.r.o sav & cur) and
BSD15B (International Banking Charges) through the Data Engine:

1. the ``row_key`` vocabulary is generated from the line maps, unique per
   sheet, and the Sample Bank CSV covers every value row of every sheet (the
   two numeric-sheet block sub-headings excepted, by design);
2. the Sample Bank CSV pushed through the REAL API (three-call push flow, the
   generic ``scripts/ingest_push.py`` reader) lands under the kind with batch
   lineage; BSD15A / BSD15B generated through ``POST /regulatory-packages``
   then carry the register's ``charge_value`` in every tariff cell —
   ``input_required`` → ``mapped`` — text on the two text sheets, cedi amounts
   on the Range sheet, and the values-only export writes them at the official
   cells;
3. the register schema rejects a malformed row honestly (missing required
   field, unknown sheet code, non-numeric floor).
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.db.session import get_sessionmaker
from app.domain.ingestion.reference_schemas import schema_for
from app.domain.ingestion.reference_schemas.tariff_schedule import SCHEMA
from app.models import Bank, CanonicalReferenceRow
from app.services.regulatory_reporting.bog_forms.linemaps import bsd15a, bsd15b, line_maps_for
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from scripts.ingest_push import read_rows
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
CSV = Path(__file__).resolve().parents[3] / "onboarding" / "sample_bank" / "tariff_schedule.csv"
DOMESTIC, RANGE, INTL = bsd15a.DOMESTIC, bsd15a.RANGE, bsd15b.SHEET
#: numeric-sheet block sub-headings ("Foreign Currency Account" / "Foreign
#: Exchange Account") take no amount — deliberately absent from the register
RANGE_HEADINGS = {"FX.1", "FX.6"}
#: the ML-ETL type coercer's null sentinels (app/etl/preprocessing/type_coercion)
NULL_SENTINELS = {
    "",
    "n/a",
    "na",
    "n.a.",
    "-",
    "--",
    "tbc",
    "tbd",
    "none",
    "null",
    "nil",
    "#n/a",
    ".",
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _prepare(db_client: TestClient) -> str:
    """Materialise the hermetic book; return the latest period end (ISO)."""
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return max(p["period_end"] for p in periods)


def _push_reference(
    db_client: TestClient, kind: str, rows: list[dict[str, Any]], *, as_of: str, key: str
) -> dict[str, Any]:
    opened = db_client.post(
        f"{BASE}/push-batches",
        headers=headers(),
        json={"as_of_date": as_of, "idempotency_key": key, "reason": f"Sample Bank {kind}"},
    )
    assert opened.status_code == 201, opened.text
    push_id = opened.json()["push_batch_id"]
    for start in range(0, len(rows), 4000):
        staged = db_client.post(
            f"{BASE}/push-batches/{push_id}/records",
            headers=headers(),
            json={"reference": {kind: rows[start : start + 4000]}},
        )
        assert staged.status_code == 200, staged.text
    committed = db_client.post(f"{BASE}/push-batches/{push_id}/commit", headers=headers())
    assert committed.status_code == 201, committed.text
    return committed.json()


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def _landed(kind: str) -> list[CanonicalReferenceRow]:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    try:
        return list(
            session.scalars(
                select(CanonicalReferenceRow)
                .where(
                    CanonicalReferenceRow.bank_id == SAMPLE_BANK_ID,
                    CanonicalReferenceRow.dataset_kind == kind,
                )
                .order_by(CanonicalReferenceRow.row_index)
            )
        )
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 1. the row_key vocabulary and the Sample Bank register
# ---------------------------------------------------------------------------


def test_row_keys_are_generated_unique_and_named_after_the_official_item() -> None:
    for keys, rows in (
        (bsd15a.DOMESTIC_KEYS, bsd15a.DOMESTIC_ROWS),
        (bsd15a.RANGE_KEYS, bsd15a.RANGE_ROWS),
        (bsd15b.KEYS, bsd15b.ROWS),
    ):
        assert set(keys) == set(rows)  # every value row has exactly one key
        assert len({k for k, _ in keys.values()}) == len(keys)  # unique per sheet
    # documented anchors
    assert bsd15a.DOMESTIC_KEYS[10] == ("1.1", "minimum")  # COT minimum
    assert bsd15a.RANGE_KEYS[13] == ("S1.1", "Initial Deposit")
    assert bsd15a.RANGE_KEYS[110] == ("FX.1", "Foreign Currency Account")
    assert bsd15b.KEYS[363] == ("17.1", "17. ACCOUNT CLOSURE")  # childless heading
    # every tariff cell is bound to the register through refs.field on charge_value
    for form, sheet in (("BSD15A", DOMESTIC), ("BSD15A", RANGE), ("BSD15B", INTL)):
        for line in line_maps_for(form)[sheet]:
            if line.source == "constant":
                continue
            assert line.source == "refs.field", line
            assert line.params["kind"] == "tariff_schedule"
            assert line.params["field"] == "charge_value"
            filters = cast("dict[str, str]", line.params["filters"])
            assert set(filters) == {"form", "sheet", "row_key"}


def test_sample_bank_register_covers_every_official_tariff_cell() -> None:
    rows = read_rows(CSV, entity=False)
    assert len(rows) == 552  # noqa: PLR2004 — 168 + 75 + 309
    keyed = {(r["form"], r["sheet"], r["row_key"]) for r in rows}
    assert len(keyed) == len(rows)  # one row per official cell
    expected = (
        {("BSD15A", "DOMESTIC", k) for k, _ in bsd15a.DOMESTIC_KEYS.values()}
        | {("BSD15A", "RANGE", k) for k, _ in bsd15a.RANGE_KEYS.values() if k not in RANGE_HEADINGS}
        | {("BSD15B", "INTL", k) for k, _ in bsd15b.KEYS.values()}
    )
    assert keyed == expected
    for row in rows:
        assert SCHEMA.validate_row(row) == [], row
        # the Data Engine's type coercer reads "N/A" / "Nil" / "-" / "none" as null
        # sentinels (→ blank cell), so a tariff register spells them out in words
        assert row["charge_value"].strip().lower() not in NULL_SENTINELS, row
        if row["sheet"] == "RANGE":
            float(row["charge_value"])  # cedi amounts on the numeric sheet


# ---------------------------------------------------------------------------
# 2. push through the real API → BSD15A / BSD15B cells flip to mapped
# ---------------------------------------------------------------------------


def test_pushed_register_lands_with_lineage_and_fills_bsd15a_and_bsd15b(
    db_client: TestClient,
) -> None:
    reporting_date = _prepare(db_client)
    rows = read_rows(CSV, entity=False)
    result = _push_reference(
        db_client, "tariff_schedule", rows, as_of=reporting_date, key="tariff-2026"
    )
    batch = result["batch"]
    assert batch["status"] in ("accepted", "accepted_with_warnings"), batch["validation_report"]
    assert batch["validation_report"]["summary"]["reference_rows"] == {"tariff_schedule": 552}

    landed = _landed("tariff_schedule")
    assert len(landed) == 552  # noqa: PLR2004
    assert {str(r.ingestion_batch_id) for r in landed} == {batch["id"]}
    assert all(r.lineage_id is not None for r in landed)
    assert landed[0].payload["row_key"] == "1.1" and landed[0].payload["form"] == "BSD15A"

    # BSD15A — every tariff cell carries the register's charge_value
    snapshot = _generate(db_client, "BSD15A", reporting_date)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    counts = payload["status_counts"]
    assert counts["mapped"] == 22 + 168 + 75  # item numbers + domestic + range  # noqa: PLR2004
    assert counts["input_required"] == len(RANGE_HEADINGS)
    assert counts["unmapped"] == 0
    cells = payload["cells"]
    assert cells[DOMESTIC]["C10"] == "GHS 5.00 per month"  # 1.1 COT minimum
    assert cells[DOMESTIC]["C11"] == "GHS 500.00 per month"  # 1.2 COT maximum
    assert cells[DOMESTIC]["C205"] == "GHS 20.00 per debit card"  # 21.1 ATM issuing
    assert cells[DOMESTIC]["C231"] == "Not applicable"  # 22.16 Sika card
    assert cells[DOMESTIC]["A9"] == 1 and cells[DOMESTIC]["A211"] == 22  # noqa: PLR2004
    assert cells[RANGE]["B13"] == 50  # noqa: PLR2004 — S1.1 initial deposit (cedis)
    assert cells[RANGE]["B99"] == 2000  # noqa: PLR2004 — C3.1 corporate current initial deposit
    assert cells[RANGE]["B110"] is None and cells[RANGE]["B116"] is None  # FX.1 / FX.6 headings
    residual = [
        (section["title"], row["cell"])
        for section in snapshot["sections"]
        for row in section["rows"]
        if row["status"] == "input_required"
    ]
    assert residual == [(RANGE, "B110"), (RANGE, "B116")]

    wb = _export("BSD15A", snapshot)
    ws = wb[DOMESTIC]
    assert ws["B9"].value == "COMMISSION ON TURNOVER (COT)"
    assert ws["C10"].value == "GHS 5.00 per month"
    assert ws["C231"].value == "Not applicable"
    ws2 = wb[RANGE[:31]]
    assert ws2["B13"].value == 50  # noqa: PLR2004 — "Amounts in Cedis", unscaled
    assert ws2["B110"].value is None
    listed = {(row[0].value, row[1].value) for row in wb["Completion notes"].iter_rows(min_row=1)}
    assert (DOMESTIC, "C10") not in listed and (RANGE, "B110") in listed

    # BSD15B — all 309 international tariff cells
    snapshot_b = _generate(db_client, "BSD15B", reporting_date)
    payload_b = snapshot_b["bog_form"]
    assert not payload_b["errors"], payload_b["errors"]
    assert payload_b["status_counts"] == {
        "mapped": 309,
        "input_required": 0,
        "unmapped": 0,
        "derived": 0,
    }
    cells_b = payload_b["cells"][INTL]
    assert cells_b["B10"] == "0.50% per quarter of value, min GHS 300 (or part thereof)"
    assert cells_b["B363"] == "GHS 50.00 if closed within 6 months of opening; nil thereafter"
    assert cells_b["B392"] == "GHS 120.00 per book"  # 21.5 corporate cheque book
    ws_b = _export("BSD15B", snapshot_b)[INTL]
    assert ws_b["B363"].value == cells_b["B363"]


def test_form_without_a_register_stays_input_required(db_client: TestClient) -> None:
    """No register ingested → every tariff cell honest input_required (unchanged)."""
    reporting_date = _prepare(db_client)
    payload = _generate(db_client, "BSD15B", reporting_date)["bog_form"]
    assert payload["status_counts"]["mapped"] == 0
    assert payload["status_counts"]["input_required"] == 309  # noqa: PLR2004


# ---------------------------------------------------------------------------
# 3. malformed rows are reported honestly
# ---------------------------------------------------------------------------


def test_schema_reports_malformed_rows() -> None:
    schema = schema_for("tariff_schedule")
    assert schema is not None and schema is SCHEMA
    assert schema.required == ("form", "sheet", "row_key", "charge_value")
    good = {"form": "BSD15A", "sheet": "DOMESTIC", "row_key": "1.1", "charge_value": "GHS 5.00"}
    assert schema.validate_row(good) == []
    problems = schema.validate_row(
        {"form": "BSD15A", "sheet": "BOGUS", "row_key": "1.1", "min_ghs": "five"}
    )
    assert "missing required field 'charge_value'" in problems
    assert any("'sheet' must be one of" in p for p in problems)
    assert any("'min_ghs' must be numeric" in p for p in problems)
    assert schema.validate_row({**good, "charge_basis": "hourly"}) == [
        "field 'charge_basis' must be one of ['flat', 'percent', 'per_item', 'range'] "
        "(got 'hourly')"
    ]
