"""Stress Test Output Report pack (product.md §Phase 2 item 6).

The STRESS-PACK generator re-tabulates ONLY stored engine state — latest
succeeded liquidity/capital runs per scenario, their persisted headline
metric results, and the reverse-stress frontier — into the standardized
Board/ALCO artifact. Nothing is recomputed, so every number must trace to a
source-run input hash.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.models import Bank, BankReportingPeriod, RegulatoryPackage
from app.schemas.regulatory_capital import CapitalScenarioBatchCreate
from app.schemas.regulatory_liquidity import RegulatoryRunCreate
from app.schemas.regulatory_reporting import RegulatoryPackageCreate
from app.schemas.reverse_stress import ReverseStressRunCreate
from app.services import regulatory_capital, regulatory_liquidity, reverse_stress
from app.services.regulatory_reporting import generation
from app.services.regulatory_reporting.exports import export_package
from app.services.sample_bank_seed import (
    DEMO_ORG_ID,
    DEMO_USER_ID,
    SAMPLE_BANK_ID,
    seed_sample_bank,
)
from tests.storage.inmemory import InMemoryStorageClient

MAKER = TenantContext(organization_id=DEMO_ORG_ID, actor_user_id=DEMO_USER_ID)
REPORTING_DATE = date(2026, 3, 31)


@pytest.fixture
def storage(monkeypatch: pytest.MonkeyPatch) -> InMemoryStorageClient:
    client = InMemoryStorageClient()
    monkeypatch.setattr(
        "app.services.regulatory_reporting.exports.get_storage_client", lambda: client
    )
    return client


def _period_id(db: Session):
    period_id = db.scalar(
        select(BankReportingPeriod.id).where(
            BankReportingPeriod.organization_id == DEMO_ORG_ID,
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == REPORTING_DATE,
        )
    )
    assert period_id is not None
    return period_id


def _seed_stress_outcomes(db: Session, *, with_reverse: bool = True) -> None:
    """Baselines + the combined liquidity scenario + all capital scenarios
    (+ the reverse-stress frontier) over the deterministic seeded book."""
    seed_sample_bank(db)
    period_id = _period_id(db)
    for scenario in ("baseline", "combined"):
        run = regulatory_liquidity.create_liquidity_run(
            db,
            MAKER,
            SAMPLE_BANK_ID,
            RegulatoryRunCreate(
                module="liquidity", reporting_period_id=period_id, scenario_code=scenario
            ),
        )
        assert run.status == "succeeded", run
    batch = regulatory_capital.run_all_capital_scenarios(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        CapitalScenarioBatchCreate(reporting_period_id=period_id),
    )
    assert all(run.status == "succeeded" for run in batch.runs)
    if with_reverse:
        reverse_stress.run_reverse_stress(
            db, MAKER, SAMPLE_BANK_ID, ReverseStressRunCreate(reporting_period_id=period_id)
        )


def _generate(db: Session) -> RegulatoryPackage:
    read = generation.generate_package(
        db,
        MAKER,
        SAMPLE_BANK_ID,
        RegulatoryPackageCreate(return_code="STRESS-PACK", reporting_date=REPORTING_DATE),
    )
    row = db.scalar(select(RegulatoryPackage).where(RegulatoryPackage.id == read.id))
    assert row is not None
    return row


def _section(snapshot: dict[str, Any], code: str) -> dict[str, Any]:
    return next(section for section in snapshot["sections"] if section["code"] == code)


def test_stress_pack_assembles_stored_runs_without_recomputation(db_session: Session) -> None:
    _seed_stress_outcomes(db_session)
    package = _generate(db_session)
    snapshot = package.snapshot
    assert snapshot["template_id"] == "aeq-stress-pack-v1"
    assert package.return_family == "stress"

    # Traffic lights: exactly the engines' stored headline metric results —
    # value, threshold and status re-tabulated for every consumed run.
    lights = _section(snapshot, "traffic_lights")["rows"]
    lcr_combined = next(r for r in lights if r["code"] == "liquidity:combined:lcr_pct")
    assert lcr_combined["status"] in ("green", "amber", "red")
    assert Decimal(lcr_combined["threshold"]) == Decimal("100")
    # The seeded combined scenario lands below the 100% floor (LCR 87.36).
    assert Decimal(lcr_combined["value"]) < Decimal("100")
    assert lcr_combined["status"] in ("amber", "red")

    # Ratio evolution: the severe capital scenario's stored quarterly path
    # (quarter 0 is the unstressed starting position).
    evolution = _section(snapshot, "ratio_evolution")["rows"]
    severe_quarters = [r for r in evolution if r["scenario_code"] == "severe"]
    assert [r["code"] for r in severe_quarters] == [
        "severe:q0",
        "severe:q1",
        "severe:q2",
        "severe:q3",
        "severe:q4",
    ]

    # Pro-forma capital: the final-quarter stack, verbatim from the run —
    # the end-state CAR and RWA must equal the path's last quarter.
    pro_forma = _section(snapshot, "pro_forma_capital")["rows"]
    severe_end = next(r for r in pro_forma if r["code"] == "severe")
    assert severe_end["end_car_pct"] == severe_quarters[-1]["value"]
    assert severe_end["total_rwa_ghs"] == severe_quarters[-1]["total_rwa_ghs"]
    assert Decimal(severe_end["value"]) > 0

    # Attribution: scenario minus baseline, computed from the stored metrics.
    attribution = _section(snapshot, "attribution")["rows"]
    lcr_attr = next(r for r in attribution if r["code"] == "liquidity:combined:lcr_pct")
    assert Decimal(lcr_attr["value"]) == Decimal(lcr_attr["stressed_pct"]) - Decimal(
        lcr_attr["baseline_pct"]
    )
    assert Decimal(lcr_attr["value"]) < 0

    # Reverse-stress frontier rides along when a frontier run exists.
    frontier = _section(snapshot, "reverse_stress_frontier")["rows"]
    liq_frontier = next(r for r in frontier if r["code"] == "liquidity_frontier")
    assert liq_frontier["breached"] == "true"
    assert Decimal(liq_frontier["value"]) < Decimal("1")
    assert snapshot["metadata"]["reverse_stress_narrative"]

    # Recommended actions include a funding-remediation line for the breached
    # combined-scenario LCR.
    actions = _section(snapshot, "recommended_actions")["rows"]
    assert any(r["code"] == "liquidity:combined:lcr_pct" for r in actions)

    # Provenance: every consumed run appears in source_runs, and the headline
    # totals summarize the stress ensemble.
    modules = {entry["module"] for entry in package.source_runs}
    assert {"liquidity", "capital", "reverse_stress"} <= modules
    totals = {row["code"]: row for row in snapshot["totals"]}
    assert Decimal(totals["worst_stressed_lcr_pct"]["value"]) == Decimal(lcr_combined["value"])
    assert "worst_end_car_pct" in totals
    assert int(totals["red_light_count"]["value"]) >= 0


def test_stress_pack_without_frontier_omits_the_optional_section(db_session: Session) -> None:
    _seed_stress_outcomes(db_session, with_reverse=False)
    package = _generate(db_session)
    snapshot = package.snapshot
    assert _section(snapshot, "reverse_stress_frontier")["rows"] == []
    assert "reverse_stress_narrative" not in snapshot["metadata"]
    assert {entry["module"] for entry in package.source_runs} == {"liquidity", "capital"}


def test_stress_pack_requires_a_stress_scenario(db_session: Session) -> None:
    seed_sample_bank(db_session)
    period_id = _period_id(db_session)
    for module, runner in (
        ("liquidity", regulatory_liquidity.create_liquidity_run),
        ("capital", regulatory_capital.create_capital_run),
    ):
        run = runner(
            db_session,
            MAKER,
            SAMPLE_BANK_ID,
            RegulatoryRunCreate(
                module=module, reporting_period_id=period_id, scenario_code="baseline"  # type: ignore[index]
            ),
        )
        assert run.status == "succeeded"
    with pytest.raises(HTTPException) as excinfo:
        _generate(db_session)
    assert excinfo.value.status_code == 409
    assert excinfo.value.detail["error_code"] == "no_stress_scenarios"  # type: ignore[index]


def test_stress_pack_exports_to_xlsx(
    db_session: Session, storage: InMemoryStorageClient
) -> None:
    _seed_stress_outcomes(db_session)
    package = _generate(db_session)
    artifact = export_package(db_session, MAKER, package, "xlsx")
    slug = db_session.scalar(select(Bank.storage_slug).where(Bank.id == SAMPLE_BANK_ID))
    assert slug
    payload = None
    for obj in storage.list(slug, "outputs"):
        if obj.location.object_path == artifact.object_path:
            _, stream = storage.read(obj.location)
            payload = stream.read()
    assert payload is not None
    workbook = load_workbook(io.BytesIO(payload))
    assert workbook.sheetnames == [
        "Return Metadata",
        "Stress Outcome Traffic Lights",
        "Capital Ratio Evolution Under S",
        "Pro-Forma Capital (Stress End-S",
        "Scenario Attribution vs Baselin",
        "Reverse-Stress Frontier",
        "Recommended Actions",
        "Fidelity & Provenance",
    ]
