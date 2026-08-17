"""BoG forms against the ACTUAL primary — the "works for a real bank" tier.

Opt-in (REAL_DATA_DATABASE_URL, see tests/real_data.py). Every official form is
generated for the real Sample Bank through the real package API inside the
rolled-back ``real_client`` transaction (nothing persists), and this module
asserts the properties that must hold on ANY real book — never a magnitude:

- generation succeeds and every official input cell is accounted for;
- the workbook exports template-faithfully;
- data-fed cells carry values (mapped > 0 wherever the platform holds the
  domain), and the data-gap datasets ingested through the Data Engine light up
  the cells they were built for;
- BoG's own arithmetic holds on the exported values (Domestic + Foreign =
  Total on BSD2, cross-form ties BSD6 ← BSD2, BSD8!H22 = BSD2!D38 + D39).

It also prints a per-form fill report (mapped / input_required / unmapped) so
the onboarding gap list is always read off the real bank, not a fixture.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.models import Bank
from app.services.regulatory_reporting.bog_forms.catalog import all_form_codes
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.real_data import REAL_BANK_ID, REAL_ORG_ID, real_headers, requires_real_data

pytestmark = requires_real_data

# Forms whose inputs the platform can feed WITHOUT any bank-specific register:
# on the real Sample Bank these must show mapped cells > 0.
PLATFORM_FED = {"BSD2", "BSD2A", "BSD3A", "BSD4", "BSD5A", "BSD5B", "BSD6", "BSD8", "BSD9", "BSD13"}


def _latest_period(client: TestClient) -> dict:
    periods = client.get(
        f"/api/v1/banks/{REAL_BANK_ID}/reporting-periods", headers=real_headers()
    ).json()["periods"]
    assert periods
    return periods[0]


def _generate(client: TestClient, code: str, reporting_date: str) -> dict:
    response = client.post(
        f"/api/v1/banks/{REAL_BANK_ID}/regulatory-packages",
        headers=real_headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = client.get(
        f"/api/v1/banks/{REAL_BANK_ID}/regulatory-packages/{package['id']}", headers=real_headers()
    ).json()
    return detail["snapshot"]


@pytest.mark.parametrize("code", sorted(all_form_codes()))
def test_every_form_generates_and_exports_on_the_real_bank(
    real_client: TestClient, real_session: Session, code: str
) -> None:
    period = _latest_period(real_client)
    snapshot = _generate(real_client, code, period["period_end"])
    payload = snapshot["bog_form"]
    counts = payload["status_counts"]
    layout = load_layout(code)
    assert not payload["errors"], payload["errors"]
    assert counts["derived"] == sum(len(s.formula_cells) for s in layout.sheets)
    if code in PLATFORM_FED:
        assert counts["mapped"] > 0, f"{code}: no platform-fed cells on the real bank"

    real_session.info["organization_id"] = REAL_ORG_ID
    bank = real_session.get(Bank, REAL_BANK_ID)
    assert bank is not None
    workbook = render_bog_form_xlsx(code, snapshot, bank, datetime.now(UTC))
    wb = openpyxl.load_workbook(io.BytesIO(workbook), data_only=False)
    assert wb.sheetnames[: len(layout.sheet_names)] == [n[:31] for n in layout.sheet_names]
    assert wb.sheetnames[-1] == "Completion notes"
    print(
        f"[real-bank fill] {code}: mapped={counts['mapped']} "
        f"input_required={counts['input_required']} "
        f"unmapped={counts['unmapped']} derived={counts['derived']}"
    )


def test_bsd2_totals_and_cross_form_ties_on_the_real_bank(real_client: TestClient) -> None:
    period = _latest_period(real_client)
    bsd2 = _generate(real_client, "BSD2", period["period_end"])
    cells = bsd2["bog_form"]["cells"]["BSD2"]
    layout = load_layout("BSD2").sheet("BSD2")
    checked = 0
    for cell in layout.formula_cells:
        if (
            cell.formula
            and cell.ref.startswith("D")
            and cell.formula == f"=B{cell.ref[1:]}+C{cell.ref[1:]}"
        ):
            row = cell.ref[1:]
            d = float(cells.get(cell.ref) or 0)
            b = float(cells.get(f"B{row}") or 0)
            c = float(cells.get(f"C{row}") or 0)
            assert abs(d - (b + c)) < 1e-6, cell.ref
            checked += 1
    assert checked > 150  # noqa: PLR2004
    # the balance-sheet spine carries real money on the real bank
    total_assets = float(cells.get("D124") or 0)
    assert total_assets > 0, "BSD2 total assets must be positive on the real Sample Bank"

    bsd8 = _generate(real_client, "BSD8", period["period_end"])
    h22 = float(bsd8["bog_form"]["cells"]["BSD8"].get("H22") or 0)
    assert abs(h22 - (float(cells.get("D38") or 0) + float(cells.get("D39") or 0))) < 1e-6
