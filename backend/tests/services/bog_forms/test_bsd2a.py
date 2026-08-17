"""BSD2A — Monthly Report on Foreign Currency Exposures (Wave 1b).

The official sheet is a BLANK 14-column grid; ``linemaps/bsd2a.py`` declares
its 24 category rows (E cedi equivalent · G net worth · H % of net worth ·
N provision) and 54 per-counterparty detail rows (all 14 columns), and binds
each category row's cedi column to the BSD2 FOREIGN-column line it analyses
(Guide BSD2A ¶1). Proves, on the hermetic book plus a small canonical slice:

1. the declared grid is exactly category rows × {E,G,H,N} + detail rows ×
   A..N, inside the sheet, no heading/spacer row bound, unique codes, and
   every BSD2 reference the map names is a real BSD2 FOREIGN-column (C) or
   shareholders'-funds cell;
2. the form generates through the real package pipeline with BSD2 computed
   first (no missing dependency, no engine errors);
3. critical totals — BSD2A category rows equal BSD2's foreign column:
   ``E13 = BSD2!C7`` (FX notes and coins), ``E27 = BSD2!C68`` gross loans
   with ``N27 = BSD2!C69`` provisions, ``E30 = C34 + C72 + C102`` securities,
   ``E39 = BSD2!C113`` other assets, ``E66 = C228 + C233``, ``G = BSD2!D135``
   (net worth) on every category row and ``H = 100 × E / G`` (unscaled %);
   the Guide-¶7 judgement rows stay input_required;
4. the values-only xlsx export carries the ¢'Million-scaled cedi equivalents,
   the UNscaled percentage, and lists the detail rows in the Completion notes;
5. the ``bsd2a.form_cells_sum`` / ``bsd2a.form_cells_ratio_pct`` resolvers
   unit-tested against a dependency payload.
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any, cast

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.db.session import get_sessionmaker
from app.models import (
    Bank,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    IngestionBatch,
    LineageRecord,
)
from app.services.regulatory_reporting.bog_forms.catalog import form_spec
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.linemaps.bsd2a import (
    BSD2_FOREIGN_CELLS,
    CATEGORY_ROWS,
    DETAIL_ROWS,
    JUDGEMENT_ROWS,
    SHEET,
)
from app.services.regulatory_reporting.bog_forms.sources import ResolveContext, get_resolver
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

M = Decimal("1000000")
CATEGORY_COLUMNS = ("E", "G", "H", "N")
DETAIL_COLUMNS = tuple("ABCDEFGHIJKLMN")


# ---------------------------------------------------------------------------
# canonical slice: FX-denominated positions so BSD2's foreign column is live
# ---------------------------------------------------------------------------


def _seed_slice(db: Session, as_of: date) -> None:
    batch = IngestionBatch(
        organization_id=ORG_1,
        bank_id=SAMPLE_BANK_ID,
        source_system="EXCEL_CSV",
        adapter_version="1.0",
        extraction_mode="full",
        status="accepted",
        as_of_date=as_of,
    )
    db.add(batch)
    db.flush()
    lineage = LineageRecord(
        organization_id=ORG_1,
        ingestion_batch_id=batch.id,
        operation_type="ADAPTER_TRANSLATE",
        operation_ref="bsd2a-fixture",
        input_lineage_ids=[],
    )
    db.add(lineage)
    db.flush()
    common: dict[str, Any] = {
        "organization_id": ORG_1,
        "bank_id": SAMPLE_BANK_ID,
        "as_of_date": as_of,
        "source_system": "EXCEL_CSV",
        "ingestion_batch_id": batch.id,
        "lineage_id": lineage.id,
        "validation_status": "accepted",
    }
    kumasi = CanonicalCounterparty(
        **common,
        source_reference="CP/KUM",
        name="Kumasi Traders Ltd",
        counterparty_type="CORPORATE",
    )
    db.add(kumasi)
    db.flush()

    def position(
        ref: str, position_type: str, amount: Decimal, *, cpty: bool, attributes: dict[str, Any]
    ) -> None:
        position = CanonicalPosition(
            **common, source_reference=ref, position_type=position_type, currency="USD"
        )
        db.add(position)
        db.flush()
        db.add(
            CanonicalPositionSnapshot(
                **common,
                source_reference=ref,
                position_id=position.id,
                counterparty_id=kumasi.id if cpty else None,
                balance=amount,
                attributes={"balance_ghs": str(amount), **attributes},
            )
        )
        db.flush()

    # BSD2 A.1 foreign currency notes and coins → BSD2A row 13
    position(
        "CASH/USD",
        "CASH",
        Decimal("200000"),
        cpty=False,
        attributes={"instrument": "fx_notes_coins"},
    )
    # BSD2 8(d) private-enterprise loan (foreign column C66 → C68) → BSD2A row 27
    position("LN/KUM/USD", "LOAN", Decimal("1500000"), cpty=True, attributes={})
    # BSD2 7(f)(i) cocoa bills in USD (C54 → C53 → C34) → BSD2A row 30 securities (a SUM)
    position(
        "SEC/COCOA/USD",
        "SECURITY_HOLDING",
        Decimal("800000"),
        cpty=False,
        attributes={"instrument": "cocoa_bill"},
    )
    db.flush()


def _session() -> Session:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    return session


def _period_end(db_client: TestClient) -> str:
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    return periods[0]["period_end"]


@pytest.fixture
def seeded_book(db_client: TestClient) -> str:
    session = _session()
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    reporting_date = _period_end(db_client)
    session = _session()
    try:
        _seed_slice(session, date.fromisoformat(reporting_date))
        session.commit()
    finally:
        session.close()
    return reporting_date


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    return detail["snapshot"]


def _f(cells: dict[str, Any], ref: str) -> float:
    return float(cells.get(ref) or 0.0)


# ---------------------------------------------------------------------------
# 1. structure
# ---------------------------------------------------------------------------


def test_bsd2a_grid_is_declared_exactly_and_inside_the_sheet() -> None:
    layout = load_layout("BSD2A")
    sheet = layout.sheet(SHEET)
    assert sheet.input_cells == ()  # the official template ships blank
    lines = line_maps_for("BSD2A")[SHEET]
    declared = {ref for line in lines for ref in line.cells.values()}
    expected = {f"{c}{r}" for r in CATEGORY_ROWS for c in CATEGORY_COLUMNS} | {
        f"{c}{r}" for r in DETAIL_ROWS for c in DETAIL_COLUMNS
    }
    assert declared == expected
    assert len(declared) == 24 * 4 + 54 * 14 == 852
    for ref in declared:
        cell = sheet.by_ref.get(ref)
        assert cell is None, f"{ref} is a template {cell and cell.kind} cell"
        row = int("".join(ch for ch in ref if ch.isdigit()))
        assert 13 <= row <= 105 <= sheet.max_row
    # category rows are the labelled item rows; heading rows never bound
    for row in CATEGORY_ROWS:
        assert sheet.label_for_row(row), f"row {row} has no official label"
    for row in DETAIL_ROWS:
        assert not sheet.label_for_row(row), f"row {row} is labelled — not a detail row"
    declared_rows = {int("".join(ch for ch in ref if ch.isdigit())) for ref in declared}
    assert declared_rows.isdisjoint(
        {11, 12, 19, 20, 21, 42, 43, 52, 53, 62, 63, 64, 65, 76, 79, 97, 98}
    )
    codes = [line.code for line in lines]
    assert len(codes) == len(set(codes))
    refs = [ref for line in lines for ref in line.cells.values()]
    assert len(refs) == len(set(refs))


def test_bsd2a_map_names_only_real_bsd2_foreign_column_cells() -> None:
    bsd2 = load_layout("BSD2").sheet("BSD2")
    for row, refs in BSD2_FOREIGN_CELLS.items():
        assert row in CATEGORY_ROWS
        for ref in refs:
            cell = bsd2.by_ref.get(ref)
            assert cell is not None and cell.kind in ("input", "formula"), (
                f"BSD2!{ref} for row {row}"
            )
            assert ref.startswith("C"), f"row {row}: {ref} is not the FOREIGN column"
    d135 = bsd2.by_ref["D135"]
    assert d135.kind == "formula" and bsd2.label_for_row(135).startswith("16.  Shareholders' Funds")
    assert bsd2.by_ref["C69"].kind == "input" and "provision" in bsd2.label_for_row(69).lower()
    assert set(JUDGEMENT_ROWS) == {71, 99, 103}
    assert set(BSD2_FOREIGN_CELLS).isdisjoint(JUDGEMENT_ROWS)


def test_bsd2a_status_split_is_what_the_doc_claims() -> None:
    spec = form_spec("BSD2A")
    assert spec.depends_on == ("BSD2",)
    sheet = spec.sheet(SHEET)
    assert sheet is not None
    mapped = sum(len(line.cells) for line in sheet.lines if line.source is not None)
    pending = sum(len(line.cells) for line in sheet.lines if line.source is None)
    # E: 21 category rows via form.cell / bsd2a.form_cells_sum; G: 24 net worth;
    # H: 21 ratios; N: the loans row → 67 mapped; 3+3+23 category + 756 detail pending
    assert (mapped, pending) == (67, 785)
    by_ref = {ref: line for line in sheet.lines for ref in line.cells.values()}
    assert by_ref["E30"].source == "bsd2a.form_cells_sum"
    assert by_ref["E30"].params["refs"] == ["C34", "C72", "C102"]
    assert by_ref["H30"].source == "bsd2a.form_cells_ratio_pct" and by_ref["H30"].unscaled
    assert by_ref["G13"].params == {"form": "BSD2", "sheet": "BSD2", "ref": "D135"}
    assert by_ref["N27"].params["ref"] == "C69" and by_ref["N13"].source is None
    assert by_ref["E99"].source is None and "¶7" in by_ref["E99"].notes
    assert by_ref["A14"].source is None and "detail schedule row 14" in by_ref["A14"].notes


# ---------------------------------------------------------------------------
# 2–3. generation with the BSD2 dependency + critical totals
# ---------------------------------------------------------------------------


def test_bsd2a_category_rows_equal_bsd2_foreign_column(
    db_client: TestClient, seeded_book: str
) -> None:
    bsd2 = _generate(db_client, "BSD2", seeded_book)["bog_form"]["cells"]["BSD2"]
    snapshot = _generate(db_client, "BSD2A", seeded_book)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["missing_dependencies"] == []
    fx = payload["cells"][SHEET]

    # every category row's cedi equivalent = the BSD2 foreign-column cell(s) it names
    for row, refs in BSD2_FOREIGN_CELLS.items():
        assert _f(fx, f"E{row}") == pytest.approx(sum(_f(bsd2, ref) for ref in refs)), row
    assert _f(fx, "E13") == 200_000.0 == _f(bsd2, "C7")  # FX notes and coins
    assert _f(fx, "E27") == 1_500_000.0 == _f(bsd2, "C68")  # gross FX loans
    assert _f(fx, "N27") == pytest.approx(_f(bsd2, "C69"))  # provisions column
    assert _f(fx, "E39") == _f(bsd2, "C113") == 0.0  # other assets (no FX fact on the book)
    assert _f(fx, "E30") == pytest.approx(_f(bsd2, "C34") + _f(bsd2, "C72") + _f(bsd2, "C102"))
    assert _f(fx, "E30") == 800_000.0 == _f(bsd2, "C34")  # the USD cocoa bill, via C54→C53→C34
    assert _f(fx, "E66") == pytest.approx(_f(bsd2, "C228") + _f(bsd2, "C233"))

    # net worth (BSD2 16 shareholders' funds) on every category row; % unscaled
    net_worth = _f(bsd2, "D135")
    assert net_worth > 0
    for row in CATEGORY_ROWS:
        assert _f(fx, f"G{row}") == pytest.approx(net_worth), row
    assert _f(fx, "H27") == pytest.approx(100 * 1_500_000.0 / net_worth)
    assert _f(fx, "H30") == pytest.approx(100 * 800_000.0 / net_worth)
    assert _f(fx, "H13") == pytest.approx(100 * 200_000.0 / net_worth)

    # Guide ¶7 / dormancy judgement rows stay blank
    for row in JUDGEMENT_ROWS:
        assert fx.get(f"E{row}") is None and fx.get(f"H{row}") is None
        assert _f(fx, f"G{row}") == pytest.approx(net_worth)
    counts = payload["status_counts"]
    assert counts == {"mapped": 67, "input_required": 785, "unmapped": 0, "derived": 0}
    section = snapshot["sections"][0]
    statuses = {row["cell"]: row["status"] for row in section["rows"]}
    assert statuses["E39"] == "mapped" and statuses["E99"] == "input_required"
    assert statuses["D14"] == "input_required"
    scaled = {row["cell"]: row["value"] for row in section["rows"]}
    assert float(scaled["E27"]) == pytest.approx(1.5)  # sections carry the sheet unit
    assert float(scaled["H27"]) == pytest.approx(100 * 1_500_000.0 / net_worth)  # unscaled


# ---------------------------------------------------------------------------
# 4. export
# ---------------------------------------------------------------------------


def test_bsd2a_export_writes_the_blank_grid_values(db_client: TestClient, seeded_book: str) -> None:
    snapshot = _generate(db_client, "BSD2A", seeded_book)
    net_worth_pct = float(
        next(r["value"] for r in snapshot["sections"][0]["rows"] if r["cell"] == "H27")
    )
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx("BSD2A", snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    ws = wb[SHEET]
    assert ws["A27"].value == "(c) Loans, Overdrafts and Other Advances"
    assert ws["E27"].value == pytest.approx(1.5)  # ¢'Million
    assert ws["E13"].value == pytest.approx(0.2)
    assert ws["E30"].value == pytest.approx(0.8)
    assert ws["H27"].value == pytest.approx(net_worth_pct)  # percentage, unscaled
    assert ws["E99"].value is None and ws["D14"].value is None
    notes = wb["Completion notes"]
    rows = [[c.value for c in row] for row in notes.iter_rows(min_row=1, max_row=notes.max_row)]
    listed = {str(r[1]) for r in rows if r and r[0] == SHEET and r[4] == "input_required"}
    assert {"E99", "E103", "E71", "N13", "A14", "N105"} <= listed
    assert "E27" not in listed and "G13" not in listed


# ---------------------------------------------------------------------------
# 5. resolvers
# ---------------------------------------------------------------------------


def _rc(dependencies: dict[str, dict[tuple[str, str], Any]]) -> ResolveContext:
    none = cast(Any, None)
    return ResolveContext(
        db=none,
        ctx=none,
        bank=none,
        period=none,
        column="cedi_equivalent",
        dependencies=dependencies,
    )


def test_bsd2a_resolvers_sum_and_ratio_over_dependency_cells() -> None:
    total = get_resolver("bsd2a.form_cells_sum")
    ratio = get_resolver("bsd2a.form_cells_ratio_pct")
    dep = {"BSD2": {("BSD2", "C34"): 2.5, ("BSD2", "C72"): Decimal("1.5"), ("BSD2", "D135"): 8.0}}
    assert total(
        _rc(dep), {"form": "BSD2", "sheet": "BSD2", "refs": ["C34", "C72", "C102"]}
    ) == Decimal("4.0")
    assert total(_rc(dep), {"form": "BSD2", "sheet": "BSD2", "refs": ["C102"]}) is None
    assert total(_rc({}), {"form": "BSD2", "sheet": "BSD2", "refs": ["C34"]}) is None
    pct = ratio(
        _rc(dep),
        {"form": "BSD2", "sheet": "BSD2", "numerator": ["C34", "C72"], "denominator": ["D135"]},
    )
    assert pct == Decimal("50")
    zero = {"BSD2": {("BSD2", "C34"): 1.0, ("BSD2", "D135"): 0}}
    assert (
        ratio(
            _rc(zero),
            {"form": "BSD2", "sheet": "BSD2", "numerator": ["C34"], "denominator": ["D135"]},
        )
        is None
    )
    assert (
        ratio(
            _rc(dep),
            {"form": "BSD2", "sheet": "BSD2", "numerator": ["C102"], "denominator": ["D135"]},
        )
        is None
    )
