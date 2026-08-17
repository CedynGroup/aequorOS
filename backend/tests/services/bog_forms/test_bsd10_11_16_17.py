"""BSD10 / BSD11 / BSD16 / BSD17 — Wave 4 statutory, capex, ATM and remittance forms.

Generates all four forms through the REAL package pipeline on the hermetic
book and proves:

1. every official data cell of every sheet is declared (bound) — the captured
   input cells AND the blank-grid data cells named from the header rows —
   nothing unmapped, no engine errors; the mapped / input_required split per
   sheet is exactly what each line-map doc claims;
2. BoG's own arithmetic over blank inputs: BSD10 row totals ``H = SUM(C:G)``
   and the 0–3/3–6-month totals, BSD16 ``F = D + E`` and ``F57 = SUM(F7:F56)``
   all evaluate to 0 when the inputs are blank (input_required → 0, never an
   invented figure); the template ordinals / item numbers are kept;
3. BSD11: directors, an officer and their facilities inserted through the
   platform's OWN related-party register API + canonical positions land in the
   right cells (Sheet-2 particulars in appointment order, Sheet-4 present
   balance / rate / type, Sheet-1 current balances, Sheet-6 ranking) and every
   percentage-to-net-worth cell equals exposure ÷ BSD2 Shareholders' Funds
   (D135) × 100 — the sheet header's relationship over BoG's BSD2 figure;
4. the values-only xlsx export writes the blank-grid cells (text + ¢'Million
   amounts), keeps unscaled percentages/ordinals and never emits a formula;
5. the ``bsd11.register`` resolver unit-tested against the inserted rows.
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.db.session import get_sessionmaker
from app.models import (
    Bank,
    BankReportingPeriod,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    CanonicalProduct,
    IngestionBatch,
    LineageRecord,
)
from app.services.regulatory_reporting.bog_forms.catalog import form_spec
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.sources import ResolveContext, get_resolver
from app.services.regulatory_reporting.bog_forms.sources_ext.bsd11 import (
    directors,
    largest_exposures,
    officers,
)
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import ORG_1, USER_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BASE = f"/api/v1/banks/{SAMPLE_BANK_ID}"
M = Decimal("1000000")

S1, S2, S3, S4 = "BSD11-Sheet-1", "BSD11-Sheet-2", "BSD11-Sheet-3", "BSD11-Sheet-4"
S5, S6, S7, S8 = "BSD11-Sheet-5", "BSD11-Sheet-6", "BSD11-Sheet-7", "BSD11- Sheet 8"
ATM = "MONTHLY ATM OPERATIONS"
R1, R2 = "BSG17-SHEET 1", "BSD17 -SHEET 2"

#: (mapped-source cells, input_required cells) per sheet — what the docs claim.
EXPECTED_SPLIT: dict[str, dict[str, tuple[int, int]]] = {
    "BSD10": {"BSD10": (50, 0)},  # capital_expenditure register (data-gap closure 2026-08-16)
    "BSD11": {
        S1: (10, 10),
        S2: (30, 0),
        S3: (0, 38),
        S4: (54, 0),
        S5: (0, 70),
        S6: (80, 0),
        S7: (0, 72),
        S8: (10, 80),
    },
    # Data-gap closure (2026-08-16): every BSD16 / BSD17 data cell is now bound to a
    # reference dataset (``atm_operations`` / ``remittance_flows``, refs.* resolvers)
    # — declared-with-source counts; the hermetic book carries neither register, so at
    # GENERATION those cells still resolve to input_required (GENERATED_WITHOUT_REGISTERS).
    "BSD16": {ATM: (252, 0), "Sheet2": (0, 0), "Sheet3": (0, 0)},
    "BSD17": {R1: (14, 0), R2: (7, 0)},
}
#: (mapped, input_required) at generation on the hermetic book — no gap register ingested.
GENERATED_WITHOUT_REGISTERS: dict[str, tuple[int, int]] = {
    "BSD10": (0, 50),
    "BSD16": (50, 202),
    "BSD17": (7, 14),
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _session() -> Session:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    return session


def _period_end(db_client: TestClient) -> str:
    periods = db_client.get(f"{BASE}/reporting-periods", headers=headers()).json()["periods"]
    return periods[0]["period_end"]


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"{BASE}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(f"{BASE}/regulatory-packages/{package['id']}", headers=headers()).json()
    return detail["snapshot"]


def _declared(code: str) -> dict[str, dict[str, tuple[bool, bool]]]:
    """sheet → cell → (has source, unscaled)."""
    out: dict[str, dict[str, tuple[bool, bool]]] = {}
    for sheet, lines in line_maps_for(code).items():
        for line in lines:
            for ref in line.cells.values():
                out.setdefault(sheet, {})[ref] = (line.source is not None, line.unscaled)
    return out


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


@pytest.fixture
def hermetic_book(db_client: TestClient) -> str:
    session = _session()
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    return _period_end(db_client)


# ---------------------------------------------------------------------------
# 1. every official data cell is declared; the split matches the docs
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("code", sorted(EXPECTED_SPLIT))
def test_every_data_cell_is_declared_and_split_matches_doc(code: str) -> None:
    layout = load_layout(code)
    declared = _declared(code)
    for sheet in layout.sheets:
        cells = declared.get(sheet.name, {})
        # every CAPTURED input cell (ordinals, BSD10 zeros) is bound …
        captured = {c.ref for c in sheet.input_cells}
        assert captured <= set(cells), (
            f"{code}/{sheet.name}: unbound {sorted(captured - set(cells))}"
        )
        # … and no bound cell is a label / formula of the official template
        for ref in cells:
            cell = sheet.by_ref.get(ref)
            assert cell is None or cell.kind == "input", f"{code}/{sheet.name}!{ref} {cell}"
        sourced = sum(1 for has_source, _ in cells.values() if has_source)
        required = sum(1 for has_source, _ in cells.values() if not has_source)
        assert (sourced, required) == EXPECTED_SPLIT[code][sheet.name], (
            f"{code}/{sheet.name}: sourced={sourced} input_required={required}"
        )
    # totals the docs quote
    total_declared = sum(len(v) for v in declared.values())
    assert total_declared == {"BSD10": 50, "BSD11": 454, "BSD16": 252, "BSD17": 21}[code]


def test_only_input_required_rows_carry_a_register_naming_note() -> None:
    for code in EXPECTED_SPLIT:
        for lines in line_maps_for(code).values():
            for line in lines:
                if line.source is None:
                    assert len(line.notes) > 20, f"{code} {line.code}: note names no dataset"
                    assert (
                        "requir" in line.notes or "supplies" in line.notes or "bank" in line.notes
                    )


# ---------------------------------------------------------------------------
# 2. structure-only forms: BoG's totals over blank inputs, ordinals kept
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("code", ["BSD10", "BSD16", "BSD17"])
def test_structure_only_forms_generate_with_every_cell_accounted(
    db_client: TestClient, hermetic_book: str, code: str
) -> None:
    snapshot = _generate(db_client, code, hermetic_book)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    counts = payload["status_counts"]
    expected_sourced, expected_required = GENERATED_WITHOUT_REGISTERS[code]
    assert counts["mapped"] == expected_sourced
    assert counts["input_required"] == expected_required


def test_bsd10_totals_are_bogs_sums_over_blank_inputs(
    db_client: TestClient, hermetic_book: str
) -> None:
    cells = _generate(db_client, "BSD10", hermetic_book)["bog_form"]["cells"]["BSD10"]
    for row in (*range(7, 15), 16, 17):
        for col in "CDEFG":
            assert cells[f"{col}{row}"] is None  # input_required → blank
        assert cells[f"H{row}"] == 0  # =SUM(C:G) over blanks
    for col in "CDEFGH":
        assert cells[f"{col}18"] == 0  # =C16+C17 …
    # the sheet's leaf rows are exactly the ten Guide items × five asset classes
    lines = line_maps_for("BSD10")["BSD10"]
    assert [line.label for line in lines][:3] == [
        "A. Purchased",
        "B. On finance-lease",
        "C. On Hire-Purchase",
    ]
    assert lines[-1].label == "3 - 6 months"
    assert all(
        set(line.cells.values()) <= {f"{c}{r}" for c in "CDEFG" for r in range(7, 18)}
        for line in lines
    )


def test_bsd16_row_and_grand_totals_and_ordinals(db_client: TestClient, hermetic_book: str) -> None:
    snapshot = _generate(db_client, "BSD16", hermetic_book)
    cells = snapshot["bog_form"]["cells"][ATM]
    for row in range(7, 57):
        assert cells[f"A{row}"] == row - 6  # template ordinal kept
        assert cells[f"B{row}"] is None and cells[f"C{row}"] is None
        assert cells[f"D{row}"] is None and cells[f"E{row}"] is None
        assert cells[f"F{row}"] == 0  # =D+E over blanks
    assert cells["F57"] == 0  # =SUM(F7:F56)
    assert cells["D57"] is None and cells["E57"] is None
    # export: ordinal is unscaled (1, not 1e-6 on the ¢'Million sheet), no formulas
    ws = _export("BSD16", snapshot)[ATM]
    assert ws["A7"].value == 1 and ws["A56"].value == 50
    assert ws["F57"].value == 0
    assert ws["B6"].value == "Station / Branch"
    assert not any(
        isinstance(c.value, str) and c.value.startswith("=") for row in ws.iter_rows() for c in row
    )
    # the two empty placeholder tabs ride along, empty
    wb = _export("BSD16", snapshot)
    assert wb.sheetnames[:3] == [ATM, "Sheet2", "Sheet3"]


def test_bsd17_item_numbers_kept_and_amounts_input_required(
    db_client: TestClient, hermetic_book: str
) -> None:
    snapshot = _generate(db_client, "BSD17", hermetic_book)
    cells = snapshot["bog_form"]["cells"]
    for row in range(8, 14):
        assert cells[R1][f"A{row}"] == row - 7
        assert cells[R1][f"C{row}"] is None
    assert cells[R1]["A15"] == 7 and cells[R1]["C15"] is None
    for row in range(6, 13):
        assert cells[R2][f"B{row}"] is None
    notes = {
        row["notes"]
        for section in snapshot["sections"]
        for row in section["rows"]
        if row["status"] == "input_required"
    }
    # data-gap closure (2026-08-16): the cells now name the ``remittance_flows`` register
    assert any("remittance_flows register required" in note for note in notes)
    assert any("docs/data_engine/datasets/remittance_flows.md" in note for note in notes)


# ---------------------------------------------------------------------------
# 3. BSD11 — registers + facilities through the platform's own services
# ---------------------------------------------------------------------------


def _party(db_client: TestClient, **payload: Any) -> dict[str, Any]:
    body = {"reason": "BSD11 register fixture", "party_type": "individual", **payload}
    response = db_client.post(f"{BASE}/related-parties", headers=headers(), json=body)
    assert response.status_code == 201, response.text
    return response.json()


def _shareholding(db_client: TestClient, party_id: str, shares: str, pct: str) -> None:
    response = db_client.post(
        f"{BASE}/related-parties/{party_id}/shareholdings",
        headers=headers(),
        json={
            "reason": "BSD11 register fixture",
            "share_type": "ordinary",
            "shareholder_rights": "voting",
            "number_of_shares": shares,
            "pct_shareholding": pct,
        },
    )
    assert response.status_code == 201, response.text


class _Seeder:
    def __init__(self, db: Session, as_of: date) -> None:
        self.db = db
        batch = IngestionBatch(
            organization_id=ORG_1,
            bank_id=SAMPLE_BANK_ID,
            source_system="EXCEL_CSV",
            adapter_version="1.0",
            extraction_mode="full",
            status="accepted",
            as_of_date=as_of,
        )
        db.add(batch)
        db.flush()
        lineage = LineageRecord(
            organization_id=ORG_1,
            ingestion_batch_id=batch.id,
            operation_type="ADAPTER_TRANSLATE",
            operation_ref="bsd11-fixture",
            input_lineage_ids=[],
        )
        db.add(lineage)
        db.flush()
        self.common: dict[str, Any] = {
            "organization_id": ORG_1,
            "bank_id": SAMPLE_BANK_ID,
            "as_of_date": as_of,
            "source_system": "EXCEL_CSV",
            "ingestion_batch_id": batch.id,
            "lineage_id": lineage.id,
            "validation_status": "accepted",
        }

    def counterparty(self, ref: str, name: str, kind: str) -> CanonicalCounterparty:
        row = CanonicalCounterparty(
            **self.common, source_reference=ref, name=name, counterparty_type=kind
        )
        self.db.add(row)
        self.db.flush()
        return row

    def product(self, code: str, name: str) -> CanonicalProduct:
        row = CanonicalProduct(
            **self.common, source_reference=f"PRODUCT/{code}", product_code=code, name=name
        )
        self.db.add(row)
        self.db.flush()
        return row

    def position(  # noqa: PLR0913 — keyword-only fixture builder
        self,
        ref: str,
        position_type: str,
        amount: Decimal,
        *,
        counterparty: CanonicalCounterparty | None,
        product: CanonicalProduct | None = None,
        currency: str = "GHS",
        rate: str | None = None,
        attributes: dict[str, Any] | None = None,
        ghs: bool = True,
        notional: bool = False,
    ) -> None:
        position = CanonicalPosition(
            **self.common, source_reference=ref, position_type=position_type, currency=currency
        )
        self.db.add(position)
        self.db.flush()
        attrs: dict[str, Any] = dict(attributes or {})
        if ghs:
            attrs["notional_ghs" if notional else "balance_ghs"] = str(amount)
        self.db.add(
            CanonicalPositionSnapshot(
                **self.common,
                source_reference=ref,
                position_id=position.id,
                counterparty_id=counterparty.id if counterparty is not None else None,
                product_id=product.id if product is not None else None,
                balance=Decimal("0") if notional else amount,
                notional=amount if notional else None,
                interest_rate=Decimal(rate) if rate else None,
                attributes=attrs,
            )
        )
        self.db.flush()


# hand-checked amounts (cedis)
KOFI_LOAN = Decimal("2500000")  # director 1: one housing loan @ 24%
ABENA_LOAN_A = Decimal("1500000")  # director 2: two facilities
ABENA_LOAN_B = Decimal("500000")
YAW_LOAN = Decimal("800000")  # officer (CFO)
VOLTA_LOAN = 30 * M  # largest customer
VOLTA_LC = 5 * M  # off-balance guarantee (notional)
VOLTA_COLLATERAL = 20 * M  # crm_collateral_ghs on the loan


@pytest.fixture
def statutory_book(db_client: TestClient, hermetic_book: str) -> str:
    """Hermetic book + three directors, one officer (register API) + facilities."""
    reporting_date = hermetic_book
    kofi = _party(
        db_client,
        full_name="Kofi Owusu",
        contact={"address": "P.O. Box 12, Accra"},
        roles=[{"role": "director", "appointed_on": "2019-03-01"}, {"role": "shareholder"}],
    )
    _shareholding(db_client, kofi["id"], "1000000", "2.5")
    _party(
        db_client,
        full_name="Abena Sarpong",
        roles=[{"role": "board_chairman", "appointed_on": "2021-07-15"}],
    )
    _party(
        db_client,
        full_name="Efua Mensah",
        roles=[{"role": "director", "appointed_on": "2024-01-10"}, {"role": "shareholder"}],
    )
    _party(
        db_client,
        full_name="Yaw Boateng",
        roles=[{"role": "chief_finance_officer", "appointed_on": "2020-05-01"}],
    )
    # an inactive former director never appears
    _party(
        db_client,
        full_name="Kwame Former",
        status="inactive",
        roles=[{"role": "director", "appointed_on": "2010-01-01"}],
    )
    session = _session()
    try:
        s = _Seeder(session, date.fromisoformat(reporting_date))
        housing = s.product("LN.STAFF.HOUSE", "Staff Housing Loan")
        vehicle = s.product("LN.STAFF.CAR", "Staff Vehicle Loan")
        personal = s.product("LN.PERS", "Personal Loan")
        corp = s.product("LN.CORP", "Corporate Term Loan")
        kofi_cp = s.counterparty("CP/KOFI", "  Kofi   OWUSU ", "RETAIL_INDIVIDUAL")
        abena_cp = s.counterparty("CP/ABENA", "Abena Sarpong", "RETAIL_INDIVIDUAL")
        yaw_cp = s.counterparty("CP/YAW", "Yaw Boateng", "RETAIL_INDIVIDUAL")
        volta = s.counterparty("CP/VOLTA", "Volta Agro Ltd", "CORPORATE")
        gog = s.counterparty("CP/GOG", "Government of Ghana", "SOVEREIGN")
        s.position(
            "LOAN/KOFI",
            "LOAN",
            KOFI_LOAN,
            counterparty=kofi_cp,
            product=housing,
            rate="0.24",
            attributes={"crm_collateral_class": "RESIDENTIAL_PROPERTY"},
        )
        s.position(
            "LOAN/ABENA/A",
            "LOAN",
            ABENA_LOAN_A,
            counterparty=abena_cp,
            product=vehicle,
            rate="0.20",
        )
        s.position(
            "LOAN/ABENA/B",
            "LOAN",
            ABENA_LOAN_B,
            counterparty=abena_cp,
            product=personal,
            rate="0.30",
        )
        s.position("LOAN/YAW", "LOAN", YAW_LOAN, counterparty=yaw_cp, product=personal, rate="0.18")
        s.position(
            "LOAN/VOLTA",
            "LOAN",
            VOLTA_LOAN,
            counterparty=volta,
            product=corp,
            attributes={
                "crm_collateral_ghs": str(VOLTA_COLLATERAL),
                "crm_collateral_class": "CASH",
            },
        )
        s.position("LC/VOLTA", "LC_GUARANTEE", VOLTA_LC, counterparty=volta, notional=True)
        # a USD loan WITHOUT an ingested cedi conversion contributes zero
        s.position("LOAN/VOLTA/USD", "LOAN", 9 * M, counterparty=volta, currency="USD", ghs=False)
        # sovereign holdings are not "customers" of the Section 47 list
        s.position("SEC/GOG", "SECURITY_HOLDING", 100 * M, counterparty=gog)
        # a deposit is not an exposure
        s.position("DEP/KOFI", "DEPOSIT", 40 * M, counterparty=kofi_cp)
        session.commit()
    finally:
        session.close()
    return reporting_date


def test_bsd11_registers_land_in_the_official_cells(  # noqa: PLR0912, PLR0915 — one linear proof over eight sheets
    db_client: TestClient, statutory_book: str
) -> None:
    snapshot = _generate(db_client, "BSD11", statutory_book)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    assert payload["missing_dependencies"] == []
    cells = payload["cells"]

    # --- Sheet-2: directors in appointment order; officer and inactive party absent
    assert cells[S2]["B8"] == "Kofi Owusu, P.O. Box 12, Accra"
    assert cells[S2]["C8"] == "2019-03-01"
    assert cells[S2]["E8"] == "1,000,000 (2.50%)"
    assert cells[S2]["B9"] == "Abena Sarpong" and cells[S2]["C9"] == "2021-07-15"
    assert cells[S2]["E9"] == "Nil"  # register: no shareholder role
    assert cells[S2]["B10"] == "Efua Mensah" and cells[S2]["C10"] == "2024-01-10"
    assert cells[S2]["E10"] is None  # shareholder role, no holding rows → bank completes
    for row in (11, 12, 13):
        assert cells[S2][f"B{row}"] is None  # only three directors registered
    for row in range(8, 14):
        assert cells[S2][f"D{row}"] is None and cells[S2][f"F{row}"] is None  # not in register
    names = " ".join(str(v) for v in cells[S2].values() if v)
    assert "Yaw Boateng" not in names and "Kwame Former" not in names

    # --- Sheet-4: facilities per director (same order); name-match is normalised
    assert cells[S4]["B12"] == "Kofi Owusu, P.O. Box 12, Accra"
    assert Decimal(str(cells[S4]["J12"])) == KOFI_LOAN
    assert Decimal(str(cells[S4]["G12"])) == Decimal("24")  # one facility → its rate, in %
    assert cells[S4]["H12"] == "Staff Housing Loan"
    assert cells[S4]["F12"] == "RESIDENTIAL_PROPERTY"
    assert Decimal(str(cells[S4]["J13"])) == ABENA_LOAN_A + ABENA_LOAN_B
    assert cells[S4]["G13"] is None  # two facilities, two rates → bank states them
    assert cells[S4]["H13"] == "Personal Loan, Staff Vehicle Loan"
    assert cells[S4]["J14"] is None  # Efua has no facility
    for row in range(12, 18):
        for col in "CDEI":  # secured/unsecured/guaranteed split, board approval date
            assert cells[S4][f"{col}{row}"] is None

    # --- Sheet-1: current balances Σ over the group; movements input_required
    assert Decimal(str(cells[S1]["F13"])) == KOFI_LOAN + ABENA_LOAN_A + ABENA_LOAN_B
    assert Decimal(str(cells[S1]["F14"])) == YAW_LOAN
    assert cells[S1]["F12"] is None and cells[S1]["F15"] is None
    for row in range(12, 16):
        for col in "BCDE":
            assert cells[S1][f"{col}{row}"] is None

    # --- Sheet-6: Section 47 ranking + percentage of net worth (BSD2 line 16)
    bsd2 = _generate(db_client, "BSD2", statutory_book)["bog_form"]["cells"]["BSD2"]
    net_worth = Decimal(str(bsd2["D135"]))
    assert net_worth > 0
    assert cells[S6]["B8"] == "Volta Agro Ltd"
    assert Decimal(str(cells[S6]["C8"])) == VOLTA_LOAN  # USD leg without conversion → 0
    assert Decimal(str(cells[S6]["D8"])) == VOLTA_LC
    assert Decimal(str(cells[S6]["E8"])) == VOLTA_LOAN + VOLTA_LC
    assert Decimal(str(cells[S6]["F8"])) == VOLTA_COLLATERAL
    assert cells[S6]["B9"] == "Kofi OWUSU"  # the counterparty as ingested (whitespace tidied)
    assert cells[S6]["B10"] == "Abena Sarpong"
    assert cells[S6]["B11"] == "Yaw Boateng"
    assert cells[S6]["B12"] is None  # sovereign excluded, deposits are not exposures
    for row in range(8, 12):
        total = Decimal(str(cells[S6][f"E{row}"]))
        assert total == Decimal(str(cells[S6][f"C{row}"])) + Decimal(str(cells[S6][f"D{row}"]))
        pct_total = Decimal(str(cells[S6][f"I{row}"]))
        assert abs(pct_total - total / net_worth * 100) < Decimal("1e-9"), row
    secured = min(VOLTA_LOAN + VOLTA_LC, VOLTA_COLLATERAL)
    assert abs(Decimal(str(cells[S6]["G8"])) - secured / net_worth * 100) < Decimal("1e-9")
    assert abs(
        Decimal(str(cells[S6]["H8"])) - (VOLTA_LOAN + VOLTA_LC - secured) / net_worth * 100
    ) < Decimal("1e-9")
    assert cells[S6]["F9"] is None and cells[S6]["G9"] is None  # no security recorded
    assert cells[S6]["I9"] is not None
    ranked = [Decimal(str(cells[S6][f"E{r}"])) for r in range(8, 12)]
    assert ranked == sorted(ranked, reverse=True)

    # --- Sheets 3, 5, 7, 8: register absent → every data cell blank + input_required
    for sheet, rows, cols in (
        (S3, range(10, 15), "BCDEFGH"),
        (S5, (11, 12, 13, 17, 18, 19, 22), "CDEFGHIJKL"),
        (S7, range(9, 17), "BCDEFGHIJ"),
        (S8, range(8, 18), "BCDEFGHI"),
    ):
        for row in rows:
            for col in cols:
                assert cells[sheet][f"{col}{row}"] is None, f"{sheet}!{col}{row}"
    for row in range(8, 18):
        assert cells[S8][f"A{row}"] == row - 7  # template ordinals kept
    statuses = {
        (section["title"], row["cell"]): (row["status"], row["notes"])
        for section in snapshot["sections"]
        for row in section["rows"]
    }
    assert statuses[(S3, "B10")][0] == "input_required"
    assert "director-interest" in statuses[(S3, "B10")][1]
    assert "staff-loan register" in statuses[(S5, "C11")][1]
    assert "relatives register" in statuses[(S8, "B8")][1]
    assert statuses[(S2, "B8")][0] == "mapped"
    assert statuses[(S2, "B11")][0] == "input_required"  # rank beyond the register

    # --- values-only export: blank-grid text + scaled amounts + unscaled % / ordinals
    wb = _export("BSD11", snapshot)
    assert wb.sheetnames[:8] == list(load_layout("BSD11").sheet_names)
    assert wb[S2]["B8"].value == "Kofi Owusu, P.O. Box 12, Accra"
    assert wb[S2]["C8"].value == "2019-03-01"
    assert wb[S4]["J12"].value == pytest.approx(float(KOFI_LOAN / M))  # ¢'Million
    assert wb[S4]["G12"].value == pytest.approx(24.0)  # percent, unscaled
    assert wb[S6]["I8"].value == pytest.approx(float((VOLTA_LOAN + VOLTA_LC) / net_worth * 100))
    assert wb[S8]["A8"].value == 1 and wb[S8]["A17"].value == 10
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                assert not (isinstance(c.value, str) and c.value.startswith("=")), c.coordinate
    assert wb.sheetnames[-1] == "Completion notes"


def test_bsd11_register_resolver_against_inserted_rows(
    db_client: TestClient, statutory_book: str
) -> None:
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        period = (
            session.query(BankReportingPeriod)
            .filter(
                BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
                BankReportingPeriod.period_end == date.fromisoformat(statutory_book),
            )
            .one()
        )
        ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1, roles=("admin",))
        cache: dict[str, Any] = {}

        def rc(
            column: str, deps: dict[str, dict[tuple[str, str], Any]] | None = None
        ) -> ResolveContext:
            return ResolveContext(
                db=session,
                ctx=ctx,
                bank=bank,
                period=period,
                column=column,
                dependencies=deps or {},
                cache=cache,
            )

        resolve = get_resolver("bsd11.register")
        assert [p.name for p in directors(rc("name"))] == [
            "Kofi Owusu",
            "Abena Sarpong",
            "Efua Mensah",
        ]
        assert [p.name for p in officers(rc("name"))] == ["Yaw Boateng"]
        assert resolve(rc("name"), {"register": "directors", "rank": 2}) == "Abena Sarpong"
        assert resolve(rc("balance"), {"register": "directors", "rank": 1}) == KOFI_LOAN
        assert resolve(rc("rate"), {"register": "directors", "rank": 1}) == Decimal("24")
        assert resolve(rc("name"), {"register": "directors", "rank": 4}) is None
        assert resolve(rc("current"), {"register": "summary", "group": "officers"}) == YAW_LOAN
        assert resolve(rc("previous"), {"register": "summary", "group": "officers"}) is None
        ranked = largest_exposures(rc("name"))
        assert [e.name for e in ranked] == [
            "Volta Agro Ltd",
            "Kofi OWUSU",
            "Abena Sarpong",
            "Yaw Boateng",
        ]
        assert ranked[0].total == VOLTA_LOAN + VOLTA_LC and ranked[0].security == VOLTA_COLLATERAL
        # without the BSD2 dependency the percentage cells stay blank …
        assert resolve(rc("pct_total"), {"register": "large_exposures", "rank": 1}) is None
        # … with it they are exposure ÷ Shareholders' Funds × 100
        deps = {"BSD2": {("BSD2", "D135"): 350 * M}}
        pct = resolve(rc("pct_total", deps), {"register": "large_exposures", "rank": 1})
        assert pct == (VOLTA_LOAN + VOLTA_LC) / (350 * M) * 100
        with pytest.raises(ValueError, match="unknown register"):
            resolve(rc("name"), {"register": "nope"})
    finally:
        session.close()


def test_bsd11_depends_on_bsd2_for_net_worth() -> None:
    spec = form_spec("BSD11")
    assert "BSD2" in spec.depends_on
    # Sheet-6 mixes ¢'Million amounts (C–F) with % (G–I): the sheet unit is
    # millions and the % lines carry unscaled=True (framework ask applied).
    assert spec.sheet(S6) is not None and spec.sheet(S6).unit == "millions"  # type: ignore[union-attr]
