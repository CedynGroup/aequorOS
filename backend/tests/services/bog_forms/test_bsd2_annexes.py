"""BSD2 annex sheets (Wave 1b) — line map + tie-outs to the BSD2 spine.

The 21 annex sheets of ``FORM BSD2 REVISED.xls`` are mostly BLANK detail
schedules (name / amount rows with no ``0`` placeholder), so the committed
layout carries input cells only for Annex 2a (5) and Annex 4 (25); every other
grid is declared explicitly by ``linemaps/bsd2.py`` via ``grid_lines``.

Proves, on the hermetic book plus a small canonical slice:

1. every annex sheet with data cells is in the line map, every CAPTURED input
   cell is declared, every declared cell lies inside the official grid and is
   never a template formula, line codes are unique, and the mapped /
   input_required split per sheet is exactly what ``bsd2_line_map.md`` claims;
2. the form generates through the real package pipeline with no engine errors;
3. critical totals — annex schedule totals equal the BSD2 spine line they
   analyse, by BoG's own arithmetic over our inputs:
   - Annex 16 ``I11`` (Σ contingent liabilities) = BSD2 ``D282`` (line 33,
     domestic + foreign) and each ``I`` row = E + F + G + H;
   - Annex 4 ``G13`` (total advances) = BSD2 ``D68`` (line 8 sub-total) and
     ``B11`` (staff advances) is a subset of ``B10``;
   - Annex 6 total row = BSD2 line 11 other assets, domestic AND foreign;
   - Annex 15 total row = BSD2 line 29 other liabilities, domestic AND foreign;
   - Annex 7 total = BSD2 15(e) other reserves; Annex 1 ``E4`` = BSD2 ``C7``;
4. the values-only xlsx export writes the blank-grid values at their official
   cells (¢'Million-scaled) and lists the detail rows in the Completion notes.
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any, cast

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.db.session import get_sessionmaker
from app.models import (
    Bank,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    IngestionBatch,
    LineageRecord,
)
from app.services.regulatory_reporting.bog_forms.catalog import form_spec
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.linemaps._common import BANK_COA_MAPPING
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

M = Decimal("1000000")
A1, A2A, A4, A6, A7, A15, A16 = (
    "BSD2-Annex 1",
    "BSD2-Annex 2a",
    "BSD2-Annex 4 ",
    "BSD2-Annex 6",
    "BSD2-Annex7",
    "BSD2-Annex 15",
    "BSD2-Annex 16",
)
ANNEX_SHEETS = tuple(
    name for name in load_layout("BSD2").sheet_names if name.startswith("BSD2-Annex")
)

#: (mapped cells, input_required cells) per annex — the doc's claim.
EXPECTED_SPLIT: dict[str, tuple[int, int]] = {
    "BSD2-Annex 1": (4, 41),
    "BSD2-Annex 2a": (0, 30),
    "BSD2-Annex 2b": (0, 43),
    "BSD2-Annex 2c": (0, 43),
    "BSD2-Annex 2d": (0, 43),
    "BSD2-Annex 3 ": (0, 88),
    "BSD2-Annex 4 ": (15, 10),
    "BSD2-Annex 5 ": (0, 88),
    "BSD2-Annex 6": (2, 88),
    "BSD2-Annex7": (1, 44),
    "BSD2-Annex 8": (0, 177),
    "BSD2-Annex 9": (0, 177),
    "BSD2-Annex 10": (0, 133),
    "BSD2-Annex 11": (0, 133),
    "BSD2-Annex 12": (0, 47),
    "BSD2-Annex 13": (0, 38),
    "BSD2-Annex 14": (0, 90),
    "BSD2-Annex 15": (2, 90),
    "BSD2-Annex 16": (20, 10),
    "BSD2-Annex 17": (0, 8),
}

# ---------------------------------------------------------------------------
# canonical slice — the attribute conventions the annex map declares
# ---------------------------------------------------------------------------


class _Seeder:
    def __init__(self, db: Session, as_of: date) -> None:
        self.db = db
        batch = IngestionBatch(
            organization_id=ORG_1,
            bank_id=SAMPLE_BANK_ID,
            source_system="EXCEL_CSV",
            adapter_version="1.0",
            extraction_mode="full",
            status="accepted",
            as_of_date=as_of,
        )
        db.add(batch)
        db.flush()
        lineage = LineageRecord(
            organization_id=ORG_1,
            ingestion_batch_id=batch.id,
            operation_type="ADAPTER_TRANSLATE",
            operation_ref="bsd2-annex-fixture",
            input_lineage_ids=[],
        )
        db.add(lineage)
        db.flush()
        self.common: dict[str, Any] = {
            "organization_id": ORG_1,
            "bank_id": SAMPLE_BANK_ID,
            "as_of_date": as_of,
            "source_system": "EXCEL_CSV",
            "ingestion_batch_id": batch.id,
            "lineage_id": lineage.id,
            "validation_status": "accepted",
        }

    def counterparty(
        self, ref: str, name: str, counterparty_type: str, *, resident: bool = True
    ) -> CanonicalCounterparty:
        row = CanonicalCounterparty(
            **self.common,
            source_reference=ref,
            name=name,
            counterparty_type=counterparty_type,
            resident=resident,
        )
        self.db.add(row)
        self.db.flush()
        return row

    def position(  # noqa: PLR0913 — keyword-only fixture builder
        self,
        ref: str,
        position_type: str,
        amount: Decimal,
        *,
        counterparty: CanonicalCounterparty | None = None,
        currency: str = "GHS",
        notional: bool = False,
        attributes: dict[str, Any] | None = None,
    ) -> None:
        position = CanonicalPosition(
            **self.common, source_reference=ref, position_type=position_type, currency=currency
        )
        self.db.add(position)
        self.db.flush()
        attrs: dict[str, Any] = {"balance_ghs": str(amount), **(attributes or {})}
        self.db.add(
            CanonicalPositionSnapshot(
                **self.common,
                source_reference=ref,
                position_id=position.id,
                counterparty_id=counterparty.id if counterparty is not None else None,
                balance=amount,
                notional=amount if notional else None,
                attributes=attrs,
            )
        )
        self.db.flush()


def _seed_slice(db: Session, as_of: date) -> None:
    s = _Seeder(db, as_of)
    gog = s.counterparty("CP/GOG", "Government of Ghana", "SOVEREIGN")
    kumasi = s.counterparty("CP/KUM", "Kumasi Traders Ltd", "CORPORATE")
    ama = s.counterparty("CP/AMA", "Ama Mensah (staff)", "RETAIL_INDIVIDUAL")
    # Annex 4 — loans by borrower class × facility_type (BSD2 line 8 rows 61/66/67)
    s.position(
        "LN/GOG", "LOAN", 10 * M, counterparty=gog, attributes={"facility_type": "scheduled"}
    )
    s.position(
        "LN/KUM/OD", "LOAN", 6 * M, counterparty=kumasi, attributes={"facility_type": "overdraft"}
    )
    s.position(
        "LN/KUM/USD",
        "LOAN",
        Decimal("1500000"),
        counterparty=kumasi,
        currency="USD",
        attributes={"facility_type": "unscheduled"},
    )
    s.position(
        "LN/AMA",
        "LOAN",
        Decimal("300000"),
        counterparty=ama,
        attributes={"facility_type": "scheduled", "scheme": "staff_advance"},
    )
    # Annex 16 — contingent liabilities by obs_category × obs_status × currency (BSD2 line 33)
    s.position(
        "LC/G1",
        "LC_GUARANTEE",
        4 * M,
        counterparty=kumasi,
        notional=True,
        attributes={"obs_category": "guarantee", "obs_status": "performing"},
    )
    s.position(
        "LC/G2",
        "LC_GUARANTEE",
        Decimal("500000"),
        counterparty=kumasi,
        notional=True,
        attributes={"obs_category": "guarantee", "obs_status": "non_performing"},
    )
    s.position(
        "LC/DC1",
        "LC_GUARANTEE",
        1 * M,
        counterparty=kumasi,
        currency="USD",
        notional=True,
        attributes={"obs_category": "letter_of_credit", "obs_status": "performing"},
    )
    # Annex 15 — other liabilities (BSD2 line 29), domestic + foreign
    s.position("OL/GHS", "OTHER_LIABILITY", 2 * M)
    s.position("OL/USD", "OTHER_LIABILITY", Decimal("700000"), currency="USD")
    # Annex 1 — foreign currency notes and coins (BSD2 A.1)
    s.position(
        "CASH/USD",
        "CASH",
        Decimal("200000"),
        currency="USD",
        attributes={"instrument": "fx_notes_coins"},
    )
    db.flush()


def _session() -> Session:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    return session


def _period_end(db_client: TestClient) -> str:
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    return periods[0]["period_end"]


@pytest.fixture
def seeded_book(db_client: TestClient) -> str:
    session = _session()
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    reporting_date = _period_end(db_client)
    session = _session()
    try:
        _seed_slice(session, date.fromisoformat(reporting_date))
        session.commit()
    finally:
        session.close()
    return reporting_date


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    return detail["snapshot"]


def _f(cells: dict[str, Any], ref: str) -> float:
    return float(cells.get(ref) or 0.0)


# ---------------------------------------------------------------------------
# 1. structure
# ---------------------------------------------------------------------------


def test_every_annex_grid_is_declared_inside_the_official_sheet() -> None:
    layout = load_layout("BSD2")
    maps = line_maps_for("BSD2")
    assert set(maps) == {"BSD2", *ANNEX_SHEETS}  # Summary is all formulas — nothing to bind
    for name in ANNEX_SHEETS:
        sheet = layout.sheet(name)
        lines = maps[name]
        declared = {ref for line in lines for ref in line.cells.values()}
        # every CAPTURED input cell (Annex 2a, Annex 4) is declared
        for cell in sheet.input_cells:
            assert cell.ref in declared, f"{name}!{cell.ref} captured but not declared"
        # every declared cell is inside the grid and never a template formula
        for ref in declared:
            cell = sheet.by_ref.get(ref)
            assert cell is None or cell.kind == "input", f"{name}!{ref} is a {cell and cell.kind}"
            col = "".join(ch for ch in ref if ch.isalpha())
            row = int("".join(ch for ch in ref if ch.isdigit()))
            assert row <= sheet.max_row, f"{name}!{ref} beyond max_row {sheet.max_row}"
            assert len(col) == 1 and ord(col) - ord("A") + 1 <= sheet.max_col, f"{name}!{ref}"
        codes = [line.code for line in lines]
        assert len(codes) == len(set(codes)), name
        # no cell declared twice across the per-column bindings
        refs = [ref for line in lines for ref in line.cells.values()]
        assert len(refs) == len(set(refs)), name


def test_annex_status_split_is_what_the_doc_claims() -> None:
    spec = form_spec("BSD2")
    split: dict[str, tuple[int, int]] = {}
    coa = 0
    for sheet in spec.sheets:
        if sheet.name not in ANNEX_SHEETS:
            continue
        mapped = sum(len(line.cells) for line in sheet.lines if line.source is not None)
        pending = sum(len(line.cells) for line in sheet.lines if line.source is None)
        coa += sum(
            len(line.cells)
            for line in sheet.lines
            if line.source is None and line.notes == BANK_COA_MAPPING.notes
        )
        split[sheet.name] = (mapped, pending)
    assert split == EXPECTED_SPLIT
    assert coa == 0  # no annex cell is a bare chart-of-accounts placeholder
    assert sum(m for m, _ in split.values()) == 44
    assert sum(p for _, p in split.values()) == 1421
    # every input_required detail row says what fills it
    for sheet in spec.sheets:
        for line in sheet.lines:
            if sheet.name in ANNEX_SHEETS and line.source is None:
                assert line.notes, f"{sheet.name} {line.code} has no note"


def test_annex_16_and_annex_4_grids_follow_the_template_columns() -> None:
    maps = line_maps_for("BSD2")
    a16 = {ref: line for line in maps[A16] for ref in line.cells.values()}
    assert set(a16) == {f"{c}{r}" for c in "CDEFGH" for r in range(6, 11)}
    # I6:I11 are the template's own row totals / grand total — never bound
    assert not any(ref.startswith("I") for ref in a16)
    fx = a16["E8"]
    assert fx.source == "positions.sum"
    assert fx.params["attribute_eq"] == {"obs_category": "guarantee", "obs_status": "performing"}
    assert fx.params["currency"] == "FX" and a16["G8"].params["currency"] == "GHS"
    h7 = cast(dict[str, Any], a16["H7"].params["attribute_eq"])
    assert h7["obs_status"] == "non_performing"
    a4 = {ref: line for line in maps[A4] for ref in line.cells.values()}
    assert set(a4) == {f"{c}{r}" for c in "BCDEF" for r in range(8, 13)}
    assert a4["D10"].params["attribute_eq"] == {"facility_type": "overdraft"}
    assert a4["B11"].params["attribute_eq"] == {
        "facility_type": "scheduled",
        "scheme": "staff_advance",
    }
    assert a4["B9"].source is None and a4["F12"].source is None  # provisions: sub-ledger


# ---------------------------------------------------------------------------
# 2–3. generation + critical totals
# ---------------------------------------------------------------------------


def test_annex_totals_tie_to_the_bsd2_spine(db_client: TestClient, seeded_book: str) -> None:
    snapshot = _generate(db_client, "BSD2", seeded_book)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["missing_dependencies"] == []
    cells = payload["cells"]
    spine, a4, a16, a6, a7, a15, a1 = (
        cells["BSD2"], cells[A4], cells[A16], cells[A6], cells[A7], cells[A15], cells[A1]
    )  # fmt: skip

    # Annex 16 — contingent liabilities: I = E+F+G+H per row (template), I11 = BSD2 line 33
    assert _f(a16, "G8") == 4_000_000.0
    assert _f(a16, "H8") == 500_000.0
    assert _f(a16, "E7") == 1_000_000.0
    for row in range(6, 11):
        assert _f(a16, f"I{row}") == pytest.approx(sum(_f(a16, f"{c}{row}") for c in "EFGH"))
    assert _f(a16, "I11") == pytest.approx(5_500_000.0)
    assert _f(a16, "I11") == pytest.approx(_f(spine, "D282"))
    assert _f(spine, "B282") == pytest.approx(4_500_000.0)
    assert _f(spine, "C282") == pytest.approx(1_000_000.0)

    # Annex 4 — advances by facility type: G13 (template) = BSD2 line 8 sub-total D68
    assert _f(a4, "B8") == 10_000_000.0
    assert _f(a4, "D10") == 6_000_000.0
    assert _f(a4, "C10") == 1_500_000.0
    assert _f(a4, "B10") == 300_000.0 == _f(a4, "B11")  # staff advance ⊂ private-sector scheduled
    assert _f(a4, "G13") == pytest.approx(17_800_000.0)
    assert _f(a4, "G13") == pytest.approx(_f(spine, "D68"))
    assert _f(spine, "C68") == pytest.approx(1_500_000.0)  # the USD loan sits in the foreign column

    # Annex 6 — other assets: total row = BSD2 line 11, domestic AND foreign (the
    # hermetic book carries a cedi other-assets plug and no FX other-assets fact)
    assert _f(a6, "B51") == pytest.approx(_f(spine, "B113"))
    assert _f(spine, "B113") > 0
    assert _f(a6, "C51") == pytest.approx(_f(spine, "C113")) == 0.0

    # Annex 15 — other liabilities: total row = BSD2 line 29, domestic AND foreign
    assert _f(a15, "C51") == pytest.approx(_f(spine, "B278")) == 2_000_000.0
    assert _f(a15, "D51") == pytest.approx(_f(spine, "C278")) == 700_000.0

    # Annex 7 — other reserves total = BSD2 15(e); Annex 1 E4 = BSD2 A.1 foreign column
    assert _f(a7, "C50") == pytest.approx(_f(spine, "B134"))
    assert _f(a7, "C50") > 0
    assert _f(a1, "E4") == pytest.approx(_f(spine, "C7")) == 200_000.0

    # detail rows stay blank + input_required, category totals are mapped
    counts = payload["status_counts"]
    assert counts["mapped"] == 280 + 44
    assert counts["input_required"] == 130 + 1421
    assert counts["unmapped"] == 0
    section = next(s for s in snapshot["sections"] if s["title"] == A16)
    statuses = {row["cell"]: row["status"] for row in section["rows"]}
    assert statuses["G8"] == "mapped" and statuses["C8"] == "input_required"


# ---------------------------------------------------------------------------
# 4. export: blank-grid values reach the official cells; detail rows are noted
# ---------------------------------------------------------------------------


def test_annex_values_and_completion_notes_reach_the_workbook(
    db_client: TestClient, seeded_book: str
) -> None:
    snapshot = _generate(db_client, "BSD2", seeded_book)
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx("BSD2", snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    a16 = wb[A16]
    assert a16["G8"].value == pytest.approx(4.0)  # ¢'Million
    assert a16["H8"].value == pytest.approx(0.5)
    assert a16["I11"].value == pytest.approx(5.5)  # template total, values-only
    assert a16["B8"].value == "Liabilities on guarantees"
    a4 = wb[A4]
    assert a4["B8"].value == pytest.approx(10.0)
    assert a4["G13"].value == pytest.approx(17.8)
    a6 = wb[A6]
    other_assets = float(snapshot["bog_form"]["cells"]["BSD2"]["B113"])
    assert a6["B51"].value == pytest.approx(other_assets / 1_000_000)
    a1 = wb[A1]
    assert a1["E4"].value == pytest.approx(0.2)
    assert a1["C4"].value is None  # foreign currency amount — input required, blank
    notes = wb["Completion notes"]
    rows = [[c.value for c in row] for row in notes.iter_rows(min_row=1, max_row=notes.max_row)]
    listed = {(r[0], r[1]) for r in rows if r and r[0] in ANNEX_SHEETS and r[4] == "input_required"}
    assert ("BSD2-Annex 12", "C8") in listed
    assert ("BSD2-Annex 13", "B39") in listed
    assert (A2A, "C10") in listed and (A2A, "C26") in listed
    assert ("BSD2-Annex 17", "C5") in listed
    assert (A16, "G8") not in listed
    detail_note = next(r[6] for r in rows if r and (r[0], r[1]) == ("BSD2-Annex 8", "C7"))
    assert "listing row 7" in str(detail_note)
