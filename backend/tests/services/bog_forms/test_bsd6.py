"""BSD6 (Maturity Analysis of Assets and Liabilities) — line map + resolver proof.

Generates BSD6 (and BSD2, its "FROM BSD2" dependency) through the real package
pipeline on the hermetic book, with a handful of positions inserted at known
maturities so every band is exercised, and proves:

1. every official input cell of BSD6A/BSD6B is bound (nothing ``unmapped``);
   the mapped / input_required split is what docs/bog_returns/bsd6_line_map.md
   claims;
2. the Total column equals the sum of the eight maturity bands on every row
   whose bands are filled (the template's Total is an INPUT cell — the
   resolver's own invariant — and BoG's section formulas re-add the bands);
3. every FROM BSD2 cell equals the corresponding BSD2 cell of the same
   reporting date, and every platform-filled Total equals its FROM BSD2 cell
   (the Guide's "totals must agree with BSD2");
4. the Guide's placement rules land where the notes say — overdue ≥ 14 days,
   < 14 days → less than 1 month, boundary dates belong to the later band,
   weekend maturities roll to Monday, behavioural maturity for savings, cash
   → Overdue, provisions → Overdue, reserves → 5 years and over (negative →
   Overdue), unplaceable fact rows leave the bands blank;
5. the calendar-month arithmetic reproduces the Guide's worked table.
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.db.session import get_sessionmaker
from app.models import (
    Bank,
    BankFinancialFact,
    BankReportingPeriod,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    IngestionBatch,
    LineageRecord,
)
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import line_maps_for
from app.services.regulatory_reporting.bog_forms.sources_ext.bsd6 import (
    BUCKETS,
    _add_months,
    band_boundaries,
    bucket_for,
)
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import headers
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

BAND_COLUMNS = dict(zip(BUCKETS, "DEFGHIJK", strict=True))
MILLION = Decimal("1000000")

# Positions inserted at the latest period end (2026-03-31). Band boundaries for
# that reporting date (Guide, calendar-month basis; month-end → month-end):
# <1m: to 2026-04-29 · 1m–<3m: 2026-04-30–2026-06-29 · 3m–<6m: 2026-06-30–
# 2026-09-29 · 6m–<1y: 2026-09-30–2027-03-30 · 1y–<3y: 2027-03-31–2029-03-30 ·
# 3y–<5y: 2029-03-31–2031-03-30 · 5y+: from 2031-03-31.
PERIOD_END = date(2026, 3, 31)


# ---------------------------------------------------------------------------
# calendar arithmetic — the Guide's worked table, verbatim
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("reporting_date", "lt_1m_last", "one_to_three", "three_to_six", "six_to_year"),
    [
        # Guide table: reporting date → last day of "<1 month" · first/last of the
        # next three bands.
        (date(2026, 3, 31), date(2026, 4, 29), (date(2026, 4, 30), date(2026, 6, 29)),
         (date(2026, 6, 30), date(2026, 9, 29)), (date(2026, 9, 30), date(2027, 3, 30))),
        (date(2026, 6, 30), date(2026, 7, 30), (date(2026, 7, 31), date(2026, 9, 29)),
         (date(2026, 9, 30), date(2026, 12, 30)), (date(2026, 12, 31), date(2027, 6, 29))),
        (date(2026, 9, 30), date(2026, 10, 30), (date(2026, 10, 31), date(2026, 12, 30)),
         (date(2026, 12, 31), date(2027, 3, 30)), (date(2027, 3, 31), date(2027, 9, 29))),
        (date(2026, 12, 31), date(2027, 1, 30), (date(2027, 1, 31), date(2027, 3, 30)),
         (date(2027, 3, 31), date(2027, 6, 29)), (date(2027, 6, 30), date(2027, 12, 30))),
    ],
)  # fmt: skip
def test_bands_follow_the_guides_calendar_month_table(
    reporting_date: date,
    lt_1m_last: date,
    one_to_three: tuple[date, date],
    three_to_six: tuple[date, date],
    six_to_year: tuple[date, date],
) -> None:
    lower_1m, lower_3m, lower_6m, lower_1y, *_ = band_boundaries(reporting_date)
    assert (lower_1m, lower_3m, lower_6m, lower_1y) == (
        one_to_three[0],
        three_to_six[0],
        six_to_year[0],
        _add_months(reporting_date, 12),
    )
    # a weekday probe on each documented boundary (weekend roll is separate):
    # a band's first day is probed forward, its last day backward
    for day, band, step in (
        (lt_1m_last, "lt_1m", -1),
        (one_to_three[0], "1m_lt_3m", 1),
        (one_to_three[1], "1m_lt_3m", -1),
        (three_to_six[0], "3m_lt_6m", 1),
        (three_to_six[1], "3m_lt_6m", -1),
        (six_to_year[0], "6m_lt_1y", 1),
        (six_to_year[1], "6m_lt_1y", -1),
    ):
        probe = day
        while probe.weekday() >= 5:  # noqa: PLR2004 — the table's dates are calendar days
            probe = date.fromordinal(probe.toordinal() + step)
        assert bucket_for(probe, reporting_date, side="asset") == band, (day, band)


def test_bucket_edges_overdue_rule_and_weekend_roll() -> None:
    pe = PERIOD_END
    # assets: due date passed by ≥ 14 days → Overdue, by < 14 days → <1 month
    assert bucket_for(date(2026, 3, 17), pe, side="asset") == "overdue"
    assert bucket_for(date(2026, 3, 18), pe, side="asset") == "lt_1m"
    assert bucket_for(date(2026, 3, 31), pe, side="asset") == "lt_1m"
    # liabilities already due → earliest repayment date → <1 month
    assert bucket_for(date(2025, 12, 31), pe, side="liability") == "lt_1m"
    # boundary date belongs to the later band
    assert bucket_for(date(2026, 4, 30), pe, side="asset") == "1m_lt_3m"
    assert bucket_for(date(2027, 3, 31), pe, side="asset") == "1y_lt_3y"
    assert bucket_for(date(2029, 3, 30), pe, side="asset") == "1y_lt_3y"
    # Saturday 2029-03-31 → Monday 2029-04-02, still 3y–<5y; Saturday 2031-03-29
    # rolls to Monday 2031-03-31 = the 5-year boundary → 5 years and over
    assert bucket_for(date(2029, 3, 31), pe, side="asset") == "3y_lt_5y"
    assert bucket_for(date(2031, 3, 29), pe, side="asset") == "5y_plus"
    assert bucket_for(date(2031, 3, 28), pe, side="asset") == "3y_lt_5y"
    # non-month-end anchor keeps the day-of-month
    assert _add_months(date(2026, 1, 15), 1) == date(2026, 2, 15)
    assert _add_months(date(2026, 1, 31), 1) == date(2026, 2, 28)
    assert _add_months(date(2026, 2, 28), 1) == date(2026, 3, 31)


# ---------------------------------------------------------------------------
# hermetic book + a bucketed position book
# ---------------------------------------------------------------------------


def _seed_positions(session: Any) -> None:  # noqa: PLR0915 — one linear fixture script
    """A handful of positions whose BSD2 filters and maturities are known, so
    each BSD6 band receives a hand-checkable amount."""
    period = session.query(BankReportingPeriod).filter_by(period_end=PERIOD_END).one()
    batch = IngestionBatch(
        organization_id=DEMO_ORG_ID,
        bank_id=SAMPLE_BANK_ID,
        source_system="EXCEL_CSV",
        adapter_version="1.0",
        extraction_mode="full",
        status="accepted",
        as_of_date=PERIOD_END,
    )
    session.add(batch)
    session.flush()
    lineage = LineageRecord(
        organization_id=DEMO_ORG_ID,
        ingestion_batch_id=batch.id,
        operation_type="ADAPTER_TRANSLATE",
        operation_ref="bsd6-test",
        input_lineage_ids=[],
    )
    session.add(lineage)
    session.flush()
    common = {
        "organization_id": DEMO_ORG_ID,
        "bank_id": SAMPLE_BANK_ID,
        "as_of_date": PERIOD_END,
        "source_system": "EXCEL_CSV",
        "ingestion_batch_id": batch.id,
        "lineage_id": lineage.id,
        "validation_status": "accepted",
    }

    def counterparty(ref: str, kind: str, *, resident: bool) -> CanonicalCounterparty:
        row = CanonicalCounterparty(
            **common,
            source_reference=f"CP/{ref}",
            name=ref,
            counterparty_type=kind,
            resident=resident,
        )
        session.add(row)
        return row

    corp = counterparty("CORP", "CORPORATE", resident=True)
    retail = counterparty("RETAIL", "RETAIL_INDIVIDUAL", resident=True)
    local_bank = counterparty("LOCALBANK", "BANK_NON_OECD", resident=True)
    foreign_bank = counterparty("NOSTRO", "BANK_OECD", resident=False)
    gog = counterparty("GOG", "SOVEREIGN", resident=True)
    session.flush()

    def position(  # noqa: PLR0913 — keyword-only fixture builder
        ref: str,
        position_type: str,
        currency: str,
        balance: str,
        *,
        counterparty: CanonicalCounterparty | None = None,
        maturity: date | None = None,
        deposit_account_type: str | None = None,
        behavioral_months: int | None = None,
        attributes: dict[str, Any] | None = None,
    ) -> None:
        row = CanonicalPosition(
            **common, source_reference=ref, position_type=position_type, currency=currency
        )
        session.add(row)
        session.flush()
        session.add(
            CanonicalPositionSnapshot(
                **common,
                source_reference=ref,
                position_id=row.id,
                counterparty_id=counterparty.id if counterparty else None,
                balance=Decimal(balance),
                contractual_maturity=maturity,
                deposit_account_type=deposit_account_type,
                behavioral_maturity_months=behavioral_months,
                attributes=attributes or {},
            )
        )

    # --- BSD6A (cedis) -------------------------------------------------------
    # BSD2 66 (d) Private enterprises → BSD6A row 24
    position(
        "LOAN/CORP/1", "LOAN", "GHS", "30000000", counterparty=corp, maturity=date(2029, 6, 30)
    )  # 3y–<5y
    position(
        "LOAN/CORP/2", "LOAN", "GHS", "4000000", counterparty=corp, maturity=date(2031, 3, 29)
    )  # Saturday → Monday 2031-03-31 → 5y+
    # BSD2 67 (e) Individuals → BSD6A row 26
    position(
        "LOAN/RET/1", "LOAN", "GHS", "8000000", counterparty=retail, maturity=date(2026, 4, 20)
    )  # <1m
    position(
        "LOAN/RET/2", "LOAN", "GHS", "3000000", counterparty=retail, maturity=date(2026, 2, 15)
    )  # 44 days past due → Overdue
    position(
        "LOAN/RET/3", "LOAN", "GHS", "2000000", counterparty=retail, maturity=date(2026, 3, 25)
    )  # 6 days past due → <1m
    # BSD2 36 (i) 91 Day T-bill → BSD6A row 13. NOTE: BSD2's attribute_eq
    # compares JSON attributes AS TEXT (positions.sum: attributes[key].as_string()
    # == str(value)); an integer 91 in the JSON does not match on SQLite (and a
    # bool never matches anywhere) — recorded as a framework ask in the BSD6 doc.
    # The fixture stores the text form so the tie-out path is exercised.
    position(
        "SEC/TBILL",
        "SECURITY_HOLDING",
        "GHS",
        "15000000",
        counterparty=gog,
        maturity=date(2026, 6, 15),
        attributes={"instrument": "tbill", "tenor_days": "91"},
    )  # 1m–<3m
    # BSD2 76 (iv) 2 year Bonds → BSD6A row 32
    position(
        "SEC/BOND2Y",
        "SECURITY_HOLDING",
        "GHS",
        "20000000",
        counterparty=gog,
        maturity=date(2028, 3, 31),
        attributes={"instrument": "gog_bond", "tenor_years": "2"},
    )  # 1y–<3y
    # BSD2 22 (i) Commercial banks (placements) → BSD6A row 9
    position(
        "IBP/1",
        "INTERBANK_PLACEMENT",
        "GHS",
        "5000000",
        counterparty=local_bank,
        maturity=date(2026, 4, 30),
    )  # boundary → 1m–<3m
    # BSD2 228 / 236 / 245 (25. deposits, block 26) → BSD6A rows 71 / 72 / 73
    position(
        "DEP/CUR",
        "DEPOSIT",
        "GHS",
        "25000000",
        counterparty=retail,
        deposit_account_type="CURRENT",
        attributes={"deposit_account_type": "CURRENT", "block": "26"},
    )  # on demand → <1m
    position(
        "DEP/SAV",
        "DEPOSIT",
        "GHS",
        "20000000",
        counterparty=retail,
        deposit_account_type="SAVINGS",
        behavioral_months=18,
        attributes={"deposit_account_type": "SAVINGS", "block": "26"},
    )  # 1y–<3y
    position(
        "DEP/FIX",
        "DEPOSIT",
        "GHS",
        "10000000",
        counterparty=corp,
        deposit_account_type="FIXED",
        maturity=date(2026, 9, 30),
        attributes={"deposit_account_type": "FIXED", "block": "26"},
    )  # boundary → 6m–<1y
    # BSD2 190 (i) Commercial banks term borrowing → BSD6A row 60
    position(
        "IBB/1",
        "INTERBANK_BORROWING",
        "GHS",
        "6000000",
        counterparty=local_bank,
        maturity=date(2026, 8, 29),
        attributes={"instrument": "term_borrowing"},
    )  # 3m–<6m
    # --- BSD6B (foreign currency) -------------------------------------------
    position(
        "LOAN/USD", "LOAN", "USD", "1000000", counterparty=corp, maturity=date(2028, 6, 30)
    )  # BSD2 66 foreign → BSD6B row 24, 1y–<3y
    position(
        "DEP/USD",
        "DEPOSIT",
        "USD",
        "200000",
        counterparty=retail,
        deposit_account_type="SAVINGS",
        attributes={"deposit_account_type": "SAVINGS"},
    )  # BSD2 153 → BSD6B row 60, <1m
    position(
        "CASH/USD", "CASH", "USD", "50000", attributes={"instrument": "fx_notes_coins"}
    )  # BSD2 7 → BSD6B row 7, Overdue
    position(
        "NOSTRO/USD",
        "INTERBANK_PLACEMENT",
        "USD",
        "300000",
        counterparty=foreign_bank,
        maturity=date(2026, 4, 10),
    )  # BSD2 8 → BSD6B row 9, <1m
    # A negative reserve (BSD2 131 revaluation reserve) → BSD6A row 44 Overdue
    session.add(
        BankFinancialFact(
            organization_id=DEMO_ORG_ID,
            bank_id=SAMPLE_BANK_ID,
            reporting_period_id=period.id,
            fact_group="capital_component",
            category="revaluation_reserve",
            amount=Decimal("-2000000"),
            currency="GHS",
        )
    )
    session.flush()


def _materialize_with_positions(db_client: TestClient) -> None:
    _ = db_client
    session = get_sessionmaker()()
    try:
        materialize_canonical_test_book(session)
        _seed_positions(session)
        session.commit()
    finally:
        session.close()


def _generate(db_client: TestClient, code: str) -> dict[str, Any]:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": PERIOD_END.isoformat()},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    return detail["snapshot"]


def _num(value: Any) -> Decimal | None:
    if value is None or isinstance(value, str):
        return None
    return Decimal(str(value))


def _leaf_rows(sheet: str) -> dict[int, Any]:
    return {int(line.cells["from_bsd2"][1:]): line for line in line_maps_for("BSD6")[sheet]}


# ---------------------------------------------------------------------------
# the proof
# ---------------------------------------------------------------------------


@pytest.fixture
def snapshots(db_client: TestClient) -> tuple[dict[str, Any], dict[str, Any]]:
    _materialize_with_positions(db_client)
    return _generate(db_client, "BSD6"), _generate(db_client, "BSD2")


def test_every_input_cell_of_both_sheets_is_bound(
    snapshots: tuple[dict[str, Any], dict[str, Any]],
) -> None:
    bsd6, _ = snapshots
    payload = bsd6["bog_form"]
    assert payload["code"] == "BSD6"
    assert not payload["errors"], payload["errors"]
    assert payload["unmapped_cells"] == []
    assert payload["missing_dependencies"] == []
    layout = load_layout("BSD6")
    counts = payload["status_counts"]
    total_inputs = sum(len(s.input_cells) for s in layout.sheets)
    assert total_inputs == 600 + 500  # noqa: PLR2004 — 60 + 50 leaf rows × 10 columns
    assert counts["mapped"] + counts["input_required"] == total_inputs
    assert counts["derived"] == 150 + 140  # noqa: PLR2004 — the template's own roll-ups
    # per-sheet split as documented in bsd6_line_map.md: BSD6A 52 rows fed (44 +
    # the 7 accrued-interest rows the interest_accruals sub-ledger feeds through
    # BSD2 since 2026-08-16 + the fixed-assets row 35 the capital_expenditure
    # register feeds) / 8 input_required, BSD6B 44 (40 + 3 + 1) / 6; the "Other
    # Assets" row (a fact with no Guide band) fills FROM BSD2 + Total only.
    lines_a = line_maps_for("BSD6")["BSD6A"]
    lines_b = line_maps_for("BSD6")["BSD6B"]
    assert (len(lines_a), len(lines_b)) == (60, 50)
    assert sum(1 for line in lines_a if line.source) == 52  # noqa: PLR2004
    assert sum(1 for line in lines_b if line.source) == 44  # noqa: PLR2004
    by_sheet: dict[str, dict[str, int]] = {}
    for section in bsd6["sections"]:
        tally = by_sheet.setdefault(section["title"], {"mapped": 0, "input_required": 0})
        for row in section["rows"]:
            tally[row["status"]] += 1
    # the hermetic book has no accruals sub-ledger and no capex register: the rows fed
    # only by those datasets (6A 11, 35, 56, 64, 66–69; 6B 35, 53, 57, 63) stay
    # input_required end to end; BSD6B's Other Assets foreign slice is 0 — nothing to
    # allocate, so its bands read 0 (mapped)
    assert by_sheet["BSD6A"] == {"mapped": 44 * 10 - 8, "input_required": 16 * 10 + 8}
    assert by_sheet["BSD6B"] == {"mapped": 40 * 10, "input_required": 10 * 10}


def test_total_equals_sum_of_bands_and_bogs_section_formulas_re_add_them(
    snapshots: tuple[dict[str, Any], dict[str, Any]],
) -> None:
    bsd6, _ = snapshots
    for sheet in ("BSD6A", "BSD6B"):
        cells = bsd6["bog_form"]["cells"][sheet]
        checked = 0
        for row in _leaf_rows(sheet):
            total = _num(cells.get(f"C{row}"))
            bands = [_num(cells.get(f"{col}{row}")) for col in "DEFGHIJK"]
            if total is None or any(b is None for b in bands):
                continue
            assert sum(b for b in bands if b is not None) == total, (sheet, row)
            checked += 1
        # every platform-fed row except BSD6A "Other Assets" (bands blank by design;
        # its foreign slice is 0 on BSD6B, so nothing is left to allocate there)
        assert checked == {"BSD6A": 43, "BSD6B": 40}[sheet], sheet
        # the template's own section arithmetic over those inputs, per column
        layout = load_layout("BSD6").sheet(sheet)
        for col in "BCDEFGHIJK":
            assert _num(cells[f"{col}6"]) == sum(
                _num(cells.get(f"{col}{r}")) or Decimal(0) for r in range(7, 12)
            ), (sheet, col)
            assert _num(cells[f"{col}28"]) == sum(
                _num(cells.get(f"{col}{r}")) or Decimal(0) for r in range(21, 28)
            ), (sheet, col)
        # Total Assets (row 36) = 1 + 2 + 3(net) + 4 + 5 + 6 + 7 — BoG's formula
        assert layout.by_ref["C36"].formula == "=C6+C12+C20+C32+C33+C34+C35"
        expected = sum(_num(cells[f"C{r}"]) or Decimal(0) for r in (6, 12, 20, 32, 33, 34, 35))
        assert _num(cells["C36"]) == expected, sheet


def test_from_bsd2_and_totals_agree_with_bsd2(
    snapshots: tuple[dict[str, Any], dict[str, Any]],
) -> None:
    """The Guide's rule: BSD6A totals = BSD2 Domestic, BSD6B totals = BSD2 Foreign."""
    bsd6, bsd2 = snapshots
    bsd2_cells = bsd2["bog_form"]["cells"]["BSD2"]
    for sheet in ("BSD6A", "BSD6B"):
        cells = bsd6["bog_form"]["cells"][sheet]
        agreed = 0
        nonzero = 0
        for row, line in _leaf_rows(sheet).items():
            if line.source is None:
                continue
            refs = line.params["bsd2_refs"]
            assert isinstance(refs, list)
            from_bsd2 = _num(cells.get(f"B{row}"))
            ref_values = [_num(bsd2_cells.get(ref)) for ref in refs]
            if from_bsd2 is None:
                # every BSD2 leaf behind the row is input_required (the accruals
                # sub-ledger is not ingested in the hermetic book): so is the BSD6 row,
                # FROM BSD2 included (BSD2's subtotal formula over blanks is not a figure)
                assert _num(cells.get(f"C{row}")) is None, (sheet, row)
                components = line.params["components"]
                assert isinstance(components, list)
                assert all(c["source"] == "refs.sum" for c in components), (sheet, row)
                continue
            expected = sum(value or Decimal(0) for value in ref_values)
            assert from_bsd2 == expected, (sheet, row, refs)
            total = _num(cells.get(f"C{row}"))
            assert total is not None, (sheet, row)
            # platform-fed Total reconciles to the BSD2 line it mirrors
            assert total == from_bsd2, (sheet, row, refs)
            agreed += 1
            nonzero += total != 0
        assert agreed >= 40, sheet  # noqa: PLR2004
        # the seeded book reaches the ladder; BSD6B carries fewer live rows now
        # that BSD2's capital lines are Domestic-only (they no longer leak cedi
        # capital into the foreign-currency sheet — the tie-out above still holds)
        assert nonzero >= {"BSD6A": 6, "BSD6B": 4}[sheet], sheet
    # named cross-form equalities (base units; ¢'Million scaling happens at export)
    a = bsd6["bog_form"]["cells"]["BSD6A"]
    b = bsd6["bog_form"]["cells"]["BSD6B"]
    assert _num(a["B24"]) == _num(bsd2_cells["B66"]) == 34 * MILLION  # private enterprises
    assert _num(a["B26"]) == _num(bsd2_cells["B67"]) == 13 * MILLION  # individuals
    assert _num(a["B13"]) == _num(bsd2_cells["B35"]) == 15 * MILLION  # GoG bills
    assert _num(a["B7"]) == _num(bsd2_cells["B14"]) != 0  # cash on hand (facts)
    assert _num(a["B41"]) == _num(bsd2_cells["B128"]) != 0  # paid-up capital (facts)
    assert _num(a["B28"]) == _num(bsd2_cells["B68"])  # loans sub-total = BSD2 sub-total
    assert _num(b["B24"]) == _num(bsd2_cells["C66"]) == 1 * MILLION  # USD loan, foreign column
    assert _num(b["B7"]) == _num(bsd2_cells["C7"]) == Decimal(50000)


def test_guide_placements_land_in_the_documented_bands(
    snapshots: tuple[dict[str, Any], dict[str, Any]],
) -> None:
    bsd6, _ = snapshots
    a = bsd6["bog_form"]["cells"]["BSD6A"]
    b = bsd6["bog_form"]["cells"]["BSD6B"]

    def band(cells: dict[str, Any], row: int, name: str) -> Decimal:
        return _num(cells[f"{BAND_COLUMNS[name]}{row}"]) or Decimal(0)

    # loans to private enterprises: 30M in 3y–<5y, 4M (Sat 2031-03-29 → Mon = 5y edge)
    assert band(a, 24, "3y_lt_5y") == 30 * MILLION
    assert band(a, 24, "5y_plus") == 4 * MILLION
    # individuals: 8M + 2M (6 days past due) in <1m, 3M (44 days past due) Overdue
    assert band(a, 26, "lt_1m") == 10 * MILLION
    assert band(a, 26, "overdue") == 3 * MILLION
    # 91-day T-bill maturing 2026-06-15 → 1m–<3m
    assert band(a, 13, "1m_lt_3m") == 15 * MILLION
    # 2-year bond 2028-03-31 → 1y–<3y (long-term securities row)
    assert band(a, 32, "1y_lt_3y") == 20 * MILLION
    # placement maturing exactly on the 1-month boundary → 1m–<3m
    assert band(a, 9, "1m_lt_3m") == 5 * MILLION
    assert band(a, 9, "lt_1m") == 0
    # deposits: current on demand → <1m; savings with 18-month behavioural
    # maturity → 1y–<3y; fixed maturing on the 6-month boundary → 6m–<1y
    assert band(a, 71, "lt_1m") == 25 * MILLION
    assert band(a, 72, "1y_lt_3y") == 20 * MILLION
    assert band(a, 73, "6m_lt_1y") == 10 * MILLION
    # term borrowing 2026-08-29 (Saturday → Monday 31 Aug) → 3m–<6m
    assert band(a, 60, "3m_lt_6m") == 6 * MILLION
    # Guide fixed bands for fact-sourced lines
    assert band(a, 7, "overdue") == _num(a["C7"]) != 0  # cash on hand → Overdue
    assert band(a, 41, "5y_plus") == _num(a["C41"]) != 0  # paid-up capital → 5y+
    assert band(a, 44, "overdue") == _num(a["C44"]) == -2 * MILLION  # negative reserve
    assert band(a, 44, "5y_plus") == 0
    # provisions → Overdue (fixture provisions facts may be zero; the band holds the total)
    assert band(a, 29, "overdue") == _num(a["C29"])
    # Other Assets: FROM BSD2 + Total filled, bands blank for the bank to allocate
    assert _num(a["C34"]) is not None and _num(a["C34"]) != 0
    assert all(a.get(f"{col}34") is None for col in "DEFGHIJK")
    # BSD6B: FX notes/coins → Overdue; nostro <1m; USD savings on demand; USD loan 1y–<3y
    assert band(b, 7, "overdue") == Decimal(50000)
    assert band(b, 9, "lt_1m") == Decimal(300000)
    assert band(b, 60, "lt_1m") == Decimal(200000)
    assert band(b, 24, "1y_lt_3y") == 1 * MILLION
    # nothing leaks across the currency split
    assert _num(a["C24"]) == 34 * MILLION
    assert _num(b["C26"]) == 0


def test_bsd6_export_scales_to_millions_and_lists_completion_notes(
    db_client: TestClient, snapshots: tuple[dict[str, Any], dict[str, Any]]
) -> None:
    """The package export path renders the ladder in ¢'Million with the
    input_required rows on the Completion notes sheet (structure never dropped)."""
    bsd6, _ = snapshots
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = DEMO_ORG_ID
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx("BSD6", bsd6, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    assert wb.sheetnames[:2] == ["BSD6A", "BSD6B"]
    ws = wb["BSD6A"]
    assert ws["J24"].value == pytest.approx(30.0)  # 30M cedis → 30 (¢'Million)
    assert ws["D26"].value == pytest.approx(3.0)
    assert ws["D34"].value is None  # blank, listed in the notes
    notes = wb["Completion notes"]
    blob = " ".join(str(c.value) for row in notes.iter_rows() for c in row if c.value)
    assert "Other Assets" in blob or "BSD6A.R34" in blob
    assert "Interest in suspense" in blob or "BSD6A.R30" in blob
    _ = db_client
