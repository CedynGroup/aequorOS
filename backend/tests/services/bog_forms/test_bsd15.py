"""BSD15A / BSD15B — Charges (Domestic / International Banking).

No tariff / fee register exists on the platform, so both forms are structure
+ honest ``input_required``:

1. every tariff row of each sheet is bound at its official value cell (BSD15A
   Domestic: 168 rows in column C + the 22 item numbers in column A kept as
   template constants; Range of products: 77 rows in column B; BSD15B: 309
   rows in column B); a heading that introduces sub-lines is not a value row,
   a childless heading is;
2. both forms generate through the REAL package pipeline with every tariff
   cell ``input_required`` ("tariff register required") and zero errors;
3. the xlsx export carries every official label verbatim, the item numbers,
   blanks in every tariff cell, and a Completion-notes sheet listing each
   tariff line the bank must supply.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

import openpyxl
from fastapi.testclient import TestClient

from app.db.session import get_sessionmaker
from app.models import Bank
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import bsd15a, bsd15b, line_maps_for
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import ORG_1, headers
from tests.fixtures.canonical_bank_fixture import (
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

DOMESTIC, RANGE, INTL = bsd15a.DOMESTIC, bsd15a.RANGE, bsd15b.SHEET
DOMESTIC_ROWS = 168
RANGE_ROWS = 77
INTL_ROWS = 309
ITEM_NUMBERS = 22


def test_line_maps_bind_every_tariff_row_and_keep_item_numbers() -> None:
    a = line_maps_for("BSD15A")
    b = line_maps_for("BSD15B")
    domestic = {ref: line for line in a[DOMESTIC] for ref in line.cells.values()}
    range_ = {ref: line for line in a[RANGE] for ref in line.cells.values()}
    intl = {ref: line for line in b[INTL] for ref in line.cells.values()}
    # value columns + row counts
    assert sum(ref.startswith("C") for ref in domestic) == DOMESTIC_ROWS
    assert sum(ref.startswith("A") for ref in domestic) == ITEM_NUMBERS
    assert len(range_) == RANGE_ROWS and all(ref.startswith("B") for ref in range_)
    assert len(intl) == INTL_ROWS and all(ref.startswith("B") for ref in intl)
    # the template's captured item numbers are the only mapped cells
    layout = load_layout("BSD15A").sheet(DOMESTIC)
    captured = {c.ref for c in layout.input_cells}
    assert captured == {ref for ref in domestic if ref.startswith("A")}
    for ref in captured:
        assert domestic[ref].source == "constant"
        assert domestic[ref].params["value"] == layout.by_ref[ref].value
    # data-gap closure (2026-08-16): every tariff cell reads the ``tariff_schedule``
    # register (refs.field on charge_value); without a register it resolves to
    # input_required at generation, which the export tests below still prove
    for line in (*range_.values(), *intl.values(), *(domestic[r] for r in domestic if r[0] == "C")):
        assert line.source == "refs.field"
        assert line.params["kind"] == "tariff_schedule"
        assert "tariff register required" in line.notes
    # a childless numbered heading takes the value; a heading with sub-lines does not
    assert "B363" in intl  # 17. ACCOUNT CLOSURE
    assert "B9" not in intl  # 1. LETTERS OF CREDIT (IMPORTS) — has sub-lines
    assert "B28" not in intl  # "Amendments:" — group label with sub-lines
    assert "C10" in domestic and "C9" not in domestic  # COT minimum vs the COT heading
    assert "B13" in range_ and "B12" not in range_  # Initial Deposit vs "1. SAVINGS A/C (S1)"


def _prepare(db_client: TestClient) -> str:
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    return periods[0]["period_end"]


def _generate(db_client: TestClient, code: str, reporting_date: str) -> dict[str, Any]:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": reporting_date},
    )
    assert response.status_code == 201, response.text[:400]
    package = response.json()
    return db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()["snapshot"]


def _export(code: str, snapshot: dict[str, Any]) -> openpyxl.Workbook:
    session = get_sessionmaker()()
    try:
        session.info["organization_id"] = ORG_1
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        payload = render_bog_form_xlsx(code, snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
    finally:
        session.close()
    return openpyxl.load_workbook(io.BytesIO(payload), data_only=False)


def test_bsd15a_exports_the_tariff_structure_with_every_cell_input_required(
    db_client: TestClient,
) -> None:
    reporting_date = _prepare(db_client)
    snapshot = _generate(db_client, "BSD15A", reporting_date)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    counts = payload["status_counts"]
    assert counts["mapped"] == ITEM_NUMBERS  # the template's own item numbers
    assert counts["input_required"] == DOMESTIC_ROWS + RANGE_ROWS
    assert counts["unmapped"] == 0
    for section in snapshot["sections"]:
        for row in section["rows"]:
            if row["cell"].startswith("A") and section["title"] == DOMESTIC:
                assert row["status"] == "mapped"
            else:
                assert row["status"] == "input_required", row
                assert row["value"] is None
                assert "tariff register required" in row["notes"]
    wb = _export("BSD15A", snapshot)
    ws = wb[DOMESTIC]
    assert ws["B9"].value == "COMMISSION ON TURNOVER (COT)"
    assert ws["A9"].value == 1 and ws["A211"].value == ITEM_NUMBERS  # noqa: PLR2004
    assert ws["C10"].value is None and ws["C231"].value is None
    ws2 = wb[RANGE[:31]]
    assert ws2["A13"].value == "    Initial Deposit"
    assert ws2["B13"].value is None
    notes = wb["Completion notes"]
    listed = {(row[0].value, row[1].value) for row in notes.iter_rows(min_row=1)}
    assert (DOMESTIC, "C10") in listed and (RANGE, "B13") in listed


def test_bsd15b_exports_the_tariff_structure_with_every_cell_input_required(
    db_client: TestClient,
) -> None:
    reporting_date = _prepare(db_client)
    snapshot = _generate(db_client, "BSD15B", reporting_date)
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    counts = payload["status_counts"]
    assert counts == {"mapped": 0, "input_required": INTL_ROWS, "unmapped": 0, "derived": 0}
    wb = _export("BSD15B", snapshot)
    ws = wb[INTL]
    assert ws["A9"].value == "1. LETTERS OF CREDIT (IMPORTS)"
    assert ws["A363"].value == "17. ACCOUNT CLOSURE"
    assert ws["B11"].value is None and ws["B363"].value is None
    notes = wb["Completion notes"]
    listed = {(row[0].value, row[1].value) for row in notes.iter_rows(min_row=1)}
    assert (INTL, "B11") in listed and (INTL, "B363") in listed
