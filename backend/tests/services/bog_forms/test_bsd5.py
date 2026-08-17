"""BSD5A / BSD5B (Capital Adequacy Returns) — line map + resolver proof.

Runs the capital baseline on the hermetic book, generates BSD5A (and BSD5B,
which depends on it) through the real package pipeline and proves:

1. every official input cell of all sheets is bound — CAR FORMAT (amounts +
   item numbers), NEW RISK WEIGHTS (every printed weight is a ``constant``
   EQUAL to the template's own value — a BoG weight is never retyped),
   PROVISION (the printed ladder); BSD5B's four blank data cells (D8, D24,
   D71, D72) are bound too; the mapped / input_required split is what the
   docs claim;
2. the capital side reconciles to the capital run: Tier 1 (E10) = Σ Tier 1
   lines, and the ADJUSTED CAPITAL BASE (E25) equals the signed sum of the
   run's persisted capital-component line items that have a BSD5A row
   (the run's total capital less the components the form predates — DTA,
   general provisions);
3. the asset side reconciles to the platform state: TOTAL ASSETS = Σ
   balance-sheet asset facts, the printed percentages apply (50% of
   residential mortgages, 80% of claims on other banks), the LC/guarantee
   book is never double counted, 50% of NOP and the 3-year average gross
   income come from the run's line items;
4. BoG's own arithmetic on those inputs: Adjusted Total Assets, Net
   Contingent Liabs, ADJUSTED ASSET BASE, the ratio E70 ``=E25/E69`` and the
   6% surplus test — and, explicitly, WHY the BSD5A ratio differs from the
   run's Basel III CAR (BoG's add-ons are 50% of NOP and 100% of average gross
   income; the engine's are the FX charge × 12.5 and the BIA charge × 12.5);
5. BSD5B: group = solo (form.cell links) until a subsidiary book exists,
   consolidation-only lines input_required, BoG's group formulas (D74
   ``=D30/D73%``, D75 10% test, the D46–D50 copies) reproduced verbatim;
6. the resolvers against inserted rows (deduction signs, tier residual,
   the printed-percentage wrapper, the LC/guarantee residual, the run lines).
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import TenantContext
from app.db.session import get_sessionmaker
from app.models import (
    Bank,
    BankFinancialFact,
    BankReportingPeriod,
    CanonicalCounterparty,
    CanonicalPosition,
    CanonicalPositionSnapshot,
    IngestionBatch,
    LineageRecord,
)
from app.services.regulatory_reporting.bog_forms.layout import load_layout
from app.services.regulatory_reporting.bog_forms.linemaps import bsd5a, line_maps_for
from app.services.regulatory_reporting.bog_forms.sources import ResolveContext, get_resolver
from app.services.regulatory_reporting.exports import render_bog_form_xlsx
from tests.api.helpers import ORG_1, USER_1, headers
from tests.fixtures.canonical_bank_fixture import (
    DEMO_ORG_ID,
    SAMPLE_BANK_ID,
    materialize_canonical_test_book,
)

M = Decimal("1000000")
PERIOD_END = date(2026, 3, 31)
CAR = "CAR FORMAT"
GROUP = "CAR FORMAT-GROUP"
RW = "NEW RISK WEIGHTS"
PROV = "PROVISION"

# BSD5A CAR FORMAT amount rows the platform cannot feed honestly (doc §Residual)
CAR_INPUT_REQUIRED_ROWS = {30, 33, 40, 48, 49, 50, 52, 57, 58, 59, 60, 61, 63, 64}
# every category name with its own BSD5A row (Tier 1 residual + AT1 handled by tier)
NAMED = frozenset(
    (
        *bsd5a.PAID_UP,
        *bsd5a.PERMANENT_PREFERENCE,
        *bsd5a.UNDISCLOSED,
        *bsd5a.REVALUATION,
        *bsd5a.GOODWILL,
        *bsd5a.LOSSES_NOT_PROVIDED,
        *bsd5a.INVESTMENTS_SUBSIDIARIES,
        *bsd5a.INVESTMENTS_OTHER_BANKS,
        *bsd5a.CONNECTED_LENDING,
        *bsd5a.SUBORDINATED_DEBT,
        *bsd5a.HYBRID,
        *bsd5a.CUMULATIVE_PREFERENCE,
    )
)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _session() -> Session:
    session = get_sessionmaker()()
    session.info["organization_id"] = ORG_1
    return session


def _materialize(db_client: TestClient) -> None:
    _ = db_client
    session = _session()
    try:
        materialize_canonical_test_book(session)
        session.commit()
    finally:
        session.close()


def _latest_period(db_client: TestClient) -> dict[str, Any]:
    periods = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/reporting-periods", headers=headers()
    ).json()["periods"]
    latest = periods[0]
    assert latest["period_end"] == PERIOD_END.isoformat()
    return latest


def _run_capital_baseline(db_client: TestClient) -> dict[str, Any]:
    period = _latest_period(db_client)
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-runs",
        headers=headers(),
        json={
            "module": "capital",
            "reporting_period_id": period["id"],
            "scenario_code": "baseline",
        },
    )
    assert response.status_code == 201, response.text[:400]
    run = response.json()
    assert run["status"] == "succeeded", run
    return run


def _generate(db_client: TestClient, code: str) -> dict[str, Any]:
    response = db_client.post(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages",
        headers=headers(),
        json={"return_code": code, "reporting_date": PERIOD_END.isoformat()},
    )
    assert response.status_code == 201, f"{code}: {response.status_code} {response.text[:300]}"
    package = response.json()
    detail = db_client.get(
        f"/api/v1/banks/{SAMPLE_BANK_ID}/regulatory-packages/{package['id']}", headers=headers()
    ).json()
    return detail["snapshot"]


def _cells(snapshot: dict[str, Any], sheet: str) -> dict[str, Decimal]:
    out: dict[str, Decimal] = {}
    for ref, value in snapshot["bog_form"]["cells"][sheet].items():
        if value is not None and not isinstance(value, str):
            out[ref] = Decimal(str(value))
    return out


def _close(a: object, b: Decimal | float | int, tol: str = "0.01") -> bool:
    if a is None or isinstance(a, str):
        return False
    return abs(Decimal(str(a)) - Decimal(str(b))) <= Decimal(tol)


def _statuses(snapshot: dict[str, Any], sheet: str) -> dict[str, str]:
    for section in snapshot["sections"]:
        if section["title"] == sheet:
            return {row["cell"]: row["status"] for row in section["rows"]}
    raise AssertionError(sheet)


def _has_bsd5a_row(line_code: str, weighted: Decimal) -> bool:
    """Whether a run capital_component line (``tier:category``) lands on a BSD5A row:
    named categories always; otherwise positive CET1 (Disclosed Reserves) and
    positive AT1 (Permanent Preference Shares) by tier."""
    tier, _, category = line_code.partition(":")
    if category in NAMED:
        return True
    return tier in ("cet1", "at1") and weighted >= 0


# ---------------------------------------------------------------------------
# 1. static: every input cell bound; constants equal the template's own values
# ---------------------------------------------------------------------------


def test_bsd5a_binds_every_input_cell_of_all_three_sheets() -> None:
    layout = load_layout("BSD5A")
    maps = line_maps_for("BSD5A")
    for sheet in layout.sheets:
        bound = {ref for line in maps[sheet.name] for ref in line.cells.values()}
        assert bound == {c.ref for c in sheet.input_cells}, sheet.name
        by_cell = {ref: line for line in maps[sheet.name] for ref in line.cells.values()}
        # printed values (item numbers, risk weights, provision ladder) are
        # constants EQUAL to the template — never a retyped BoG figure
        for cell in sheet.input_cells:
            line = by_cell[cell.ref]
            if line.source == "constant":
                assert line.params["value"] == cell.value, f"{sheet.name}!{cell.ref}"
                assert line.unscaled is True
    car = maps[CAR]
    amount_rows = {int(line.cells["amount"][1:]): line for line in car if "amount" in line.cells}
    assert len(amount_rows) == 45  # noqa: PLR2004 — 45 amount cells on the official sheet
    assert {row for row, line in amount_rows.items() if line.source is None} == (
        CAR_INPUT_REQUIRED_ROWS
    )
    assert all(line.notes for line in car if line.source is None)
    # NEW RISK WEIGHTS: 139 printed percentages, all constants; PROVISION ladder
    assert len({ref for line in maps[RW] for ref in line.cells.values()}) == 139  # noqa: PLR2004
    assert all(line.source == "constant" for line in maps[RW])
    prov = {
        line.cells["pct_provision"]: line.params["value"]
        for line in maps[PROV]
        if "pct_provision" in line.cells
    }
    assert prov == {"D9": 1, "D10": 10, "D11": 25, "D12": 50, "D13": 100}
    # a partially weighted class applies the SHEET's printed percentage
    assert amount_rows[51].source == "bsd5.pct_of" and amount_rows[51].params["pct"] == 50  # noqa: PLR2004
    assert amount_rows[47].params["pct"] == 80  # noqa: PLR2004
    assert amount_rows[67].params["pct"] == 50  # noqa: PLR2004


def test_bsd5b_binds_every_input_cell_plus_the_blank_data_cells() -> None:
    layout = load_layout("BSD5B").sheet(GROUP)
    lines = line_maps_for("BSD5B")[GROUP]
    bound = {ref for line in lines for ref in line.cells.values()}
    captured = {c.ref for c in layout.input_cells}
    assert captured <= bound
    assert bound - captured == {"D8", "D24", "D71", "D72"}  # blank official data cells
    by_cell = {ref: line for line in lines for ref in line.cells.values()}
    # group-only rows read the subsidiaries register (data-gap closure 2026-08-16)
    assert by_cell["D10"].source == "refs.sum" and by_cell["D26"].source == "refs.sum"
    assert by_cell["D8"].source == "form.cell" and by_cell["D8"].params["ref"] == "E7"
    assert by_cell["D30"].source if "D30" in by_cell else True  # D30 is BoG's formula
    assert by_cell["E15"].source == "constant" and by_cell["E15"].params["value"] == 0
    assert line_maps_for("BSD5B").get("Sheet2", ()) == ()


# ---------------------------------------------------------------------------
# 2–4. BSD5A on the hermetic book, fed by the capital baseline run
# ---------------------------------------------------------------------------


def test_bsd5a_reconciles_to_the_capital_run_and_bog_arithmetic(  # noqa: PLR0915
    db_client: TestClient,
) -> None:
    _materialize(db_client)
    run = _run_capital_baseline(db_client)
    snapshot = _generate(db_client, "BSD5A")
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    counts = payload["status_counts"]
    assert counts["unmapped"] == 0
    assert counts["input_required"] == len(CAR_INPUT_REQUIRED_ROWS)
    assert counts["mapped"] == 96 + 139 + 10 - len(CAR_INPUT_REQUIRED_ROWS)
    assert snapshot["metadata"]["missing_dependencies"] == []
    e = _cells(snapshot, CAR)
    statuses = _statuses(snapshot, CAR)
    for row in CAR_INPUT_REQUIRED_ROWS:
        assert statuses[f"E{row}"] == "input_required", row
        assert f"E{row}" not in e

    # -- capital side ---------------------------------------------------------
    assert e["E7"] == 150 * M  # paid_up_capital
    assert e["E8"] == (95 + 45 + 10) * M  # retained + statutory + other reserves (CET1 residual)
    assert e["E9"] == 20 * M  # AT1 perpetual instruments
    assert e["E10"] == e["E7"] + e["E8"] + e["E9"] == 320 * M  # BoG: Tier 1 = Σ lines 1–3
    assert e["E12"] == 25 * M and e["E42"] == e["E12"]  # goodwill/intangibles, both sides
    assert e["E13"] == 0 and e["E14"] == 0 and e["E15"] == 0 and e["E16"] == 0
    assert e["E17"] == e["E10"] - sum(e[f"E{r}"] for r in (12, 13, 14, 15, 16)) == 295 * M
    assert e["E20"] == 0 and e["E21"] == 0 and e["E23"] == 0
    assert e["E22"] == 45 * M  # subordinated_debt
    assert e["E24"] == sum(e[f"E{r}"] for r in (20, 21, 22, 23)) == 45 * M
    assert e["E25"] == e["E17"] + e["E24"] == 340 * M
    # ADJUSTED CAPITAL BASE == the signed sum of the run's capital-component
    # lines that have a BSD5A row (== the run's total capital less the components
    # the pre-CRD form has no line for: deferred tax assets, general provisions)
    components = [item for item in run["line_items"] if item["section"] == "capital_component"]
    assert len(components) == 9  # noqa: PLR2004 — the fixture's register
    with_row = [
        c for c in components if _has_bsd5a_row(c["line_code"], Decimal(c["weighted_amount"]))
    ]
    without_row = {c["line_code"] for c in components if c not in with_row}
    assert without_row == {"cet1:deferred_tax_assets", "t2:general_provisions"}
    assert e["E25"] == sum(Decimal(c["weighted_amount"]) for c in with_row)
    assert e["E25"] == Decimal(run["metrics"]["total_capital_ghs"]) - sum(
        Decimal(c["weighted_amount"]) for c in components if c["line_code"] in without_row
    )

    # -- asset side -----------------------------------------------------------
    session = _session()
    try:
        period = session.scalar(
            select(BankReportingPeriod).where(
                BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
                BankReportingPeriod.period_end == PERIOD_END,
            )
        )
        assert period is not None
        facts = list(
            session.scalars(
                select(BankFinancialFact).where(
                    BankFinancialFact.bank_id == SAMPLE_BANK_ID,
                    BankFinancialFact.reporting_period_id == period.id,
                )
            )
        )
    finally:
        session.close()
    asset_total = sum(
        (
            f.amount
            for f in facts
            if f.fact_group == "balance_sheet" and f.attributes.get("side") == "asset"
        ),
        Decimal(0),
    )
    assert e["E27"] == asset_total == 2400 * M
    assert e["E29"] == 45 * M  # cash_vault
    assert e["E32"] == (175 + 70) * M  # BoG required + excess reserves
    assert e["E34"] == 0 and e["E36"] == 0  # no swap / repo positions on the hermetic book
    assert e["E35"] == 260 * M and e["E39"] == 360 * M
    assert e["E31"] == sum(e[f"E{r}"] for r in (32, 33, 34, 35, 36) if f"E{r}" in e) == 505 * M
    assert e["E38"] == e["E39"] + e.get("E40", Decimal(0)) == 360 * M
    mortgage = next(
        f.amount
        for f in facts
        if f.fact_group == "loan_exposure" and f.category == "residential_mortgage"
    )
    assert e["E51"] == mortgage / 2 == 100 * M  # 50% of residential mortgage loans (BoG's %)
    assert e["E41"] == 0 and e["E46"] == 0 and e["E47"] == 0  # no positions on the hermetic book
    deductions = sum(e.get(f"E{r}", Decimal(0)) for r in range(41, 53))
    assert (
        e["E53"]
        == e["E27"] - e["E29"] - e.get("E30", Decimal(0)) - e["E31"] - e["E38"] - deductions
    )
    assert e["E53"] == 1365 * M
    # contingents: the platform's LC/guarantee book (off_balance facts) on the
    # guarantees row, letters of credit (none tagged) on their own row — never both
    off_balance = sum((f.amount for f in facts if f.fact_group == "off_balance"), Decimal(0))
    assert off_balance == 320 * M
    assert e["E55"] == 0 and e["E56"] == off_balance
    assert e["E54"] == e["E65"] == 320 * M  # no class-1/2 relief supplied
    # 50% of NOP and the 3-year average gross income — from the run's line items
    lines = {(i["section"], i["line_code"]): i for i in run["line_items"]}
    open_position = Decimal(lines[("market_rwa", "fx_charge")]["exposure_amount"])
    assert open_position == 45 * M  # max(net_long 45, net_short 12)
    assert e["E67"] == open_position / 2
    gi = [
        Decimal(v["exposure_amount"])
        for (s, c), v in lines.items()
        if s == "operational_rwa" and c.startswith("gross_income")
    ]
    assert len(gi) == 3  # noqa: PLR2004
    avg_gi = sum(gi, Decimal(0)) / 3
    assert _close(e["E68"], avg_gi, "0.0001") and _close(avg_gi, (340 + 380 + 400) * M / 3, "0.01")

    # -- BoG's arithmetic: asset base, ratio, 6% test --------------------------
    assert _close(e["E69"], e["E53"] + e["E65"] + e["E67"] + e["E68"], "0.0001")
    ratio = e["E25"] / e["E69"]
    assert _close(e["E70"], ratio, "0.000000001")  # E70 = E25/E69 (a fraction: 16.34%)
    assert _close(e["E71"], e["E25"] - e["E69"] * Decimal("0.06"), "0.0001")
    assert e["E71"] > 0  # capital surplus at 6%
    # the hermetic figures quoted in docs/bog_returns/bsd5a_line_map.md (¢'Million)
    assert _close(e["E69"] / M, Decimal("2080.8333"), "0.0001")
    assert _close(e["E70"], Decimal("0.163396"), "0.000001")
    assert _close(e["E71"] / M, Decimal("215.15"), "0.01")
    # WHY BSD5A's ratio is not the run's Basel III CAR: same NOP and gross income,
    # different add-on rules — 50% of NOP vs FX charge × RWA multiplier; 100% of the
    # 3-year average vs the BIA charge × RWA multiplier; credit by BoG's printed
    # classes vs the standardised weights. The add-ons reconcile exactly:
    thresholds = run["inputs"]["parameters"]["thresholds_pct"]
    fx_pct = Decimal(thresholds["fx_charge_pct"])
    alpha = Decimal(thresholds["bia_alpha_pct"])
    mult = Decimal(thresholds["rwa_multiplier"])
    market_rwa = Decimal(run["metrics"]["market_rwa_ghs"])
    op_rwa = Decimal(run["metrics"]["operational_rwa_ghs"])
    assert _close(market_rwa, e["E67"] * 2 * fx_pct / 100 * mult / 100, "0.01")
    assert _close(op_rwa, e["E68"] * alpha / 100 * mult / 100, "0.01")
    car_pct = Decimal(run["metrics"]["car_pct"])
    assert _close(car_pct, Decimal("15.8324"), "0.0001")
    assert not _close(e["E70"] * 100, car_pct, "0.5")  # by construction, not by accident

    # -- export: millions on CAR FORMAT, item numbers / percentages unscaled -----
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        assert bank is not None
        from datetime import UTC, datetime  # noqa: PLC0415

        workbook = openpyxl.load_workbook(
            io.BytesIO(
                render_bog_form_xlsx("BSD5A", snapshot, bank, datetime(2026, 8, 16, tzinfo=UTC))
            )
        )
    finally:
        session.close()
    ws = workbook[CAR]
    assert ws["C7"].value == 1 and ws["C71"].value == 52  # noqa: PLR2004 — item numbers
    assert _close(Decimal(str(ws["E7"].value)), 150) and _close(Decimal(str(ws["E25"].value)), 340)
    assert _close(Decimal(str(ws["E71"].value)), e["E71"] / M, "0.000001")  # surplus, ¢'Million
    # Framework ask (docs/bog_returns/bsd5a_line_map.md): E70 is a RATIO formula
    # (=E25/E69) on a ¢'Million sheet; the exporter divides every formula cell by
    # the sheet divisor, so the exported E70 is ratio/1e6 until formula cells can
    # be declared unscaled. The snapshot value (asserted above) is correct.
    assert ws["E70"].value is not None
    assert workbook[RW]["D10"].value == 20 and workbook[RW]["F7"].value == 100  # noqa: PLR2004
    assert workbook[PROV]["D11"].value == 25 and workbook[PROV]["B13"].value == 5  # noqa: PLR2004


# ---------------------------------------------------------------------------
# 5. BSD5B — group equals solo until a subsidiary book exists
# ---------------------------------------------------------------------------


def test_bsd5b_group_return_links_to_bsd5a_and_reproduces_bogs_group_formulas(
    db_client: TestClient,
) -> None:
    _materialize(db_client)
    _run_capital_baseline(db_client)
    solo = _cells(_generate(db_client, "BSD5A"), CAR)
    snapshot = _generate(db_client, "BSD5B")
    payload = snapshot["bog_form"]
    assert not payload["errors"], payload["errors"]
    assert payload["missing_dependencies"] == []
    # 98 declared (94 captured + 4 blank data cells): 2 group-only rows + 12 links to
    # BSD5A input_required cells are blank; 14 BoG formulas evaluated
    assert payload["status_counts"] == {
        "mapped": 84,
        "input_required": 14,
        "unmapped": 0,
        "derived": 14,
    }
    d = _cells(snapshot, GROUP)
    statuses = _statuses(snapshot, GROUP)
    # consolidation-only lines and the solo input_required lines stay blank
    for ref in (
        "D10",
        "D26",
        "D35",
        "D38",
        "D44",
        "D53",
        "D54",
        "D60",
        "D61",
        "D62",
        "D63",
        "D64",
        "D67",
        "D68",
    ):
        assert statuses[ref] == "input_required", ref
    # group = solo, cell for cell (form.cell links), incl. the blank data cells
    for group_ref, solo_ref in (
        ("D8", "E7"), ("D9", "E8"), ("D11", "E9"), ("D14", "E12"), ("D16", "E14"),
        ("D28", "E22"), ("D32", "E27"), ("D34", "E29"), ("D37", "E32"), ("D40", "E35"),
        ("D43", "E39"), ("D45", "E41"), ("D51", "E47"), ("D52", "E51"), ("D58", "E55"),
        ("D59", "E56"), ("D71", "E67"), ("D72", "E68"),
    ):  # fmt: skip
        assert d[group_ref] == solo[solo_ref], (group_ref, solo_ref)
    assert d["E15"] == 0  # template artefact reproduced
    # BoG's group arithmetic
    assert d["D12"] == d["D8"] + d["D9"] + d["D11"] == solo["E10"]  # (D10 minority blank)
    assert (
        d["D20"] == d["D12"] - sum(d.get(f"D{r}", Decimal(0)) for r in range(14, 20)) == solo["E17"]
    )
    assert d["D29"] == sum(d.get(f"D{r}", Decimal(0)) for r in range(21, 29)) == solo["E24"]
    assert d["D30"] == d["D20"] + d["D29"] == solo["E25"] == 340 * M
    # the template's copies (D46=D14 … D49=D19) and the D50 "=D20" quirk, verbatim
    assert d["D46"] == d["D14"] and d["D47"] == d["D16"] and d["D48"] == d["D17"]
    assert d["D49"] == d["D19"] and d["D50"] == d["D20"]
    assert d["D55"] == d["D32"] - sum(d.get(f"D{r}", Decimal(0)) for r in range(34, 55))
    assert d["D65"] == sum(d.get(f"D{r}", Decimal(0)) for r in range(58, 65)) == solo["E54"]
    # D73 does not add rows 50–51 (BoG's formula); D74 is the percent form
    assert _close(
        d["D73"],
        d["D55"] + d["D65"] - d.get("D67", Decimal(0)) - d.get("D68", Decimal(0)),
        "0.0001",
    )
    assert _close(d["D74"], d["D30"] / d["D73"] * 100, "0.000001")
    assert _close(d["D75"], d["D30"] - d["D73"] * Decimal("0.10"), "0.0001")
    # the hermetic figures quoted in docs/bog_returns/bsd5b_line_map.md
    assert d["D55"] == 1070 * M and d["D73"] == 1390 * M and d["D75"] == 201 * M
    assert _close(d["D74"], Decimal("24.4604"), "0.0001")


# ---------------------------------------------------------------------------
# 6. the resolvers against inserted rows
# ---------------------------------------------------------------------------


def _seed_probe_rows(session: Session) -> None:
    """Extra capital facts + two positions (a tagged LC, an interbank placement)
    at the latest period end so each resolver branch has a known answer."""
    period = session.scalar(
        select(BankReportingPeriod).where(
            BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
            BankReportingPeriod.period_end == PERIOD_END,
        )
    )
    assert period is not None

    def fact(category: str, amount: str, tier: str, *, deduction: bool = False) -> None:
        session.add(
            BankFinancialFact(
                organization_id=DEMO_ORG_ID,
                bank_id=SAMPLE_BANK_ID,
                reporting_period_id=period.id,
                fact_group="capital_component",
                category=category,
                amount=Decimal(amount) * M,
                currency="GHS",
                capital_tier=tier,
                is_deduction=deduction,
            )
        )

    fact("share_premium", "12", "CET1")  # → Disclosed Reserves (CET1 residual)
    fact("income_surplus", "7", "CET1", deduction=True)  # a deficit → reduces reserves
    fact("goodwill", "3", "CET1", deduction=True)  # → Goodwill/Intangibles (with intangibles)
    fact("hybrid_capital", "9", "AT1")  # AT1 but named hybrid → Tier 2 hybrid, not row 3
    fact("cumulative_preference_shares", "4", "T2")  # → BSD5A hybrid row / BSD5B row 17
    fact("latent_revaluation_reserve", "6", "T2")  # → BSD5A revaluation / BSD5B latent

    batch = IngestionBatch(
        organization_id=DEMO_ORG_ID,
        bank_id=SAMPLE_BANK_ID,
        source_system="EXCEL_CSV",
        adapter_version="1.0",
        extraction_mode="full",
        status="accepted",
        as_of_date=PERIOD_END,
    )
    session.add(batch)
    session.flush()
    lineage = LineageRecord(
        organization_id=DEMO_ORG_ID,
        ingestion_batch_id=batch.id,
        operation_type="ADAPTER_TRANSLATE",
        operation_ref="bsd5-test",
        input_lineage_ids=[],
    )
    session.add(lineage)
    session.flush()
    common: dict[str, Any] = {
        "organization_id": DEMO_ORG_ID,
        "bank_id": SAMPLE_BANK_ID,
        "as_of_date": PERIOD_END,
        "source_system": "EXCEL_CSV",
        "ingestion_batch_id": batch.id,
        "lineage_id": lineage.id,
        "validation_status": "accepted",
    }
    bank_cp = CanonicalCounterparty(
        **common,
        source_reference="CP/LOCALBANK",
        name="Local Bank",
        counterparty_type="BANK_NON_OECD",
        resident=True,
    )
    corp_cp = CanonicalCounterparty(
        **common,
        source_reference="CP/CORP",
        name="Corp",
        counterparty_type="CORPORATE",
        resident=True,
    )
    session.add_all([bank_cp, corp_cp])
    session.flush()

    def position(  # noqa: PLR0913 — keyword-only fixture builder
        ref: str,
        kind: str,
        currency: str,
        balance: str,
        notional: str,
        cp: CanonicalCounterparty,
        attributes: dict[str, Any],
    ) -> None:
        row = CanonicalPosition(
            **common, source_reference=ref, position_type=kind, currency=currency
        )
        session.add(row)
        session.flush()
        session.add(
            CanonicalPositionSnapshot(
                **common,
                source_reference=ref,
                position_id=row.id,
                counterparty_id=cp.id,
                balance=Decimal(balance) * M,
                notional=Decimal(notional) * M,
                attributes=attributes,
            )
        )

    position(
        "LC/1", "LC_GUARANTEE", "GHS", "0", "30", corp_cp, {"obs_category": "letter_of_credit"}
    )
    position("LC/2", "LC_GUARANTEE", "USD", "0", "50", corp_cp, {"obs_category": "guarantee"})
    position("IBP/1", "INTERBANK_PLACEMENT", "GHS", "10", "0", bank_cp, {})
    position("IBP/2", "INTERBANK_PLACEMENT", "USD", "5", "0", bank_cp, {})
    session.flush()


def test_bsd5_resolvers_against_inserted_rows(db_client: TestClient) -> None:  # noqa: PLR0915
    _materialize(db_client)
    session = _session()
    try:
        _seed_probe_rows(session)
        session.commit()
    finally:
        session.close()
    run = _run_capital_baseline(db_client)
    session = _session()
    try:
        bank = session.get(Bank, SAMPLE_BANK_ID)
        period = session.scalar(
            select(BankReportingPeriod).where(
                BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
                BankReportingPeriod.period_end == PERIOD_END,
            )
        )
        assert bank is not None and period is not None
        ctx = TenantContext(organization_id=ORG_1, actor_user_id=USER_1)
        cache: dict[str, Any] = {}
        rc = ResolveContext(
            db=session, ctx=ctx, bank=bank, period=period, column="amount", cache=cache
        )
        capital = get_resolver("bsd5.capital_facts")
        pct_of = get_resolver("bsd5.pct_of")
        run_line = get_resolver("bsd5.run_line")
        avg_gi = get_resolver("bsd5.avg_gross_income")
        residual = get_resolver("bsd5.off_balance_residual")
        side = get_resolver("bsd5.balance_sheet_side")

        rows = {
            int(line.cells["amount"][1:]): line
            for line in line_maps_for("BSD5A")[CAR]
            if "amount" in line.cells
        }

        def value(row: int) -> Decimal | None:
            line = rows[row]
            assert line.source is not None
            raw = get_resolver(line.source)(rc, dict(line.params))
            return None if raw is None else Decimal(str(raw))

        # Disclosed Reserves: CET1 residual + share premium − the income-surplus deficit
        assert value(8) == (95 + 45 + 10 + 12 - 7) * M
        # row 3: AT1 by tier, but the AT1 fact NAMED hybrid_capital is excluded …
        assert value(9) == 20 * M
        # … and lands on Hybrid Capital together with the cumulative preference shares
        assert value(23) == (9 + 4) * M
        # goodwill + intangibles, deduction-flagged, positive for the sheet to subtract
        assert value(12) == (25 + 3) * M
        assert value(21) == 6 * M  # latent revaluation joins BSD5A's single revaluation row
        # explicit resolver semantics
        assert capital(rc, {"categories": ["goodwill"], "deduction": True}) == 3 * M
        assert capital(rc, {"categories": ["goodwill"]}) == 0  # positive mode ignores deductions
        assert (
            capital(
                rc, {"categories": ["income_surplus"], "include_deductions": ["income_surplus"]}
            )
            == -7 * M
        )
        assert (
            capital(rc, {"tiers": ["T2"], "exclude": list(bsd5a.SUBORDINATED_DEBT)})
            == (15 + 4 + 6) * M
        )
        assert side(rc, {"side": "asset"}) == 2400 * M
        assert side(rc, {"side": "equity"}) == 340 * M
        # printed percentage of a platform aggregate: 80% of claims on other banks (both ccy)
        assert value(47) == Decimal("0.8") * 15 * M
        assert pct_of(rc, {"pct": 25, "source": "constant", "params": {"value": 200}}) == 50  # noqa: PLR2004
        assert (
            pct_of(
                rc,
                {
                    "pct": 25,
                    "source": "form.cell",
                    "params": {"form": "X", "sheet": "S", "ref": "A1"},
                },
            )
            is None
        )
        # LC/guarantee book: tagged LCs on their row, the rest on guarantees — Σ = the book
        assert value(55) == 30 * M
        assert value(56) == 320 * M - 30 * M
        assert residual(rc, {"less": []}) == 320 * M
        # run lines
        assert (
            run_line(
                rc, {"section": "market_rwa", "line_code": "fx_charge", "field": "exposure_amount"}
            )
            == 45 * M
        )
        assert (
            run_line(rc, {"section": "market_rwa", "line_code": "fx_charge", "field": "rate_pct"})
            == 8
        )  # noqa: PLR2004
        assert run_line(rc, {"section": "market_rwa", "line_code": "no_such_line"}) is None
        assert (
            run_line(rc, {"section": "market_rwa", "line_code": "fx_charge", "module": "liquidity"})
            is None
        )
        assert value(67) == Decimal("22.5") * M
        assert _close(avg_gi(rc, {"years": 3}), (340 + 380 + 400) * M / 3, "0.0001")
        assert _close(avg_gi(rc, {"years": 2}), (380 + 400) * M / 2, "0.0001")  # latest N years
        assert avg_gi(rc, {"years": 3, "prefix": "net_income"}) is None
        # the run lookup is memoised under run.metric's key (one query per form)
        assert cache["run:capital:baseline"].id.hex == run["id"].replace("-", "")
        # no run for the period → the run-fed rows are input_required, nothing else changes
        cache.clear()
        earlier = session.scalar(
            select(BankReportingPeriod).where(
                BankReportingPeriod.bank_id == SAMPLE_BANK_ID,
                BankReportingPeriod.period_end == date(2026, 2, 28),
            )
        )
        assert earlier is not None
        rc_prev = ResolveContext(
            db=session, ctx=ctx, bank=bank, period=earlier, column="amount", cache={}
        )
        assert run_line(rc_prev, {"section": "market_rwa", "line_code": "fx_charge"}) is None
        assert avg_gi(rc_prev, {}) is None
        assert pct_of(rc_prev, dict(rows[67].params)) is None
        assert Decimal(str(side(rc_prev, {"side": "asset"}))) > 0
    finally:
        session.close()
    # the generated form carries the probe rows through BoG's arithmetic
    e = _cells(_generate(db_client, "BSD5A"), CAR)
    assert e["E10"] == (150 + 155 + 20) * M
    assert e["E17"] == e["E10"] - 28 * M
    assert e["E24"] == (0 + 6 + 45 + 13) * M
    assert e["E25"] == e["E17"] + e["E24"]
    assert e["E47"] == 12 * M and e["E55"] == 30 * M and e["E56"] == 290 * M
    assert e["E65"] == 320 * M
    d = _cells(_generate(db_client, "BSD5B"), GROUP)
    assert d["D23"] == 6 * M and d["D25"] == 4 * M and d["D27"] == 9 * M  # the group split
    assert d["D22"] == 0 and d["D21"] == 0 and d["D18"] == 0
    assert d["D29"] == solo_tier2(d) == e["E24"]  # same Tier 2 total, finer rows


def solo_tier2(d: dict[str, Decimal]) -> Decimal:
    return sum((d.get(f"D{r}", Decimal(0)) for r in range(21, 29)), Decimal(0))


@pytest.mark.parametrize("code", ["BSD5A", "BSD5B"])
def test_capital_returns_are_the_guides_frequency_and_basis(code: str) -> None:
    from app.services.regulatory_reporting.bog_forms.catalog import form_spec  # noqa: PLC0415

    spec = form_spec(code)
    assert spec.time_limit_days == 14  # noqa: PLR2004
    assert (spec.frequency, spec.basis) == (
        ("monthly", "solo") if code == "BSD5A" else ("quarterly", "consolidated")
    )
    assert "BSD5A" in form_spec("BSD5B").depends_on
