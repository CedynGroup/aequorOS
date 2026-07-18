"""End-to-end Data Engine journeys: workbook in, canonical state and lineage out."""

from __future__ import annotations

import json as jsonlib
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import select
from sqlalchemy import text as sql_text

from app.db.session import get_sessionmaker
from app.domain.ingestion.contracts import EntityMapping, MappingConfig, ReferenceMapping
from app.models import CanonicalReferenceRow
from app.storage.client import StorageLocation
from tests.adapters.excel_csv import fixtures
from tests.api.helpers import ORG_2, headers

AS_OF = str(fixtures.AS_OF)

FULL_MAPPING = MappingConfig(
    field_mappings={
        "gl_account": EntityMapping(
            source_table="GL",
            fields={
                "source_reference": "Code",
                "account_code": "Code",
                "name": "Label",
                "account_class": "Class",
            },
        ),
        "counterparty": EntityMapping(
            source_table="Customers",
            fields={
                "source_reference": "CustomerId",
                "name": "CustomerName",
                "counterparty_type": "Segment",
                "country_code": "Country",
            },
        ),
        "product": EntityMapping(
            source_table="Products",
            fields={
                "source_reference": "ProductCode",
                "product_code": "ProductCode",
                "name": "ProductName",
            },
        ),
        "position": EntityMapping(
            source_table="Loans",
            fields={
                "source_reference": "AccountRef",
                "position_type": "Type",
                "currency": "Ccy",
                "balance": "Outstanding",
                "counterparty_reference": "Customer",
                "product_code": "Product",
                "interest_rate": "Rate",
                "rate_type": "RateKind",
                "contractual_maturity": "Maturity",
            },
        ),
    },
    enum_mappings={
        "counterparty_type": {"RETAIL": "RETAIL_INDIVIDUAL", "CORP": "CORPORATE"},
        "rate_type": {"F": "FIXED", "V": "FLOATING", "FLOAT": "FLOATING"},
    },
    product_mappings={"LN.CORP.5Y": "CORPORATE_LOAN_UNRATED_100RW"},
)

RECON_MAPPING = MappingConfig(
    field_mappings={
        "gl_account": EntityMapping(
            source_table="GL",
            fields={
                "source_reference": "Code",
                "account_code": "Code",
                "name": "Label",
                "account_class": "Class",
                "balance": "Balance",
            },
        ),
        "position": EntityMapping(
            source_table="Loans",
            fields={
                "source_reference": "AccountRef",
                "position_type": "Type",
                "currency": "Ccy",
                "balance": "Outstanding",
                "gl_account_code": "GLAccount",
            },
        ),
    },
)


def seed_bank(client: TestClient) -> str:
    response = client.post("/api/v1/banks/seed-demo", headers=headers())
    assert response.status_code == 200, response.text
    return response.json()["bank_id"]


def activate_mapping(client: TestClient, bank_id: str, mapping: MappingConfig) -> str:
    response = client.post(
        f"/api/v1/banks/{bank_id}/mapping-configs",
        headers=headers(),
        json={
            "source_system": "EXCEL_CSV",
            "name": "Workbook mapping",
            "config": mapping.model_dump(mode="json"),
            "activate": True,
            "reason": "Onboarding mapping for tests.",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]


def start_batch(client: TestClient, bank_id: str, location: Path) -> dict[str, Any]:
    response = client.post(
        f"/api/v1/banks/{bank_id}/ingestion-batches",
        headers=headers(),
        json={
            "source_system": "EXCEL_CSV",
            "as_of_date": AS_OF,
            "location": str(location),
            "reason": "Month-end ingestion.",
        },
    )
    assert response.status_code == 201, response.text
    return response.json()


def create_scoped_mapping(
    client: TestClient,
    bank_id: str,
    *,
    source_ref: str,
    name: str,
    mapping: MappingConfig = FULL_MAPPING,
) -> dict[str, Any]:
    response = client.post(
        f"/api/v1/banks/{bank_id}/mapping-configs",
        headers=headers(),
        json={
            "source_system": "DB_DIRECT",
            "source_ref": source_ref,
            "name": name,
            "config": mapping.model_dump(mode="json"),
            "activate": True,
            "reason": "Per-source onboarding mapping.",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


class TestSourceRefSeparation:
    """Two sources of one source_system at one bank each keep their own mapping.

    A Merchant Bank running an Oracle FLEXCUBE core and a Snowflake warehouse —
    both ``DB_DIRECT`` — must not have one connection's mapping clobber the other's.
    """

    def test_two_db_direct_sources_stay_independently_active(
        self, db_client: TestClient
    ) -> None:
        bank_id = seed_bank(db_client)
        oracle = create_scoped_mapping(
            db_client, bank_id, source_ref="conn-oracle", name="Oracle FLEXCUBE"
        )
        snowflake = create_scoped_mapping(
            db_client, bank_id, source_ref="conn-snowflake", name="Snowflake warehouse"
        )

        # Activating the second source must NOT retire the first: the single-active
        # constraint keys on (bank, source_system, source_ref), so both stay active.
        assert oracle["source_ref"] == "conn-oracle"
        assert oracle["status"] == "active"
        assert snowflake["source_ref"] == "conn-snowflake"
        assert snowflake["status"] == "active"

        listed = db_client.get(
            f"/api/v1/banks/{bank_id}/mapping-configs", headers=headers()
        ).json()["configs"]
        active_refs = {
            config["source_ref"]
            for config in listed
            if config["status"] == "active" and config["source_system"] == "DB_DIRECT"
        }
        assert active_refs == {"conn-oracle", "conn-snowflake"}

    def test_reactivating_one_source_leaves_the_other_untouched(
        self, db_client: TestClient
    ) -> None:
        bank_id = seed_bank(db_client)
        create_scoped_mapping(db_client, bank_id, source_ref="conn-oracle", name="Oracle v1")
        snow_v1 = create_scoped_mapping(
            db_client, bank_id, source_ref="conn-snowflake", name="Snowflake v1"
        )
        snow_v2 = create_scoped_mapping(
            db_client, bank_id, source_ref="conn-snowflake", name="Snowflake v2"
        )

        # Re-activating Snowflake bumps ITS version within ITS scope; Oracle's active
        # mapping is versioned independently and is unaffected.
        assert snow_v2["version"] == snow_v1["version"] + 1
        listed = db_client.get(
            f"/api/v1/banks/{bank_id}/mapping-configs", headers=headers()
        ).json()["configs"]
        active = {
            config["source_ref"]: config
            for config in listed
            if config["status"] == "active" and config["source_system"] == "DB_DIRECT"
        }
        assert set(active) == {"conn-oracle", "conn-snowflake"}
        assert active["conn-snowflake"]["id"] == snow_v2["id"]


class TestHappyPath:
    def test_workbook_to_accepted_canonical_state(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")

        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]
        assert started["reused"] is False
        assert batch["status"] == "accepted"
        assert batch["records_extracted"] == 8
        assert batch["records_translated"] == 8
        assert batch["records_accepted"] == 8
        assert batch["records_error"] == 0
        assert batch["validation_report"]["summary"]["overall_status"] == "ACCEPTED"
        assert batch["content_hash"]

        assert batch["raw_artifact_path"] == f"excel_csv/{AS_OF}/{batch['id']}/bank.xlsx"
        assert (
            batch["report_artifact_path"]
            == f"validation_reports/{AS_OF}/{batch['id']}/validation_report.json"
        )

        positions = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions",
            headers=headers(),
            params={"as_of_date": AS_OF},
        ).json()["positions"]
        assert len(positions) == 2
        by_reference = {position["source_reference"]: position for position in positions}
        assert Decimal(by_reference["LN-0001"]["balance"]) == Decimal("1500000.50")
        assert by_reference["LN-0001"]["rate_type"] == "FIXED"
        assert by_reference["LN-0001"]["validation_status"] == "accepted"

    def test_identical_rerun_is_idempotent(self, db_client: TestClient, tmp_path: Path) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")

        first = start_batch(db_client, bank_id, workbook)
        second = start_batch(db_client, bank_id, workbook)
        assert second["reused"] is True
        assert second["batch"]["id"] == first["batch"]["id"]

        positions = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        assert len(positions) == 2

    def test_restatement_supersedes_prior_snapshots(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")
        first = start_batch(db_client, bank_id, workbook)

        original = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        original_snapshot = {
            position["source_reference"]: position["snapshot_id"] for position in original
        }

        loaded = openpyxl.load_workbook(workbook)
        loaded["Loans"]["D2"] = 1600000.00  # restate LN-0001's balance
        loaded.save(workbook)
        second = start_batch(db_client, bank_id, workbook)
        assert second["reused"] is False
        assert second["batch"]["id"] != first["batch"]["id"]

        restated = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        assert len(restated) == 2
        by_reference = {position["source_reference"]: position for position in restated}
        assert Decimal(by_reference["LN-0001"]["balance"]) == Decimal("1600000")
        assert by_reference["LN-0001"]["snapshot_id"] != original_snapshot["LN-0001"]

    def test_lineage_walks_from_position_to_extraction(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        start_batch(db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx"))

        position = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"][0]
        walk = db_client.get(f"/api/v1/lineage/{position['lineage_id']}", headers=headers()).json()
        operations = [node["operation_type"] for node in walk["nodes"]]
        assert operations == [
            "VALIDATION",
            "ADAPTER_TRANSLATE",
            "ML_ETL_DEDUP",
            "ML_ETL_PREPROCESS",
            "ADAPTER_EXTRACT",
        ]
        assert walk["nodes"][-1]["input_lineage_ids"] == []


class TestGatingAndFailure:
    def test_reconciliation_break_rejects_the_batch(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, RECON_MAPPING)
        workbook = fixtures.build_reconciliation_workbook(
            tmp_path / "recon.xlsx", gl_balance="1500"
        )

        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]
        assert batch["status"] == "rejected"
        assert batch["records_blocked"] == 3
        report = batch["validation_report"]
        assert report["summary"]["overall_status"] == "REJECTED"
        assert report["reconciliation"]["gl_vs_subledger"]["1000"]["within_tolerance"] is False

        positions = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        assert positions == []

    def test_reconciled_workbook_is_accepted(self, db_client: TestClient, tmp_path: Path) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, RECON_MAPPING)
        workbook = fixtures.build_reconciliation_workbook(
            tmp_path / "recon.xlsx", gl_balance="1000"
        )
        started = start_batch(db_client, bank_id, workbook)
        assert started["batch"]["status"] == "accepted"

    def test_missing_file_persists_a_failed_batch(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        started = start_batch(db_client, bank_id, tmp_path / "nowhere.xlsx")
        batch = started["batch"]
        assert batch["status"] == "failed"
        assert batch["error_code"] == "connection_failed"
        assert "does not exist" in batch["error_message"]
        assert batch["raw_artifact_path"] is None
        assert batch["report_artifact_path"] is None

    def test_untranslatable_rows_are_preserved_for_review(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_dirty_cells(tmp_path / "dirty.xlsx")
        # The dirty fixture has only a Loans sheet; other mapped sheets are absent.
        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]

        failures = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-batches/{batch['id']}/translation-failures",
            headers=headers(),
        ).json()["failures"]
        assert len(failures) == 1
        assert failures[0]["raw_record"]["Outstanding"] == "TBC"
        assert "balance" in failures[0]["error_message"]

    def test_ingestion_requires_an_active_mapping_config(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")
        response = db_client.post(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            json={
                "source_system": "EXCEL_CSV",
                "as_of_date": AS_OF,
                "location": str(workbook),
                "reason": "No mapping yet.",
            },
        )
        assert response.status_code == 422
        assert "active mapping config" in response.json()["error"]["message"]


class TestStorageArtifacts:
    def test_raw_source_and_report_land_in_tiered_storage(
        self, db_client: TestClient, tmp_path: Path, storage_engine
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")
        batch = start_batch(db_client, bank_id, workbook)["batch"]

        slugs = {
            key[0].split("-", 2)[2].rsplit("-", 1)[0]
            for key in storage_engine._objects  # noqa: SLF001
        }
        assert len(slugs) == 1
        slug = slugs.pop()

        raw_location = StorageLocation(slug, "raw", batch["raw_artifact_path"])
        descriptor, stream = storage_engine.read(raw_location)
        assert stream.read() == workbook.read_bytes()
        assert descriptor.metadata.ingestion_batch_id == batch["id"]
        assert descriptor.metadata.checksum_sha256 == batch["content_hash"]
        assert descriptor.metadata.lineage_node_id is not None

        report_location = StorageLocation(slug, "outputs", batch["report_artifact_path"])
        _, report_stream = storage_engine.read(report_location)
        report = jsonlib.loads(report_stream.read())
        assert report["summary"]["overall_status"] == "ACCEPTED"

    def test_identical_rerun_stores_no_duplicate_raw_versions(
        self, db_client: TestClient, tmp_path: Path, storage_engine
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")
        start_batch(db_client, bank_id, workbook)
        second = start_batch(db_client, bank_id, workbook)
        assert second["reused"] is True
        raw_objects = [
            key
            for key in storage_engine._objects
            if "-raw" in key[0]  # noqa: SLF001
        ]
        assert len(raw_objects) == 1


class TestUploadFlow:
    def upload(self, client: TestClient, bank_id: str, path: Path) -> dict[str, Any]:
        with path.open("rb") as handle:
            response = client.post(
                f"/api/v1/banks/{bank_id}/ingestion-uploads",
                headers=headers(),
                files={"file": (path.name, handle, "application/octet-stream")},
            )
        assert response.status_code == 201, response.text
        return response.json()

    def test_upload_then_ingest_from_temp_tier(
        self, db_client: TestClient, tmp_path: Path, storage_engine
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "uploaded.xlsx")

        staged = self.upload(db_client, bank_id, workbook)
        assert staged["filename"] == "uploaded.xlsx"
        assert staged["byte_size"] == workbook.stat().st_size
        assert staged["location"].startswith("temp://uploads/")

        started = db_client.post(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            json={
                "source_system": "EXCEL_CSV",
                "as_of_date": AS_OF,
                "location": staged["location"],
                "reason": "Ingest uploaded workbook.",
            },
        )
        assert started.status_code == 201, started.text
        batch = started.json()["batch"]
        assert batch["status"] == "accepted"
        assert batch["records_accepted"] == 8
        assert batch["raw_artifact_path"].endswith("/uploaded.xlsx")
        assert batch["content_hash"] == staged["checksum_sha256"]

    def test_missing_staged_object_fails_the_batch_loudly(self, db_client: TestClient) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        started = db_client.post(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            json={
                "source_system": "EXCEL_CSV",
                "as_of_date": AS_OF,
                "location": "temp://uploads/never-staged/ghost.xlsx",
                "reason": "Ingest a ghost.",
            },
        )
        assert started.status_code == 201
        batch = started.json()["batch"]
        assert batch["status"] == "failed"
        assert batch["error_code"] == "storage_error"

    def test_empty_upload_is_rejected(self, db_client: TestClient, tmp_path: Path) -> None:
        bank_id = seed_bank(db_client)
        empty = tmp_path / "empty.xlsx"
        empty.write_bytes(b"")
        with empty.open("rb") as handle:
            response = db_client.post(
                f"/api/v1/banks/{bank_id}/ingestion-uploads",
                headers=headers(),
                files={"file": (empty.name, handle, "application/octet-stream")},
            )
        assert response.status_code == 422


class TestTenantIsolation:
    def test_other_tenants_see_nothing(self, db_client: TestClient, tmp_path: Path) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        started = start_batch(
            db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx")
        )
        position = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"][0]

        foreign = headers(ORG_2)
        assert (
            db_client.get(f"/api/v1/banks/{bank_id}/ingestion-batches", headers=foreign).status_code
            == 404
        )
        assert (
            db_client.get(
                f"/api/v1/banks/{bank_id}/ingestion-batches/{started['batch']['id']}",
                headers=foreign,
            ).status_code
            == 404
        )
        assert (
            db_client.get(f"/api/v1/lineage/{position['lineage_id']}", headers=foreign).status_code
            == 404
        )


class TestManualOverride:
    def test_override_supersedes_with_provenance_and_lineage(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        start_batch(db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx"))
        position = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"][0]

        response = db_client.post(
            f"/api/v1/banks/{bank_id}/position-snapshots/{position['snapshot_id']}/override",
            headers=headers(),
            json={
                "field": "behavioral_maturity_months",
                "value": 48,
                "reason": "Bank policy: cap NMD duration at 4 years.",
            },
        )
        assert response.status_code == 200, response.text
        overridden = response.json()
        assert overridden["behavioral_maturity_months"] == 48
        assert overridden["superseded_snapshot_id"] == position["snapshot_id"]
        provenance = overridden["enrichment_provenance"]["behavioral_maturity_months"]
        assert provenance["source"] == "MANUAL_OVERRIDE"
        assert provenance["override"]["reason"].startswith("Bank policy")

        walk = db_client.get(
            f"/api/v1/lineage/{overridden['lineage_id']}", headers=headers()
        ).json()
        operations = [node["operation_type"] for node in walk["nodes"]]
        assert operations == [
            "HUMAN_OVERRIDE",
            "VALIDATION",
            "ADAPTER_TRANSLATE",
            "ML_ETL_DEDUP",
            "ML_ETL_PREPROCESS",
            "ADAPTER_EXTRACT",
        ]

        current = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        by_reference = {item["source_reference"]: item for item in current}
        assert by_reference[position["source_reference"]]["snapshot_id"] == overridden["id"]

    def test_superseded_snapshot_cannot_be_overridden(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        start_batch(db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx"))
        position = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"][0]
        url = f"/api/v1/banks/{bank_id}/position-snapshots/{position['snapshot_id']}/override"

        first = db_client.post(
            url,
            headers=headers(),
            json={"field": "ifrs9_stage", "value": 2, "reason": "Credit review outcome."},
        )
        assert first.status_code == 200
        stale = db_client.post(
            url,
            headers=headers(),
            json={"field": "ifrs9_stage", "value": 3, "reason": "Second thoughts."},
        )
        assert stale.status_code == 409


class TestT24ThroughTheApi:
    def test_t24_ingestion_requires_a_staged_bundle(self, db_client: TestClient) -> None:
        # T24 ingests through the stage-then-ingest path: a pull stages a bundle
        # to the temp tier and passes its temp:// location. A raw non-bundle
        # location fails connection honestly rather than crashing.
        bank_id = seed_bank(db_client)
        response = db_client.post(
            f"/api/v1/banks/{bank_id}/mapping-configs",
            headers=headers(),
            json={
                "source_system": "T24",
                "name": "T24 placeholder mapping",
                "config": {},
                "activate": True,
                "reason": "Prepare for onboarding.",
            },
        )
        assert response.status_code == 200
        started = db_client.post(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            json={
                "source_system": "T24",
                "as_of_date": AS_OF,
                "location": "tafj://bank-t24.internal",
                "reason": "Attempt direct T24 ingestion.",
            },
        )
        assert started.status_code == 201
        batch = started.json()["batch"]
        assert batch["status"] == "failed"
        assert batch["error_code"] == "connection_failed"
        assert "bundle" in batch["error_message"].lower()


class TestMappingConfigVersioning:
    def test_versions_increment_and_single_active_is_enforced(self, db_client: TestClient) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        second_id = activate_mapping(db_client, bank_id, FULL_MAPPING)

        configs = db_client.get(
            f"/api/v1/banks/{bank_id}/mapping-configs", headers=headers()
        ).json()["configs"]
        assert [config["version"] for config in configs] == [2, 1]
        assert [config["status"] for config in configs] == ["active", "retired"]
        assert configs[0]["id"] == second_id


MISMATCHED_MAPPING = MappingConfig(
    field_mappings={
        "gl_account": EntityMapping(
            source_table="03_gl_accounts",
            fields={"source_reference": "gl_code", "account_code": "gl_code"},
        ),
        "position": EntityMapping(
            source_table="06_loans",
            fields={"source_reference": "position_id", "balance": "balance_ccy"},
        ),
    },
)

ALIASED_REFERENCE_MAPPING = MappingConfig(
    field_mappings={
        "gl_account": EntityMapping(
            source_table="03_gl_accounts",
            source_table_aliases=["General_Ledger"],
            fields={
                "source_reference": "Code",
                "account_code": "Code",
                "name": "Label",
                "account_class": "Class",
            },
        ),
    },
    reference_mappings={
        "yield_curves": ReferenceMapping(
            source_table="13_yield_curves",
            source_table_aliases=["Yield_Curves"],
            dataset_kind="yield_curve",
        ),
        "fx_rates": ReferenceMapping(
            source_table="14_fx_rates_current",
            source_table_aliases=["FX_Rates_Current"],
            dataset_kind="fx_rates_current",
        ),
    },
)


class TestHonestZeroExtraction:
    def test_mismatched_mapping_rejects_with_found_versus_expected(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, MISMATCHED_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")

        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]
        assert batch["status"] == "rejected"
        assert batch["records_extracted"] == 0
        assert batch["records_accepted"] == 0

        report = batch["validation_report"]
        assert report["summary"]["overall_status"] == "REJECTED"
        blocker = next(f for f in report["failures"] if f["rule"] == "no_tables_matched")
        assert blocker["severity"] == "BLOCKER"
        # The message names what the source actually contains, with row counts...
        assert "GL (2 rows)" in blocker["detail"]
        assert "Loans (2 rows)" in blocker["detail"]
        # ...and what the mapping expected to find.
        assert "03_gl_accounts" in blocker["detail"]
        assert "06_loans" in blocker["detail"]

        warnings = [f for f in report["failures"] if f["rule"] == "table_not_found"]
        assert {w["entity_type"] for w in warnings} == {"gl_account", "position"}

        positions = db_client.get(
            f"/api/v1/banks/{bank_id}/canonical-positions", headers=headers()
        ).json()["positions"]
        assert positions == []

    def test_partial_match_is_accepted_with_table_warnings(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        # The dirty fixture only has a Loans sheet; GL/Customers/Products miss.
        workbook = fixtures.build_dirty_cells(tmp_path / "dirty.xlsx")

        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]
        assert batch["status"] == "accepted_with_warnings"
        assert batch["records_extracted"] > 0

        report = batch["validation_report"]
        warnings = [f for f in report["failures"] if f["rule"] == "table_not_found"]
        assert {w["entity_type"] for w in warnings} == {"gl_account", "counterparty", "product"}
        assert all(f["rule"] != "no_tables_matched" for f in report["failures"])


NEAR_MISS_MAPPING = MappingConfig(
    field_mappings={
        "gl_account": FULL_MAPPING.field_mappings["gl_account"],
        "product": FULL_MAPPING.field_mappings["product"],
        "position": FULL_MAPPING.field_mappings["position"],
        # Typo'd table: the workbook's sheet is "Customers".
        "counterparty": EntityMapping(
            source_table="Customer_Master",
            fields=FULL_MAPPING.field_mappings["counterparty"].fields,
        ),
    },
    enum_mappings=FULL_MAPPING.enum_mappings,
)


class TestTablesBreakdown:
    def test_every_sheet_in_a_multi_tab_workbook_is_listed(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")

        batch = start_batch(db_client, bank_id, workbook)["batch"]
        tables = {entry["source_table"]: entry for entry in batch["validation_report"]["tables"]}
        assert set(tables) == {"GL", "Customers", "Products", "Loans"}
        assert tables["GL"]["resolved_to"] == "gl_account"
        assert tables["Customers"]["resolved_to"] == "counterparty"
        assert tables["Products"]["resolved_to"] == "product"
        assert tables["Loans"]["resolved_to"] == "position"
        for entry in tables.values():
            assert entry["rows_extracted"] == 2
            assert entry["rows_accepted"] == 2
            assert entry["rows_warning"] == 0
            assert entry["rows_error"] == 0
            assert entry["suggestion"] is None

    def test_unmatched_sheet_is_flagged_with_the_near_miss_mapping(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, NEAR_MISS_MAPPING)
        workbook = fixtures.build_well_formed(tmp_path / "bank.xlsx")

        batch = start_batch(db_client, bank_id, workbook)["batch"]
        assert batch["status"] == "accepted_with_warnings"
        tables = {entry["source_table"]: entry for entry in batch["validation_report"]["tables"]}
        assert set(tables) == {"GL", "Customers", "Products", "Loans"}

        unmatched = tables["Customers"]
        assert unmatched["resolved_to"] is None
        assert unmatched["rows_extracted"] == 0
        assert unmatched["rows_accepted"] == 0
        assert "counterparty" in unmatched["suggestion"]
        assert "Customer_Master" in unmatched["suggestion"]
        # The matched tabs still load in full.
        assert tables["Loans"]["resolved_to"] == "position"
        assert tables["Loans"]["rows_extracted"] == 2


class TestReferenceDatasetIngestion:
    def test_aliased_entities_and_reference_rows_ingest_together(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, ALIASED_REFERENCE_MAPPING)
        workbook = fixtures.build_bank_realistic(tmp_path / "bank.xlsx")

        started = start_batch(db_client, bank_id, workbook)
        batch = started["batch"]
        # fx_rates table is absent from this workbook -> warning, not failure.
        assert batch["status"] == "accepted_with_warnings"
        # 2 GL rows via the General_Ledger alias + 2 yield-curve reference rows.
        assert batch["records_extracted"] == 4
        assert batch["records_translated"] == 4
        assert batch["records_accepted"] == 4

        summary = batch["validation_report"]["summary"]
        assert summary["reference_rows"] == {"yield_curve": 2}
        warnings = [
            f for f in batch["validation_report"]["failures"] if f["rule"] == "table_not_found"
        ]
        assert [w["entity_type"] for w in warnings] == ["reference:fx_rates"]

    def test_reference_rows_land_in_the_canonical_reference_table(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, ALIASED_REFERENCE_MAPPING)
        workbook = fixtures.build_bank_realistic(tmp_path / "bank.xlsx")
        batch = start_batch(db_client, bank_id, workbook)["batch"]

        session = get_sessionmaker()()
        try:
            if session.get_bind().dialect.name == "postgresql":
                session.execute(
                    sql_text("SELECT set_config('app.organization_id', :org, true)"),
                    {"org": "11111111-1111-4111-8111-111111111111"},
                )
            rows = session.scalars(
                select(CanonicalReferenceRow).order_by(CanonicalReferenceRow.row_index)
            ).all()
        finally:
            session.close()
        assert [row.row_index for row in rows] == [1, 2]
        assert {row.dataset_kind for row in rows} == {"yield_curve"}
        assert rows[0].payload["curve_name"] == "GHS_SOVEREIGN"
        assert rows[0].payload["quote_date"] == "2026-06-01"
        assert str(rows[0].ingestion_batch_id) == batch["id"]
        assert rows[0].source_reference.startswith("bank.xlsx#Yield_Curves!R")


class TestSourceFilterAndSummary:
    def test_batch_history_filters_by_source_system(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        start_batch(db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx"))

        unfiltered = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-batches", headers=headers()
        ).json()["batches"]
        assert len(unfiltered) == 1

        excel = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            params={"source_system": "EXCEL_CSV"},
        ).json()["batches"]
        assert [batch["source_system"] for batch in excel] == ["EXCEL_CSV"]

        pushed = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            params={"source_system": "API_PUSH"},
        ).json()["batches"]
        assert pushed == []

        unknown = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-batches",
            headers=headers(),
            params={"source_system": "NOT_A_SOURCE"},
        )
        assert unknown.status_code == 422

    def test_summary_rolls_up_sources_and_canonical_counts(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        started = start_batch(
            db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx")
        )
        batch = started["batch"]

        summary = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-summary", headers=headers()
        ).json()
        assert summary["bank_id"] == bank_id

        assert len(summary["sources"]) == 1
        source = summary["sources"][0]
        assert source["source_system"] == "EXCEL_CSV"
        assert source["batches"] == 1
        assert source["last_status"] == batch["status"]
        assert source["last_batch_at"] is not None
        assert source["records_accepted_total"] == batch["records_accepted"]
        assert source["records_warning_total"] == batch["records_warning"]

        counts = summary["canonical_counts"]
        # The well-formed workbook: 2 loans, 2 customers, 2 products, 2 GL rows.
        assert counts["positions"] == 2
        assert counts["position_snapshots"] == 2
        assert counts["counterparties"] == 2
        assert counts["gl_accounts"] == 2
        assert counts["products"] == 2
        assert counts["reference_rows"] == 0

        assert summary["activations_count"] == 0
        assert summary["last_activation_at"] is None

    def test_summary_counts_only_current_generation_and_latest_reference_batch(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, ALIASED_REFERENCE_MAPPING)
        workbook = fixtures.build_bank_realistic(tmp_path / "bank.xlsx")
        start_batch(db_client, bank_id, workbook)

        # Restate the same workbook (one changed cell) -> a second batch whose
        # reference rows replace, not add to, the first batch's in the summary.
        loaded = openpyxl.load_workbook(workbook)
        sheet = loaded["General_Ledger"]
        sheet.cell(row=2, column=2, value="Cash and balances (restated)")
        loaded.save(workbook)
        second = start_batch(db_client, bank_id, workbook)
        assert second["reused"] is False

        summary = db_client.get(
            f"/api/v1/banks/{bank_id}/ingestion-summary", headers=headers()
        ).json()
        source = summary["sources"][0]
        assert source["batches"] == 2
        # Superseded GL generations are excluded; yield_curve rows count once.
        assert summary["canonical_counts"]["gl_accounts"] == 2
        assert summary["canonical_counts"]["reference_rows"] == 2

    def test_summary_and_filter_are_tenant_scoped(
        self, db_client: TestClient, tmp_path: Path
    ) -> None:
        bank_id = seed_bank(db_client)
        activate_mapping(db_client, bank_id, FULL_MAPPING)
        start_batch(db_client, bank_id, fixtures.build_well_formed(tmp_path / "bank.xlsx"))

        foreign = headers(ORG_2)
        assert (
            db_client.get(f"/api/v1/banks/{bank_id}/ingestion-summary", headers=foreign).status_code
            == 404
        )
        assert (
            db_client.get(
                f"/api/v1/banks/{bank_id}/ingestion-batches",
                headers=foreign,
                params={"source_system": "EXCEL_CSV"},
            ).status_code
            == 404
        )
