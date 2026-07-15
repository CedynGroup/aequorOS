"""Synthetic workbook fixtures covering the ways bank spreadsheets go wrong.

Fixtures are generated, not committed as binaries, so each pattern is
readable, reviewable, and cheap to extend.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.adapters.excel_csv.type_coercion import excel_serial_for

AS_OF = date(2026, 6, 30)


def _active_sheet(workbook: Workbook) -> Worksheet:
    sheet = workbook.active
    assert sheet is not None  # a fresh Workbook always has an active sheet
    return sheet


def build_well_formed(path: Path) -> Path:
    workbook = Workbook()

    gl = _active_sheet(workbook)
    gl.title = "GL"
    gl.append(["Code", "Label", "Class"])
    gl.append(["1000", "Cash and balances", "ASSET"])
    gl.append(["2000", "Customer deposits", "LIABILITY"])

    customers = workbook.create_sheet("Customers")
    customers.append(["CustomerId", "CustomerName", "Segment", "Country"])
    customers.append(["C-001", "Kojo Mensah", "RETAIL", "GH"])
    customers.append(["C-002", "Volta Agro Ltd", "CORP", "GH"])

    products = workbook.create_sheet("Products")
    products.append(["ProductCode", "ProductName"])
    products.append(["LN.CORP.5Y", "5y corporate loan"])
    products.append(["DP.CURRENT", "Current account"])

    loans = workbook.create_sheet("Loans")
    loans.append(
        [
            "AccountRef",
            "Type",
            "Ccy",
            "Outstanding",
            "Customer",
            "Product",
            "Rate",
            "RateKind",
            "Maturity",
        ]
    )
    loans.append(
        ["LN-0001", "LOAN", "GHS", 1500000.50, "C-002", "LN.CORP.5Y", 0.245, "F", date(2031, 3, 15)]
    )
    loans.append(
        ["LN-0002", "LOAN", "GHS", 250000, "C-001", "LN.CORP.5Y", 0.31, "V", date(2028, 9, 1)]
    )

    workbook.save(path)
    return path


def build_merged_headers(path: Path) -> Path:
    """A title banner merged across the top, blank spacer, then the table."""
    workbook = Workbook()
    sheet = _active_sheet(workbook)
    sheet.title = "Loans"
    sheet["A1"] = "Sample Bank Limited — Loan Book as of 30 June 2026"
    sheet.merge_cells("A1:E1")
    # Row 2 intentionally blank.
    sheet.append([])
    sheet.append(["AccountRef", "Type", "Ccy", "Outstanding", "Rate"])
    sheet.append(["LN-0001", "LOAN", "GHS", 1500000.50, 0.245])
    sheet.append(["LN-0002", "LOAN", "USD", 74000, 0.081])
    workbook.save(path)
    return path


def build_multiple_tables_per_sheet(path: Path) -> Path:
    """Two unrelated tables stacked in one sheet, separated by blank rows."""
    workbook = Workbook()
    sheet = _active_sheet(workbook)
    sheet.title = "Data"
    sheet.append(["ProductCode", "ProductName"])
    sheet.append(["LN.CORP.5Y", "5y corporate loan"])
    sheet.append([])
    sheet.append([])
    sheet.append(["AccountRef", "Type", "Ccy", "Outstanding"])
    sheet.append(["LN-0009", "LOAN", "GHS", 90000])
    workbook.save(path)
    return path


def build_dirty_cells(path: Path) -> Path:
    """Currency noise, percent-typed rates, serial dates, and TBC balances."""
    workbook = Workbook()
    sheet = _active_sheet(workbook)
    sheet.title = "Loans"
    sheet.append(["AccountRef", "Type", "Ccy", "Outstanding", "Rate", "RateKind", "Maturity"])
    sheet.append(
        [
            "LN-0001",
            "LOAN",
            "GHS",
            "GHS 1,500,000.50",
            "24.5%",
            "FIXED",
            excel_serial_for(date(2031, 3, 15)),
        ]
    )
    sheet.append(["LN-0002", "LOAN", "GHS", "(2,500.00)", 24.5, "F", "15/03/2031"])
    sheet.append(["LN-0003", "LOAN", "GHS", 74000, 0.245, "FLOAT", "2031-03-15"])
    sheet.append(["LN-0004", "LOAN", "GHS", "TBC", "N/A", "F", "-"])
    workbook.save(path)
    return path


def build_reconciliation_workbook(path: Path, *, gl_balance: str) -> Path:
    """GL control account plus two loans; the GL balance decides reconciliation."""
    workbook = Workbook()
    gl = _active_sheet(workbook)
    gl.title = "GL"
    gl.append(["Code", "Label", "Class", "Balance"])
    gl.append(["1000", "Loans control", "ASSET", gl_balance])

    loans = workbook.create_sheet("Loans")
    loans.append(["AccountRef", "Type", "Ccy", "Outstanding", "GLAccount"])
    loans.append(["LN-0001", "LOAN", "GHS", 600, "1000"])
    loans.append(["LN-0002", "LOAN", "GHS", 400, "1000"])
    workbook.save(path)
    return path


def build_positions_csv(path: Path) -> Path:
    path.write_text(
        "AccountRef,Type,Ccy,Outstanding,Rate,RateKind\n"
        'LN-0001,LOAN,GHS,"1,500,000.50",24.5%,F\n'
        "LN-0002,DEPOSIT,GHS,250000,0.05,V\n",
        encoding="utf-8",
    )
    return path


def build_bank_realistic(path: Path) -> Path:
    """Sheets named the way a bank Treasury team names them, plus a reference
    table — exercises alias resolution and reference-dataset extraction."""
    workbook = Workbook()
    gl = _active_sheet(workbook)
    gl.title = "General_Ledger"
    gl.append(["Code", "Label", "Class"])
    gl.append(["1000", "Cash and balances", "ASSET"])
    gl.append(["2000", "Customer deposits", "LIABILITY"])

    curves = workbook.create_sheet("Yield_Curves")
    curves.append(["curve_name", "tenor_months", "rate", "quote_date"])
    curves.append(["GHS_SOVEREIGN", 3, 0.158, date(2026, 6, 1)])
    curves.append(["GHS_SOVEREIGN", 12, 0.181, date(2026, 6, 1)])

    workbook.save(path)
    return path


def build_hedge_and_swap_book(path: Path) -> Path:
    """The treasury FX hedge book and IRS blotter, headers as in data/18+19.

    Exercises the starter template's carriage design: canonical fields from
    the shared position mapping (trade/maturity date fallbacks, notional_ccy
    balance fallback) and instrument specifics via attribute columns.
    """
    workbook = Workbook()
    hedges = _active_sheet(workbook)
    hedges.title = "FX_Hedges"
    hedges.append(
        [
            "position_id",
            "hedge_id",
            "position_type",
            "instrument",
            "currency_pair",
            "buy_currency",
            "sell_currency",
            "currency",
            "notional_ccy",
            "contract_rate",
            "trade_date",
            "maturity_date",
            "mtm_ghs",
            "prospective_r2",
            "dollar_offset_ratio",
        ]
    )
    hedges.append(
        [
            "SBL-FXH-000001",
            "FXH-USD-001",
            "FX_HEDGE",
            "FORWARD",
            "USD/GHS",
            "GHS",
            "USD",
            "USD",
            3000000,
            13.05,
            date(2026, 1, 30),
            date(2026, 7, 29),
            510000,
            0.94,
            1.02,
        ]
    )

    swaps = workbook.create_sheet("Interest_Rate_Swaps")
    swaps.append(
        [
            "position_id",
            "swap_id",
            "position_type",
            "direction",
            "currency",
            "notional_ghs",
            "notional_ccy",
            "interest_rate",
            "pay_rate_pct",
            "receive_index",
            "trade_date",
            "maturity_date",
            "tenor_years",
        ]
    )
    swaps.append(
        [
            "SBL-IRS-000001",
            "IRS-2026-001",
            "INTEREST_RATE_SWAP",
            "PAY_FIXED",
            "GHS",
            60000000,
            60000000,
            0.2475,
            24.75,
            "91D_TBILL",
            date(2026, 4, 28),
            date(2028, 4, 28),
            2.0,
        ]
    )
    workbook.save(path)
    return path


def build_position_variants(path: Path) -> Path:
    """Two position sheets with diverging headers (loan book + OBS register),
    exercising multi-table extraction, fallback columns, and attribute columns."""
    workbook = Workbook()
    loans = _active_sheet(workbook)
    loans.title = "Loans"
    loans.append(["AccountRef", "Type", "Ccy", "Outstanding", "Rate", "RateKind"])
    loans.append(["LN-0001", "LOAN", "GHS", 1000, 0.2, "F"])

    obs = workbook.create_sheet("LC_and_Guarantees")
    obs.append(["AccountRef", "Type", "Ccy", "NotionalCcy", "CCF"])
    obs.append(["OBS-0001", "LC", "USD", 500, 0.2])

    workbook.save(path)
    return path
