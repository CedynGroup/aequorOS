"""Builders and canonical-count helpers for manual upload adapter tests.

``build_full_coverage_workbook`` produces one multi-sheet workbook covering
every scope the adapter advertises, so the contract suite's "every listed
scope is pullable" tests have a single staged fixture file.
"""

from __future__ import annotations

import hashlib
import io
from datetime import UTC, date, datetime
from typing import Any
from uuid import UUID, uuid4

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.adapters.market_data.base import CredentialSet
from app.adapters.market_data.manual_upload.templates import TEMPLATE_HEADERS
from app.adapters.market_data.scope_taxonomy import DataScope, ScopeCategory, category_of
from app.models import (
    Bank,
    CanonicalCounterpartyRating,
    CanonicalFxRate,
    CanonicalMarketIndex,
    CanonicalYieldCurve,
    CanonicalYieldCurvePoint,
)
from app.storage.client import ObjectMetadata, StorageClient, StorageLocation

FIXTURE_AS_OF = date(2026, 6, 30)

# Percent yields per currency and tenor — plausible mid-2026 sovereign levels.
CURVE_RATES: dict[str, dict[int, float]] = {
    "GHS": {1: 15.20, 3: 15.80, 6: 16.40},
    "USD": {1: 4.30, 3: 4.45, 6: 4.60},
    "EUR": {1: 2.10, 3: 2.25, 6: 2.40},
    "GBP": {1: 3.80, 3: 3.95, 6: 4.05},
    "NGN": {1: 21.50, 3: 22.10, 6: 22.80},
    "KES": {1: 12.40, 3: 12.90, 6: 13.30},
    "ZAR": {1: 7.60, 3: 7.85, 6: 8.10},
}

# Spot levels: quote currency per 1 unit of base currency.
SPOT_RATES: dict[tuple[str, str], float] = {
    ("USD", "GHS"): 12.85,
    ("EUR", "GHS"): 13.90,
    ("GBP", "GHS"): 16.20,
    ("USD", "NGN"): 1450.00,
}

# USD/GHS forward points per tenor.
FORWARD_RATES: dict[int, float] = {1: 12.95, 3: 13.10, 6: 13.40, 12: 13.95}

RATING_ROWS: tuple[tuple[str, str, str, str, date], ...] = (
    ("GHANA_SOVEREIGN", "fitch", "B-", "stable", date(2026, 6, 15)),
    ("NIGERIA_SOVEREIGN", "moodys", "Caa1", "positive", date(2026, 6, 10)),
)

MACRO_ROWS: tuple[tuple[str, float, str, int], ...] = (
    ("GHANA_GDP_FORECAST", 5.80, "base", 12),
    ("GHANA_INFLATION_FORECAST", 18.20, "base", 12),
    ("GHANA_POLICY_RATE_PATH", 26.00, "base", 6),
)


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_full_coverage_workbook(as_of: date = FIXTURE_AS_OF) -> bytes:
    """A four-sheet workbook with rows for every non-security-master scope."""
    workbook = Workbook()
    curves = workbook.active
    assert curves is not None
    curves.title = "yield_curve"
    curves.append(list(TEMPLATE_HEADERS["yield_curve"]))
    for currency, tenors in CURVE_RATES.items():
        for tenor, percent in tenors.items():
            curves.append([currency, f"{currency}_GOV_BOND", as_of.isoformat(), tenor, percent])

    fx = workbook.create_sheet("fx_rates")
    fx.append(list(TEMPLATE_HEADERS["fx_rates"]))
    for (base, quote), rate in SPOT_RATES.items():
        fx.append([base, quote, "spot", None, rate, as_of.isoformat()])
    for tenor, rate in FORWARD_RATES.items():
        fx.append(["USD", "GHS", "forward", tenor, rate, as_of.isoformat()])

    ratings = workbook.create_sheet("credit_ratings")
    ratings.append(list(TEMPLATE_HEADERS["credit_ratings"]))
    for issuer, agency, rating, watch, rating_date in RATING_ROWS:
        ratings.append([issuer, agency, rating, watch, rating_date.isoformat()])

    macro = workbook.create_sheet("macro_forecasts")
    macro.append(list(TEMPLATE_HEADERS["macro_forecasts"]))
    for index_code, value, scenario, horizon in MACRO_ROWS:
        macro.append([index_code, value, scenario, horizon, as_of.isoformat()])

    return workbook_bytes(workbook)


def build_yield_curve_workbook(rows: list[list[Any]], *, sheet_title: str = "yield_curve") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_title
    sheet.append(list(TEMPLATE_HEADERS["yield_curve"]))
    for row in rows:
        sheet.append(row)
    return workbook_bytes(workbook)


def stage_upload(storage: StorageClient, slug: str, content: bytes, filename: str) -> str:
    """Write bytes to the bank's temp tier and return the temp:// handle."""
    storage.ensure_institution(slug)
    object_path = f"uploads/{uuid4().hex}/{filename}"
    storage.write(
        StorageLocation(institution_slug=slug, tier="temp", object_path=object_path),
        io.BytesIO(content),
        ObjectMetadata(
            institution_slug=slug,
            tier="temp",
            checksum_sha256=hashlib.sha256(content).hexdigest(),
            written_at=datetime.now(UTC),
            written_by="manual-upload-tests",
        ),
    )
    return f"temp://{object_path}"


def credentials_for(bank: Bank, staged_location: str) -> CredentialSet:
    return CredentialSet(
        institution_id=str(bank.id),
        vendor="manual_upload",
        credentials={"staged_location": staged_location},
        issued_at=datetime.now(UTC),
        expires_at=None,
    )


_CANONICAL_MODELS = (
    CanonicalYieldCurve,
    CanonicalYieldCurvePoint,
    CanonicalFxRate,
    CanonicalMarketIndex,
    CanonicalCounterpartyRating,
)


def produced_batch_records(db: Session, batch_id: UUID) -> list[Any]:
    """Every canonical market data row a batch inserted, across entity types."""
    records: list[Any] = []
    for model in _CANONICAL_MODELS:
        records.extend(db.scalars(select(model).where(model.ingestion_batch_id == batch_id)))
    return records


def count_current_canonical(db: Session, bank: Bank, scope: DataScope, as_of: date) -> int:
    """Current-generation canonical rows for one scope and business date.

    Curve points carry no supersession chain of their own; their current
    generation is reached through non-superseded curve headers.
    """
    category = category_of(scope)
    if category is ScopeCategory.YIELD_CURVE:
        currency = scope.value.removeprefix("YIELD_CURVE_")
        curve_ids = list(
            db.scalars(
                select(CanonicalYieldCurve.id).where(
                    CanonicalYieldCurve.organization_id == bank.organization_id,
                    CanonicalYieldCurve.bank_id == bank.id,
                    CanonicalYieldCurve.as_of_date == as_of,
                    CanonicalYieldCurve.currency == currency,
                    CanonicalYieldCurve.superseded_by.is_(None),
                )
            )
        )
        if not curve_ids:
            return 0
        points = db.scalar(
            select(func.count())
            .select_from(CanonicalYieldCurvePoint)
            .where(CanonicalYieldCurvePoint.yield_curve_id.in_(curve_ids))
        )
        return len(curve_ids) + int(points or 0)
    if category in (ScopeCategory.FX_SPOT, ScopeCategory.FX_FORWARD):
        parts = scope.value.split("_")
        if category is ScopeCategory.FX_SPOT:
            base, quote, tenor = parts[2], parts[3], None
        else:
            base, quote, tenor = parts[2], parts[3], int(parts[4].removesuffix("M"))
        count = db.scalar(
            select(func.count())
            .select_from(CanonicalFxRate)
            .where(
                CanonicalFxRate.organization_id == bank.organization_id,
                CanonicalFxRate.bank_id == bank.id,
                CanonicalFxRate.as_of_date == as_of,
                CanonicalFxRate.base_currency == base,
                CanonicalFxRate.quote_currency == quote,
                CanonicalFxRate.tenor_months.is_(None)
                if tenor is None
                else CanonicalFxRate.tenor_months == tenor,
                CanonicalFxRate.superseded_by.is_(None),
            )
        )
        return int(count or 0)
    if category is ScopeCategory.CREDIT_RATING:
        issuer = scope.value.removeprefix("CREDIT_RATING_")
        count = db.scalar(
            select(func.count())
            .select_from(CanonicalCounterpartyRating)
            .where(
                CanonicalCounterpartyRating.organization_id == bank.organization_id,
                CanonicalCounterpartyRating.bank_id == bank.id,
                CanonicalCounterpartyRating.as_of_date == as_of,
                CanonicalCounterpartyRating.issuer == issuer,
                CanonicalCounterpartyRating.superseded_by.is_(None),
            )
        )
        return int(count or 0)
    if category is ScopeCategory.MACRO_FORECAST:
        index_code = scope.value.removeprefix("MACRO_")
        count = db.scalar(
            select(func.count())
            .select_from(CanonicalMarketIndex)
            .where(
                CanonicalMarketIndex.organization_id == bank.organization_id,
                CanonicalMarketIndex.bank_id == bank.id,
                CanonicalMarketIndex.as_of_date == as_of,
                CanonicalMarketIndex.index_code == index_code,
                CanonicalMarketIndex.superseded_by.is_(None),
            )
        )
        return int(count or 0)
    msg = f"count_current_canonical has no handler for {scope.value}"
    raise AssertionError(msg)
