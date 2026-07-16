"""Template generation (§8.2): every kind builds, headers match the parser,
and a filled template round-trips through the parser."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
import pytest

from app.adapters.market_data.manual_upload.parser import parse_upload
from app.adapters.market_data.manual_upload.templates import (
    COMMENT_PREFIX,
    TEMPLATE_HEADERS,
    TEMPLATE_KINDS,
    TemplateKind,
    build_template,
    template_filename,
)
from app.adapters.market_data.scope_taxonomy import DataScope

AS_OF = "2026-06-30"


def _load_grid(content: bytes) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    return [list(row) for row in sheet.iter_rows(values_only=True)]


@pytest.mark.parametrize("kind", TEMPLATE_KINDS)
def test_every_kind_builds_with_exact_headers(kind: TemplateKind) -> None:
    grid = _load_grid(build_template(kind))
    assert tuple(grid[0]) == TEMPLATE_HEADERS[kind]


@pytest.mark.parametrize("kind", TEMPLATE_KINDS)
def test_example_row_and_legend_are_comment_marked(kind: TemplateKind) -> None:
    grid = _load_grid(build_template(kind))
    example = grid[1]
    assert isinstance(example[0], str) and example[0].startswith(COMMENT_PREFIX)
    legend_cells = [
        row[0]
        for row in grid[2:]
        if row and isinstance(row[0], str) and row[0].startswith(COMMENT_PREFIX)
    ]
    assert legend_cells, "template must carry a legend block"
    legend_text = " ".join(legend_cells)
    for column in TEMPLATE_HEADERS[kind]:
        assert column in legend_text, f"legend must explain column {column!r}"
    # Data-validation hints live in the legend.
    assert "ISO 8601" in legend_text
    assert "ISO 4217" in legend_text


@pytest.mark.parametrize("kind", TEMPLATE_KINDS)
def test_pristine_template_parses_to_nothing(kind: TemplateKind) -> None:
    parsed = parse_upload(build_template(kind), template_filename(kind))
    assert parsed.scopes == {}
    assert parsed.problems == []


def _fill(content: bytes, rows: list[list[Any]]) -> bytes:
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_yield_curve_round_trip() -> None:
    filled = _fill(
        build_template("yield_curve"),
        [
            ["GHS", "GHS_GOV_BOND", AS_OF, 3, 15.80],
            ["GHS", "GHS_GOV_BOND", AS_OF, 6, 16.40],
        ],
    )
    parsed = parse_upload(filled, "yield_curve_template.xlsx", expected_as_of=date(2026, 6, 30))
    assert parsed.problems == []
    assert set(parsed.scopes) == {DataScope.YIELD_CURVE_GHS}
    bundle = parsed.scopes[DataScope.YIELD_CURVE_GHS].bundle
    assert len(bundle.curves) == 1
    curve = bundle.curves[0]
    assert curve.curve_name == "GHS_GOV_BOND"
    assert [(point.tenor_months, point.rate) for point in curve.points] == [
        (3, Decimal("0.158")),
        (6, Decimal("0.164")),
    ]
    assert bundle.sample_values["GHS 3M"] == "15.80%"


def test_fx_rates_round_trip() -> None:
    filled = _fill(
        build_template("fx_rates"),
        [
            ["USD", "GHS", "spot", None, 12.85, AS_OF],
            ["USD", "GHS", "forward", 3, 13.10, AS_OF],
        ],
    )
    parsed = parse_upload(filled, "fx_rates_template.xlsx")
    assert parsed.problems == []
    assert set(parsed.scopes) == {
        DataScope.FX_SPOT_USD_GHS,
        DataScope.FX_FORWARD_USD_GHS_3M,
    }
    spot = parsed.scopes[DataScope.FX_SPOT_USD_GHS].bundle.fx_rates[0]
    assert (spot.rate, spot.tenor_months) == (Decimal("12.85"), None)
    forward = parsed.scopes[DataScope.FX_FORWARD_USD_GHS_3M].bundle.fx_rates[0]
    assert (forward.rate, forward.tenor_months) == (Decimal("13.1"), 3)


def test_credit_ratings_round_trip() -> None:
    filled = _fill(
        build_template("credit_ratings"),
        [
            ["GHANA_SOVEREIGN", "fitch", "B-", "stable", "2026-06-15"],
            ["NIGERIA_SOVEREIGN", "moodys", "Caa1", None, "2026-06-10"],
        ],
    )
    parsed = parse_upload(filled, "credit_ratings_template.xlsx")
    assert parsed.problems == []
    assert set(parsed.scopes) == {
        DataScope.CREDIT_RATING_GHANA_SOVEREIGN,
        DataScope.CREDIT_RATING_NIGERIA_SOVEREIGN,
    }
    ghana = parsed.scopes[DataScope.CREDIT_RATING_GHANA_SOVEREIGN].bundle.ratings[0]
    assert (ghana.rating, ghana.watch_status, ghana.rating_date) == (
        "B-",
        "stable",
        date(2026, 6, 15),
    )
    nigeria = parsed.scopes[DataScope.CREDIT_RATING_NIGERIA_SOVEREIGN].bundle.ratings[0]
    assert nigeria.watch_status is None


def test_macro_forecasts_round_trip() -> None:
    filled = _fill(
        build_template("macro_forecasts"),
        [
            ["GHANA_GDP_FORECAST", 5.80, "base", 12, AS_OF],
            ["GHANA_INFLATION_FORECAST", 18.20, "adverse", 12, AS_OF],
        ],
    )
    parsed = parse_upload(filled, "macro_forecasts_template.xlsx")
    assert parsed.problems == []
    assert set(parsed.scopes) == {
        DataScope.MACRO_GHANA_GDP_FORECAST,
        DataScope.MACRO_GHANA_INFLATION_FORECAST,
    }
    gdp = parsed.scopes[DataScope.MACRO_GHANA_GDP_FORECAST].bundle.indices[0]
    assert (gdp.value, gdp.scenario, gdp.horizon_months) == (Decimal("5.8"), "base", 12)
